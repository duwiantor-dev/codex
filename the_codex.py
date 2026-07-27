import io
import os
import re
import time
import zipfile
import importlib.util
import html as html_lib
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
from urllib.parse import parse_qs, quote_plus, unquote, urljoin, urlparse

import pandas as pd
try:
    import requests
except ModuleNotFoundError:
    requests = None

import streamlit as st

try:
    from bs4 import BeautifulSoup
except ModuleNotFoundError:
    BeautifulSoup = None
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
try:
    from rapidfuzz import fuzz
except ModuleNotFoundError:
    fuzz = None


# ============================================================
# APP CONFIG
# ============================================================
APP_TITLE = "Codexid version.9"
MAX_MASS_FILES = 50
MAX_TOTAL_UPLOAD_MB = 200
BIGSELLER_MAX_ROWS_PER_FILE = 10000
DEFAULT_TONGLE_GUDANGS = ["JKT-1A", "JKT-3B", "JKT-3C", "JKT-4B"]
STOCK_PRICELIST_SHEETS = ["LAPTOP", "TELCO", "PC HOM ELE", "SOF COM SUP", "ACC"]
ORACLE_API_BASE_DEFAULT = os.environ.get("ORACLE_API_BASE", "http://127.0.0.1:8000")

st.set_page_config(page_title=APP_TITLE, layout="wide")

# Global style: blue loading bar
st.markdown("""
<style>
    div.stProgress > div > div > div > div {
        background-color: #2563eb;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# SESSION STATE
# ============================================================
SESSION_DEFAULTS = {
    "download_cache": {},
    "summary_cache": {},
    "issue_preview_cache": {},
    "stock_shopee_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_tiktokshop_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_mwh_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_bigseller_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_blibli_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "stock_akulaku_areas_loaded": {"area_options": [], "gudang_options": [], "default_gudang_options": []},
    "issue_output_mode": "Info saja",
}
for _k, _v in SESSION_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ============================================================
# GENERIC HELPERS
# ============================================================
def s(x) -> str:
    if x is None:
        return ""
    return str(x).strip()


def s_clean(x) -> str:
    if x is None:
        return ""
    txt = str(x).replace("\xa0", " ")
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


def su(x) -> str:
    return s_clean(x).upper()


def norm_sku(v) -> str:
    txt = su(v)
    if not txt:
        return ""
    if re.fullmatch(r"\d+\.0", txt):
        txt = txt[:-2]
    txt = re.sub(r"\s+", "", txt)
    return txt


def split_sku_addons(full_sku: str) -> Tuple[str, List[str]]:
    parts = [p.strip() for p in s_clean(full_sku).split("+") if p and s_clean(p)]
    if not parts:
        return "", []
    return parts[0], parts[1:]


def normalize_addon_code(x) -> str:
    return su(x)


def parse_number_like_id(x) -> str:
    if x is None:
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if pd.isna(x):
            return ""
        if x.is_integer():
            return str(int(x))
        return str(x)
    return s_clean(x)


def to_int_or_none(v) -> Optional[int]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if pd.isna(v):
            return None
        return int(round(v))
    digits = re.findall(r"\d+", s_clean(v))
    if not digits:
        return None
    return int("".join(digits))


def parse_price_cell(val) -> Optional[int]:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        try:
            if isinstance(val, float) and pd.isna(val):
                return None
            return int(round(float(val)))
        except Exception:
            return None

    txt = s_clean(val)
    if not txt:
        return None
    txt = txt.replace("Rp", "").replace("rp", "").replace(" ", "")

    if "." in txt and "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "." in txt and "," not in txt:
        txt = txt.replace(".", "")
    elif "," in txt and "." not in txt:
        txt = txt.replace(",", "")

    try:
        return int(round(float(txt)))
    except Exception:
        return None


def apply_multiplier_if_needed(x: Optional[int], threshold: int = 1_000_000, multiplier: int = 1000) -> int:
    if x is None:
        return 0
    if x < threshold:
        return int(x) * multiplier
    return int(x)


def total_upload_size_mb(files: List[Any]) -> float:
    total = 0
    for f in files:
        try:
            total += len(f.getvalue())
        except Exception:
            pass
    return total / (1024 * 1024)


def lower_map_headers(ws: Worksheet, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = s_clean(v).lower()
        if key and key not in m:
            m[key] = c
    return m


def get_header_col_fuzzy(ws: Worksheet, header_row: int, candidates: List[str]) -> Optional[int]:
    m = lower_map_headers(ws, header_row)
    normalized = {re.sub(r"[^a-z0-9]", "", k): v for k, v in m.items()}
    for cand in candidates:
        target = re.sub(r"[^a-z0-9]", "", s_clean(cand).lower())
        if target in normalized:
            return normalized[target]
    return None



def header_contains_any(header_value, keywords: List[str], *, exclude_keywords: Optional[List[str]] = None) -> bool:
    """True jika header mengandung salah satu keyword, dengan guard supaya tidak salah ambil kolom ID."""
    header = su(header_value)
    if not header:
        return False
    if exclude_keywords and any(su(x) in header for x in exclude_keywords):
        return False
    return any(su(k) in header for k in keywords if su(k))


SKU_HEADER_PRIORITY_GROUPS = [
    ["SELLER SKU", "SKU PENJUAL", "SKU PRODUK", "SKU REF", "SKU REF NO", "SKU REF. NO", "KODEBARANG", "KODE BARANG"],
    ["SKU"],
]
SKU_HEADER_EXCLUDES = ["SKU ID", "ID SKU", "PRODUCT ID", "ID PRODUK"]
PRICE_HEADER_KEYWORDS = ["HARGA", "PRICE", "RETAIL PRICE", "HARGA RITEL", "LOCAL CURRENCY", "MATA UANG LOKAL"]
QTY_HEADER_KEYWORDS = ["STOK", "STOCK", "QTY", "QUANTITY", "KUANTITAS", "JUMLAH"]


def find_col_contains_in_row(
    ws: Worksheet,
    header_row: int,
    keywords: List[str],
    *,
    exclude_keywords: Optional[List[str]] = None,
) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        if header_contains_any(ws.cell(header_row, c).value, keywords, exclude_keywords=exclude_keywords):
            return c
    return None


def find_col_contains_any_row(
    ws: Worksheet,
    keywords: List[str],
    *,
    scan_rows: int = 10,
    exclude_keywords: Optional[List[str]] = None,
) -> Tuple[Optional[int], Optional[int]]:
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        col = find_col_contains_in_row(ws, r, keywords, exclude_keywords=exclude_keywords)
        if col is not None:
            return r, col
    return None, None


def find_sku_col_any_row(ws: Worksheet, *, scan_rows: int = 10) -> Tuple[Optional[int], Optional[int]]:
    # Prioritaskan Seller SKU / SKU Penjual. Jangan sampai salah ambil SKU ID.
    for group in SKU_HEADER_PRIORITY_GROUPS:
        for r in range(1, min(scan_rows, ws.max_row) + 1):
            col = find_col_contains_in_row(ws, r, group, exclude_keywords=SKU_HEADER_EXCLUDES)
            if col is not None:
                return r, col
    return None, None


def find_tiktok_style_columns(
    ws: Worksheet,
    *,
    need_price: bool = False,
    need_qty: bool = False,
    scan_rows: int = 10,
    data_offset: int = 3,
) -> Tuple[int, int, Optional[int], Optional[int]]:
    header_rows: List[int] = []
    sku_header_row, sku_col = find_sku_col_any_row(ws, scan_rows=scan_rows)
    if sku_header_row is not None:
        header_rows.append(sku_header_row)

    price_col = None
    if need_price:
        price_header_row, price_col = find_col_contains_any_row(ws, PRICE_HEADER_KEYWORDS, scan_rows=scan_rows)
        if price_header_row is not None:
            header_rows.append(price_header_row)

    qty_col = None
    if need_qty:
        qty_header_row, qty_col = find_col_contains_any_row(ws, QTY_HEADER_KEYWORDS, scan_rows=scan_rows)
        if qty_header_row is not None:
            header_rows.append(qty_header_row)

    if sku_col is None:
        raise ValueError("Kolom Seller SKU / SKU Penjual tidak ditemukan.")
    if need_price and price_col is None:
        raise ValueError("Kolom harga / price tidak ditemukan.")
    if need_qty and qty_col is None:
        raise ValueError("Kolom stok / quantity tidak ditemukan.")

    header_row = max(header_rows) if header_rows else 3
    data_start = header_row + data_offset
    return data_start, sku_col, price_col, qty_col


def find_header_row_by_exact(ws: Worksheet, header_text: str, scan_rows: int = 150) -> Optional[int]:
    target = su(header_text)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            if su(ws.cell(r, c).value) == target:
                return r
    return None


def find_row_contains(ws: Worksheet, needle: str, scan_rows: int = 300) -> Optional[int]:
    target = su(needle)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            v = su(ws.cell(r, c).value)
            if v and (target == v or target in v):
                return r
    return None


def get_first_sheet(wb) -> Worksheet:
    return wb[wb.sheetnames[0]]


def reset_readonly_dimensions_if_needed(ws):
    """Fix template Excel yang dimension metadata-nya salah (sering terbaca hanya kolom A di read_only)."""
    try:
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()
    except Exception:
        pass


def get_row_values_readonly(ws, row_num: int) -> List[Any]:
    try:
        return list(next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True)))
    except StopIteration:
        return []


def build_merged_lookup_map(ws: Worksheet) -> Dict[Tuple[int, int], object]:
    merged_map: Dict[Tuple[int, int], object] = {}
    for mr in ws.merged_cells.ranges:
        top_left_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = top_left_val
    return merged_map


def get_cell_or_merged_value(ws: Worksheet, merged_map: Dict[Tuple[int, int], object], row: int, col: int):
    v = ws.cell(row, col).value
    if v not in (None, ""):
        return v
    return merged_map.get((row, col))


def safe_set_cell_value(ws: Worksheet, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        coord = cell.coordinate
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
        return
    cell.value = value


def workbook_to_bytes(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def zip_named_files(named_files: List[Tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for fname, fbytes in named_files:
            zf.writestr(fname, fbytes)
    return buf.getvalue()


def make_issues_workbook(issues: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "issues_report"
    headers = ["file", "row", "sku_full", "old_value", "new_value", "reason"]
    ws.append(headers)
    for item in issues:
        ws.append([
            item.get("file", ""),
            item.get("row", ""),
            item.get("sku_full", ""),
            item.get("old_value", ""),
            item.get("new_value", ""),
            item.get("reason", ""),
        ])
    return workbook_to_bytes(wb)


def render_summary(title: str, summary: Dict[str, Any]):
    st.subheader(title)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Jumlah File", int(summary.get("files_total", 0)))
    c2.metric("Baris Diperiksa", int(summary.get("rows_scanned", 0)))
    c3.metric("Baris Diupdate", int(summary.get("rows_written", 0)))
    c4.metric("Skip / Issue", int(summary.get("rows_unmatched", 0) + summary.get("issues_count", 0)))


def init_summary(files_total: int, *, include_unchanged: bool = False) -> Dict[str, int]:
    summary = {
        "files_total": int(files_total),
        "rows_scanned": 0,
        "rows_written": 0,
        "rows_unmatched": 0,
        "issues_count": 0,
    }
    if include_unchanged:
        summary["rows_unchanged"] = 0
    return summary


def merge_summary_stats(summary: Dict[str, Any], stats: Dict[str, Any], keys: Tuple[str, ...]):
    for key in keys:
        summary[key] = int(summary.get(key, 0)) + int(stats.get(key, 0))


def normalize_issue_reason(reason: str) -> str:
    reason = s_clean(reason)
    if not reason:
        return "Issue tidak diketahui"
    if "tidak ada di Pricelist" in reason and reason.startswith("Base SKU"):
        return "Base SKU tidak ada di Pricelist"
    if "tidak ada di file Addon Mapping" in reason and reason.startswith("Addon"):
        return "Addon tidak ada di file Addon Mapping"
    if reason.startswith("Harga hasil") and "<= 0" in reason:
        return "Harga hasil <= 0"
    if reason.startswith("Harga ") and "kosong di Pricelist" in reason:
        return "Harga kosong di Pricelist"
    return reason


def summarize_issues_text(issues: List[Dict[str, Any]], max_items: int = 50) -> List[str]:
    counter: Dict[str, int] = {}
    for item in issues or []:
        reason = normalize_issue_reason(item.get("reason", ""))
        counter[reason] = counter.get(reason, 0) + 1
    return [f"{count}x {reason}" for reason, count in sorted(counter.items(), key=lambda x: (-x[1], x[0]))[:max_items]]


def issues_workbook_or_none(issues: List[Dict[str, Any]]) -> Optional[bytes]:
    # Issue selalu tampil sebagai info/ringkasan saja.
    # Opsi Excel di sidebar sudah dihapus supaya proses lebih ringan dan UI lebih bersih.
    st.session_state["_last_issue_info"] = summarize_issues_text(issues)
    return None


def package_output_files(output_files: List[Tuple[str, bytes]], zip_name: str):
    if len(output_files) == 1:
        single_name, single_bytes = output_files[0]
        return single_bytes, single_name
    return zip_named_files(output_files), zip_name


def prune_worksheet_to_changed_rows(ws: Worksheet, changed_rows: List[int], data_start: int):
    if changed_rows:
        rewrite_worksheet_with_rows(ws, data_start, collect_changed_row_values(ws, changed_rows))


def append_no_changes_issue(issues: List[Dict[str, Any]], file_name: str):
    issues.append({"file": file_name, "reason": "Tidak ada baris berubah pada file ini."})




def progress_tick(progress_callback, start_pct: int, end_pct: int, current: int, total: int, message: str):
    """Update progress aman untuk processor besar."""
    if not progress_callback:
        return
    total = max(1, int(total or 1))
    current = max(0, min(int(current or 0), total))
    pct = int(start_pct + (end_pct - start_pct) * (current / total))
    progress_callback(pct, message)


def rewrite_worksheet_with_rows(ws: Worksheet, data_start: int, row_values: List[List[Any]]):
    """Lebih cepat daripada delete_rows satu-per-satu: hapus blok data sekali lalu tulis rows hasil."""
    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)
    for out_r, row_vals in enumerate(row_values, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=out_r, column=c).value = val


def collect_changed_row_values(ws: Worksheet, changed_rows: List[int], max_col: Optional[int] = None) -> List[List[Any]]:
    max_col = max_col or ws.max_column
    return [[ws.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in changed_rows]

def process_marketplace_stock_sheet(
    ws: Worksheet,
    *,
    file_name: str,
    stock_lookup: Dict[str, Dict],
    selected_modes: Set[str],
    chosen_areas: Set[str],
    chosen_gudangs: Set[str],
    zero_below: int,
    zero_if_missing: bool = False,
    data_start: int,
    sku_col: int,
    qty_col: int,
    summary: Dict[str, Any],
    issues: List[Dict[str, Any]],
    progress_callback=None,
    progress_label: str = "Memproses stok",
    progress_start: int = 15,
    progress_end: int = 82,
) -> List[int]:
    changed_rows: List[int] = []
    total_rows = max(1, ws.max_row - data_start + 1)
    for r in range(data_start, ws.max_row + 1):
        if progress_callback and (r - data_start) % 2000 == 0:
            progress_tick(progress_callback, progress_start, progress_end, r - data_start, total_rows, f"{progress_label}: {r - data_start} baris...")
        sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
        if not sku_full:
            continue
        summary["rows_scanned"] += 1
        old_qty = to_int_or_none(ws.cell(row=r, column=qty_col).value)
        new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing)
        if new_qty is None:
            summary["rows_unmatched"] += 1
            issues.append({
                "file": file_name,
                "row": r,
                "sku_full": sku_full,
                "old_value": old_qty,
                "new_value": "",
                "reason": "SKU tidak ditemukan di Pricelist stok",
            })
            continue
        if old_qty is not None and int(old_qty) == int(new_qty):
            continue
        safe_set_cell_value(ws, r, qty_col, int(new_qty))
        changed_rows.append(r)
        summary["rows_written"] += 1
    return changed_rows


def process_marketplace_price_sheet(
    ws: Worksheet,
    *,
    file_name: str,
    price_map: Dict[str, Dict[str, int]],
    addon_map: Dict[str, int],
    discount_rp: int,
    price_key: str,
    data_start: int,
    sku_col: int,
    target_price_cols: List[int],
    summary: Dict[str, Any],
    issues: List[Dict[str, Any]],
    progress_callback=None,
    progress_label: str = "Memproses stok",
    progress_start: int = 15,
    progress_end: int = 82,
) -> List[int]:
    changed_rows: List[int] = []
    total_rows = max(1, ws.max_row - data_start + 1)
    for r in range(data_start, ws.max_row + 1):
        if progress_callback and (r - data_start) % 2000 == 0:
            progress_tick(progress_callback, progress_start, progress_end, r - data_start, total_rows, f"{progress_label}: {r - data_start} baris...")
        sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
        if not sku_full:
            continue
        summary["rows_scanned"] += 1
        old_values = [parse_price_cell(ws.cell(row=r, column=col).value) for col in target_price_cols]
        new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, price_key, discount_rp)
        if new_price is None:
            summary["rows_unmatched"] += 1
            issues.append({
                "file": file_name,
                "row": r,
                "sku_full": sku_full,
                "old_value": old_values[0] if old_values else "",
                "new_value": "",
                "reason": reason,
            })
            continue
        if old_values and all(v is not None and int(v) == int(new_price) for v in old_values):
            continue
        for col in target_price_cols:
            safe_set_cell_value(ws, r, col, int(new_price))
        changed_rows.append(r)
        summary["rows_written"] += 1
    return changed_rows


def build_issue_preview_from_bytes(issues_bytes: Optional[bytes], max_items: int = 30) -> List[str]:
    if not issues_bytes:
        return []
    try:
        wb = load_workbook(io.BytesIO(issues_bytes), data_only=True, read_only=True)
        ws = wb.active
        headers = [s_clean(ws.cell(1, c).value).lower() for c in range(1, ws.max_column + 1)]
        reason_col = headers.index("reason") + 1 if "reason" in headers else ws.max_column
        counter: Dict[str, int] = {}
        for r in range(2, ws.max_row + 1):
            reason = s_clean(ws.cell(r, reason_col).value)
            if not reason:
                continue
            counter[reason] = counter.get(reason, 0) + 1
        return [f"{count}x {reason}" for reason, count in sorted(counter.items(), key=lambda x: (-x[1], x[0]))[:max_items]]
    except Exception as e:
        return [f"Gagal membaca detail issues: {e}"]


def cache_downloads(
    cache_key: str,
    result_name: str,
    result_bytes: Optional[bytes],
    issues_bytes: Optional[bytes],
    summary: Optional[Dict[str, Any]] = None,
    issues_name: str = "issues_report.xlsx",
):
    st.session_state.download_cache[cache_key] = {
        "result_name": result_name,
        "result_bytes": result_bytes,
        "issues_name": issues_name,
        "issues_bytes": issues_bytes,
    }
    if summary is not None:
        st.session_state.summary_cache[cache_key] = summary
    if issues_bytes:
        st.session_state.issue_preview_cache[cache_key] = build_issue_preview_from_bytes(issues_bytes)
    else:
        st.session_state.issue_preview_cache[cache_key] = list(st.session_state.get("_last_issue_info", []))

    # Simpan pointer hasil terakhir per halaman/fitur.
    # Jadi hasil tidak hilang saat rerun di fitur yang sama, tapi tidak ikut muncul di fitur lain.
    st.session_state.last_result_cache_key = cache_key
    st.session_state.last_result_title = result_name or "Hasil terakhir"
    st.session_state.last_result_route = st.session_state.get("current_route")


def render_issue_preview(cache_key: str):
    items = st.session_state.issue_preview_cache.get(cache_key, [])
    if not items:
        return
    with st.expander("Detail Issue", expanded=True):
        for item in items:
            st.write(f"> {item}")


def render_downloads(cache_key: str):
    st.session_state.setdefault("_rendered_cache_keys", set()).add(cache_key)
    payload = st.session_state.download_cache.get(cache_key)
    if not payload:
        return
    if payload.get("result_bytes"):
        st.download_button(
            "Download Hasil",
            payload["result_bytes"],
            file_name=payload["result_name"],
            key=f"dl_{cache_key}_result",
            on_click="ignore",
        )
    if payload.get("issues_bytes"):
        st.download_button(
            "Download Issues",
            payload["issues_bytes"],
            file_name=payload["issues_name"],
            key=f"dl_{cache_key}_issues",
            on_click="ignore",
        )


def render_cached_summary(cache_key: str, title: str = "Ringkasan Hasil"):
    st.session_state.setdefault("_rendered_cache_keys", set()).add(cache_key)
    summary = st.session_state.summary_cache.get(cache_key)
    if summary:
        render_summary(title, summary)
        render_issue_preview(cache_key)


def render_last_result_panel():
    """Tampilkan hasil proses terakhir walaupun user pindah fitur/menu.

    Streamlit selalu rerun dari atas saat radio/sidebar berubah. Tanpa panel global,
    hasil hanya terlihat di halaman asal karena render_downloads dipanggil di halaman itu saja.
    """
    cache_key = st.session_state.get("last_result_cache_key")
    if not cache_key:
        return
    if cache_key in st.session_state.get("_rendered_cache_keys", set()):
        return
    if cache_key not in st.session_state.download_cache and cache_key not in st.session_state.summary_cache:
        return

    # Jangan tampilkan hasil dari fitur lain.
    # Contoh: hasil Update Stok TikTokShop tidak muncul saat user pindah ke MWH/BigSeller.
    current_route = st.session_state.get("current_route")
    last_route = st.session_state.get("last_result_route")
    if last_route and current_route and last_route != current_route:
        return

    st.markdown("---")
    with st.expander("Hasil Terakhir", expanded=True):
        title = st.session_state.get("last_result_title") or "Hasil terakhir"
        st.caption(f"Hasil masih tersimpan selama aplikasi belum direstart: {title}")
        render_cached_summary(cache_key, title="Ringkasan Hasil Terakhir")
        render_downloads(cache_key)


def get_change_sheet(wb):
    for sname in wb.sheetnames:
        if su(sname) == "CHANGE":
            return wb[sname]
    raise ValueError("Sheet 'CHANGE' tidak ditemukan di Pricelist.")


def find_header_row_by_candidates(
    ws: Worksheet,
    required_candidates: Dict[str, List[str]],
    scan_rows: int = 10
) -> Tuple[int, Dict[str, int]]:
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        found: Dict[str, int] = {}
        ok = True
        for key, candidates in required_candidates.items():
            col = get_header_col_fuzzy(ws, r, candidates)
            if col is None:
                ok = False
                break
            found[key] = col
        if ok:
            return r, found
    raise ValueError("Header tidak ditemukan. Pastikan file memiliki kolom yang dibutuhkan.")


def normalize_header_token(v) -> str:
    return re.sub(r"[^a-z0-9]", "", s_clean(v).lower())


def _header_match_score(header_token: str, candidates: List[str], *, excludes: Optional[List[str]] = None) -> int:
    if not header_token:
        return 0
    excludes_norm = [normalize_header_token(x) for x in (excludes or []) if normalize_header_token(x)]
    if any(x and x in header_token for x in excludes_norm):
        return 0
    best = 0
    for cand in candidates:
        cand_token = normalize_header_token(cand)
        if not cand_token:
            continue
        if header_token == cand_token:
            best = max(best, 100)
        elif cand_token in header_token:
            best = max(best, 70)
        elif header_token in cand_token and len(header_token) >= 3:
            best = max(best, 55)
    return best


def find_header_row_by_candidates_flexible(
    ws: Worksheet,
    required_candidates: Dict[str, List[str]],
    *,
    scan_rows: int = 30,
    excludes: Optional[Dict[str, List[str]]] = None,
) -> Tuple[int, Dict[str, int]]:
    # BigSeller export kadang punya worksheet dimension metadata salah,
    # sehingga read_only hanya membaca kolom A. Reset dulu agar kolom H/K ikut terbaca.
    reset_readonly_dimensions_if_needed(ws)

    excludes = excludes or {}
    last_headers: List[str] = []
    max_scan = min(scan_rows, ws.max_row or scan_rows)

    for r in range(1, max_scan + 1):
        headers = get_row_values_readonly(ws, r)
        headers_clean = [s_clean(v) for v in headers]
        if any(headers_clean):
            last_headers = [h for h in headers_clean if h]

        found: Dict[str, int] = {}
        used_cols: Set[int] = set()

        for key, candidates in required_candidates.items():
            best_col = None
            best_score = 0
            for idx, header in enumerate(headers_clean, start=1):
                if idx in used_cols:
                    continue
                token = normalize_header_token(header)
                if not token:
                    continue

                # Exact match harus menang. Jangan sampai cell warning seperti
                # "Peringatan: Jangan Mengubah Product_ID dan Sku_ID!" dianggap header.
                cand_tokens = [normalize_header_token(c) for c in candidates if normalize_header_token(c)]
                if token in cand_tokens:
                    score = 100
                else:
                    score = _header_match_score(token, candidates, excludes=excludes.get(key))

                if score > best_score:
                    best_score = score
                    best_col = idx

            if best_col is not None and best_score > 0:
                found[key] = best_col
                used_cols.add(best_col)

        if set(required_candidates.keys()).issubset(found.keys()):
            return r, found

    preview = last_headers[:12]
    raise ValueError(f"Header tidak ditemukan. Header terbaca: {preview}")


def find_bigseller_omnichannel_columns(ws: Worksheet, *, mode: str) -> Tuple[int, Dict[str, int]]:
    sku_candidates = [
        "Seller SKU", "SKU Penjual", "SKU", "Product SKU", "Merchant SKU",
        "SKU Ref. No.(Optional)", "SKU Ref No Optional", "SKU Ref No", "Kode Barang", "Kodebarang",
    ]
    sku_excludes = ["SKU ID", "ID SKU", "Product ID", "ID Produk", "Variation ID", "Variant ID", "Goods ID", "Item ID"]

    if mode == "stock":
        target_candidates = ["Stock", "Stok", "Quantity", "Qty", "Available Stock", "Warehouse Stock", "Jumlah", "Inventory"]
        target_excludes = ["Stock ID", "Warehouse ID", "SKU ID"]
        target_key = "qty"
    else:
        target_candidates = [
            "Price", "Harga", "Harga Jual", "Selling Price", "Sale Price", "Retail Price",
            "Normal Price", "Harga Normal", "Original Price", "Local Currency", "Mata Uang Lokal",
            "Harga Produk", "Product Price", "Unit Price",
        ]
        target_excludes = ["Price ID", "Harga ID", "Discount", "Diskon", "Promo", "Campaign"]
        target_key = "price"

    return find_header_row_by_candidates_flexible(
        ws,
        {"sku": sku_candidates, target_key: target_candidates},
        scan_rows=40,
        excludes={"sku": sku_excludes, target_key: target_excludes},
    )


# ============================================================
# STOCK PRICELIST HELPERS
# ============================================================
def sheet_range_between(sheetnames: List[str], start: str, end: str) -> List[str]:
    up = [su(x) for x in sheetnames]
    if start.upper() not in up or end.upper() not in up:
        raise ValueError(f"Sheet range tidak valid. Pastikan ada '{start}' dan '{end}'.")
    i0 = up.index(start.upper())
    i1 = up.index(end.upper())
    if i0 > i1:
        i0, i1 = i1, i0
    return sheetnames[i0:i1 + 1]


def delete_coming_block_in_laptop(ws: Worksheet):
    r_start = find_row_contains(ws, "COMING", scan_rows=600)
    r_end = find_row_contains(ws, "END COMING", scan_rows=1200)
    if r_start and r_end and r_end >= r_start:
        ws.delete_rows(r_start, r_end - r_start + 1)


def find_tot_col(ws: Worksheet, header_row_hint: int) -> Tuple[int, int]:
    for c in range(1, ws.max_column + 1):
        if su(ws.cell(header_row_hint, c).value) == "TOT":
            return header_row_hint, c
    for r in range(1, min(12, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if su(ws.cell(r, c).value) == "TOT":
                return r, c
    raise ValueError("Kolom 'TOT' tidak ketemu.")


def build_area_warehouse_meta(
    ws: Worksheet,
    merged_map: Dict[Tuple[int, int], object],
    area_row: int,
    warehouse_row: int,
    start_col: int,
) -> Dict[int, Dict[str, str]]:
    col_area_wh: Dict[int, Dict[str, str]] = {}
    for c in range(start_col, ws.max_column + 1):
        area_raw = get_cell_or_merged_value(ws, merged_map, area_row, c)
        warehouse_raw = get_cell_or_merged_value(ws, merged_map, warehouse_row, c)
        area_name = su(area_raw)
        warehouse_name = su(warehouse_raw)
        if not area_name or not warehouse_name:
            continue
        col_area_wh[c] = {
            "area": area_name,
            "warehouse": warehouse_name,
            "area_wh": f"{area_name}-{warehouse_name}",
        }
    return col_area_wh


def build_stock_lookup_from_sheet_fast(ws: Worksheet, sheet_name: str, progress_callback=None, progress_start: int = 4, progress_end: int = 12):
    if progress_callback:
        progress_callback(progress_start, f"[{sheet_name}] mencari header KODEBARANG...")

    header_row = find_header_row_by_exact(ws, "KODEBARANG", scan_rows=200)
    if header_row is None:
        header_row = find_header_row_by_exact(ws, "KODE BARANG", scan_rows=200)
    if header_row is None:
        raise ValueError(f"[{sheet_name}] Header 'KODEBARANG' tidak ketemu.")

    sku_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v in ("KODEBARANG", "KODE BARANG"):
            sku_col = c
            break
    if sku_col is None:
        raise ValueError(f"[{sheet_name}] Kolom 'KODEBARANG' / 'KODE BARANG' tidak ditemukan.")

    if progress_callback:
        progress_callback(progress_start + 1, f"[{sheet_name}] membaca struktur area/gudang...")

    header_row_used, tot_col = find_tot_col(ws, header_row)
    merged_map = build_merged_lookup_map(ws)
    area_row = header_row_used + 1
    warehouse_row = header_row_used + 2
    col_area_wh = build_area_warehouse_meta(
        ws,
        merged_map,
        area_row=area_row,
        warehouse_row=warehouse_row,
        start_col=tot_col + 1,
    )

    if progress_callback:
        progress_callback(
            progress_start + 2,
            f"[{sheet_name}] area/gudang terbaca: {len(col_area_wh)} kolom stok. Mulai scan SKU...",
        )

    sku_map: Dict[str, Dict[str, Any]] = {}
    area_names: Set[str] = set()
    area_warehouses: Set[str] = set()
    for meta in col_area_wh.values():
        area_names.add(meta["area"])
        area_warehouses.add(meta["area_wh"])

    data_start = max(header_row, warehouse_row) + 1
    total_rows = max(1, ws.max_row - data_start + 1)
    for idx, r in enumerate(range(data_start, ws.max_row + 1), start=1):
        if progress_callback and (idx == 1 or idx % 500 == 0):
            progress_tick(
                progress_callback,
                progress_start + 2,
                progress_end,
                idx,
                total_rows,
                f"[{sheet_name}] scan Pricelist stok: {idx}/{total_rows} baris, SKU terbaca {len(sku_map)}...",
            )

        sku = s_clean(ws.cell(r, sku_col).value)
        if not sku:
            continue
        sku_key = norm_sku(sku)
        if sku_key in ("TOTAL", "KODEBARANG", "KODE BARANG", "KODEBARANG."):
            continue

        tot_val = to_int_or_none(ws.cell(r, tot_col).value)
        by_area_wh: Dict[str, int] = {}
        by_area: Dict[str, int] = {}
        for c, meta in col_area_wh.items():
            v = to_int_or_none(ws.cell(r, c).value)
            if v is None:
                continue
            area_wh_name = meta["area_wh"]
            area_name = meta["area"]
            by_area_wh[area_wh_name] = by_area_wh.get(area_wh_name, 0) + int(v)
            by_area[area_name] = by_area.get(area_name, 0) + int(v)
        sku_map[sku_key] = {"TOT": tot_val, "by_area_wh": by_area_wh, "by_area": by_area}

    if progress_callback:
        progress_callback(progress_end, f"[{sheet_name}] selesai: {len(sku_map)} SKU stok terbaca.")

    return sku_map, {"area_options": sorted(area_names), "gudang_options": sorted(area_warehouses), "default_gudang_options": sorted(area_warehouses)}


def build_stock_lookup_from_pricelist_bytes(pl_bytes: bytes, progress_callback=None, progress_start: int = 3, progress_end: int = 12):
    if progress_callback:
        progress_callback(progress_start, "Membuka workbook Pricelist stok...")

    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True, read_only=False)

    if progress_callback:
        progress_callback(progress_start + 1, f"Workbook terbuka. Sheet terdeteksi: {len(wb.sheetnames)} sheet...")

    for sname in wb.sheetnames:
        if su(sname) == "LAPTOP":
            if progress_callback:
                progress_callback(progress_start + 1, "Membersihkan blok COMING di sheet LAPTOP...")
            delete_coming_block_in_laptop(wb[sname])
            break

    requested = {su(name): name for name in STOCK_PRICELIST_SHEETS}
    target_sheets = [sname for sname in wb.sheetnames if su(sname) in requested]
    if not target_sheets:
        raise ValueError(
            f"Sheet stok tidak ditemukan. Pricelist harus punya minimal salah satu sheet: {', '.join(STOCK_PRICELIST_SHEETS)}"
        )

    if progress_callback:
        progress_callback(
            progress_start + 2,
            f"Sheet stok yang diproses: {', '.join(target_sheets)}",
        )

    merged_lookup: Dict[str, Dict[str, Any]] = {}
    area_options_all: Set[str] = set()
    gudang_options_all: Set[str] = set()
    default_gudang_options_all: Set[str] = set()

    total_sheets = max(1, len(target_sheets))
    span = max(1, progress_end - (progress_start + 2))
    for idx, sname in enumerate(target_sheets, start=1):
        sheet_start = progress_start + 2 + int((idx - 1) / total_sheets * span)
        sheet_end = progress_start + 2 + int(idx / total_sheets * span)
        sheet_end = max(sheet_start + 1, sheet_end)

        if progress_callback:
            progress_callback(sheet_start, f"Membaca sheet stok {idx}/{total_sheets}: {sname}...")

        sku_map, meta = build_stock_lookup_from_sheet_fast(
            wb[sname],
            sname,
            progress_callback=progress_callback,
            progress_start=sheet_start,
            progress_end=min(progress_end, sheet_end),
        )
        merged_lookup.update(sku_map)
        area_options_all |= set(meta.get("area_options", []))
        gudang_options_all |= set(meta.get("gudang_options", []))
        default_gudang_options_all |= set(meta.get("default_gudang_options", []))

        if progress_callback:
            progress_callback(
                min(progress_end, sheet_end),
                f"Sheet {idx}/{total_sheets} selesai: {sname}, total SKU terkumpul {len(merged_lookup)}...",
            )

    if not merged_lookup:
        raise ValueError("Pricelist terbaca, tapi lookup stok kosong.")

    if progress_callback:
        progress_callback(
            progress_end,
            f"Pricelist stok selesai: {len(merged_lookup)} SKU, {len(area_options_all)} area, {len(gudang_options_all)} gudang.",
        )

    return merged_lookup, {
        "area_options": sorted(area_options_all),
        "gudang_options": sorted(gudang_options_all),
        "default_gudang_options": sorted(default_gudang_options_all),
    }

def apply_stock_floor_rule(qty: Optional[int], zero_below: int = 0) -> Optional[int]:
    if qty is None:
        return None
    qty = int(qty)
    if zero_below > 0 and qty < int(zero_below):
        return 0
    return qty


def get_default_tongle_gudangs(area_warehouses: List[str]) -> List[str]:
    area_wh_set = {su(a) for a in area_warehouses}
    return [area for area in DEFAULT_TONGLE_GUDANGS if su(area) in area_wh_set]


def pick_stock_value(
    sku_full: str,
    stock_lookup: Dict[str, Dict],
    selected_modes: Set[str],
    chosen_areas: Set[str],
    chosen_gudangs: Set[str],
    zero_below: int = 0,
    zero_if_missing: bool = False,
) -> Optional[int]:
    base, _ = split_sku_addons(sku_full)
    base_key = norm_sku(base)
    if not base_key or base_key not in stock_lookup:
        return 0 if zero_if_missing else None

    rec = stock_lookup[base_key]
    tot = rec.get("TOT")
    by_area = rec.get("by_area", {}) or {}
    by_area_wh = rec.get("by_area_wh", {}) or {}

    if "Stok Nasional (TOT)" in selected_modes:
        return apply_stock_floor_rule(tot if tot is not None else None, zero_below)

    picked_area_whs: Set[str] = set()
    if "Default" in selected_modes:
        picked_area_whs |= {a for a in DEFAULT_TONGLE_GUDANGS if a in by_area_wh}

    total = 0
    counted_area_whs: Set[str] = set()

    if "Area" in selected_modes:
        for area_name in chosen_areas:
            total += int(by_area.get(area_name, 0) or 0)
            counted_area_whs |= {k for k in by_area_wh.keys() if k.startswith(f"{area_name}-")}

    if "Gudang" in selected_modes:
        picked_area_whs |= {a for a in chosen_gudangs if a in by_area_wh}

    for area_wh in picked_area_whs:
        if area_wh in counted_area_whs:
            continue
        total += int(by_area_wh.get(area_wh, 0) or 0)

    if not selected_modes or total == 0 and not counted_area_whs and not picked_area_whs:
        return None

    return apply_stock_floor_rule(total, zero_below)


# ============================================================
# STOCK PROCESSORS
# ============================================================
def find_shopee_columns_readonly(ws) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 7
    row_vals = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    sku_col = None
    qty_col = None
    for idx, val in enumerate(row_vals, start=1):
        v = su(val)
        if v == "SKU":
            sku_col = idx
        if v == "STOK":
            qty_col = idx
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/Stok tidak ketemu pada template Shopee (Mall & Star).")
    return data_start, sku_col, qty_col


def find_shopee_columns_normal(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 3
    data_start = 7
    sku_col = None
    qty_col = None
    for c in range(1, ws.max_column + 1):
        v = su(ws.cell(header_row, c).value)
        if v == "SKU":
            sku_col = c
        if v == "STOK":
            qty_col = c
    if not sku_col or not qty_col:
        raise ValueError("Kolom SKU/Stok tidak ketemu pada template Shopee (Mall & Star).")
    return data_start, sku_col, qty_col


def collect_changed_rows_stock_shopee(file_bytes: bytes, stock_lookup: Dict[str, Dict], selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None, progress_label: str = "Proses stok Shopee", progress_start: int = 15, progress_end: int = 82):
    stats = {"rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0}
    changed_rows: List[List[Any]] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    data_start, sku_col, qty_col = find_shopee_columns_readonly(ws)

    total_rows = max(1, ws.max_row - data_start + 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=0):
        if progress_callback and row_idx % 2000 == 0:
            progress_tick(progress_callback, progress_start, progress_end, row_idx, total_rows, f"{progress_label}: {row_idx} baris...")
        row_list = list(row)
        sku_full = s_clean(row_list[sku_col - 1] if len(row_list) >= sku_col else None)
        if not sku_full:
            continue
        stats["rows_scanned"] += 1
        old_qty = to_int_or_none(row_list[qty_col - 1] if len(row_list) >= qty_col else None)
        new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing)
        if new_qty is None:
            stats["rows_unmatched"] += 1
            continue
        if old_qty is not None and int(old_qty) == int(new_qty):
            stats["rows_unchanged"] += 1
            continue
        if len(row_list) < qty_col:
            row_list.extend([None] * (qty_col - len(row_list)))
        row_list[qty_col - 1] = int(new_qty)
        changed_rows.append(row_list)
        stats["rows_written"] += 1
    wb.close()
    return changed_rows, stats


def write_stock_shopee_output(template_bytes: bytes, changed_rows_all: List[List[Any]]) -> bytes:
    out_wb = load_workbook(io.BytesIO(template_bytes))
    out_ws = get_first_sheet(out_wb)
    data_start, _, _ = find_shopee_columns_normal(out_ws)
    if out_ws.max_row >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)
    for idx, row_vals in enumerate(changed_rows_all, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            out_ws.cell(idx, c).value = val
    return workbook_to_bytes(out_wb)


def process_shopee_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist stok dan area/gudang...")
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue(), progress_callback=progress_callback)
    changed_rows_all: List[List[Any]] = []
    issues: List[Dict[str, Any]] = []
    summary = init_summary(len(mass_files), include_unchanged=True)

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        try:
            if progress_callback:
                progress_callback(10 + int((file_idx - 1) / total_files * 70), f"Membaca file Shopee {file_idx}/{total_files}: {mf.name}")
            rows, stats = collect_changed_rows_stock_shopee(mf.getvalue(), stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback, f"Proses stok Shopee {file_idx}/{total_files}")
            changed_rows_all.extend(rows)
            merge_summary_stats(summary, stats, ("rows_scanned", "rows_written", "rows_unchanged", "rows_unmatched"))
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    if progress_callback:
        progress_callback(88, f"Membuat output Shopee ({len(changed_rows_all)} baris berubah)...")
    result_bytes = write_stock_shopee_output(mass_files[0].getvalue(), changed_rows_all)
    summary["issues_count"] = len(issues)
    return result_bytes, issues_workbook_or_none(issues), summary


def find_tiktokshop_columns_readonly(ws) -> Tuple[int, int, int]:
    data_start, sku_col, _, qty_col = find_tiktok_style_columns(ws, need_qty=True, scan_rows=10, data_offset=3)
    if qty_col is None:
        raise ValueError("Kolom SKU/stok tidak ketemu pada template TikTokShop.")
    return data_start, sku_col, qty_col


def find_tiktokshop_columns_normal(ws: Worksheet) -> Tuple[int, int, int]:
    data_start, sku_col, _, qty_col = find_tiktok_style_columns(ws, need_qty=True, scan_rows=10, data_offset=3)
    if qty_col is None:
        raise ValueError("Kolom SKU/stok tidak ketemu pada template TikTokShop.")
    return data_start, sku_col, qty_col


def collect_changed_rows_stock_tiktokshop(file_bytes: bytes, stock_lookup: Dict[str, Dict], selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None, progress_label: str = "Proses stok TikTokShop", progress_start: int = 15, progress_end: int = 82):
    stats = {"rows_scanned": 0, "rows_written": 0, "rows_unchanged": 0, "rows_unmatched": 0}
    changed_rows: List[List[Any]] = []
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    data_start, sku_col, qty_col = find_tiktokshop_columns_readonly(ws)

    total_rows = max(1, ws.max_row - data_start + 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=0):
        if progress_callback and row_idx % 2000 == 0:
            progress_tick(progress_callback, progress_start, progress_end, row_idx, total_rows, f"{progress_label}: {row_idx} baris...")
        row_list = list(row)
        sku_full = s_clean(row_list[sku_col - 1] if len(row_list) >= sku_col else None)
        if not sku_full:
            continue
        stats["rows_scanned"] += 1
        old_qty = to_int_or_none(row_list[qty_col - 1] if len(row_list) >= qty_col else None)
        new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing)
        if new_qty is None:
            stats["rows_unmatched"] += 1
            continue
        if old_qty is not None and int(old_qty) == int(new_qty):
            stats["rows_unchanged"] += 1
            continue
        if len(row_list) < qty_col:
            row_list.extend([None] * (qty_col - len(row_list)))
        row_list[qty_col - 1] = int(new_qty)
        changed_rows.append(row_list)
        stats["rows_written"] += 1
    wb.close()
    return changed_rows, stats


def write_stock_tiktokshop_output(template_bytes: bytes, changed_rows_all: List[List[Any]]) -> bytes:
    out_wb = load_workbook(io.BytesIO(template_bytes))
    out_ws = get_first_sheet(out_wb)
    data_start, _, _ = find_tiktokshop_columns_normal(out_ws)
    if out_ws.max_row >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)
    for idx, row_vals in enumerate(changed_rows_all, start=data_start):
        for c, val in enumerate(row_vals, start=1):
            out_ws.cell(idx, c).value = val
    return workbook_to_bytes(out_wb)


def process_tiktokshop_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist stok dan area/gudang...")
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue(), progress_callback=progress_callback)
    changed_rows_all: List[List[Any]] = []
    issues: List[Dict[str, Any]] = []
    summary = init_summary(len(mass_files), include_unchanged=True)

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        try:
            if progress_callback:
                progress_callback(10 + int((file_idx - 1) / total_files * 70), f"Membaca file TikTokShop {file_idx}/{total_files}: {mf.name}")
            rows, stats = collect_changed_rows_stock_tiktokshop(mf.getvalue(), stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback, f"Proses stok TikTokShop {file_idx}/{total_files}")
            changed_rows_all.extend(rows)
            merge_summary_stats(summary, stats, ("rows_scanned", "rows_written", "rows_unchanged", "rows_unmatched"))
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    if progress_callback:
        progress_callback(88, f"Membuat output TikTokShop ({len(changed_rows_all)} baris berubah)...")
    result_bytes = write_stock_tiktokshop_output(mass_files[0].getvalue(), changed_rows_all)
    summary["issues_count"] = len(issues)
    return result_bytes, issues_workbook_or_none(issues), summary


def find_mwh_stock_columns(ws: Worksheet) -> Tuple[int, int, int]:
    data_start, sku_col, _, qty_col = find_tiktok_style_columns(ws, need_qty=True, scan_rows=10, data_offset=3)
    if qty_col is None:
        raise ValueError("Kolom SKU/Jumlah tidak ketemu pada template Mwh.")
    return data_start, sku_col, qty_col


def process_mwh_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist stok dan area/gudang...")
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue(), progress_callback=progress_callback)
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files), include_unchanged=True)

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        try:
            if progress_callback:
                progress_callback(10 + int((file_idx - 1) / total_files * 70), f"Membaca file Mwh {file_idx}/{total_files}: {mf.name}")
            wb = load_workbook(io.BytesIO(mf.getvalue()))
            ws = wb.active
            data_start, sku_col, qty_col = find_mwh_stock_columns(ws)

            total_rows = max(1, ws.max_row - data_start + 1)
            for r in range(data_start, ws.max_row + 1):
                if progress_callback and (r - data_start) % 2000 == 0:
                    progress_tick(progress_callback, 15, 82, r - data_start, total_rows, f"Memproses baris Mwh {file_idx}/{total_files}: {r - data_start} baris...")
                sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
                if not sku_full:
                    continue
                summary["rows_scanned"] += 1
                old_qty = to_int_or_none(ws.cell(row=r, column=qty_col).value)
                new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing)

                if new_qty is None:
                    summary["rows_unmatched"] += 1
                    issues.append({
                        "file": mf.name,
                        "row": r,
                        "sku_full": sku_full,
                        "old_value": old_qty,
                        "new_value": "",
                        "reason": "SKU tidak ditemukan di Pricelist stok",
                    })
                    continue

                if old_qty is not None and int(old_qty) == int(new_qty):
                    summary["rows_unchanged"] += 1
                    continue

                safe_set_cell_value(ws, r, qty_col, int(new_qty))
                summary["rows_written"] += 1

            output_files.append((f"hasil_update_stok_mwh_{mf.name}", workbook_to_bytes(wb)))
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    if progress_callback:
        progress_callback(88, f"Membuat output Mwh ({len(output_files)} file)...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, "hasil_update_stok_mwh.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary


def find_bigseller_stock_columns(ws: Worksheet) -> Tuple[int, int, int]:
    header_row, found_cols = find_bigseller_omnichannel_columns(ws, mode="stock")
    return header_row + 1, found_cols["sku"], found_cols["qty"]


def process_bigseller_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None):
    # Optimized BigSeller stock:
    # - read_only=True + iter_rows(values_only=True) untuk file besar
    # - progress dibagi fase: pricelist, proses file, buat output, zip
    # - mode "Info saja" tidak menyimpan puluhan ribu detail issue ke memory/workbook
    if progress_callback:
        progress_callback(3, "Membaca Pricelist stok dan area/gudang...")
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue(), progress_callback=progress_callback)

    issue_excel_mode = st.session_state.get("issue_output_mode", "Info saja") == "Excel"
    issues: List[Dict[str, Any]] = []
    issue_counter: Dict[str, int] = {}

    def add_issue(reason: str, *, file_name: str = "", row: Any = "", sku_full: str = "", old_value: Any = "", new_value: Any = ""):
        normalized = normalize_issue_reason(reason)
        issue_counter[normalized] = issue_counter.get(normalized, 0) + 1
        if issue_excel_mode:
            issues.append({
                "file": file_name,
                "row": row,
                "sku_full": sku_full,
                "old_value": old_value,
                "new_value": new_value,
                "reason": reason,
            })

    def issue_info_from_counter() -> List[str]:
        return [
            f"{count}x {reason}"
            for reason, count in sorted(issue_counter.items(), key=lambda x: (-x[1], x[0]))[:50]
        ]

    summary = init_summary(len(mass_files))
    output_parts: List[Tuple[str, bytes]] = []
    current_rows: List[List[Any]] = []
    current_part = 1
    output_header: List[Any] = []
    header_len = 0

    def flush_part():
        nonlocal current_rows, current_part, output_parts, output_header, header_len
        if not current_rows:
            return
        if progress_callback:
            progress_callback(88, f"Membuat output BigSeller part {current_part} ({len(current_rows)} baris)...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c, val in enumerate(output_header, start=1):
            ws.cell(row=1, column=c).value = val
        for r_idx, row_vals in enumerate(current_rows, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws.cell(row=r_idx, column=c).value = val
        output_parts.append((f"hasil_update_stok_bigseller_part_{current_part}.xlsx", workbook_to_bytes(wb)))
        current_rows = []
        current_part += 1

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        wb = None
        try:
            if progress_callback:
                progress_callback(10 + int((file_idx - 1) / total_files * 75), f"Membaca BigSeller {file_idx}/{total_files}: {mf.name}")
            wb = load_workbook(io.BytesIO(mf.getvalue()), read_only=True, data_only=False)
            ws = wb.worksheets[0]
            reset_readonly_dimensions_if_needed(ws)
            data_start, sku_col, qty_col = find_bigseller_stock_columns(ws)

            if not output_header:
                header_row = data_start - 1
                output_header = get_row_values_readonly(ws, header_row)
                header_len = max(len(output_header), sku_col, qty_col)
                if len(output_header) < header_len:
                    output_header.extend([None] * (header_len - len(output_header)))

            file_rows_scanned = 0
            for r, row_vals_raw in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
                row_vals = list(row_vals_raw[:header_len])
                if len(row_vals) < header_len:
                    row_vals.extend([None] * (header_len - len(row_vals)))
                sku_full = s_clean(row_vals[sku_col - 1] if len(row_vals) >= sku_col else None)
                if not sku_full:
                    continue

                summary["rows_scanned"] += 1
                file_rows_scanned += 1
                old_qty = to_int_or_none(row_vals[qty_col - 1] if len(row_vals) >= qty_col else None)
                new_qty = pick_stock_value(sku_full, stock_lookup, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing)

                if new_qty is None:
                    summary["rows_unmatched"] += 1
                    add_issue(
                        "SKU tidak ditemukan di Pricelist stok",
                        file_name=mf.name,
                        row=r,
                        sku_full=sku_full,
                        old_value=old_qty,
                        new_value="",
                    )
                    continue

                if old_qty is not None and int(old_qty) == int(new_qty):
                    continue

                row_vals[qty_col - 1] = int(new_qty)
                current_rows.append(row_vals)
                summary["rows_written"] += 1

                if len(current_rows) >= BIGSELLER_MAX_ROWS_PER_FILE:
                    flush_part()

                if progress_callback and file_rows_scanned % 3000 == 0:
                    progress_callback(
                        10 + int((file_idx - 0.25) / total_files * 75),
                        f"Proses file {file_idx}/{total_files}: {file_rows_scanned} baris file ini, total {summary['rows_scanned']} baris...",
                    )

        except Exception as e:
            add_issue(f"Gagal proses file: {e}", file_name=mf.name)
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    if progress_callback:
        progress_callback(90, "Menyimpan output terakhir...")
    flush_part()

    summary["issues_count"] = sum(issue_counter.values())
    if not output_parts:
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_ws.title = "Sheet1"
        if output_header:
            for c, val in enumerate(output_header, start=1):
                empty_ws.cell(row=1, column=c).value = val
        output_parts.append(("hasil_update_stok_bigseller_part_1.xlsx", workbook_to_bytes(empty_wb)))

    if issue_excel_mode:
        issues_bytes = make_issues_workbook(issues) if issues else None
        st.session_state["_last_issue_info"] = summarize_issues_text(issues)
    else:
        issues_bytes = None
        st.session_state["_last_issue_info"] = issue_info_from_counter()

    if len(output_parts) == 1:
        if progress_callback:
            progress_callback(97, "Menyiapkan file download...")
        return output_parts[0][1], output_parts[0][0], issues_bytes, summary

    if progress_callback:
        progress_callback(95, f"Membuat ZIP dari {len(output_parts)} part...")
    return zip_named_files(output_parts), "hasil_update_stok_bigseller.zip", issues_bytes, summary




def find_blibli_stock_columns(ws: Worksheet) -> Tuple[int, int, int, int]:
    header_row = 1
    data_start = 5
    sku_col = get_header_col_fuzzy(ws, header_row, ["Seller SKU"])
    qty_col = get_header_col_fuzzy(ws, header_row, ["Stok", "Stock"])
    sheet_col = None
    for idx, sname in enumerate(ws.parent.sheetnames):
        if ws.parent[sname] is ws:
            sheet_col = idx
            break
    if sku_col is None or qty_col is None:
        raise ValueError("Kolom Seller SKU/Stok tidak ketemu pada template Blibli.")
    return data_start, sku_col, qty_col, 0


def process_blibli_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist stok dan area/gudang...")
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue(), progress_callback=progress_callback)
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files))

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(10 + int((file_idx - 1) / total_files * 70), f"Membaca file Blibli {file_idx}/{total_files}: {mf.name}")
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
        data_start, sku_col, qty_col, _ = find_blibli_stock_columns(ws)
        changed_rows = process_marketplace_stock_sheet(
            ws,
            file_name=mf.name,
            stock_lookup=stock_lookup,
            selected_modes=selected_modes,
            chosen_areas=chosen_areas,
            chosen_gudangs=chosen_gudangs,
            zero_below=zero_below,
            zero_if_missing=zero_if_missing,
            data_start=data_start,
            sku_col=sku_col,
            qty_col=qty_col,
            summary=summary,
            issues=issues,
            progress_callback=progress_callback,
            progress_label=f"Proses stok Blibli {file_idx}/{total_files}",
        )

        if changed_rows:
            prune_worksheet_to_changed_rows(ws, changed_rows, data_start)
        else:
            append_no_changes_issue(issues, mf.name)

        output_files.append((f"hasil_update_stok_blibli_{mf.name}", workbook_to_bytes(wb)))

    if progress_callback:
        progress_callback(88, f"Membuat output Blibli ({len(output_files)} file)...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, "hasil_update_stok_blibli.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary


def find_akulaku_stock_columns(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 1
    data_start = 2
    sku_col = get_header_col_fuzzy(ws, header_row, ["SKU Produk"])
    qty_col = get_header_col_fuzzy(ws, header_row, ["Stok", "Stock"])
    if sku_col is None or qty_col is None:
        raise ValueError("Kolom SKU Produk/Stok tidak ketemu pada template Akulaku.")
    return data_start, sku_col, qty_col


def process_akulaku_stock(mass_files: List[Any], pricelist_file: Any, selected_modes: Set[str], chosen_areas: Set[str], chosen_gudangs: Set[str], zero_below: int = 0, zero_if_missing: bool = False, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist stok dan area/gudang...")
    stock_lookup, _ = build_stock_lookup_from_pricelist_bytes(pricelist_file.getvalue(), progress_callback=progress_callback)
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files))

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(10 + int((file_idx - 1) / total_files * 70), f"Membaca file Akulaku {file_idx}/{total_files}: {mf.name}")
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        data_start, sku_col, qty_col = find_akulaku_stock_columns(ws)
        changed_rows = process_marketplace_stock_sheet(
            ws,
            file_name=mf.name,
            stock_lookup=stock_lookup,
            selected_modes=selected_modes,
            chosen_areas=chosen_areas,
            chosen_gudangs=chosen_gudangs,
            zero_below=zero_below,
            zero_if_missing=zero_if_missing,
            data_start=data_start,
            sku_col=sku_col,
            qty_col=qty_col,
            summary=summary,
            issues=issues,
            progress_callback=progress_callback,
            progress_label=f"Proses stok Akulaku {file_idx}/{total_files}",
        )

        if changed_rows:
            prune_worksheet_to_changed_rows(ws, changed_rows, data_start)
        else:
            append_no_changes_issue(issues, mf.name)

        output_files.append((f"hasil_update_stok_akulaku_{mf.name}", workbook_to_bytes(wb)))

    if progress_callback:
        progress_callback(88, f"Membuat output Akulaku ({len(output_files)} file)...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, "hasil_update_stok_akulaku.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary

# ============================================================
# PRICE LOADERS
# ============================================================
def load_addon_map_generic(addon_bytes: bytes) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active
    code_candidates = ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON", "Standarisasi Kode SKU di Varian"]
    price_candidates = ["harga", "HARGA", "Price", "PRICE", "Harga"]

    header_row = None
    code_col = None
    price_col = None
    for r in range(1, 30):
        code_col = get_header_col_fuzzy(ws, r, code_candidates)
        price_col = get_header_col_fuzzy(ws, r, price_candidates)
        if code_col and price_col:
            header_row = r
            break
    if header_row is None or code_col is None or price_col is None:
        raise ValueError("Header Addon Mapping tidak ketemu. Pastikan ada kolom addon_code & harga (atau setara).")

    addon_map: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = normalize_addon_code(ws.cell(row=r, column=code_col).value)
        if not code:
            continue
        price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
        if price_raw is None:
            continue
        addon_map[code] = int(apply_multiplier_if_needed(price_raw))
    return addon_map


def find_header_row_and_cols_pricelist_fixed(ws: Worksheet, required_price_cols: List[str]) -> Tuple[int, int, Dict[str, int]]:
    sku_candidates = ["KODEBARANG", "KODE BARANG", "SKU", "SKU NO", "SKU_NO", "KODEBARANG "]

    for header_row in range(1, min(25, ws.max_row) + 1):
        sku_col = get_header_col_fuzzy(ws, header_row, sku_candidates)
        if sku_col is None:
            continue

        price_cols: Dict[str, int] = {}
        all_found = True
        for p in required_price_cols:
            col = get_header_col_fuzzy(ws, header_row, [p])
            if col is None:
                all_found = False
                break
            price_cols[p] = col

        if all_found:
            return header_row, sku_col, price_cols

    raise ValueError(
        f"Header Pricelist tidak ketemu. Pastikan ada kolom SKU/KODEBARANG dan kolom harga {required_price_cols}."
    )


def load_pricelist_price_map(pl_bytes: bytes, needed_cols: List[str]) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)
    ws = get_change_sheet(wb)
    header_row, sku_col, price_cols = find_header_row_and_cols_pricelist_fixed(ws, needed_cols)
    result: Dict[str, Dict[str, int]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku = norm_sku(ws.cell(row=r, column=sku_col).value)
        if not sku:
            continue
        result[sku] = {}
        for label, col in price_cols.items():
            raw = parse_price_cell(ws.cell(row=r, column=col).value)
            if raw is not None:
                result[sku][label] = int(apply_multiplier_if_needed(raw))
    return result


def load_pricelist_price_map_multisheet(
    pl_bytes: bytes,
    needed_cols: List[str],
    start_sheet: str = "LAPTOP",
    end_sheet: str = "ACC",
) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)

    for sname in wb.sheetnames:
        if su(sname) == "LAPTOP":
            delete_coming_block_in_laptop(wb[sname])
            break

    target_sheets = sheet_range_between(wb.sheetnames, start_sheet, end_sheet)
    result: Dict[str, Dict[str, int]] = {}
    parsed_sheets = 0

    for sname in target_sheets:
        ws = wb[sname]
        try:
            header_row, sku_col, price_cols = find_header_row_and_cols_pricelist_fixed(ws, needed_cols)
        except Exception:
            continue

        parsed_sheets += 1
        for r in range(header_row + 1, ws.max_row + 1):
            sku = norm_sku(ws.cell(row=r, column=sku_col).value)
            if not sku:
                continue
            if sku not in result:
                result[sku] = {}
            for label, col in price_cols.items():
                raw = parse_price_cell(ws.cell(row=r, column=col).value)
                if raw is not None:
                    result[sku][label] = int(apply_multiplier_if_needed(raw))

    if parsed_sheets == 0:
        raise ValueError(
            f"Tidak ada sheet pricelist yang valid pada range '{start_sheet}' s/d '{end_sheet}' untuk kolom {needed_cols}."
        )
    if not result:
        raise ValueError(
            f"Pricelist multi-sheet '{start_sheet}' s/d '{end_sheet}' terbaca, tapi data harga kosong."
        )
    return result


def compute_price_from_maps(sku_full: str, price_map: Dict[str, Dict[str, int]], addon_map: Dict[str, int], price_key: str, discount_rp: int) -> Tuple[Optional[int], str]:
    base_sku, addons = split_sku_addons(sku_full)
    base_sku = norm_sku(base_sku)
    if not base_sku:
        return None, "SKU kosong"
    pl = price_map.get(base_sku)
    if not pl:
        return None, f"Base SKU '{base_sku}' tidak ada di Pricelist"
    base_price = pl.get(price_key)
    if base_price is None:
        return None, f"Harga {price_key} kosong di Pricelist untuk SKU '{base_sku}'"

    base_price_int = int(base_price)
    if base_price_int <= 0:
        return None, f"Harga {price_key} 0/kosong di Pricelist untuk SKU '{base_sku}'"

    has_aff = False
    addon_total = 0
    for addon in addons:
        code = normalize_addon_code(addon)
        if not code:
            continue

        # Special trigger addon:
        # KODEBARANG+AFF = harga barang naik 1%, lalu discount_rp tetap dikurangkan.
        # Nilai AFF di file Addon Mapping tidak dipakai sebagai nominal tambah.
        if code == "AFF":
            has_aff = True
            continue

        if code not in addon_map:
            return None, f"Addon '{code}' tidak ada di file Addon Mapping"
        addon_total += int(addon_map.get(code, 0))

    if has_aff:
        # Integer rounding: base_price * 1.01, dibulatkan ke rupiah terdekat.
        price_before_discount = (base_price_int * 101 + 50) // 100
    else:
        price_before_discount = base_price_int

    final_price = int(price_before_discount) + addon_total - int(discount_rp)
    if final_price <= 0:
        return None, f"Harga hasil {price_key} <= 0 untuk SKU '{base_sku}'"

    if has_aff:
        return final_price, f"{price_key} + AFF 1% + addon lain - diskon"
    return final_price, f"{price_key} + addon - diskon"


# ============================================================
# PRICE PROCESSORS
# ============================================================
def _process_shopee_price_common(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, price_key: str, page_title: str, mode: str, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist harga multi-sheet...")
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    if progress_callback:
        progress_callback(8, "Membaca Addon Mapping...")
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files))

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(12 + int((file_idx - 1) / total_files * 70), f"Membaca file {page_title} {file_idx}/{total_files}: {mf.name}")
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active

        if mode == "normal":
            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "SKU Ref. No.(Optional)", "SKU Ref No Optional", "SKU Penjual"],
                    "price": ["Harga", "Harga Diskon", "Harga diskon"],
                },
                scan_rows=10,
            )
        else:
            header_row, found_cols = find_header_row_by_candidates(
                ws,
                {
                    "sku": ["SKU", "SKU Ref. No.(Optional)", "SKU Ref No Optional", "SKU Penjual"],
                    "price": ["Harga Diskon", "Harga diskon", "Harga"],
                },
                scan_rows=10,
            )

        sku_col = found_cols["sku"]
        price_col = found_cols["price"]
        data_start_fixed = header_row + 1

        changed_rows: List[int] = []
        total_rows = max(1, ws.max_row - data_start_fixed + 1)
        for r in range(data_start_fixed, ws.max_row + 1):
            if progress_callback and (r - data_start_fixed) % 2000 == 0:
                progress_tick(progress_callback, 15, 82, r - data_start_fixed, total_rows, f"Proses harga {page_title} {file_idx}/{total_files}: {r - data_start_fixed} baris...")
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue

            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, price_key, discount_rp)

            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({
                    "file": mf.name,
                    "row": r,
                    "sku_full": sku_full,
                    "old_value": old_price,
                    "new_value": "",
                    "reason": reason,
                })
                continue

            if old_price is not None and int(old_price) == int(new_price):
                continue

            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            # Lebih cepat daripada delete row satu-per-satu.
            # Ambil hanya baris berubah, bersihkan blok data sekali, lalu tulis ulang hasilnya.
            max_col = ws.max_column
            changed_row_values = [
                [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
                for row_idx in changed_rows
            ]
            if ws.max_row >= data_start_fixed:
                ws.delete_rows(data_start_fixed, ws.max_row - data_start_fixed + 1)
            for out_r, row_vals in enumerate(changed_row_values, start=data_start_fixed):
                for c, val in enumerate(row_vals, start=1):
                    ws.cell(row=out_r, column=c).value = val
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})

        if progress_callback:
            progress_callback(12 + int(file_idx / total_files * 70), f"Menyimpan output {page_title} {file_idx}/{total_files}...")
        output_files.append((f"hasil_{page_title.lower().replace(' ', '_')}_{mf.name}", workbook_to_bytes(wb)))

    if progress_callback:
        progress_callback(90, f"Menyiapkan download {page_title}...")
    if progress_callback:
        progress_callback(90, f"Menyiapkan download {page_title}...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, f"hasil_{page_title.lower().replace(' ', '_')}.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary




# PRICE PROCESSORS - GROUPED BY TYPE
# ============================================================
# Normal Price
def process_shopee_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, progress_callback=None):
    return _process_shopee_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "M4",
        "Harga Normal Shopee",
        "normal",
        progress_callback,
    )


def process_tiktokshop_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist harga TikTokShop...")
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files))

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(10 + int((file_idx - 1) / total_files * 75), f"Membaca file TikTokShop {file_idx}/{total_files}: {mf.name}")
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        try:
            data_start, sku_col, price_col, _ = find_tiktok_style_columns(ws, need_price=True, scan_rows=10, data_offset=3)
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Header mass update TikTokShop tidak sesuai: {e}"})
            continue
        if price_col is None:
            issues.append({"file": mf.name, "reason": "Header mass update TikTokShop tidak sesuai."})
            continue

        changed_rows: List[int] = []
        total_rows = max(1, ws.max_row - data_start + 1)
        for r in range(data_start, ws.max_row + 1):
            if progress_callback and (r - data_start) % 2000 == 0:
                progress_tick(progress_callback, 15, 82, r - data_start, total_rows, f"Memproses baris TikTokShop {file_idx}/{total_files}: {r - data_start} baris...")
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M3", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and int(old_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            rewrite_worksheet_with_rows(ws, data_start, collect_changed_row_values(ws, changed_rows))
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})
        output_files.append((f"hasil_harga_normal_tiktokshop_{mf.name}", workbook_to_bytes(wb)))

    if progress_callback:
        progress_callback(90, "Menyiapkan download Harga Normal TikTokShop...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, "hasil_harga_normal_tiktokshop.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary


def process_mwh_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, price_key: str, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist harga Mwh...")
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files), include_unchanged=True)

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        try:
            if progress_callback:
                progress_callback(10 + int((file_idx - 1) / total_files * 75), f"Membaca file Mwh {file_idx}/{total_files}: {mf.name}")
            wb = load_workbook(io.BytesIO(mf.getvalue()))
            ws = wb.active
            data_start, sku_col, price_col, _ = find_tiktok_style_columns(ws, need_price=True, scan_rows=10, data_offset=3)
            if price_col is None:
                issues.append({"file": mf.name, "reason": "Header mass update Mwh tidak sesuai."})
                continue

            for r in range(data_start, ws.max_row + 1):
                sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
                if not sku_full:
                    continue
                summary["rows_scanned"] += 1
                old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
                new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, price_key, discount_rp)
                if new_price is None:
                    summary["rows_unmatched"] += 1
                    issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                    continue
                if old_price is not None and int(old_price) == int(new_price):
                    summary["rows_unchanged"] += 1
                    continue
                safe_set_cell_value(ws, r, price_col, int(new_price))
                summary["rows_written"] += 1

            output_files.append((f"hasil_harga_normal_mwh_{mf.name}", workbook_to_bytes(wb)))
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})

    if summary["rows_written"] == 0 and not issues:
        issues.append({"file": "", "reason": "Tidak ada baris berubah / tidak ada SKU yang match."})

    if progress_callback:
        progress_callback(90, "Menyiapkan download Harga Normal Mwh...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, "hasil_harga_normal_mwh.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary


def _process_powemerchant_price_common(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, page_title: str, progress_callback=None):
    if progress_callback:
        progress_callback(3, f"Membaca Pricelist {page_title}...")
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files))

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(10 + int((file_idx - 1) / total_files * 75), f"Memproses {page_title} {file_idx}/{total_files}: {mf.name}")
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        try:
            data_start, sku_col, price_col, _ = find_tiktok_style_columns(ws, need_price=True, scan_rows=10, data_offset=3)
        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Header mass update PowerMerchant tidak sesuai: {e}"})
            continue
        if price_col is None:
            issues.append({"file": mf.name, "reason": "Header mass update PowerMerchant tidak sesuai."})
            continue

        changed_rows: List[int] = []
        total_rows = max(1, ws.max_row - data_start + 1)
        for r in range(data_start, ws.max_row + 1):
            if progress_callback and (r - data_start) % 2000 == 0:
                progress_tick(progress_callback, 15, 82, r - data_start, total_rows, f"Memproses baris {page_title} {file_idx}/{total_files}: {r - data_start} baris...")
            sku_full = s_clean(ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            old_price = parse_price_cell(ws.cell(row=r, column=price_col).value)
            new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, "M4", discount_rp)
            if new_price is None:
                summary["rows_unmatched"] += 1
                issues.append({"file": mf.name, "row": r, "sku_full": sku_full, "old_value": old_price, "new_value": "", "reason": reason})
                continue
            if old_price is not None and int(old_price) == int(new_price):
                continue
            safe_set_cell_value(ws, r, price_col, int(new_price))
            changed_rows.append(r)
            summary["rows_written"] += 1

        if changed_rows:
            rewrite_worksheet_with_rows(ws, data_start, collect_changed_row_values(ws, changed_rows))
        else:
            issues.append({"file": mf.name, "reason": "Tidak ada baris berubah pada file ini."})
        output_files.append((f"hasil_{page_title.lower().replace(' ', '_')}_{mf.name}", workbook_to_bytes(wb)))

    if progress_callback:
        progress_callback(90, f"Menyiapkan download {page_title}...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, f"hasil_{page_title.lower().replace(' ', '_')}.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary


def process_powemerchant_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, progress_callback=None):
    return _process_powemerchant_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "Harga Normal PowerMerchant",
        progress_callback,
    )


def process_bigseller_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, price_key: str, progress_callback=None):
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    summary = init_summary(len(mass_files))
    output_parts: List[Tuple[str, bytes]] = []
    current_rows: List[List[Any]] = []
    current_part = 1
    output_header: List[Any] = []
    header_len = 0

    def flush_part():
        nonlocal current_rows, current_part, output_parts, output_header, header_len
        if not current_rows:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c, val in enumerate(output_header, start=1):
            ws.cell(row=1, column=c).value = val
        for r_idx, row_vals in enumerate(current_rows, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws.cell(row=r_idx, column=c).value = val
        output_parts.append((f"hasil_harga_normal_bigseller_part_{current_part}.xlsx", workbook_to_bytes(wb)))
        current_rows = []
        current_part += 1

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        wb = None
        try:
            if progress_callback:
                progress_callback(10 + int((file_idx - 1) / total_files * 80), f"Membaca BigSeller {file_idx}/{total_files}...")
            wb = load_workbook(io.BytesIO(mf.getvalue()), read_only=True, data_only=False)
            ws = wb.worksheets[0]
            reset_readonly_dimensions_if_needed(ws)

            header_row, found_cols = find_bigseller_omnichannel_columns(ws, mode="price")

            sku_col = found_cols["sku"]
            harga_col = found_cols["price"]

            if not output_header:
                output_header = get_row_values_readonly(ws, header_row)
                header_len = max(len(output_header), sku_col, harga_col)
                if len(output_header) < header_len:
                    output_header.extend([None] * (header_len - len(output_header)))

            for offset, row_vals_raw in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                row_vals = list(row_vals_raw[:header_len])
                if len(row_vals) < header_len:
                    row_vals.extend([None] * (header_len - len(row_vals)))

                sku_full = s_clean(row_vals[sku_col - 1] if len(row_vals) >= sku_col else None)
                if not sku_full:
                    continue

                summary["rows_scanned"] += 1
                old_price = parse_price_cell(row_vals[harga_col - 1] if len(row_vals) >= harga_col else None)
                new_price, reason = compute_price_from_maps(sku_full, price_map, addon_map, price_key, discount_rp)

                if new_price is None:
                    summary["rows_unmatched"] += 1
                    issues.append({
                        "file": mf.name,
                        "row": offset,
                        "sku_full": sku_full,
                        "old_value": old_price,
                        "new_value": "",
                        "reason": reason,
                    })
                    continue

                if old_price is not None and int(old_price) == int(new_price):
                    continue

                row_vals[harga_col - 1] = int(new_price)
                current_rows.append(row_vals)
                summary["rows_written"] += 1

                if len(current_rows) >= BIGSELLER_MAX_ROWS_PER_FILE:
                    flush_part()

                if progress_callback and summary["rows_scanned"] % 2000 == 0:
                    progress_callback(10 + int((file_idx - 0.5) / total_files * 80), f"Memproses BigSeller {file_idx}/{total_files}: {summary['rows_scanned']} baris...")

        except Exception as e:
            issues.append({"file": mf.name, "reason": f"Gagal proses file: {e}"})
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    flush_part()
    summary["issues_count"] = len(issues)
    if not output_parts:
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_ws.title = "Sheet1"
        if output_header:
            for c, val in enumerate(output_header, start=1):
                empty_ws.cell(row=1, column=c).value = val
        output_parts.append(("hasil_harga_normal_bigseller_part_1.xlsx", workbook_to_bytes(empty_wb)))
    result_bytes, result_name = package_output_files(output_parts, "hasil_harga_normal_bigseller.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary




def find_blibli_price_columns(ws: Worksheet) -> Tuple[int, int, int, int]:
    header_row = 1
    data_start = 5
    sku_col = get_header_col_fuzzy(ws, header_row, ["Seller SKU"])
    price_col = get_header_col_fuzzy(ws, header_row, ["Harga (Rp)", "Harga"])
    sale_price_col = get_header_col_fuzzy(ws, header_row, ["Harga Penjualan (Rp)", "Harga Penjualan"])
    if sku_col is None or price_col is None or sale_price_col is None:
        raise ValueError("Kolom Seller SKU/Harga/Harga Penjualan tidak ketemu pada template Blibli.")
    return data_start, sku_col, price_col, sale_price_col


def process_blibli_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist harga Blibli...")
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files))

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(10 + int((file_idx - 1) / total_files * 75), f"Membaca file Blibli {file_idx}/{total_files}: {mf.name}")
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
        data_start, sku_col, price_col, sale_price_col = find_blibli_price_columns(ws)
        changed_rows = process_marketplace_price_sheet(
            ws,
            file_name=mf.name,
            price_map=price_map,
            addon_map=addon_map,
            discount_rp=discount_rp,
            price_key="M3",
            data_start=data_start,
            sku_col=sku_col,
            target_price_cols=[price_col, sale_price_col],
            summary=summary,
            issues=issues,
            progress_callback=progress_callback,
            progress_label=f"Proses harga Blibli {file_idx}/{total_files}",
        )

        if changed_rows:
            prune_worksheet_to_changed_rows(ws, changed_rows, data_start)
        else:
            append_no_changes_issue(issues, mf.name)

        output_files.append((f"hasil_harga_normal_blibli_{mf.name}", workbook_to_bytes(wb)))

    if progress_callback:
        progress_callback(90, "Menyiapkan download Harga Normal Blibli...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, "hasil_harga_normal_blibli.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary


def find_akulaku_price_columns(ws: Worksheet) -> Tuple[int, int, int]:
    header_row = 1
    data_start = 2
    sku_col = get_header_col_fuzzy(ws, header_row, ["SKU Produk"])
    price_col = get_header_col_fuzzy(ws, header_row, ["Harga", "Price"])
    if sku_col is None or price_col is None:
        raise ValueError("Kolom SKU Produk/Harga tidak ketemu pada template Akulaku.")
    return data_start, sku_col, price_col


def process_akulaku_price(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist harga Akulaku...")
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3", "M4"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = init_summary(len(mass_files))

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(10 + int((file_idx - 1) / total_files * 75), f"Membaca file Akulaku {file_idx}/{total_files}: {mf.name}")
        wb = load_workbook(io.BytesIO(mf.getvalue()))
        ws = wb.active
        data_start, sku_col, price_col = find_akulaku_price_columns(ws)
        changed_rows = process_marketplace_price_sheet(
            ws,
            file_name=mf.name,
            price_map=price_map,
            addon_map=addon_map,
            discount_rp=discount_rp,
            price_key="M3",
            data_start=data_start,
            sku_col=sku_col,
            target_price_cols=[price_col],
            summary=summary,
            issues=issues,
            progress_callback=progress_callback,
            progress_label=f"Proses harga Akulaku {file_idx}/{total_files}",
        )

        if changed_rows:
            prune_worksheet_to_changed_rows(ws, changed_rows, data_start)
        else:
            append_no_changes_issue(issues, mf.name)

        output_files.append((f"hasil_harga_normal_akulaku_{mf.name}", workbook_to_bytes(wb)))

    if progress_callback:
        progress_callback(90, "Menyiapkan download Harga Normal Akulaku...")
    summary["issues_count"] = len(issues)
    result_bytes, result_name = package_output_files(output_files, "hasil_harga_normal_akulaku.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary

# ============================================================
# SUBMIT CAMPAIGN PROCESSORS
# ============================================================

# Discount Price
def process_shopee_discount(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, progress_callback=None):
    return _process_shopee_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "M4",
        "Harga Coret Shopee",
        "coret",
        progress_callback,
    )


def process_tiktokshop_discount(input_file: Any, pricelist_file: Any, addon_file: Any, discount_rp: int, only_changed: bool = True, progress_callback=None):
    if progress_callback:
        progress_callback(3, "Membaca Pricelist harga coret TikTokShop...")
    price_map = load_pricelist_price_map_multisheet(pricelist_file.getvalue(), ["M3"])
    addon_map = load_addon_map_generic(addon_file.getvalue())
    wb_in = load_workbook(io.BytesIO(input_file.getvalue()), data_only=True)
    ws_in = wb_in.active

    out_wb = Workbook()
    ws_out = out_wb.active
    ws_out.title = "Sheet1"
    headers = [
        "Product_id (wajib)",
        "SKU_id (wajib)",
        "Harga Penawaran (wajib)",
        "Total Stok Promosi (opsional)\n1. Total Stok Promosi≤ Stok \n2. Jika tidak diisi artinya tidak terbatas",
        "Batas Pembelian (opsional)\n1. 1 ≤ Batas pembelian≤ 99\n2. Jika tidak diisi artinya tidak terbatas",
    ]
    for i, h in enumerate(headers, start=1):
        ws_out.cell(row=1, column=i).value = h

    issues: List[Dict[str, Any]] = []
    summary = {"files_total": 1, "rows_scanned": 0, "rows_written": 0, "rows_unmatched": 0, "issues_count": 0}
    row_out = 2

    total_rows = max(1, ws_in.max_row - 5)
    for r in range(6, ws_in.max_row + 1):
        if progress_callback and (r - 6) % 2000 == 0:
            progress_tick(progress_callback, 15, 85, r - 6, total_rows, f"Memproses harga coret TikTokShop: {r - 6} baris...")
        product_id = parse_number_like_id(ws_in.cell(row=r, column=1).value)
        sku_id = parse_number_like_id(ws_in.cell(row=r, column=4).value)
        old_price = parse_price_cell(ws_in.cell(row=r, column=6).value)
        stock = to_int_or_none(ws_in.cell(row=r, column=7).value)
        seller_sku = s_clean(ws_in.cell(row=r, column=8).value or ws_in.cell(row=r, column=5).value)
        if not seller_sku:
            continue
        summary["rows_scanned"] += 1
        new_price, reason = compute_price_from_maps(seller_sku, price_map, addon_map, "M3", discount_rp)
        if new_price is None:
            summary["rows_unmatched"] += 1
            issues.append({"file": input_file.name, "row": r, "sku_full": seller_sku, "old_value": old_price, "new_value": "", "reason": reason})
            continue
        if only_changed and old_price is not None and int(old_price) == int(new_price):
            continue
        ws_out.cell(row=row_out, column=1).value = product_id
        ws_out.cell(row=row_out, column=2).value = sku_id
        ws_out.cell(row=row_out, column=3).value = new_price
        ws_out.cell(row=row_out, column=4).value = stock if stock is not None else ""
        row_out += 1
        summary["rows_written"] += 1

    if progress_callback:
        progress_callback(92, "Menyimpan output Harga Coret TikTokShop...")
    summary["issues_count"] = len(issues)
    return workbook_to_bytes(out_wb), "hasil_harga_coret_tiktokshop.xlsx", issues_workbook_or_none(issues), summary


def process_powemerchant_discount(mass_files: List[Any], pricelist_file: Any, addon_file: Any, discount_rp: int, progress_callback=None):
    return _process_powemerchant_price_common(
        mass_files,
        pricelist_file,
        addon_file,
        discount_rp,
        "Harga Coret PowerMerchant",
        progress_callback,
    )




def find_col_by_contains_on_row(
    ws: Worksheet,
    header_row: int,
    keywords: List[str],
    *,
    exclude_keywords: Optional[List[str]] = None,
) -> Optional[int]:
    return find_col_contains_in_row(ws, header_row, keywords, exclude_keywords=exclude_keywords)


def find_header_row_with_columns(
    ws: Worksheet,
    required: Dict[str, Tuple[List[str], Optional[List[str]]]],
    *,
    scan_rows: int = 10,
) -> Tuple[int, Dict[str, int]]:
    """Cari header secara global/contains untuk beberapa kolom wajib."""
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        found: Dict[str, int] = {}
        ok = True
        for key, cfg in required.items():
            keywords, excludes = cfg
            col = find_col_by_contains_on_row(ws, r, keywords, exclude_keywords=excludes)
            if col is None:
                ok = False
                break
            found[key] = col
        if ok:
            return r, found
    raise ValueError("Header tidak ditemukan. Pastikan kolom wajib tersedia.")


def get_optional_col_contains(ws: Worksheet, header_row: int, keywords: List[str], *, exclude_keywords: Optional[List[str]] = None) -> Optional[int]:
    return find_col_by_contains_on_row(ws, header_row, keywords, exclude_keywords=exclude_keywords)


def get_row_values(ws: Worksheet, row: int, max_col: Optional[int] = None) -> List[Any]:
    max_col = max_col or ws.max_column
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]



def process_shopee_campaign(mass_file: Any, campaign_file: Any, progress_callback=None):
    issues: List[Dict[str, Any]] = []
    summary = init_summary(1)

    try:
        if progress_callback:
            progress_callback(5, "Membaca file mass update Shopee...")
        mass_wb = load_workbook(io.BytesIO(mass_file.getvalue()), data_only=False)
        mass_ws = mass_wb.active

        mass_header_row, mass_cols = find_header_row_with_columns(
            mass_ws,
            {
                "product_id": (["KODE PRODUK", "PRODUCT ID"], None),
                "product_name": (["NAMA PRODUK", "PRODUCT NAME"], None),
                "variation_id": (["KODE VARIASI", "VARIATION ID"], None),
                "variation_name": (["NAMA VARIASI", "VARIATION NAME"], None),
                "sku": (["SKU"], ["SKU INDUK", "PARENT SKU", "SKU ID", "ID SKU"]),
            },
            scan_rows=10,
        )

        price_col = get_optional_col_contains(mass_ws, mass_header_row, ["HARGA", "PRICE"])
        stock_col = get_optional_col_contains(mass_ws, mass_header_row, ["STOK", "STOCK", "QUANTITY", "KUANTITAS", "JUMLAH"])
        mass_data_start = mass_header_row + 4 if mass_header_row <= 3 else mass_header_row + 1

        selected_by_variant: Dict[str, Dict[str, Any]] = {}
        total_mass_rows = max(1, mass_ws.max_row - mass_data_start + 1)
        for r in range(mass_data_start, mass_ws.max_row + 1):
            if progress_callback and (r - mass_data_start) % 2000 == 0:
                progress_tick(progress_callback, 15, 45, r - mass_data_start, total_mass_rows, f"Scan SKU campaign Shopee: {r - mass_data_start} baris...")
            sku_full = s_clean(mass_ws.cell(row=r, column=mass_cols["sku"]).value)
            if not sku_full:
                continue
            summary["rows_scanned"] += 1
            if su(sku_full) != "ND-ALL-CAMPAIGN":
                summary["rows_unmatched"] += 1
                continue

            variation_id = parse_number_like_id(mass_ws.cell(row=r, column=mass_cols["variation_id"]).value)
            if not variation_id:
                summary["rows_unmatched"] += 1
                issues.append({
                    "file": mass_file.name,
                    "row": r,
                    "sku_full": sku_full,
                    "reason": "SKU ND-ALL-CAMPAIGN ditemukan, tapi Kode Variasi kosong.",
                })
                continue

            selected_by_variant[variation_id] = {
                "product_id": parse_number_like_id(mass_ws.cell(row=r, column=mass_cols["product_id"]).value),
                "product_name": s_clean(mass_ws.cell(row=r, column=mass_cols["product_name"]).value),
                "variation_id": variation_id,
                "variation_name": s_clean(mass_ws.cell(row=r, column=mass_cols["variation_name"]).value),
                "sku": sku_full,
                "price": parse_price_cell(mass_ws.cell(row=r, column=price_col).value) if price_col else None,
                "stock": to_int_or_none(mass_ws.cell(row=r, column=stock_col).value) if stock_col else None,
            }

        if not selected_by_variant:
            issues.append({
                "file": mass_file.name,
                "reason": "Tidak ada SKU yang sama persis dengan ND-ALL-CAMPAIGN di file mass update.",
            })

        if progress_callback:
            progress_callback(50, "Membaca file campaign Shopee...")
        campaign_wb = load_workbook(io.BytesIO(campaign_file.getvalue()))
        campaign_ws = campaign_wb.active
        campaign_header_row, campaign_cols = find_header_row_with_columns(
            campaign_ws,
            {
                "variation_id": (["KODE VARIASI", "KODE KRITERIA", "VARIATION ID", "VARIATION CODE"], None),
            },
            scan_rows=10,
        )
        campaign_data_start = campaign_header_row + 3 if campaign_header_row <= 1 else campaign_header_row + 1

        existing_data_rows = []
        for r in range(campaign_data_start, campaign_ws.max_row + 1):
            variant = parse_number_like_id(campaign_ws.cell(row=r, column=campaign_cols["variation_id"]).value)
            if variant:
                existing_data_rows.append(r)

        if existing_data_rows:
            keep_rows_data = []
            for r in range(campaign_data_start, campaign_ws.max_row + 1):
                variant = parse_number_like_id(campaign_ws.cell(row=r, column=campaign_cols["variation_id"]).value)
                if variant in selected_by_variant:
                    keep_rows_data.append(get_row_values(campaign_ws, r))
                    summary["rows_written"] += 1
            rewrite_worksheet_with_rows(campaign_ws, campaign_data_start, keep_rows_data)
            if summary["rows_written"] == 0:
                issues.append({
                    "file": campaign_file.name,
                    "reason": "Kode Variasi campaign tidak ada yang cocok dengan SKU ND-ALL-CAMPAIGN dari mass update.",
                })
        else:
            # Jika file campaign masih berupa template kosong, isi dari data mass update yang SKU-nya ND-ALL-CAMPAIGN.
            optional_cols = {
                "product_name": get_optional_col_contains(campaign_ws, campaign_header_row, ["NAMA PRODUK", "PRODUCT NAME"]),
                "product_id": get_optional_col_contains(campaign_ws, campaign_header_row, ["KODE PRODUK", "PRODUCT ID"]),
                "variation_name": get_optional_col_contains(campaign_ws, campaign_header_row, ["NAMA VARIASI", "VARIATION NAME"]),
                "original_price": get_optional_col_contains(campaign_ws, campaign_header_row, ["HARGA AWAL", "ORIGINAL PRICE"]),
                "current_price": get_optional_col_contains(campaign_ws, campaign_header_row, ["HARGA SAAT INI", "CURRENT PRICE"]),
                "discount_price": get_optional_col_contains(campaign_ws, campaign_header_row, ["HARGA DISKON", "DISCOUNT PRICE"]),
            }
            out_r = campaign_data_start
            for rec in selected_by_variant.values():
                if optional_cols["product_name"]:
                    campaign_ws.cell(row=out_r, column=optional_cols["product_name"]).value = rec["product_name"]
                if optional_cols["product_id"]:
                    campaign_ws.cell(row=out_r, column=optional_cols["product_id"]).value = rec["product_id"]
                if optional_cols["variation_name"]:
                    campaign_ws.cell(row=out_r, column=optional_cols["variation_name"]).value = rec["variation_name"]
                campaign_ws.cell(row=out_r, column=campaign_cols["variation_id"]).value = rec["variation_id"]
                if rec.get("price") is not None:
                    for key in ("original_price", "current_price", "discount_price"):
                        if optional_cols[key]:
                            campaign_ws.cell(row=out_r, column=optional_cols[key]).value = int(rec["price"])
                out_r += 1
                summary["rows_written"] += 1

        if summary["rows_written"] == 0 and not issues:
            issues.append({"file": campaign_file.name, "reason": "Tidak ada baris yang masuk hasil Submit Campaign Shopee."})

        if progress_callback:
            progress_callback(92, "Menyimpan output Submit Campaign Shopee...")
        summary["issues_count"] = len(issues)
        return workbook_to_bytes(campaign_wb), f"hasil_submit_campaign_shopee_{campaign_file.name}", issues_workbook_or_none(issues), summary

    except Exception as e:
        issues.append({"file": getattr(mass_file, "name", ""), "reason": f"Gagal proses file: {e}"})
        summary["issues_count"] = len(issues)
        empty_wb = Workbook()
        return workbook_to_bytes(empty_wb), "hasil_submit_campaign_shopee.xlsx", issues_workbook_or_none(issues), summary


def process_tiktokshop_campaign(mass_files: List[Any], progress_callback=None):
    issues: List[Dict[str, Any]] = []
    output_files: List[Tuple[str, bytes]] = []
    summary = {
        "files_total": len(mass_files),
        "rows_scanned": 0,
        "rows_written": 0,
        "rows_unmatched": 0,
        "issues_count": 0,
    }

    total_files = max(1, len(mass_files))
    for file_idx, mf in enumerate(mass_files, start=1):
        if progress_callback:
            progress_callback(10 + int((file_idx - 1) / total_files * 75), f"Memfilter campaign TikTokShop {file_idx}/{total_files}: {mf.name}")
        src_wb = load_workbook(io.BytesIO(mf.getvalue()))
        src_ws = src_wb.active

        header_row = 2
        data_start = 3

        sku_col = get_header_col_fuzzy(src_ws, header_row, [
            "SKU Name",
            "Nama SKU",
            "Seller SKU",
            "SKU Penjual",
        ])

        if sku_col is None:
            issues.append({
                "file": mf.name,
                "reason": "Header Submit Campaign tidak sesuai. Pastikan row 2 berisi header SKU Name.",
            })
            continue

        filtered_rows_data: List[List[Any]] = []

        for r in range(data_start, src_ws.max_row + 1):
            sku_full = s_clean(src_ws.cell(row=r, column=sku_col).value)
            if not sku_full:
                continue

            summary["rows_scanned"] += 1
            if "ND-ALL-CAMPAIGN" in su(sku_full):
                row_vals = [src_ws.cell(row=r, column=c).value for c in range(1, src_ws.max_column + 1)]
                filtered_rows_data.append(row_vals)
                summary["rows_written"] += 1
            else:
                summary["rows_unmatched"] += 1

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = src_ws.title

        for c in range(1, src_ws.max_column + 1):
            out_ws.cell(row=1, column=c, value=src_ws.cell(row=1, column=c).value)
            out_ws.cell(row=2, column=c, value=src_ws.cell(row=2, column=c).value)

        if filtered_rows_data:
            for out_r, row_vals in enumerate(filtered_rows_data, start=data_start):
                for c, val in enumerate(row_vals, start=1):
                    out_ws.cell(row=out_r, column=c, value=val)
        else:
            issues.append({
                "file": mf.name,
                "reason": "Tidak ada baris dengan SKU Name yang mengandung 'ND-ALL-CAMPAIGN'.",
            })

        output_files.append((f"hasil_submit_campaign_tiktokshop_{mf.name}", workbook_to_bytes(out_wb)))

    if progress_callback:
        progress_callback(90, "Menyiapkan download Submit Campaign TikTokShop...")
    summary["issues_count"] = len(issues)

    result_bytes, result_name = package_output_files(output_files, "hasil_submit_campaign_tiktokshop.zip")
    return result_bytes, result_name, issues_workbook_or_none(issues), summary


# ============================================================
# UI HELPERS
# ============================================================
def page_header(title: str, desc: str, requirements: List[str]):
    st.title(title)
    st.caption(desc)
    with st.expander("Kebutuhan File", expanded=True):
        for item in requirements:
            st.write(f"- {item}")


def render_stock_controls(area_key_prefix: str, pricelist_file: Any, mode_key: str, loaded_areas_key: str, load_button_key: str):
    selected_modes = set(st.multiselect(
        "Mode Stok",
        ["Stok Nasional (TOT)", "Default", "Area", "Gudang"],
        default=["Default"],
        key=mode_key,
        help="Selain Stok Nasional (TOT), mode stok bisa dipilih lebih dari 1. Jika pilih TOT bersamaan dengan mode lain, hasil stok akan pakai TOT.",
    ))
    zero_below = st.number_input("Stok < angka ini jadi 0", min_value=0, value=0, step=1, key=f"{area_key_prefix}_zero_below")
    zero_if_missing = st.toggle(
        "SKU tidak ditemukan di Pricelist = 0",
        value=False,
        key=f"{area_key_prefix}_zero_if_missing",
        help="Jika aktif, SKU yang tidak ditemukan di Pricelist stok akan diisi 0. Jika nonaktif, SKU tetap di-skip.",
    )

    needs_lookup_data = bool(selected_modes & {"Default", "Area", "Gudang"})

    if st.button("Load Data Area / Gudang", key=load_button_key):
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
        else:
            progress = st.progress(0, text="Menyiapkan Pricelist stok...")
            status = st.empty()
            try:
                def area_progress_callback(pct, msg):
                    pct = max(0, min(100, int(pct)))
                    progress.progress(pct, text=msg)
                    status.caption(msg)

                area_progress_callback(3, "Membuka Pricelist stok...")
                _, meta = build_stock_lookup_from_pricelist_bytes(
                    pricelist_file.getvalue(),
                    progress_callback=area_progress_callback,
                    progress_start=3,
                    progress_end=85,
                )
                progress.progress(90, text="Menyusun daftar area dan gudang...")
                st.session_state[loaded_areas_key] = meta
                default_gudangs = get_default_tongle_gudangs(meta.get("gudang_options", []))
                if default_gudangs:
                    st.session_state[f"{area_key_prefix}_gudangs"] = default_gudangs
                progress.progress(100, text="Selesai load area/gudang.")
                status.caption("Area dan gudang siap dipilih.")
                st.success(
                    f"Data berhasil dimuat: {len(meta.get('area_options', []))} area, {len(meta.get('gudang_options', []))} kombinasi area-gudang"
                )
            except Exception as e:
                progress.progress(100, text="Gagal load area/gudang.")
                status.caption("Gagal membaca Pricelist stok.")
                st.error(f"Gagal load data area / gudang: {e}")

    meta = st.session_state.get(loaded_areas_key, {}) or {}
    areas = meta.get("area_options", [])
    gudangs = meta.get("gudang_options", [])
    default_gudangs = st.session_state.get(f"{area_key_prefix}_gudangs", get_default_tongle_gudangs(gudangs))

    if "Default" in selected_modes:
        st.caption("Mode Default mengunci gudang: JKT-1A, JKT-3B, JKT-3C, JKT-4B")

    chosen_areas: Set[str] = set()
    chosen_gudangs: Set[str] = set()
    if "Area" in selected_modes:
        chosen_areas = set(st.multiselect(
            "Pilih Area",
            areas,
            default=st.session_state.get(f"{area_key_prefix}_areas", []),
            key=f"{area_key_prefix}_areas",
        ))

    if "Gudang" in selected_modes:
        chosen_gudangs = set(st.multiselect(
            "Pilih Gudang (format Area-Gudang)",
            gudangs,
            default=default_gudangs if "Default" in selected_modes else st.session_state.get(f"{area_key_prefix}_gudangs", []),
            key=f"{area_key_prefix}_gudangs",
        ))

    process_disabled = False
    if not selected_modes:
        process_disabled = True
    elif needs_lookup_data and not (areas or gudangs):
        process_disabled = True
    elif "Area" in selected_modes and not chosen_areas:
        process_disabled = True
    elif "Gudang" in selected_modes and not chosen_gudangs:
        process_disabled = True

    return selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, process_disabled


def validate_mass_uploads(mass_files: List[Any]) -> Optional[str]:
    if not mass_files:
        return "Upload file mass update minimal 1 file."
    if len(mass_files) > MAX_MASS_FILES:
        return f"Maksimal {MAX_MASS_FILES} file per proses."
    if total_upload_size_mb(mass_files) > MAX_TOTAL_UPLOAD_MB:
        return f"Total upload melebihi {MAX_TOTAL_UPLOAD_MB} MB."
    return None


def run_with_loading(process_fn, loading_text: str = "Memproses..."):
    # Progress global untuk SEMUA fitur lama.
    # Catatan: fitur lama belum punya progress per-baris, jadi progress dibuat per tahap umum.
    # Progress sengaja tidak di-empty langsung agar tidak terlihat hilang lalu muncul loading ulang
    # saat Streamlit sedang render summary/download button.
    progress = st.progress(0, text=loading_text)
    status = st.empty()
    try:
        progress.progress(5, text=loading_text)
        status.caption("Menyiapkan data...")
        progress.progress(20, text="Membaca file dan memproses data...")
        result = process_fn()
        progress.progress(95, text="Menyusun output dan tombol download...")
        status.caption("Menyusun hasil...")
        progress.progress(100, text="Selesai. Hasil siap didownload.")
        status.caption("Selesai.")
        return result
    except Exception:
        progress.progress(100, text="Proses gagal. Cek pesan error di bawah.")
        status.caption("Gagal memproses.")
        raise


def run_with_loading_callback(process_fn, loading_text: str = "Memproses..."):
    # Progress detail untuk fitur yang support progress_callback.
    # Dipakai di BigSeller dan fitur lain yang nanti di-upgrade per processor.
    progress = st.progress(0, text=loading_text)
    status = st.empty()

    def update_progress(value: int, text: Optional[str] = None):
        value = max(0, min(100, int(value)))
        msg = text or loading_text
        progress.progress(value, text=msg)
        status.caption(msg)

    try:
        update_progress(5, loading_text)
        result = process_fn(update_progress)
        update_progress(100, "Selesai. Hasil siap didownload.")
        return result
    except Exception:
        update_progress(100, "Proses gagal. Cek pesan error di bawah.")
        raise




# ============================================================
# EMBEDDED PROCESSORS
# ============================================================

def render_analisa_penjualan_app():
    """Embedded processor: Analisa Penjualan"""
    import io
    import re
    from datetime import date, timedelta
    from typing import Optional, Tuple, Dict, List

    import numpy as np
    import pandas as pd
    import importlib.util
    import importlib
    px = importlib.import_module("plotly.express") if importlib.util.find_spec("plotly.express") else None
    import streamlit as st


    APP_TITLE = "Analisa Penjualan"


    # =========================
    # CSS (Light + Badge + Clip + SKU column widths + Smaller headers)
    # =========================
    st.markdown(
        """
    <style>
    .block-container { padding-top: 0.8rem; padding-bottom: 1rem; }
    section[data-testid="stSidebar"] .block-container { padding-top: 0.6rem; }

    /* Header (large -> small) */
    .header-wrap { display:flex; align-items:center; gap:12px; margin: 0.2rem 0 0.7rem 0; }
    .header-title { font-size: 22px; font-weight: 900; margin:0; line-height:1.2; color: #111827; }

    /* Make section headers smaller */
    h2 { font-size: 18px !important; }
    h3 { font-size: 16px !important; }

    /* Cards */
    .kpi-grid { display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 12px; }

    .card {
      border: 1px solid rgba(17,24,39,0.08);
      background: #ffffff;
      border-radius: 14px;
      padding: 12px 12px 10px 12px;
      box-shadow: 0 10px 20px rgba(17,24,39,0.04);
    }
    .card-title { font-size: 12px; color: rgba(17,24,39,0.7); margin-bottom: 6px; }
    .card-value { font-size: 18px; font-weight: 900; line-height: 1.15; color: #111827; }
    .card-sub { font-size: 11px; margin-top: 6px; font-weight: 800; }

    .pos { color: #16a34a; }
    .neg { color: #dc2626; }
    .na  { color: #64748b; }

    hr { border: none; border-top: 1px solid rgba(17,24,39,0.10); margin: 14px 0; }

    /* Small headings (for long titles) */
    .small-h { font-size: 16px; font-weight: 900; margin: 0 0 6px 0; }
    .small-h .muted { color: rgba(17,24,39,0.6); font-weight: 800; }

    /* HTML table styling + CLIP */
    table {
      width: 100%;
      border-collapse: separate;
      border-spacing: 0;
      overflow: hidden;
      border-radius: 12px;
      border: 1px solid rgba(17,24,39,0.08);
      background: #fff;
      table-layout: fixed;
    }
    thead th {
      background: rgba(17,24,39,0.03);
      font-size: 12px;
      color: rgba(17,24,39,0.8);
      padding: 10px 10px;
      border-bottom: 1px solid rgba(17,24,39,0.08);
      text-align: left;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    tbody td {
      padding: 9px 10px;
      font-size: 12px;
      border-bottom: 1px solid rgba(17,24,39,0.06);
      color: #111827;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    tbody tr:hover td { background: rgba(225,29,46,0.04); }

    /* Growth badges: merah/hijau */
    .badge-pos, .badge-neg, .badge-na {
      display:inline-block;
      font-size: 11px;
      font-weight: 900;
      padding: 2px 8px;
      border-radius: 999px;
      white-space: nowrap;
    }
    .badge-pos { background: rgba(22,163,74,0.12); color:#16a34a; }
    .badge-neg { background: rgba(220,38,38,0.12); color:#dc2626; }
    .badge-na  { background: rgba(100,116,139,0.12); color:#64748b; }

    /* SKU table column widths (SPESIFIKASI wider, Growth wider so header & % visible) */
    table.sku-table th:nth-child(1), table.sku-table td:nth-child(1) { width: 50%; }
    table.sku-table th:nth-child(2), table.sku-table td:nth-child(2) { width: 14%; }
    table.sku-table th:nth-child(3), table.sku-table td:nth-child(3) { width: 14%; }
    table.sku-table th:nth-child(4), table.sku-table td:nth-child(4) { width: 10%; } /* Delta */
    table.sku-table th:nth-child(5), table.sku-table td:nth-child(5) { width: 12%; } /* Growth */
    </style>
    """,
        unsafe_allow_html=True,
    )

    DEFAULT_DATE_FORMAT_HINT = "Format TGL: dd-mm-yyyy / dd/mm/yyyy / yyyy-mm-dd"


    def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]
        return df


    def coerce_numeric_series(s: pd.Series) -> pd.Series:
        def to_num(x):
            if pd.isna(x):
                return np.nan
            if isinstance(x, (int, float, np.integer, np.floating)):
                return float(x)
            txt = str(x).strip()
            txt = re.sub(r"[^0-9\-\.,]", "", txt)
            if txt in ("", "-", ".", ","):
                return np.nan
            txt = txt.replace(" ", "").replace(".", "").replace(",", "")
            try:
                return float(txt)
            except Exception:
                return np.nan

        return s.map(to_num)


    def parse_tgl(df: pd.DataFrame, col: str = "TGL") -> pd.Series:
        s = df[col]
        if pd.api.types.is_datetime64_any_dtype(s):
            return pd.to_datetime(s, errors="coerce").dt.date
        parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)
        return parsed.dt.date


    def ensure_required_columns(df: pd.DataFrame) -> Tuple[bool, str]:
        required = [
            "STATUS", "TGL", "TRANSAKSI", "TEAM",
            "PRODUCT", "BRAND", "QTY", "JUMLAH",
            "SO NO",
            "COUNTRY",
            "SPESIFIKASI",
            "NAMA CUSTOMER",
            "OTO",
        ]
        missing = [c for c in required if c not in df.columns]
        if missing:
            return False, f"Kolom wajib tidak ditemukan: {', '.join(missing)}"
        return True, ""


    def drop_total_rows(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        if "NO" in out.columns:
            out = out[~out["NO"].astype(str).str.strip().str.upper().eq("TOTAL")].copy()
        for c in ["STATUS", "TGL"]:
            if c in out.columns:
                out = out[~out[c].astype(str).str.strip().str.upper().eq("TOTAL")].copy()
        return out.dropna(how="all")


    def format_idr(x: float) -> str:
        if pd.isna(x):
            return "-"
        n = int(round(float(x)))
        s = f"{n:,}".replace(",", ".")
        return f"IDR {s}"


    def format_int_id(x: float) -> str:
        if pd.isna(x):
            return "-"
        return f"{int(round(float(x))):,}".replace(",", ".")


    def compact_number(x: float) -> str:
        if pd.isna(x):
            return ""
        x = float(x)
        ax = abs(x)
        if ax >= 1_000_000_000:
            return f"{x/1_000_000_000:.2f}B".replace(".", ",")
        if ax >= 1_000_000:
            return f"{x/1_000_000:.2f}M".replace(".", ",")
        if ax >= 1_000:
            return f"{x/1_000:.2f}K".replace(".", ",")
        return str(int(round(x)))


    def safe_growth_pct(this_val: float, last_val: float) -> Optional[float]:
        if last_val is None or pd.isna(last_val):
            return None
        last_val = float(last_val)
        if last_val == 0.0:
            return None
        return (float(this_val) - last_val) / last_val * 100.0


    def growth_label(g: Optional[float]) -> str:
        if g is None or pd.isna(g):
            return "N/A"
        s = f"{g:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")
        return ("▲ " + s) if g >= 0 else ("▼ " + s)


    def growth_badge_html(g: Optional[float]) -> str:
        if g is None or pd.isna(g):
            return '<span class="badge-na">N/A</span>'
        s = f"{g:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")
        if g >= 0:
            return f'<span class="badge-pos">▲ {s}</span>'
        return f'<span class="badge-neg">▼ {s}</span>'


    def kpi_delta_class(g: Optional[float]) -> str:
        if g is None or pd.isna(g):
            return "na"
        return "pos" if g >= 0 else "neg"


    @st.cache_data(show_spinner=False)
    def read_excel_cached(file_bytes: bytes, sheet_name: str, header_row_1based: int) -> pd.DataFrame:
        bio = io.BytesIO(file_bytes)
        header_idx = int(header_row_1based) - 1
        if sheet_name.strip() == "":
            return pd.read_excel(bio, header=header_idx)
        return pd.read_excel(bio, sheet_name=sheet_name.strip(), header=header_idx)


    @st.cache_data(show_spinner=False)
    def clean_sales_df_cached(df_raw: pd.DataFrame):
        df = normalize_columns(df_raw)
        ok, msg = ensure_required_columns(df)
        if not ok:
            raise ValueError(msg)

        # Template penjualan dapat memakai header DIVISI, DIVISION, atau DIV.
        # Disatukan ke satu kolom agar filter dashboard selalu konsisten.
        if "DIVISI" not in df.columns:
            for alias in ["DIVISION", "DIV", "DIVISI PENJUALAN"]:
                if alias in df.columns:
                    df = df.rename(columns={alias: "DIVISI"})
                    break
        if "DIVISI" not in df.columns:
            df["DIVISI"] = "(Tidak tersedia)"

        df = drop_total_rows(df)

        keep_cols = [
            "STATUS", "TGL", "TRANSAKSI", "TEAM",
            "PRODUCT", "BRAND", "QTY", "JUMLAH",
            "SO NO",
            "COUNTRY",
            "DIVISI",
            "SPESIFIKASI",
            "NAMA CUSTOMER",
            "OTO",
            "AREA",
        ]
        df = df[keep_cols].copy()

        df["TGL"] = parse_tgl(df, "TGL")
        if df["TGL"].isna().any():
            bad = df[df["TGL"].isna()].head(8)
            raise ValueError(
                f"Ada TGL gagal diparse.\n\nContoh:\n{bad[['TGL']].to_string(index=False)}\n\n{DEFAULT_DATE_FORMAT_HINT}"
            )

        df["QTY"] = pd.to_numeric(coerce_numeric_series(df["QTY"]), errors="coerce").fillna(0.0)
        df["JUMLAH"] = pd.to_numeric(coerce_numeric_series(df["JUMLAH"]), errors="coerce").fillna(0.0)

        for c in [
            "STATUS", "TRANSAKSI", "TEAM", "PRODUCT", "BRAND", "DIVISI",
            "SO NO", "COUNTRY", "SPESIFIKASI", "NAMA CUSTOMER", "OTO"
        ]:
            df[c] = df[c].astype(str).str.strip()

        df["ROW_TYPE"] = np.where(df["STATUS"].str.upper().str.contains("RETUR"), "RETUR", "SO_OUT")
        df["OTO_YES"] = df["OTO"].str.upper().eq("YES")
        df["PLATFORM"] = df["NAMA CUSTOMER"].astype(str).str.strip()

        df = df[df["STATUS"].str.strip().ne("")].copy()
        return df, df["TGL"].min(), df["TGL"].max()


    @st.cache_data(show_spinner=False)
    def compute_kpis_cached(df: pd.DataFrame) -> Dict[str, float]:
        sales = float(df["JUMLAH"].sum())
        qty = float(df["QTY"].sum())

        so = df[df["ROW_TYPE"] == "SO_OUT"]
        orders = float(so["SO NO"].nunique())

        returns = float(len(df[df["ROW_TYPE"] == "RETUR"]))
        aov = sales / orders if orders else np.nan
        return {"sales": sales, "qty": qty, "orders": orders, "returns": returns, "aov": float(aov) if not pd.isna(aov) else np.nan}


    def get_week_start(d: date) -> date:
        return d - timedelta(days=d.weekday())


    def month_start(d: date) -> date:
        return date(d.year, d.month, 1)


    def prev_month_same_day(d: date) -> date:
        first = month_start(d)
        prev_last = first - timedelta(days=1)
        day = min(d.day, prev_last.day)
        return date(prev_last.year, prev_last.month, day)


    def slice_period(df: pd.DataFrame, start: date, end_inclusive: date) -> pd.DataFrame:
        return df[(df["TGL"] >= start) & (df["TGL"] <= end_inclusive)].copy()


    def build_period_frames(df_all: pd.DataFrame, mode: str, df_upload_a: pd.DataFrame, df_upload_b: pd.DataFrame):
        if mode == "UPLOAD":
            return df_upload_a, df_upload_b, "Periode A", "Periode B"

        anchor = df_all["TGL"].max()

        if mode == "WOW":
            this_start = get_week_start(anchor)
            this_end = anchor
            last_end = this_start - timedelta(days=1)
            last_start = get_week_start(last_end)
            return (
                slice_period(df_all, last_start, last_end),
                slice_period(df_all, this_start, this_end),
                "Week Lalu",
                "Week Ini",
            )

        if mode == "MOM":
            this_start = month_start(anchor)
            this_end = anchor
            last_anchor = prev_month_same_day(anchor)
            last_start = month_start(last_anchor)
            last_month_mask = (df_all["TGL"] >= last_start) & (df_all["TGL"] < this_start)
            last_end = df_all.loc[last_month_mask, "TGL"].max() if last_month_mask.any() else last_start
            return (
                slice_period(df_all, last_start, last_end),
                slice_period(df_all, this_start, this_end),
                "Bulan Lalu",
                "Bulan Ini",
            )

        if mode == "MTD":
            this_start = month_start(anchor)
            this_end = anchor
            last_anchor = prev_month_same_day(anchor)
            last_start = month_start(last_anchor)

            next_month_first = month_start(anchor)
            last_month_last_day = next_month_first - timedelta(days=1)
            last_end_candidate = date(last_month_last_day.year, last_month_last_day.month, min(anchor.day, last_month_last_day.day))

            last_month_mask = (df_all["TGL"] >= last_start) & (df_all["TGL"] < this_start)
            if last_month_mask.any():
                last_end_data = df_all.loc[last_month_mask, "TGL"].max()
                last_end = min(last_end_candidate, last_end_data)
            else:
                last_end = last_end_candidate

            return (
                slice_period(df_all, last_start, last_end),
                slice_period(df_all, this_start, this_end),
                "MTD Bulan Lalu",
                "MTD Bulan Ini",
            )

        return df_upload_a, df_upload_b, "Periode Lalu", "Periode Ini"


    def options_for(df_all: pd.DataFrame, col: str) -> List[str]:
        return sorted([v for v in df_all[col].dropna().unique().tolist() if str(v).strip() != ""])


    def apply_multifilter(df: pd.DataFrame, col: str, selected: List[str]) -> pd.DataFrame:
        if not selected:
            return df
        return df[df[col].isin(selected)].copy()


    @st.cache_data(show_spinner=False)
    def top_table_cached(df_this: pd.DataFrame, df_last: pd.DataFrame, by_col: str, metric: str, top_n: int) -> pd.DataFrame:
        agg_this = df_this.groupby(by_col, as_index=False).agg(THIS=(metric, "sum"))
        agg_last = df_last.groupby(by_col, as_index=False).agg(LAST=(metric, "sum"))
        merged = agg_this.merge(agg_last, on=by_col, how="outer").fillna(0.0)
        merged["DELTA"] = merged["THIS"] - merged["LAST"]
        merged["GROWTH_NUM"] = merged.apply(lambda r: safe_growth_pct(r["THIS"], r["LAST"]), axis=1)
        merged = merged.sort_values("THIS", ascending=False).head(top_n)

        if metric == "JUMLAH":
            merged["Periode Ini"] = merged["THIS"].map(format_idr)
            merged["Periode Lalu"] = merged["LAST"].map(format_idr)
            merged["Delta"] = merged["DELTA"].map(format_idr)
        else:
            merged["Periode Ini"] = merged["THIS"].map(format_int_id)
            merged["Periode Lalu"] = merged["LAST"].map(format_int_id)
            merged["Delta"] = merged["DELTA"].map(format_int_id)

        merged["Growth"] = merged["GROWTH_NUM"].apply(growth_badge_html)
        merged = merged[[by_col, "Periode Ini", "Periode Lalu", "Delta", "Growth"]]
        return merged


    def render_html_table(df: pd.DataFrame, table_class: str = ""):
        html = df.to_html(escape=False, index=False)
        if table_class:
            html = html.replace("<table ", f'<table class="{table_class}" ', 1)
        st.markdown(html, unsafe_allow_html=True)


    def small_title(text: str, hint: str = ""):
        hint_html = f' <span class="muted">{hint}</span>' if hint else ""
        st.markdown(f'<div class="small-h">{text}{hint_html}</div>', unsafe_allow_html=True)


    def render_header():
        st.markdown(
            """
    <div class="header-wrap">
      <div>
        <div class="header-title">Growth Dashboard Agres</div>
      </div>
    </div>
    """,
            unsafe_allow_html=True,
        )


    def style_growth_pct_df(df_in: pd.DataFrame):
        df = df_in.copy()

        def color_growth(val):
            try:
                if pd.isna(val):
                    return "color: #64748b;"
                return "color: #dc2626; font-weight: 800;" if float(val) < 0 else "color: #16a34a; font-weight: 800;"
            except Exception:
                return "color: #64748b;"

        return (
            df.style
            .format({"Growth %": "{:.2f}%"})
            .map(color_growth, subset=["Growth %"])
        )


    # ===== NEW: Team Down % Table =====
    @st.cache_data(show_spinner=False)
    def team_down_ratio_table_cached(df_last: pd.DataFrame, df_this: pd.DataFrame) -> pd.DataFrame:
        last = df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum"))
        this = df_this.groupby("TEAM", as_index=False).agg(QTY_INI=("QTY", "sum"))
        t = last.merge(this, on="TEAM", how="outer").fillna(0.0)
        t["DELTA"] = t["QTY_INI"] - t["QTY_LALU"]

        # total team aktif: punya activity di salah satu periode
        active = t[(t["QTY_LALU"] > 0) | (t["QTY_INI"] > 0)].copy()
        total = int(len(active))
        turun = int((active["DELTA"] < 0).sum())
        naik = int((active["DELTA"] > 0).sum())
        flat = int((active["DELTA"] == 0).sum())
        pct_turun = (turun / total * 100.0) if total else 0.0

        out = pd.DataFrame(
            {
                "Total TEAM aktif": [total],
                "TEAM Turun": [turun],
                "TEAM Naik": [naik],
                "TEAM Tetap": [flat],
                "% TEAM Turun": [pct_turun],
            }
        )
        return out


    def _drivers_as_text(df_delta: pd.DataFrame, team_dir: str, top_k: int = 3) -> Dict[str, str]:
        """
        df_delta columns: TEAM, DIM, DELTA
        team_dir: mapping TEAM -> +1 (naik) or -1 (turun)
        """
        # Join direction for filtering
        d = df_delta.copy()
        d["DIR"] = d["TEAM"].map(team_dir).fillna(0).astype(int)

        # For naik: keep DELTA > 0, take top_k biggest
        # For turun: keep DELTA < 0, take top_k most negative
        naik_df = d[(d["DIR"] > 0) & (d["DELTA"] > 0)].copy()
        turun_df = d[(d["DIR"] < 0) & (d["DELTA"] < 0)].copy()

        naik_df = naik_df.sort_values(["TEAM", "DELTA"], ascending=[True, False]).groupby("TEAM").head(top_k)
        turun_df = turun_df.sort_values(["TEAM", "DELTA"], ascending=[True, True]).groupby("TEAM").head(top_k)

        # Build strings
        out: Dict[str, str] = {}

        def fmt_row(dim: str, delta: float) -> str:
            sign = "+" if delta > 0 else ""
            return f"{dim} ({sign}{int(delta):,})".replace(",", ".")

        for team, g in naik_df.groupby("TEAM"):
            out[team] = ", ".join([fmt_row(r["DIM"], r["DELTA"]) for _, r in g.iterrows()])

        for team, g in turun_df.groupby("TEAM"):
            out[team] = ", ".join([fmt_row(r["DIM"], r["DELTA"]) for _, r in g.iterrows()])

        return out


    @st.cache_data(show_spinner=False)
    def team_driver_analysis_table_cached(df_last: pd.DataFrame, df_this: pd.DataFrame, top_k: int = 3) -> pd.DataFrame:
        # team totals
        last_t = df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum"))
        this_t = df_this.groupby("TEAM", as_index=False).agg(QTY_INI=("QTY", "sum"))
        team = last_t.merge(this_t, on="TEAM", how="outer").fillna(0.0)
        team = team[(team["QTY_LALU"] > 0) | (team["QTY_INI"] > 0)].copy()
        team["DELTA_QTY"] = team["QTY_INI"] - team["QTY_LALU"]
        team["GROWTH_PCT"] = team.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

        # direction mapping (+1 naik, -1 turun, 0 flat)
        team_dir = {r["TEAM"]: (1 if r["DELTA_QTY"] > 0 else (-1 if r["DELTA_QTY"] < 0 else 0)) for _, r in team.iterrows()}

        # build delta per TEAM x DIM for PRODUCT / BRAND / SKU(SPESIFIKASI)
        def build_delta(dim_col: str) -> pd.DataFrame:
            a = df_this.groupby(["TEAM", dim_col], as_index=False).agg(THIS=("QTY", "sum"))
            b = df_last.groupby(["TEAM", dim_col], as_index=False).agg(LAST=("QTY", "sum"))
            m = a.merge(b, on=["TEAM", dim_col], how="outer").fillna(0.0)
            m["DELTA"] = m["THIS"] - m["LAST"]
            m = m.rename(columns={dim_col: "DIM"})
            return m[["TEAM", "DIM", "DELTA"]]

        prod_delta = build_delta("PRODUCT")
        brand_delta = build_delta("BRAND")
        sku_delta = build_delta("SPESIFIKASI")

        prod_map = _drivers_as_text(prod_delta, team_dir, top_k=top_k)
        brand_map = _drivers_as_text(brand_delta, team_dir, top_k=top_k)
        sku_map = _drivers_as_text(sku_delta, team_dir, top_k=top_k)

        team["Arah"] = team["DELTA_QTY"].apply(lambda x: "NAIK" if x > 0 else ("TURUN" if x < 0 else "TETAP"))
        team["Produk (driver)"] = team["TEAM"].map(prod_map).fillna("-")
        team["Brand (driver)"] = team["TEAM"].map(brand_map).fillna("-")
        team["SKU/Spesifikasi (driver)"] = team["TEAM"].map(sku_map).fillna("-")

        # pretty
        out = team.copy()
        out["QTY Lalu"] = out["QTY_LALU"].map(format_int_id)
        out["QTY Ini"] = out["QTY_INI"].map(format_int_id)
        out["Delta"] = out["DELTA_QTY"].map(lambda x: f"{int(x):,}".replace(",", "."))
        out["Growth %"] = out["GROWTH_PCT"].apply(lambda x: float(x) if (x is not None and not pd.isna(x)) else np.nan)

        out = out.sort_values(["DELTA_QTY"], ascending=True)  # yang turun paling parah di atas (biar langsung keliatan)
        out = out[
            [
                "TEAM", "Arah", "QTY Lalu", "QTY Ini", "Delta", "Growth %",
                "Produk (driver)", "Brand (driver)", "SKU/Spesifikasi (driver)"
            ]
        ]
        return out




    @st.cache_data(show_spinner=False)
    def brand_delta_analysis_table_cached(df_last: pd.DataFrame, df_this: pd.DataFrame) -> pd.DataFrame:
        category_rules = [
            ("delta semua produk", None),
            ("delta laptop 2nd", ["LAPTOP 2ND", "LAPTOP SECOND", "LAPTOP 2NDHAND", "LAPTOP 2ND HAND"]),
            ("delta laptop d", ["LAPTOP D", "LAPTOP-D", "LAPTOP D "]),
            ("delta laptop R", ["LAPTOP R", "LAPTOP-R", "LAPTOP R "]),
            ("delta aio", ["AIO", "ALL IN ONE", "ALL-IN-ONE"]),
            ("delta pcdesktop", ["PCDESKTOP", "PC DESKTOP", "DESKTOP"]),
            ("delta pcmini", ["PCMINI", "PC MINI", "MINI PC"]),
            ("delta phone", ["PHONE", "SMARTPHONE", "HP ", "HANDPHONE"]),
            ("delta tablet", ["TABLET", "TAB"]),
        ]

        def add_delta_col(base: pd.DataFrame, label: str, product_keywords: Optional[List[str]]) -> pd.DataFrame:
            if product_keywords is None:
                last_src = df_last
                this_src = df_this
            else:
                pattern = "|".join([re.escape(k) for k in product_keywords])
                last_src = df_last[df_last["PRODUCT"].astype(str).str.upper().str.contains(pattern, na=False, regex=True)]
                this_src = df_this[df_this["PRODUCT"].astype(str).str.upper().str.contains(pattern, na=False, regex=True)]

            last_agg = last_src.groupby("BRAND", as_index=False).agg(LAST=("QTY", "sum"))
            this_agg = this_src.groupby("BRAND", as_index=False).agg(THIS=("QTY", "sum"))
            delta = this_agg.merge(last_agg, on="BRAND", how="outer").fillna(0.0)
            delta[label] = delta["THIS"] - delta["LAST"]
            delta = delta[["BRAND", label]]
            return base.merge(delta, on="BRAND", how="outer")

        brand_base = pd.DataFrame({"BRAND": sorted(set(df_last["BRAND"].dropna().astype(str)) | set(df_this["BRAND"].dropna().astype(str)))})
        brand_base = brand_base[brand_base["BRAND"].astype(str).str.strip().ne("")].copy()

        out = brand_base
        for label, keywords in category_rules:
            out = add_delta_col(out, label, keywords)

        if out.empty:
            return out

        value_cols = [label for label, _ in category_rules]
        out[value_cols] = out[value_cols].fillna(0.0)
        out = out.sort_values("delta semua produk", ascending=False).copy()
        for col in value_cols:
            out[col] = out[col].map(format_int_id)
        return out

    @st.cache_data(show_spinner=False)
    def brand_delta_analysis_raw_cached(df_last: pd.DataFrame, df_this: pd.DataFrame) -> pd.DataFrame:
        category_rules = [
            ("delta semua produk", None),
            ("delta laptop 2nd", ["LAPTOP 2ND", "LAPTOP SECOND", "LAPTOP 2NDHAND", "LAPTOP 2ND HAND"]),
            ("delta laptop d", ["LAPTOP D", "LAPTOP-D", "LAPTOP D "]),
            ("delta laptop R", ["LAPTOP R", "LAPTOP-R", "LAPTOP R "]),
            ("delta aio", ["AIO", "ALL IN ONE", "ALL-IN-ONE"]),
            ("delta pcdesktop", ["PCDESKTOP", "PC DESKTOP", "DESKTOP"]),
            ("delta pcmini", ["PCMINI", "PC MINI", "MINI PC"]),
            ("delta phone", ["PHONE", "SMARTPHONE", "HP ", "HANDPHONE"]),
            ("delta tablet", ["TABLET", "TAB"]),
        ]

        def add_delta_col(base: pd.DataFrame, label: str, product_keywords: Optional[List[str]]) -> pd.DataFrame:
            if product_keywords is None:
                last_src = df_last
                this_src = df_this
            else:
                pattern = "|".join([re.escape(k) for k in product_keywords])
                last_src = df_last[df_last["PRODUCT"].astype(str).str.upper().str.contains(pattern, na=False, regex=True)]
                this_src = df_this[df_this["PRODUCT"].astype(str).str.upper().str.contains(pattern, na=False, regex=True)]

            last_agg = last_src.groupby("BRAND", as_index=False).agg(LAST=("QTY", "sum"))
            this_agg = this_src.groupby("BRAND", as_index=False).agg(THIS=("QTY", "sum"))
            delta = this_agg.merge(last_agg, on="BRAND", how="outer").fillna(0.0)
            delta[label] = delta["THIS"] - delta["LAST"]
            delta = delta[["BRAND", label]]
            return base.merge(delta, on="BRAND", how="outer")

        brand_base = pd.DataFrame({"BRAND": sorted(set(df_last["BRAND"].dropna().astype(str)) | set(df_this["BRAND"].dropna().astype(str)))})
        brand_base = brand_base[brand_base["BRAND"].astype(str).str.strip().ne("")].copy()

        out = brand_base
        for label, keywords in category_rules:
            out = add_delta_col(out, label, keywords)

        if out.empty:
            return out

        value_cols = [label for label, _ in category_rules]
        out[value_cols] = out[value_cols].fillna(0.0)
        out = out.sort_values("BRAND", ascending=True).copy()
        return out



    def proper_delta_header(col_name: str) -> str:
        mapping = {
            "delta semua produk": "Delta All",
            "delta laptop 2nd": "Delta Laptop 2nd",
            "delta laptop d": "Delta Laptop D",
            "delta laptop R": "Delta Laptop R",
            "delta aio": "Delta AIO",
            "delta pcdesktop": "Delta PC Desktop",
            "delta pcmini": "Delta PC Mini",
            "delta phone": "Delta Phone",
            "delta tablet": "Delta Tablet",
        }
        return mapping.get(col_name, str(col_name).title())


    def render_clickable_brand_delta_table(raw_df: pd.DataFrame):
        if raw_df.empty:
            st.info("Tidak ada data brand pada filter & periode saat ini.")
            return

        value_cols = [c for c in raw_df.columns if c != "BRAND"]

        header_cols = st.columns([1.25] + [1.15] * len(value_cols), gap="small")
        header_cols[0].markdown("**BRAND**")
        for i, col_name in enumerate(value_cols, start=1):
            header_cols[i].markdown(f"**{proper_delta_header(col_name)}**")

        table_box = st.container(height=420)
        with table_box:
            for r_idx, row in raw_df.reset_index(drop=True).iterrows():
                cols = st.columns([1.25] + [1.15] * len(value_cols), gap="small")
                cols[0].markdown(str(row["BRAND"]))

                for c_idx, col_name in enumerate(value_cols, start=1):
                    val = float(row[col_name])
                    label = format_int_id(val)

                    clicked = cols[c_idx].button(
                        label,
                        key=f"brand_delta_click_{r_idx}_{c_idx}_{row['BRAND']}_{col_name}",
                        use_container_width=True,
                    )
                    if clicked:
                        st.session_state["selected_brand_delta_cell"] = {
                            "brand": str(row["BRAND"]),
                            "category": str(col_name),
                            "value": val,
                        }


    def render_brand_team_detail_card(df_last: pd.DataFrame, df_this: pd.DataFrame):
        selected = st.session_state.get("selected_brand_delta_cell")
        if not selected:
            st.info("Klik salah satu angka pada tabel Analisa Brand untuk menampilkan detail TEAM.")
            return

        selected_brand = selected.get("brand", "")
        selected_kategori = selected.get("category", "")
        selected_value = float(selected.get("value", 0))

        category_map = {
            "delta semua produk": None,
            "delta laptop 2nd": ["LAPTOP 2ND", "LAPTOP SECOND", "LAPTOP 2NDHAND", "LAPTOP 2ND HAND"],
            "delta laptop d": ["LAPTOP D", "LAPTOP-D", "LAPTOP D "],
            "delta laptop R": ["LAPTOP R", "LAPTOP-R", "LAPTOP R "],
            "delta aio": ["AIO", "ALL IN ONE", "ALL-IN-ONE"],
            "delta pcdesktop": ["PCDESKTOP", "PC DESKTOP", "DESKTOP"],
            "delta pcmini": ["PCMINI", "PC MINI", "MINI PC"],
            "delta phone": ["PHONE", "SMARTPHONE", "HP ", "HANDPHONE"],
            "delta tablet": ["TABLET", "TAB"],
        }

        keywords = category_map.get(selected_kategori)

        if keywords is None:
            df_last_team = df_last[df_last["BRAND"] == selected_brand].copy()
            df_this_team = df_this[df_this["BRAND"] == selected_brand].copy()
        else:
            pattern = "|".join([re.escape(k) for k in keywords])
            df_last_team = df_last[
                (df_last["BRAND"] == selected_brand)
                & (df_last["PRODUCT"].astype(str).str.upper().str.contains(pattern, na=False, regex=True))
            ].copy()
            df_this_team = df_this[
                (df_this["BRAND"] == selected_brand)
                & (df_this["PRODUCT"].astype(str).str.upper().str.contains(pattern, na=False, regex=True))
            ].copy()

        last_team = df_last_team.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum"))
        this_team = df_this_team.groupby("TEAM", as_index=False).agg(QTY_INI=("QTY", "sum"))
        team_detail = this_team.merge(last_team, on="TEAM", how="outer").fillna(0.0)

        st.markdown("<hr/>", unsafe_allow_html=True)
        st.subheader("👥 Detail TEAM dari Analisa Brand")
        st.caption(f"BRAND: {selected_brand} | Kolom: {proper_delta_header(selected_kategori)} | Delta: {format_int_id(selected_value)}")

        if team_detail.empty:
            st.info("Tidak ada data TEAM untuk cell ini.")
            return

        team_detail["DELTA"] = team_detail["QTY_INI"] - team_detail["QTY_LALU"]
        team_detail["GROWTH_PCT"] = team_detail.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

        # Delta total TEAM dari semua product/brand, supaya konteks performa TEAM tetap kelihatan.
        total_last_team = df_last.groupby("TEAM", as_index=False).agg(QTY_TOTAL_LALU=("QTY", "sum"))
        total_this_team = df_this.groupby("TEAM", as_index=False).agg(QTY_TOTAL_INI=("QTY", "sum"))
        total_team = total_this_team.merge(total_last_team, on="TEAM", how="outer").fillna(0.0)
        total_team["DELTA_TOTAL"] = total_team["QTY_TOTAL_INI"] - total_team["QTY_TOTAL_LALU"]
        team_detail = team_detail.merge(total_team[["TEAM", "DELTA_TOTAL"]], on="TEAM", how="left").fillna({"DELTA_TOTAL": 0.0})

        def _top_delta_driver(df_last_src: pd.DataFrame, df_this_src: pd.DataFrame, dim_col: str, selected_team: str, direction_value: float, top_k: int = 3) -> str:
            last_src = df_last_src[df_last_src["TEAM"] == selected_team].copy()
            this_src = df_this_src[df_this_src["TEAM"] == selected_team].copy()

            a = this_src.groupby(dim_col, as_index=False).agg(THIS=("QTY", "sum"))
            b = last_src.groupby(dim_col, as_index=False).agg(LAST=("QTY", "sum"))
            m = a.merge(b, on=dim_col, how="outer").fillna(0.0)
            m[dim_col] = m[dim_col].astype(str).str.strip()
            m = m[m[dim_col].ne("")].copy()
            if m.empty:
                return "-"

            m["DELTA_DRIVER"] = m["THIS"] - m["LAST"]
            if direction_value < 0:
                m = m[m["DELTA_DRIVER"] < 0].sort_values("DELTA_DRIVER", ascending=True)
            elif direction_value > 0:
                m = m[m["DELTA_DRIVER"] > 0].sort_values("DELTA_DRIVER", ascending=False)
            else:
                m = m[m["DELTA_DRIVER"] != 0].copy()
                m["ABS_DELTA_DRIVER"] = m["DELTA_DRIVER"].abs()
                m = m.sort_values("ABS_DELTA_DRIVER", ascending=False)

            if m.empty:
                return "-"

            def fmt_driver(row):
                delta = float(row["DELTA_DRIVER"])
                sign = "+" if delta > 0 else ""
                return f"{row[dim_col]} ({sign}{int(delta):,})".replace(",", ".")

            return ", ".join([fmt_driver(row) for _, row in m.head(top_k).iterrows()])

        team_detail["Delta Platform"] = team_detail.apply(
            lambda r: _top_delta_driver(df_last_team, df_this_team, "PLATFORM", r["TEAM"], r["DELTA"]),
            axis=1,
        )
        team_detail["SKU/Spesifikasi (driver)"] = team_detail.apply(
            lambda r: _top_delta_driver(df_last_team, df_this_team, "SPESIFIKASI", r["TEAM"], r["DELTA"]),
            axis=1,
        )

        if selected_value < 0:
            team_detail = team_detail.sort_values("DELTA", ascending=True)
        else:
            team_detail = team_detail.sort_values("DELTA", ascending=False)

        team_detail["QTY Lalu"] = team_detail["QTY_LALU"].astype(int)
        team_detail["QTY Ini"] = team_detail["QTY_INI"].astype(int)
        team_detail["Delta"] = team_detail["DELTA"].astype(int)
        team_detail["Growth %"] = team_detail["GROWTH_PCT"].apply(lambda x: float(x) if (x is not None and not pd.isna(x)) else np.nan)

        detail_view = team_detail[
            [
                "TEAM",
                "QTY Lalu",
                "QTY Ini",
                "Delta",
                "Growth %",
                "Delta Platform",
                "SKU/Spesifikasi (driver)",
            ]
        ].copy()

        st.dataframe(
            style_growth_pct_df(detail_view),
            use_container_width=True,
            height=520,
            hide_index=True,
            column_config={
                "TEAM": st.column_config.TextColumn("TEAM", width="small"),
                "QTY Lalu": st.column_config.NumberColumn("QTY Lalu", width="small", format="%d"),
                "QTY Ini": st.column_config.NumberColumn("QTY Ini", width="small", format="%d"),
                "Delta": st.column_config.NumberColumn("Delta", width="small", format="%d"),
                "Growth %": st.column_config.NumberColumn("Growth %", width="small", format="%.2f%%"),
                "Delta Platform": st.column_config.TextColumn("Delta Platform", width="medium"),
                "SKU/Spesifikasi (driver)": st.column_config.TextColumn("SKU/Spesifikasi (driver)", width="large"),
            },
        )




    def _render_analisa_penjualan_app_inner():
        render_header()

        header_row_a = 2
        header_row_b = 2
        sheet_a = ""
        sheet_b = ""

        st.subheader("Upload Data")
        upload_col1, upload_col2 = st.columns(2, gap="large")

        with upload_col1:
            file_a = st.file_uploader("Excel A (.xlsx) — periode lama", type=["xlsx"], key="a")

        with upload_col2:
            file_b = st.file_uploader("Excel B (.xlsx) — periode baru", type=["xlsx"], key="b")

        if not file_a or not file_b:
            st.info("Upload 2 file Excel dulu.")
            st.stop()

        with st.spinner("Membaca & membersihkan Excel (sekali di awal)..."):
            df_a_raw = read_excel_cached(file_a.getvalue(), sheet_a, header_row_a)
            df_b_raw = read_excel_cached(file_b.getvalue(), sheet_b, header_row_b)
            df_a, a_date_min, a_date_max = clean_sales_df_cached(df_a_raw)
            df_b, b_date_min, b_date_max = clean_sales_df_cached(df_b_raw)

        df_all = pd.concat([df_a, df_b], ignore_index=True)

        st.markdown("---")
        st.subheader("Filter & Kontrol")
        ctl1, ctl2, ctl3 = st.columns([1.0, 1.0, 2.2], gap="large")

        with ctl1:
            compare_mode = st.selectbox("Pilih periode", ["MOM", "WOW", "MTD", "UPLOAD"], 0)
            metric_choice = st.selectbox("Metric", ["Qty (QTY)", "Sales (JUMLAH)"], 0)
            show_point_labels = st.toggle("Tampilkan angka di titik grafik", value=False)

        with ctl2:
            top_n = st.slider("Top N", 5, 30, 10, 1)

        with ctl3:
            with st.form("filter_form", clear_on_submit=False):
                filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4, gap="medium")
                with filter_col1:
                    divisi_sel = st.multiselect("DIVISI", options_for(df_all, "DIVISI"), default=[])
                    category_sel = st.multiselect("CATEGORY (COUNTRY)", options_for(df_all, "COUNTRY"), default=[])
                with filter_col2:
                    transaksi_sel = st.multiselect("TRANSAKSI", options_for(df_all, "TRANSAKSI"), default=[])
                    team_sel = st.multiselect("TEAM", options_for(df_all, "TEAM"), default=[])
                with filter_col3:
                    product_sel = st.multiselect("PRODUCT", options_for(df_all, "PRODUCT"), default=[])
                    brand_sel = st.multiselect("BRAND", options_for(df_all, "BRAND"), default=[])
                with filter_col4:
                    platform_sel = st.multiselect("PLATFORM (NAMA CUSTOMER)", options_for(df_all, "PLATFORM"), default=[])
                apply_clicked = st.form_submit_button("Apply Filter", use_container_width=True)

        filter_defaults = {"DIVISI": [], "COUNTRY": [], "TRANSAKSI": [], "TEAM": [], "PRODUCT": [], "BRAND": [], "PLATFORM": []}
        if "filters" not in st.session_state:
            st.session_state["filters"] = filter_defaults
        else:
            st.session_state["filters"] = {**filter_defaults, **st.session_state["filters"]}
        if apply_clicked:
            st.session_state["filters"] = {
                "DIVISI": divisi_sel,
                "COUNTRY": category_sel,
                "TRANSAKSI": transaksi_sel,
                "TEAM": team_sel,
                "PRODUCT": product_sel,
                "BRAND": brand_sel,
                "PLATFORM": platform_sel,
            }

        flt = st.session_state["filters"]


        def apply_all_filters(df: pd.DataFrame) -> pd.DataFrame:
            out = df
            out = apply_multifilter(out, "DIVISI", flt["DIVISI"])
            out = apply_multifilter(out, "COUNTRY", flt["COUNTRY"])
            out = apply_multifilter(out, "TRANSAKSI", flt["TRANSAKSI"])
            out = apply_multifilter(out, "TEAM", flt["TEAM"])
            out = apply_multifilter(out, "PRODUCT", flt["PRODUCT"])
            out = apply_multifilter(out, "BRAND", flt["BRAND"])
            out = apply_multifilter(out, "PLATFORM", flt["PLATFORM"])
            return out


        df_all_f = apply_all_filters(df_all)
        df_a_f = apply_all_filters(df_a)
        df_b_f = apply_all_filters(df_b)

        df_last, df_this, label_last, label_this = build_period_frames(df_all_f, compare_mode, df_a_f, df_b_f)

        metric_col = "QTY" if metric_choice.startswith("Qty") else "JUMLAH"
        metric_name = "Qty" if metric_col == "QTY" else "Sales (IDR)"

        # =========================
        # KPI
        # =========================
        k_last = compute_kpis_cached(df_last)
        k_this = compute_kpis_cached(df_this)

        sales_g = safe_growth_pct(k_this["sales"], k_last["sales"])
        orders_g = safe_growth_pct(k_this["orders"], k_last["orders"])
        qty_g = safe_growth_pct(k_this["qty"], k_last["qty"])
        aov_g = safe_growth_pct(k_this["aov"], k_last["aov"]) if (not pd.isna(k_this["aov"]) and not pd.isna(k_last["aov"])) else None

        st.subheader("Ringkasan Periode")
        c1, c2, c3, c4 = st.columns(4)


        def summary_card(title: str, value: str):
            st.markdown(
                f"""
        <div class="card">
          <div class="card-title">{title}</div>
          <div class="card-value" style="font-size:15px">{value}</div>
        </div>
        """,
                unsafe_allow_html=True,
            )


        with c1:
            summary_card(label_last, f"{df_last['TGL'].min()} → {df_last['TGL'].max()}" if len(df_last) else "-")
        with c2:
            summary_card(label_this, f"{df_this['TGL'].min()} → {df_this['TGL'].max()}" if len(df_this) else "-")
        with c3:
            summary_card("Rows Periode Lalu", f"{len(df_last):,}".replace(",", "."))
        with c4:
            summary_card("Rows Periode Ini", f"{len(df_this):,}".replace(",", "."))

        st.markdown("<hr/>", unsafe_allow_html=True)

        st.subheader("KPI Utama")
        kpi_html = f"""
        <div class="kpi-grid">
          <div class="card">
            <div class="card-title">Total Sales (Periode Ini)</div>
            <div class="card-value">{format_idr(k_this["sales"])}</div>
            <div class="card-sub {kpi_delta_class(sales_g)}">{growth_label(sales_g)}</div>
          </div>
          <div class="card">
            <div class="card-title">Orders</div>
            <div class="card-value">{format_int_id(k_this["orders"])}</div>
            <div class="card-sub {kpi_delta_class(orders_g)}">{growth_label(orders_g)}</div>
          </div>
          <div class="card">
            <div class="card-title">Total Qty</div>
            <div class="card-value">{format_int_id(k_this["qty"])}</div>
            <div class="card-sub {kpi_delta_class(qty_g)}">{growth_label(qty_g)}</div>
          </div>
          <div class="card">
            <div class="card-title">AOV</div>
            <div class="card-value">{format_idr(k_this["aov"])}</div>
            <div class="card-sub {kpi_delta_class(aov_g)}">{growth_label(aov_g)}</div>
          </div>
          <div class="card">
            <div class="card-title">Retur (lines)</div>
            <div class="card-value">{format_int_id(k_this["returns"])}</div>
          </div>
        </div>
        """.replace(",", ".")
        st.markdown(kpi_html, unsafe_allow_html=True)

        # =========================
        # NEW TABLE #1: % TEAM TURUN
        # =========================
        st.markdown("<hr/>", unsafe_allow_html=True)
        st.subheader("📉 Ringkasan Pergerakan TEAM (QTY)")
        down_tbl = team_down_ratio_table_cached(df_last, df_this)
        st.dataframe(
            down_tbl,
            use_container_width=True,
            hide_index=True,
            column_config={"% TEAM Turun": st.column_config.NumberColumn(format="%.2f")},
        )

        st.markdown("<hr/>", unsafe_allow_html=True)

        COLOR_MAP_PERIOD = {
            "Bulan Ini": "#1f77b4",
            "Bulan Lalu": "#aec7e8",
        }

        # =========================
        # Trend chart (Day-of-Month comparison)
        # =========================
        st.subheader(f"Tren Harian ({metric_name})")


        def day_of_month_series(df: pd.DataFrame, label: str) -> pd.DataFrame:
            tmp = df.copy()
            tmp["DAY"] = pd.to_datetime(tmp["TGL"]).dt.day
            g = tmp.groupby("DAY", as_index=False).agg(VALUE=(metric_col, "sum"))
            g["PERIODE"] = label
            return g


        all_days = pd.DataFrame({"DAY": list(range(1, 32))})
        trend_dom = pd.concat(
            [
                all_days.merge(day_of_month_series(df_last, label_last), on="DAY", how="left").assign(PERIODE=label_last),
                all_days.merge(day_of_month_series(df_this, label_this), on="DAY", how="left").assign(PERIODE=label_this),
            ],
            ignore_index=True,
        )
        trend_dom["VALUE"] = trend_dom["VALUE"].fillna(0.0)

        if px is not None:
            if show_point_labels:
                trend_dom["LABEL"] = trend_dom["VALUE"].apply(compact_number)
                fig = px.line(
                    trend_dom,
                    x="DAY",
                    y="VALUE",
                    color="PERIODE",
                    markers=True,
                    text="LABEL",
                    color_discrete_map=COLOR_MAP_PERIOD,
                )
                fig.update_traces(textposition="top center")
            else:
                fig = px.line(
                    trend_dom,
                    x="DAY",
                    y="VALUE",
                    color="PERIODE",
                    markers=True,
                    color_discrete_map=COLOR_MAP_PERIOD,
                )

            fig.update_layout(
                xaxis_title="Tanggal (Day of Month)",
                yaxis_title=metric_name,
                legend_title_text="",
                xaxis=dict(dtick=1),
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            pivot_trend = trend_dom.pivot_table(index="DAY", columns="PERIODE", values="VALUE", aggfunc="sum").fillna(0.0)
            st.line_chart(pivot_trend, use_container_width=True)

        # =========================
        # Cumulative chart
        # =========================
        st.markdown("<hr/>", unsafe_allow_html=True)
        st.subheader(f"Statistik Kumulatif ({metric_name})")

        trend_cum = trend_dom.copy()
        trend_cum = trend_cum.sort_values(["PERIODE", "DAY"]).copy()
        trend_cum["CUM_VALUE"] = trend_cum.groupby("PERIODE")["VALUE"].cumsum()

        # super clean compare hover: tampilkan Bulan Ini + Bulan Lalu + Delta dalam satu hover
        trend_compare = (
            trend_cum.pivot(index="DAY", columns="PERIODE", values="CUM_VALUE")
            .reset_index()
            .rename_axis(None, axis=1)
        )

        if "Bulan Ini" not in trend_compare.columns:
            trend_compare["Bulan Ini"] = 0
        if "Bulan Lalu" not in trend_compare.columns:
            trend_compare["Bulan Lalu"] = 0

        trend_compare["Bulan Ini"] = trend_compare["Bulan Ini"].fillna(0)
        trend_compare["Bulan Lalu"] = trend_compare["Bulan Lalu"].fillna(0)
        trend_compare["DELTA"] = trend_compare["Bulan Ini"] - trend_compare["Bulan Lalu"]

        trend_cum = trend_cum.merge(
            trend_compare[["DAY", "Bulan Ini", "Bulan Lalu", "DELTA"]],
            on="DAY",
            how="left"
        )

        trend_cum["BULAN_INI_TXT"] = trend_cum["Bulan Ini"].apply(compact_number)
        trend_cum["BULAN_LALU_TXT"] = trend_cum["Bulan Lalu"].apply(compact_number)
        trend_cum["DELTA_TXT"] = trend_cum["DELTA"].apply(
            lambda x: f"+{compact_number(x)}" if x > 0 else compact_number(x)
        )

        if px is not None:
            fig_cum = px.line(
                trend_cum,
                x="DAY",
                y="CUM_VALUE",
                color="PERIODE",
                markers=True,
                color_discrete_map=COLOR_MAP_PERIOD,
                custom_data=["BULAN_INI_TXT", "BULAN_LALU_TXT", "DELTA_TXT"],
            )

            fig_cum.update_traces(
                hovertemplate=(
                    "<b>Hari %{x}</b><br>"
                    "Bulan Ini: %{customdata[0]}<br>"
                    "Bulan Lalu: %{customdata[1]}<br>"
                    "Delta: %{customdata[2]}<extra>%{fullData.name}</extra>"
                )
            )

            fig_cum.update_layout(
                xaxis_title="Tanggal (Day of Month)",
                yaxis_title=f"Kumulatif {metric_name}",
                legend_title_text="",
                xaxis=dict(dtick=1),
            )
            st.plotly_chart(fig_cum, use_container_width=True)
        else:
            pivot_cum = trend_cum.pivot_table(index="DAY", columns="PERIODE", values="CUM_VALUE", aggfunc="sum").fillna(0.0)
            st.line_chart(pivot_cum, use_container_width=True)


        # =========================
        # Pareto + Delta + Comparison
        # =========================
        st.markdown("<hr/>", unsafe_allow_html=True)

        pareto_dim = st.selectbox(
            "Filter Pareto",
            ["PLATFORM", "TEAM", "PRODUCT", "COUNTRY", "BRAND", "TRANSAKSI", "AREA"],
            index=1,
        )

        pareto_top_n = 30

        def build_pareto_comparison(df_this: pd.DataFrame, df_last: pd.DataFrame, dim_col: str, value_col: str, top_n: int = 10):
            this_agg = (
                df_this.groupby(dim_col, as_index=False)
                .agg(THIS_VALUE=(value_col, "sum"))
            )
            last_agg = (
                df_last.groupby(dim_col, as_index=False)
                .agg(LAST_VALUE=(value_col, "sum"))
            )

            comp = this_agg.merge(last_agg, on=dim_col, how="outer").fillna(0.0)
            comp[dim_col] = comp[dim_col].astype(str).str.strip()
            comp = comp[comp[dim_col].ne("")].copy()
            comp = comp.sort_values("THIS_VALUE", ascending=False).head(top_n).copy()

            if comp.empty:
                return comp

            total_this = comp["THIS_VALUE"].sum()
            total_last = comp["LAST_VALUE"].sum()

            comp["THIS_SHARE"] = np.where(total_this != 0, comp["THIS_VALUE"] / total_this * 100.0, 0.0)
            comp["LAST_SHARE"] = np.where(total_last != 0, comp["LAST_VALUE"] / total_last * 100.0, 0.0)
            comp["PARETO_THIS"] = comp["THIS_SHARE"].cumsum()
            comp["PARETO_LAST"] = comp["LAST_SHARE"].cumsum()
            comp["CUM_STORE_COUNT"] = np.arange(1, len(comp) + 1)
            comp["DELTA_SHARE"] = comp["THIS_SHARE"] - comp["LAST_SHARE"]
            comp["DELTA_LABEL"] = comp["DELTA_SHARE"].map(lambda x: f"{x:+.1f}%")

            comp["BAR_HOVER"] = comp.apply(
                lambda r: (
                    f"<b>{r[dim_col]}</b><br>"
                    f"{label_this}: {compact_number(r['THIS_VALUE'])}<br>"
                    f"{label_last}: {compact_number(r['LAST_VALUE'])}<br>"
                    f"Kontribusi {label_this}: {r['THIS_SHARE']:.2f}%<br>"
                    f"Kontribusi {label_last}: {r['LAST_SHARE']:.2f}%<br>"
                    f"Delta kontribusi: {r['DELTA_SHARE']:+.2f}%"
                ),
                axis=1,
            )

            return comp

        pareto_df = build_pareto_comparison(df_this, df_last, pareto_dim, metric_col, pareto_top_n)

        if pareto_df.empty:
            st.info("Belum ada data untuk grafik Pareto pada filter saat ini.")
        else:
            from plotly.subplots import make_subplots
            import plotly.graph_objects as go

            fig_pareto = make_subplots(specs=[[{"secondary_y": True}]])

            fig_pareto.add_trace(
                go.Bar(
                    x=pareto_df[pareto_dim],
                    y=pareto_df["THIS_VALUE"],
                    name=label_this,
                    marker_color="#8ecae6",
                    text=pareto_df["DELTA_LABEL"],
                    textposition="outside",
                    hovertext=pareto_df["BAR_HOVER"],
                    hovertemplate="%{hovertext}<extra></extra>",
                ),
                secondary_y=False,
            )

            fig_pareto.add_trace(
                go.Scatter(
                    x=pareto_df[pareto_dim],
                    y=pareto_df["PARETO_THIS"],
                    name=f"Pareto {label_this}",
                    mode="lines+markers",
                    line=dict(color="#1f77b4", width=3),
                    marker=dict(symbol="circle", size=7, color="#1f77b4"),
                    customdata=pareto_df[["CUM_STORE_COUNT"]],
                    hovertemplate=(
                        "jumlah toko: %{customdata[0]}<br>"
                        f"Pareto {label_this}: %{{y:.2f}}%<extra></extra>"
                    ),
                ),
                secondary_y=True,
            )

            fig_pareto.add_trace(
                go.Scatter(
                    x=pareto_df[pareto_dim],
                    y=pareto_df["PARETO_LAST"],
                    name=f"Pareto {label_last}",
                    mode="lines+markers",
                    line=dict(color="#f59e0b", width=2.5, dash="dash"),
                    marker=dict(symbol="x", size=8, color="#f59e0b"),
                    customdata=pareto_df[["CUM_STORE_COUNT"]],
                    hovertemplate=(
                        "jumlah toko: %{customdata[0]}<br>"
                        f"Pareto {label_last}: %{{y:.2f}}%<extra></extra>"
                    ),
                ),
                secondary_y=True,
            )

            fig_pareto.add_hline(
                y=80,
                line_width=1.5,
                line_dash="solid",
                line_color="#3b82f6",
                opacity=0.85,
                secondary_y=True,
            )

            fig_pareto.update_layout(
                title="Pareto + Delta + Comparison (This Vs Last Month)",
                hovermode="x unified",
                legend_title_text="",
                xaxis_title=pareto_dim.title(),
                yaxis_title=f"{metric_name} ({label_this})",
                margin=dict(l=40, r=40, t=70, b=40),
            )

            fig_pareto.update_yaxes(
                title_text=f"{metric_name} ({label_this})",
                secondary_y=False,
                showgrid=True,
                gridcolor="rgba(0,0,0,0.08)",
            )
            fig_pareto.update_yaxes(
                title_text="Cumulative (%)",
                range=[0, 105],
                ticksuffix="%",
                secondary_y=True,
                showgrid=False,
            )

            st.plotly_chart(fig_pareto, use_container_width=True)


        st.markdown("<hr/>", unsafe_allow_html=True)

        # =========================
        # NEW TABLE #2: TEAM DRIVER ANALYSIS
        # =========================
        st.markdown("<hr/>", unsafe_allow_html=True)
        st.subheader("🧠 Analisis Penyebab Perubahan")
        st.caption("Untuk TEAM yang TURUN: ditampilkan top driver yang paling narik turun. Untuk TEAM yang NAIK: top driver yang paling narik naik.")

        topk = st.slider("Top driver per kategori", 1, 10, 3, 1)
        analysis_df = team_driver_analysis_table_cached(df_last, df_this, top_k=topk)

        st.dataframe(
            style_growth_pct_df(analysis_df),
            use_container_width=True,
            height=520,
        )


        # =========================
        # NEW CARD: BRAND DELTA ANALYSIS
        # =========================
        st.markdown("<hr/>", unsafe_allow_html=True)
        st.subheader("🏷️ Analisa Brand")
        st.caption("Delta QTY per BRAND dibanding periode lalu, dibagi berdasarkan kategori PRODUCT.")

        brand_analysis_df = brand_delta_analysis_table_cached(df_last, df_this)
        brand_analysis_raw_df = brand_delta_analysis_raw_cached(df_last, df_this)
        render_clickable_brand_delta_table(brand_analysis_raw_df)
        render_brand_team_detail_card(df_last, df_this)


        # =========================
        # Top tables
        # =========================
        c1, c2 = st.columns(2)
        with c1:
            small_title(f"Top {top_n} TEAM", f"(by {metric_col})")
            render_html_table(top_table_cached(df_this, df_last, "TEAM", metric_col, top_n))
        with c2:
            small_title(f"Top {top_n} PRODUCT", f"(by {metric_col})")
            render_html_table(top_table_cached(df_this, df_last, "PRODUCT", metric_col, top_n))

        c3, c4 = st.columns(2)
        with c3:
            small_title(f"Top {top_n} BRAND", f"(by {metric_col})")
            render_html_table(top_table_cached(df_this, df_last, "BRAND", metric_col, top_n))
        with c4:
            small_title(f"Top {top_n} TRANSAKSI", f"(by {metric_col})")
            render_html_table(top_table_cached(df_this, df_last, "TRANSAKSI", metric_col, top_n))

        c5, c6 = st.columns(2)
        with c5:
            small_title(f"Top {top_n} SKU", "(SPESIFIKASI)")
            render_html_table(top_table_cached(df_this, df_last, "SPESIFIKASI", metric_col, top_n), table_class="sku-table")
        with c6:
            small_title(f"Top {top_n} PLATFORM", "(sumber: NAMA CUSTOMER)")
            render_html_table(top_table_cached(df_this, df_last, "PLATFORM", metric_col, top_n))

        st.markdown("<hr/>", unsafe_allow_html=True)

        # =========================
        # TEAM PERFORMANCE (3 columns)
        # =========================
        st.subheader("📊 Team Performance Insight (QTY)")

        team_last = df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum"))
        team_this = df_this.groupby("TEAM", as_index=False).agg(QTY_INI=("QTY", "sum"))
        team = team_last.merge(team_this, on="TEAM", how="outer").fillna(0.0)
        team["DELTA_QTY"] = team["QTY_INI"] - team["QTY_LALU"]
        team["GROWTH_PCT"] = team.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

        under = team[team["QTY_INI"] < 30].copy().sort_values(["QTY_INI", "GROWTH_PCT"], ascending=[True, True])

        oto_team = (
            df_this.groupby("TEAM", as_index=False)
            .agg(
                OTO_YES_LINES=("OTO_YES", "sum"),
                TOTAL_LINES=("OTO_YES", "count"),
                QTY_INI=("QTY", "sum"),
            )
        )
        oto_team["OTO_RATE"] = np.where(oto_team["TOTAL_LINES"] > 0, oto_team["OTO_YES_LINES"] / oto_team["TOTAL_LINES"] * 100.0, 0.0)
        oto_team = oto_team.sort_values(["OTO_YES_LINES", "OTO_RATE"], ascending=[False, False])

        top_all = team.copy().sort_values(["GROWTH_PCT", "QTY_INI"], ascending=[False, False])


        def prep_team_view(df_in: pd.DataFrame) -> pd.DataFrame:
            d = df_in.copy()
            d["QTY Lalu"] = d["QTY_LALU"].map(format_int_id)
            d["QTY Ini"] = d["QTY_INI"].map(format_int_id)
            d["Delta"] = d["DELTA_QTY"].map(format_int_id)
            d["Growth %"] = d["GROWTH_PCT"].apply(lambda x: float(x) if (x is not None and not pd.isna(x)) else np.nan)
            return d[["TEAM", "QTY Lalu", "QTY Ini", "Delta", "Growth %"]]


        def prep_oto_view(df_in: pd.DataFrame) -> pd.DataFrame:
            d = df_in.copy()
            d["OTO YES (lines)"] = d["OTO_YES_LINES"].astype(int)
            d["OTO Rate %"] = d["OTO_RATE"]
            d["QTY (Periode Ini)"] = d["QTY_INI"].map(format_int_id)
            return d[["TEAM", "OTO YES (lines)", "OTO Rate %", "QTY (Periode Ini)"]]


        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("### 🏆 Top Performer (All TEAM)")
            st.caption("Scroll & klik header kolom untuk sort (Growth% / QTY).")
            df_top = prep_team_view(top_all).copy()
            st.dataframe(style_growth_pct_df(df_top), use_container_width=True, height=420)

        with col2:
            st.markdown("### ⚠️ Under Perform (QTY < 30)")
            st.caption("Team dengan QTY periode ini di bawah 30.")
            df_under = prep_team_view(under).copy()
            st.dataframe(style_growth_pct_df(df_under), use_container_width=True, height=420)

        with col3:
            st.markdown('### 🚫 Team sering OTO "YES"')
            st.caption('Urut berdasarkan jumlah OTO == "YES" (periode ini).')
            st.dataframe(
                prep_oto_view(oto_team),
                use_container_width=True,
                height=420,
                column_config={"OTO Rate %": st.column_config.NumberColumn(format="%.2f")},
                hide_index=True,
            )

        # =========================
        # EXTRA INSIGHT (tambahan)
        # =========================
        st.markdown("<hr/>", unsafe_allow_html=True)
        st.subheader("🧠 Insight Tambahan (QTY / Retur / AREA)")

        # ---------- 1) TEAM qty besar tapi turun ----------
        st.markdown("### 1) TEAM QTY besar tapi turun")

        team_qty = (
            df_this.groupby("TEAM", as_index=False)
            .agg(QTY_INI=("QTY", "sum"))
            .merge(
                df_last.groupby("TEAM", as_index=False).agg(QTY_LALU=("QTY", "sum")),
                on="TEAM",
                how="left",
            )
            .fillna(0.0)
        )
        team_qty["DELTA"] = team_qty["QTY_INI"] - team_qty["QTY_LALU"]
        team_qty["GROWTH_PCT"] = team_qty.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

        # ambil kandidat TEAM dengan QTY ini terbesar, lalu filter yang turun
        TOP_BIG = 30  # bisa kamu ubah 20/50
        big_down = (
            team_qty.sort_values("QTY_INI", ascending=False)
            .head(TOP_BIG)
            .query("DELTA < 0")
            .copy()
            .sort_values("DELTA", ascending=True)
            .head(20)
        )

        if len(big_down) == 0:
            st.info("Tidak ada TEAM 'QTY besar tapi turun' pada filter & periode saat ini.")
        else:
            big_down_view = big_down.copy()
            big_down_view["QTY Ini"] = big_down_view["QTY_INI"].map(format_int_id)
            big_down_view["QTY Lalu"] = big_down_view["QTY_LALU"].map(format_int_id)
            big_down_view["Delta"] = big_down_view["DELTA"].map(format_int_id)
            big_down_view["Growth"] = big_down_view["GROWTH_PCT"].apply(growth_badge_html)
            big_down_view = big_down_view[["TEAM", "QTY Ini", "QTY Lalu", "Delta", "Growth"]]
            render_html_table(big_down_view)

        # ---------- 2) TEAM paling banyak retur ----------
        st.markdown("### 2) TEAM paling banyak retur")

        ret_this = df_this[df_this["ROW_TYPE"] == "RETUR"].copy()

        if len(ret_this) == 0:
            st.info("Tidak ada data RETUR pada periode ini (berdasarkan kolom STATUS).")
        else:
            ret_team = (
                ret_this.groupby("TEAM", as_index=False)
                .agg(
                    Retur_Lines=("TEAM", "count"),
                    Retur_QTY=("QTY", "sum"),
                )
                .sort_values(["Retur_Lines", "Retur_QTY"], ascending=[False, True])
                .head(20)
                .copy()
            )
            # QTY retur biasanya negatif, biar enak lihat pakai ABS
            ret_team["Retur_QTY (abs)"] = ret_team["Retur_QTY"].abs().map(format_int_id)
            ret_team["Retur_Lines"] = ret_team["Retur_Lines"].map(format_int_id)
            ret_team = ret_team[["TEAM", "Retur_Lines", "Retur_QTY (abs)"]]
            render_html_table(ret_team)

        # ---------- 3) Perform AREA (Naik/Turun) ----------
        st.markdown("### 3) Perform AREA (Naik / Turun)")

        if "AREA" not in df_this.columns or "AREA" not in df_last.columns:
            st.warning("Kolom 'AREA' tidak ditemukan di data yang terbaca. (Pastikan Excel punya kolom AREA & tidak terbuang saat cleaning).")
        else:
            area_this = df_this.groupby("AREA", as_index=False).agg(QTY_INI=("QTY", "sum"))
            area_last = df_last.groupby("AREA", as_index=False).agg(QTY_LALU=("QTY", "sum"))
            area = area_this.merge(area_last, on="AREA", how="outer").fillna(0.0)

            area["DELTA"] = area["QTY_INI"] - area["QTY_LALU"]
            area["GROWTH_PCT"] = area.apply(lambda r: safe_growth_pct(r["QTY_INI"], r["QTY_LALU"]), axis=1)

            # tampilkan top naik & top turun
            top_up = area.sort_values("DELTA", ascending=False).head(10).copy()
            top_dn = area.sort_values("DELTA", ascending=True).head(10).copy()

            colA, colB = st.columns(2)
            with colA:
                st.markdown("#### 🔼 Top AREA Naik (QTY)")
                v = top_up.copy()
                v["QTY Ini"] = v["QTY_INI"].map(format_int_id)
                v["QTY Lalu"] = v["QTY_LALU"].map(format_int_id)
                v["Delta"] = v["DELTA"].map(format_int_id)
                v["Growth"] = v["GROWTH_PCT"].apply(growth_badge_html)
                v = v[["AREA", "QTY Ini", "QTY Lalu", "Delta", "Growth"]]
                render_html_table(v)

            with colB:
                st.markdown("#### 🔽 Top AREA Turun (QTY)")
                v = top_dn.copy()
                v["QTY Ini"] = v["QTY_INI"].map(format_int_id)
                v["QTY Lalu"] = v["QTY_LALU"].map(format_int_id)
                v["Delta"] = v["DELTA"].map(format_int_id)
                v["Growth"] = v["GROWTH_PCT"].apply(growth_badge_html)
                v = v[["AREA", "QTY Ini", "QTY Lalu", "Delta", "Growth"]]
                render_html_table(v)

    _render_analisa_penjualan_app_inner()


def render_analisa_produk_app():
    """Embedded processor: Analisa Produk"""
    import io
    import re
    from typing import List

    import numpy as np
    import pandas as pd
    import streamlit as st


    APP_TITLE = "Analisa Produk"


    PERIODS = ["7DAY", "14DAY", "30DAY"]
    DIVISIONS = ["DIV03", "DIV04", "DIV05"]

    MPLSSR_DIV_COLS = {
        "7DAY": {"DIV03": "03 OLP", "DIV04": "04 MOD", "DIV05": "05 OLR"},
        "14DAY": {"DIV03": "03 OLP.1", "DIV04": "04 MOD.1", "DIV05": "05 OLR.1"},
        "30DAY": {"DIV03": "03 OLP.2", "DIV04": "04 MOD.2", "DIV05": "05 OLR.2"},
    }

    VALID_PRICELIST_SHEETS = [
        "LAPTOP",
        "TELCO",
        "PC HOM ELE",
        "SOF COM SUP",
        "ACC",
        "SER OTH CON",
    ]

    PRICE_SEGMENTS = [
        (0, 1_000_000, "< 1 JUTA"),
        (1_000_000, 1_500_000, "1 - 1.5 JUTA"),
        (1_500_000, 2_000_000, "1.5 - 2 JUTA"),
        (2_000_000, 2_500_000, "2 - 2.5 JUTA"),
        (2_500_000, 3_000_000, "2.5 - 3 JUTA"),
        (3_000_000, 4_000_000, "3 - 4 JUTA"),
        (4_000_000, 5_000_000, "4 - 5 JUTA"),
        (5_000_000, 7_000_000, "5 - 7 JUTA"),
        (7_000_000, 10_000_000, "7 - 10 JUTA"),
        (10_000_000, 12_500_000, "10 - 12.5 JUTA"),
        (12_500_000, 15_000_000, "12.5 - 15 JUTA"),
        (15_000_000, 20_000_000, "15 - 20 JUTA"),
        (20_000_000, 25_000_000, "20 - 25 JUTA"),
        (25_000_000, 30_000_000, "25 - 30 JUTA"),
        (30_000_000, 40_000_000, "30 - 40 JUTA"),
        (40_000_000, np.inf, "40 JUTA - UP"),
    ]

    SEGMENT_ORDER = [label for _, _, label in PRICE_SEGMENTS]

    st.markdown(
        """
        <style>
        .block-container {padding-top: 1rem; padding-bottom: 1rem;}
        .upload-card-wrap {
            border: 1px solid #e5e7eb;
            border-radius: 12px;
            background: #ffffff;
            padding: 14px;
            margin-bottom: 16px;
        }
        .table-card {
            border: 1px solid #d9d9d9;
            border-radius: 8px;
            background: #fff;
            padding: 10px;
            margin-bottom: 14px;
        }
        .alert-wrap {
            border: 1px solid #d9d9d9;
            border-radius: 8px;
            background: #fff;
            overflow: auto;
            max-height: 520px;
        }
        table.alert-table {
            border-collapse: collapse;
            width: max-content;
            min-width: 100%;
            table-layout: fixed;
            font-size: 12px;
        }
        table.alert-table th, table.alert-table td {
            border: 1px solid #e5e7eb;
            padding: 6px 8px;
            text-align: left;
            white-space: nowrap;
        }
        table.alert-table thead th {
            position: sticky;
            top: 0;
            background: #f8fafc;
            z-index: 2;
        }
        .main-fixed-wrap {
            border: 1px solid #d9d9d9;
            border-radius: 8px;
            background: #fff;
            overflow: auto;
            max-height: 520px;
        }
        table.main-fixed {
            border-collapse: collapse;
            width: max-content;
            min-width: 100%;
            table-layout: fixed;
            font-size: 12px;
        }
        table.main-fixed th, table.main-fixed td {
            border: 1px solid #e5e7eb;
            padding: 6px 8px;
            text-align: left;
            white-space: nowrap;
        }
        table.main-fixed thead th {
            position: sticky;
            top: 0;
            background: #f8fafc;
            z-index: 2;
        }
        table.main-fixed th:nth-child(1),
        table.main-fixed td:nth-child(1) { min-width: 180px; }
        table.main-fixed th:nth-child(2),
        table.main-fixed td:nth-child(2) { min-width: 90px; }
        table.main-fixed th:nth-child(3),
        table.main-fixed td:nth-child(3) { min-width: 80px; }
        table.main-fixed th:nth-child(4),
        table.main-fixed td:nth-child(4) {
            min-width: 420px;
            max-width: 420px;
            white-space: normal;
            word-break: break-word;
            line-height: 1.25;
        }
        table.main-fixed th:nth-child(5),
        table.main-fixed td:nth-child(5) { min-width: 80px; }
        table.main-fixed th:nth-child(6),
        table.main-fixed td:nth-child(6),
        table.main-fixed th:nth-child(7),
        table.main-fixed td:nth-child(7),
        table.main-fixed th:nth-child(8),
        table.main-fixed td:nth-child(8),
        table.main-fixed th:nth-child(9),
        table.main-fixed td:nth-child(9) { min-width: 70px; }
        .bg-red {
            background: #ffebee;
            color: #c62828;
            font-weight: 700;
        }
        .status-refill {
            background: #ffebee;
            color: #c62828;
            font-weight: 700;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    def normalize_text(series: pd.Series) -> pd.Series:
        return (
            series.astype(str)
            .str.strip()
            .str.upper()
            .replace({"NAN": np.nan, "NONE": np.nan, "": np.nan})
        )

    def to_num(series: pd.Series) -> pd.Series:
        return pd.to_numeric(series, errors="coerce")

    def format_thousands_id(value):
        try:
            num = pd.to_numeric(value, errors="coerce")
            if pd.isna(num):
                return ""
            return f"{int(round(float(num) / 1000)):,}".replace(",", ".")
        except Exception:
            return ""

    def format_units_id(value):
        try:
            num = pd.to_numeric(value, errors="coerce")
            if pd.isna(num):
                return ""
            return f"{int(round(float(num))):,}".replace(",", ".")
        except Exception:
            return ""

    def price_segment(price: float) -> str:
        if pd.isna(price):
            return "UNKNOWN"
        for low, high, label in PRICE_SEGMENTS:
            if low <= float(price) < high:
                return label
        return "UNKNOWN"

    def segment_sort_key(label: str) -> int:
        try:
            return SEGMENT_ORDER.index(label)
        except ValueError:
            return len(SEGMENT_ORDER)

    def first_row_contains_text(df: pd.DataFrame, text: str):
        target = str(text).strip().upper()
        for idx in range(len(df)):
            row_text = df.iloc[idx].astype(str).str.upper()
            if row_text.str.contains(target, na=False).any():
                return idx
        return None

    def _ffill_header(values: List) -> List:
        out = []
        last = None
        for v in values:
            if pd.notna(v) and str(v).strip() != "":
                last = str(v).strip().upper()
                out.append(last)
            else:
                out.append(last)
        return out


    def _norm_header_cell(value) -> str:
        if pd.isna(value) or str(value).strip() == "":
            return ""
        return str(value).strip().upper()

    def find_div05_stock_columns(columns: List[str], excel_row3_raw: List, excel_row4_raw: List) -> List[str]:
        # Header bisa merge cell, jadi harus di-forward fill ke kanan.
        row3_vals = [_norm_header_cell(v) for v in _ffill_header(excel_row3_raw)]
        row4_vals = [_norm_header_cell(v).replace(" ", "") for v in _ffill_header(excel_row4_raw)]

        # 05 OLR dimulai dari area RAM di row 3 dan dihitung sampai kolom paling kanan.
        ram_start = next((i for i, v in enumerate(row3_vals) if "RAM" in v), None)

        # Fallback: kalau RAM tidak ditemukan, mulai dari kolom pertama 5B di row 4.
        if ram_start is None:
            ram_start = next((i for i, v in enumerate(row4_vals) if v == "5B"), None)

        if ram_start is None:
            return []

        return [columns[i] for i in range(ram_start, len(columns)) if 0 <= i < len(columns)]

    def area_code_matches(value, prefixes: List[str]) -> bool:
        if pd.isna(value):
            return False
        txt = str(value).strip().upper().replace(" ", "")
        return any(txt.startswith(p) for p in prefixes)

    def normalize_warehouse_code(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace(" ", "")
        txt = re.sub(r"0+(?=\d)", "", txt)
        txt = txt.replace("0A", "A").replace("0B", "B").replace("0C", "C")
        return txt


    def normalize_sales_pivot_gudang(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace(" ", "")
        m = re.match(r"([A-Z]+)0*(\d+[A-Z]?)$", txt)
        if m:
            prefix = m.group(1)
            suffix = m.group(2)
            return f"{prefix} {suffix}"
        return txt


    def normalize_display_team_code(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace("_", "").replace(" ", "")
        m = re.match(r"([A-Z]+)0*(\d+)([A-Z])?$", txt)
        if m:
            prefix = m.group(1)
            number = int(m.group(2))
            suffix = m.group(3) if m.group(3) else "A"
            return f"{prefix} {number}{suffix}"
        return txt


    def normalize_team_code(value) -> str:
        if pd.isna(value):
            return np.nan
        txt = str(value).strip().upper()
        if "-" in txt:
            txt = txt.split("-", 1)[1]
        txt = txt.replace(" ", "")
        m = re.match(r"([A-Z]+)0*(\d+)$", txt)
        if m:
            return f"{m.group(1)} {int(m.group(2))}A"
        m = re.match(r"([A-Z]+)0*(\d+)([A-Z])$", txt)
        if m:
            return f"{m.group(1)} {int(m.group(2))}{m.group(3)}"
        return txt


    def normalize_combined_warehouse_code(group, wh=None) -> str:
        parts = []
        for val in [group, wh]:
            if pd.isna(val) or str(val).strip() == "":
                continue
            txt = str(val).strip().upper()
            txt = txt.replace(" ", "")
            parts.append(txt)
        return "".join(parts)

    def normalize_team_lookup_key(value) -> str:
        if pd.isna(value):
            return ""
        return str(value).strip().upper().replace(" ", "").replace("-", "")

    def ensure_datetime(series: pd.Series) -> pd.Series:
        return pd.to_datetime(series, errors="coerce", dayfirst=True)


    def get_period_date_range(df: pd.DataFrame, period: str):
        if df is None or df.empty or "TGL" not in df.columns:
            return pd.NaT, pd.NaT

        tgl = pd.to_datetime(df["TGL"], errors="coerce").dropna()
        if tgl.empty:
            return pd.NaT, pd.NaT

        max_date = tgl.max().normalize()
        days_map = {"7DAY": 7, "14DAY": 14, "30DAY": 30}
        days = days_map.get(period, 7)
        start_date = max_date - pd.Timedelta(days=days - 1)
        return start_date, max_date


    # =========================================================
    # MPLSSR
    # =========================================================
    def load_mplssr(file) -> pd.DataFrame:
        df = pd.read_excel(file, sheet_name="ALL", header=1)
        df = df.iloc[4:].copy().reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.duplicated()]

        base_cols = ["PRODUCT", "BRAND", "KODE BARANG", "SPESIFIKASI"]
        for c in base_cols:
            if c not in df.columns:
                df[c] = np.nan

        div_cols = []
        for p in PERIODS:
            for col in MPLSSR_DIV_COLS[p].values():
                if col in df.columns:
                    div_cols.append(col)

        df = df[base_cols + div_cols].copy()

        for c in base_cols:
            df[c] = normalize_text(df[c])

        df = df[df["KODE BARANG"].notna()].copy()
        df = df[~df["KODE BARANG"].isin(["TOTAL", "SHARE%"])]

        rows = []
        for period in PERIODS:
            for div, col in MPLSSR_DIV_COLS[period].items():
                if col not in df.columns:
                    continue
                tmp = df[["PRODUCT", "BRAND", "KODE BARANG", "SPESIFIKASI"]].copy()
                tmp["PERIOD"] = period
                tmp["DIVISION"] = div
                tmp["QTY"] = to_num(df[col]).fillna(0)
                rows.append(tmp)

        out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(
            columns=["PRODUCT", "BRAND", "KODE BARANG", "SPESIFIKASI", "PERIOD", "DIVISION", "QTY"]
        )
        out["MERGE_KEY"] = normalize_text(out["KODE BARANG"])
        out["SKU NO"] = out["MERGE_KEY"]
        return out

    # =========================================================
    # PRICELIST
    # =========================================================

    def parse_pricelist_sheet(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
        raw = xls.parse(sheet_name=sheet_name, header=None).copy()

        if sheet_name.upper() == "LAPTOP":
            coming_idx = first_row_contains_text(raw, "COMING")
            end_coming_idx = first_row_contains_text(raw, "END COMING")
            if coming_idx is not None and end_coming_idx is not None and end_coming_idx >= coming_idx:
                raw = raw.drop(index=range(coming_idx, end_coming_idx + 1)).reset_index(drop=True)

        row1 = raw.iloc[1].tolist()
        row2_raw = raw.iloc[2].tolist()
        row2 = _ffill_header(row2_raw)
        row3_raw = raw.iloc[3].tolist()
        row3 = [str(x).strip().upper() if pd.notna(x) and str(x).strip() != "" else None for x in row3_raw]

        columns = []
        for i, v in enumerate(row1):
            v1 = str(v).strip().upper() if pd.notna(v) and str(v).strip() != "" else None
            if v1 is not None:
                columns.append(v1)
            elif row2[i] is not None and row3[i] is not None:
                columns.append(f"{row2[i]}__{row3[i]}")
            elif row3[i] is not None:
                columns.append(row3[i])
            else:
                columns.append(f"COL_{i}")

        df = raw.iloc[5:].copy().reset_index(drop=True)
        df.columns = columns

        for c in ["SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "TOT", "M3"]:
            if c not in df.columns:
                df[c] = np.nan

        df["SKU NO"] = normalize_text(df["SKU NO"])
        df["PRODUCT"] = normalize_text(df["PRODUCT"])
        df["KODEBARANG"] = normalize_text(df["KODEBARANG"])
        df["SPESIFIKASI"] = normalize_text(df["SPESIFIKASI"])

        df = df[df["KODEBARANG"].notna()].copy()
        df = df[~df["KODEBARANG"].isin(["TOTAL"])]

        stock03_cols = [columns[i] for i, area in enumerate(row3) if area_code_matches(area, ["3", "03"])]
        stock04_cols = [columns[i] for i, area in enumerate(row3) if area_code_matches(area, ["4", "04"])]
        stock05_cols = find_div05_stock_columns(columns, row2_raw, row3_raw)

        df["PRICE"] = to_num(df["M3"]) * 1000
        df["STOK_DIV03"] = df[stock03_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) if stock03_cols else 0
        df["STOK_DIV04"] = df[stock04_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) if stock04_cols else 0
        df["STOK_DIV05"] = df[stock05_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) if stock05_cols else 0
        df["CATEGORY"] = sheet_name.upper()
        df["PRICE_SEGMENT"] = df["PRICE"].apply(price_segment)
        df["MERGE_KEY"] = df["KODEBARANG"]

        return df[[
            "SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "PRICE",
            "STOK_DIV03", "STOK_DIV04", "STOK_DIV05",
            "CATEGORY", "PRICE_SEGMENT", "MERGE_KEY"
        ]]

    def load_pricelist(file) -> pd.DataFrame:
        xls = pd.ExcelFile(file)
        sheets = [s for s in xls.sheet_names if s.upper() in VALID_PRICELIST_SHEETS]
        frames = [parse_pricelist_sheet(xls, s) for s in sheets]
        if not frames:
            return pd.DataFrame(columns=[
                "SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "PRICE",
                "STOK_DIV03", "STOK_DIV04", "STOK_DIV05",
                "CATEGORY", "PRICE_SEGMENT", "MERGE_KEY"
            ])
        out = pd.concat(frames, ignore_index=True)
        out = out.loc[:, ~out.columns.duplicated()]
        out = out.drop_duplicates(subset=["MERGE_KEY"], keep="first")
        return out

    # =========================================================
    # PRICELIST WITH WAREHOUSES
    # =========================================================



    def parse_pricelist_sheet_with_warehouses(xls: pd.ExcelFile, sheet_name: str):
        raw = xls.parse(sheet_name=sheet_name, header=None).copy()

        if sheet_name.upper() == "LAPTOP":
            col_c = raw.iloc[:, 2].astype(str).str.upper().str.strip() if raw.shape[1] > 2 else pd.Series(dtype=str)
            coming_candidates = col_c[col_c.str.contains("COMING", na=False)]
            end_coming_candidates = col_c[col_c.str.contains("END COMING", na=False)]
            if not coming_candidates.empty and not end_coming_candidates.empty:
                coming_idx = int(coming_candidates.index.min())
                end_coming_idx = int(end_coming_candidates.index.max())
                if end_coming_idx >= coming_idx:
                    raw = raw.drop(index=range(coming_idx, end_coming_idx + 1)).reset_index(drop=True)

        # Struktur pricelist:
        # - Row 2: header utama
        # - Row 3: kode gudang utama (mis. RIF, OA)
        # - Row 4: sub-kode gudang (mis. 1A, 2A)
        row1 = raw.iloc[1].tolist()
        row2 = _ffill_header(raw.iloc[2].tolist())
        row3 = _ffill_header(raw.iloc[2].tolist())
        row4 = _ffill_header(raw.iloc[3].tolist()) if len(raw) > 3 else [None] * len(row1)

        columns = []
        warehouse_meta = []

        for i, v in enumerate(row1):
            v1 = str(v).strip().upper() if pd.notna(v) and str(v).strip() != "" else None
            group = str(row3[i]).strip().upper() if pd.notna(row3[i]) and str(row3[i]).strip() != "" else None
            wh = str(row4[i]).strip().upper() if pd.notna(row4[i]) and str(row4[i]).strip() != "" else None

            if v1 is not None:
                columns.append(v1)
                warehouse_meta.append((None, None, None, None))
                continue

            group_clean = _norm_header_cell(group)
            wh_clean = _norm_header_cell(wh).replace(" ", "")

            combined_key = normalize_team_lookup_key(f"{group_clean} {wh_clean}") if group_clean and wh_clean else ""
            combined_label = f"{group_clean} {wh_clean}".strip() if group_clean and wh_clean else (group_clean or "")

            if combined_key:
                columns.append(f"{combined_key}__{i}")
                warehouse_meta.append((group_clean, wh_clean, combined_label, combined_key))
            elif group_clean:
                columns.append(f"{group_clean}__{i}")
                warehouse_meta.append((group_clean, None, group_clean, normalize_team_lookup_key(group_clean)))
            else:
                columns.append(f"COL_{i}")
                warehouse_meta.append((None, None, None, None))

        df = raw.iloc[4:].copy().reset_index(drop=True)
        df.columns = columns

        for c in ["SKU NO", "PRODUCT", "KODEBARANG", "SPESIFIKASI", "M3"]:
            if c not in df.columns:
                df[c] = np.nan

        df["SKU NO"] = normalize_text(df["SKU NO"])
        df["PRODUCT"] = normalize_text(df["PRODUCT"])
        df["KODEBARANG"] = normalize_text(df["KODEBARANG"])
        df["SPESIFIKASI"] = normalize_text(df["SPESIFIKASI"])

        df = df[df["KODEBARANG"].notna()].copy()
        df = df[~df["KODEBARANG"].isin(["TOTAL"])].copy()
        df["PRICE"] = to_num(df["M3"]) * 1000

        warehouse_stock_cols = {}
        default_stock_cols = []
        jkt_stock_cols = []

        for i, col in enumerate(columns):
            group, wh, combined_label, combined_key = warehouse_meta[i]
            if group is None:
                continue

            group_key = normalize_team_lookup_key(group)

            if group_key == "DEFAULT":
                default_stock_cols.append(col)
            if group_key == "JKT":
                jkt_stock_cols.append(col)
            if combined_key:
                warehouse_stock_cols.setdefault(combined_key, []).append(col)

        if "BRAND" not in df.columns:
            df["BRAND"] = np.nan
        df["BRAND"] = normalize_text(df["BRAND"])
        df["PRICE_SEGMENT"] = df["PRICE"].apply(price_segment)

        mapped_cols = []
        for cols in warehouse_stock_cols.values():
            mapped_cols.extend(cols)

        keep_cols = [
            "SKU NO", "PRODUCT", "BRAND", "KODEBARANG", "SPESIFIKASI", "PRICE", "PRICE_SEGMENT"
        ] + list(dict.fromkeys(default_stock_cols + jkt_stock_cols + mapped_cols))

        out = df[keep_cols].copy()
        out = out.loc[:, ~out.columns.duplicated()].copy()
        out["DEFAULT_STOCK_TOTAL"] = (
            df.loc[:, ~df.columns.duplicated()][default_stock_cols]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .sum(axis=1)
            if default_stock_cols else 0
        )
        return out, warehouse_stock_cols

    def load_pricelist_with_warehouses(file):
        xls = pd.ExcelFile(file)
        sheets = [s for s in xls.sheet_names if s.upper() in VALID_PRICELIST_SHEETS]
        frames = []
        merged_map = {}
        for s in sheets:
            part, part_map = parse_pricelist_sheet_with_warehouses(xls, s)
            frames.append(part)
            merged_map.update(part_map)

        if not frames:
            return pd.DataFrame(), {}

        out = pd.concat(frames, ignore_index=True)
        out = out.loc[:, ~out.columns.duplicated()]
        out = out.drop_duplicates(subset=["SKU NO", "KODEBARANG"], keep="first").reset_index(drop=True)
        return out, merged_map

    # =========================================================
    # SALES PIVOT
    # =========================================================

    def load_sales_pivot(file) -> pd.DataFrame:
        raw = pd.read_excel(file, header=1).copy()
        raw = raw.iloc[1:].copy().reset_index(drop=True)
        raw.columns = [str(c).strip().upper() for c in raw.columns]
        raw = raw.loc[:, ~pd.Index(raw.columns).duplicated()].copy()

        team_col = next((c for c in raw.columns if c == "TEAM" or "TEAM" in c), None)
        kode_barang_col = next((c for c in raw.columns if "KODE BARANG" in c or "KODEBARANG" in c), None)
        spesifikasi_col = next((c for c in raw.columns if "SPESIFIKASI" in c), None)
        qty_col = next((c for c in raw.columns if c == "QTY" or "QTY" in c or "PCS" in c or "TERJUAL" in c), None)
        tgl_col = next((c for c in raw.columns if c == "TGL" or "TGL" in c or "DATE" in c), None)
        country_col = next((c for c in raw.columns if c == "COUNTRY" or "COUNTRY" in c), None)
        product_col = next((c for c in raw.columns if c == "PRODUCT" or "PRODUCT" in c), None)
        gp_m0_col = next((c for c in raw.columns if c == "GP M0" or "GPM0" in c or "GP M0" in c.replace("_", " ")), None)
        m0_col = next((c for c in raw.columns if c == "M0" or "M0" in c), None)
        m3_col = next((c for c in raw.columns if c == "M3" or "M3" in c), None)
        harga_akhir_col = next((c for c in raw.columns if "HARGA AKHIR" in c or "HARGAAKHIR" in c), None)

        required = {"TEAM": team_col, "KODE BARANG": kode_barang_col, "QTY": qty_col, "TGL": tgl_col}
        missing = [label for label, col in required.items() if col is None]
        if missing:
            raise ValueError(
                f"Format SALES PIVOT tidak cocok. Kolom wajib tidak ditemukan: {missing}. "
                f"Kolom yang terbaca: {list(raw.columns)}. "
                "Pastikan header ada di row 2 dan data mulai row 4."
            )

        use_cols = [team_col, kode_barang_col, qty_col, tgl_col]
        for extra_col in [country_col, product_col, gp_m0_col, spesifikasi_col, m0_col, m3_col, harga_akhir_col]:
            if extra_col is not None and extra_col not in use_cols:
                use_cols.append(extra_col)

        df = raw[use_cols].copy()
        rename_map = {
            team_col: "TEAM_RAW",
            kode_barang_col: "KODE BARANG",
            qty_col: "QTY",
            tgl_col: "TGL",
        }
        if country_col is not None:
            rename_map[country_col] = "COUNTRY"
        if product_col is not None:
            rename_map[product_col] = "PRODUCT"
        if gp_m0_col is not None:
            rename_map[gp_m0_col] = "GP M0"
        if spesifikasi_col is not None:
            rename_map[spesifikasi_col] = "SPESIFIKASI"
        if m0_col is not None:
            rename_map[m0_col] = "M0"
        if m3_col is not None:
            rename_map[m3_col] = "M3"
        if harga_akhir_col is not None:
            rename_map[harga_akhir_col] = "HARGA AKHIR"
        df = df.rename(columns=rename_map)

        if "COUNTRY" not in df.columns:
            df["COUNTRY"] = np.nan
        if "PRODUCT" not in df.columns:
            df["PRODUCT"] = np.nan
        if "GP M0" not in df.columns:
            df["GP M0"] = np.nan
        if "SPESIFIKASI" not in df.columns:
            df["SPESIFIKASI"] = np.nan
        if "M0" not in df.columns:
            df["M0"] = np.nan
        if "M3" not in df.columns:
            df["M3"] = np.nan
        if "HARGA AKHIR" not in df.columns:
            df["HARGA AKHIR"] = np.nan

        df["TEAM"] = normalize_text(df["TEAM_RAW"])
        df["TEAM_KEY"] = df["TEAM_RAW"].apply(normalize_team_code)
        df["COUNTRY"] = normalize_text(df["COUNTRY"])
        df["PRODUCT"] = normalize_text(df["PRODUCT"])
        df["KODE BARANG"] = normalize_text(df["KODE BARANG"])
        df["SPESIFIKASI"] = normalize_text(df["SPESIFIKASI"])
        df["QTY"] = to_num(df["QTY"]).fillna(0)
        df["GP M0"] = to_num(df["GP M0"])
        df["M0"] = to_num(df["M0"])
        df["M3"] = to_num(df["M3"])
        df["HARGA AKHIR"] = to_num(df["HARGA AKHIR"])
        df["TGL"] = ensure_datetime(df["TGL"]).dt.normalize()

        df = df[df["TEAM"].notna()].copy()
        df = df[df["TEAM_KEY"].notna()].copy()
        df = df[df["KODE BARANG"].notna()].copy()
        df = df[df["QTY"] > 0].copy()
        df = df[df["TGL"].notna()].copy()

        if df.empty:
            return pd.DataFrame(columns=["TEAM", "TEAM_KEY", "COUNTRY", "PRODUCT", "KODE BARANG", "SPESIFIKASI", "QTY", "TGL", "GP M0", "M0", "M3", "HARGA AKHIR"])

        return df[["TEAM", "TEAM_KEY", "COUNTRY", "PRODUCT", "KODE BARANG", "SPESIFIKASI", "QTY", "TGL", "GP M0", "M0", "M3", "HARGA AKHIR"]].sort_values(
            ["TGL", "TEAM", "KODE BARANG"], ascending=[False, True, True]
        ).reset_index(drop=True)


    def build_sales_pivot_alerts(
        sales_pivot: pd.DataFrame,
        pricelist_wh: pd.DataFrame,
        warehouse_stock_cols: dict,
        period: str,
        selected_products=None,
        selected_brands=None,
        selected_segments=None,
        selected_kode_barang=None,
        selected_teams=None,
        start_date=None,
        end_date=None,
    ) -> pd.DataFrame:
        empty_cols = ["TEAM", "KODE BARANG", "SPESIFIKASI", "QTY", "STOK", "KET", "GUDANG READY"]
        if sales_pivot.empty or pricelist_wh.empty:
            return pd.DataFrame(columns=empty_cols)

        base = sales_pivot.copy()
        base["TGL"] = pd.to_datetime(base["TGL"], errors="coerce").dt.normalize()
        base = base[base["TGL"].notna()].copy()

        default_start_ts, default_end_ts = get_period_date_range(base, period)
        start_ts = pd.to_datetime(start_date, errors="coerce") if start_date is not None else pd.NaT
        end_ts = pd.to_datetime(end_date, errors="coerce") if end_date is not None else pd.NaT

        start_ts = default_start_ts if pd.isna(start_ts) else start_ts.normalize()
        end_ts = default_end_ts if pd.isna(end_ts) else end_ts.normalize()

        if pd.notna(start_ts):
            base = base[base["TGL"] >= start_ts]
        if pd.notna(end_ts):
            base = base[base["TGL"] <= end_ts]

        pl = pricelist_wh.copy()
        if "KODEBARANG" not in pl.columns:
            return pd.DataFrame(columns=empty_cols)

        if selected_products:
            pl = pl[pl["PRODUCT"].isin(selected_products)]
        if selected_brands:
            pl = pl[pl["BRAND"].isin(selected_brands)]
        if selected_segments:
            pl = pl[pl["PRICE_SEGMENT"].isin(selected_segments)]
        if pl.empty:
            return pd.DataFrame(columns=empty_cols)

        allowed_codes = set(normalize_text(pl["KODEBARANG"]).dropna().tolist())
        base["KODE BARANG"] = normalize_text(base["KODE BARANG"])
        base = base[base["KODE BARANG"].isin(allowed_codes)].copy()

        if selected_kode_barang:
            selected_kode_barang_norm = set(normalize_text(pd.Series(selected_kode_barang)).dropna().tolist())
            base = base[base["KODE BARANG"].isin(selected_kode_barang_norm)]
        if selected_teams:
            selected_teams_norm = set(normalize_text(pd.Series(selected_teams)).dropna().tolist())
            base = base[base["TEAM"].isin(selected_teams_norm)]

        if base.empty:
            return pd.DataFrame(columns=empty_cols)

        base = (
            base.groupby(["TEAM", "TEAM_KEY", "KODE BARANG"], as_index=False)
            .agg(SPESIFIKASI=("SPESIFIKASI", "first"), QTY=("QTY", "sum"))
        )

        merged = base.merge(pl, how="left", left_on="KODE BARANG", right_on="KODEBARANG")

        allowed_ready_team_keys = {"JKT1A", "JKT3A", "JKT3B", "JKT3C", "JKT4A", "JKT4B"}

        def get_stock_cols_by_team(team_key):
            if pd.isna(team_key):
                return []
            lookup_key = normalize_team_lookup_key(team_key)

            exact_cols = warehouse_stock_cols.get(lookup_key, [])
            if isinstance(exact_cols, str):
                exact_cols = [exact_cols]
            exact_cols = [c for c in exact_cols if c in merged.columns]
            if exact_cols:
                return exact_cols

            matched = []
            for wh_key, cols in warehouse_stock_cols.items():
                if normalize_team_lookup_key(wh_key) == lookup_key:
                    matched.extend(cols if isinstance(cols, list) else [cols])
            return [c for c in matched if c in merged.columns]

        def sum_stock_from_cols(row, cols):
            if not cols:
                return 0.0
            values = pd.to_numeric(pd.Series([row.get(c, 0) for c in cols]), errors="coerce").fillna(0)
            return float(values.sum())

        team_keys = sorted([
            str(k).strip().upper()
            for k in warehouse_stock_cols.keys()
            if str(k).strip().upper() != "DEFAULT"
        ])

        def get_current_stock(row):
            cols = get_stock_cols_by_team(row.get("TEAM_KEY"))
            return sum_stock_from_cols(row, cols)

        def get_ready_warehouses(row):
            current_team_key = normalize_team_lookup_key(row.get("TEAM_KEY"))
            ready_list = []

            for team_code in team_keys:
                team_key = normalize_team_lookup_key(team_code)
                if team_key == current_team_key:
                    continue
                if team_key not in allowed_ready_team_keys:
                    continue

                cols = warehouse_stock_cols.get(team_code, [])
                cols = cols if isinstance(cols, list) else [cols]
                cols = [c for c in cols if c in merged.columns]
                total_stock = sum_stock_from_cols(row, cols)
                if total_stock > 0:
                    ready_list.append((team_code.replace(" ", ""), int(round(total_stock))))

            ready_list = sorted(ready_list, key=lambda x: x[0])
            return ", ".join([f"{code} ({qty})" for code, qty in ready_list])

        merged["SPESIFIKASI"] = merged["SPESIFIKASI_x"].fillna(merged.get("SPESIFIKASI_y", ""))
        merged["STOK"] = merged.apply(get_current_stock, axis=1)
        merged["GUDANG READY"] = merged.apply(get_ready_warehouses, axis=1)
        merged["KET"] = np.where(
            (to_num(merged["QTY"]).fillna(0) > to_num(merged["STOK"]).fillna(0)) &
            (to_num(merged["STOK"]).fillna(0) <= 0) &
            (merged["GUDANG READY"].astype(str).str.strip() != ""),
            "REFILL",
            np.where(
                (to_num(merged["QTY"]).fillna(0) > to_num(merged["STOK"]).fillna(0)),
                "CEK",
                ""
            )
        )

        merged["KEBUTUHAN_STOK"] = to_num(merged["QTY"]).fillna(0) - to_num(merged["STOK"]).fillna(0)
        merged["PRIORITAS_STOK"] = np.where(
            merged["KET"].astype(str).str.upper().eq("REFILL"),
            2,
            np.where(merged["KET"].astype(str).str.upper().eq("CEK"), 1, 0)
        )

        out = merged[["TEAM", "KODE BARANG", "SPESIFIKASI", "QTY", "STOK", "KET", "GUDANG READY", "KEBUTUHAN_STOK", "PRIORITAS_STOK"]].copy()
        out["QTY"] = pd.to_numeric(out["QTY"], errors="coerce").fillna(0).round(0).astype(int)
        out["STOK"] = pd.to_numeric(out["STOK"], errors="coerce").fillna(0).round(0).astype(int)
        out["KEBUTUHAN_STOK"] = pd.to_numeric(out["KEBUTUHAN_STOK"], errors="coerce").fillna(0)

        out = out.sort_values(
            ["PRIORITAS_STOK", "KEBUTUHAN_STOK", "QTY", "TEAM", "KODE BARANG"],
            ascending=[False, False, False, True, True]
        ).reset_index(drop=True)

        return out[["TEAM", "KODE BARANG", "SPESIFIKASI", "QTY", "STOK", "KET", "GUDANG READY"]]

    def render_sales_pivot_alert_table(df: pd.DataFrame):
        if df.empty:
            st.info("Analisa Sales vs Stok 05 OLR belum menemukan data.")
            return

        show_df = df.copy()
        for col in ["QTY", "STOK"]:
            show_df[col] = pd.to_numeric(show_df[col], errors="coerce").fillna(0).round(0).astype(int)

        html = []
        html.append('<div class="alert-wrap"><table class="alert-table"><thead><tr>')
        for col in show_df.columns:
            html.append(f"<th>{col}</th>")
        html.append("</tr></thead><tbody>")

        for _, row in show_df.iterrows():
            html.append("<tr>")
            for col in show_df.columns:
                cls = ""
                if col in ["STOK", "KET"] and str(row.get("KET", "")).strip().upper() in ["REFILL", "CEK"]:
                    cls = ' class="bg-red"'
                html.append(f"<td{cls}>{row[col]}</td>")
            html.append("</tr>")
        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)


    def build_sku_gp_besar_table(sales_pivot: pd.DataFrame, stock: pd.DataFrame, selected_products=None) -> pd.DataFrame:
        columns = ["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3", "M0", "GP", "STOK"]
        if stock.empty:
            return pd.DataFrame(columns=columns)

        stock_base = stock.copy()
        for col in ["KODEBARANG", "SPESIFIKASI", "PRODUCT", "PRICE", "STOK_DIV03", "STOK_DIV04", "STOK_DIV05"]:
            if col not in stock_base.columns:
                stock_base[col] = np.nan if col in ["SPESIFIKASI", "PRODUCT"] else 0

        if selected_products:
            stock_base = stock_base[stock_base["PRODUCT"].isin(selected_products)].copy()

        stock_base["M3_VAL"] = to_num(stock_base.get("PRICE", np.nan)).fillna(0)
        stock_base["M0_VAL"] = 0
        if not sales_pivot.empty and "KODE BARANG" in sales_pivot.columns:
            sales_m0_lookup = sales_pivot.copy()
            sales_m0_lookup["M0_VAL"] = to_num(sales_m0_lookup.get("M0", np.nan)).fillna(0)
            sales_m0_lookup["KODE BARANG"] = normalize_text(sales_m0_lookup["KODE BARANG"])
            sales_m0_lookup = (
                sales_m0_lookup.groupby("KODE BARANG", as_index=False)
                .agg(M0_VAL=("M0_VAL", "max"))
            )
            stock_base = stock_base.merge(
                sales_m0_lookup,
                how="left",
                left_on="KODEBARANG",
                right_on="KODE BARANG",
                suffixes=("", "_sales")
            )
            stock_base["M0_VAL"] = to_num(stock_base.get("M0_VAL_sales", stock_base.get("M0_VAL"))).fillna(
                to_num(stock_base.get("M0_VAL")).fillna(0)
            )
            drop_cols = [c for c in ["KODE BARANG", "M0_VAL_sales"] if c in stock_base.columns]
            if drop_cols:
                stock_base = stock_base.drop(columns=drop_cols)

        stock_base["STOK"] = (
            to_num(stock_base.get("STOK_DIV03", 0)).fillna(0) +
            to_num(stock_base.get("STOK_DIV04", 0)).fillna(0) +
            to_num(stock_base.get("STOK_DIV05", 0)).fillna(0)
        )
        stock_base["GP"] = stock_base["M3_VAL"] - stock_base["M0_VAL"]

        stock_base["KODE BARANG"] = normalize_text(stock_base["KODEBARANG"])
        stock_base["SPESIFIKASI"] = normalize_text(stock_base["SPESIFIKASI"])
        stock_base["PRODUCT"] = normalize_text(stock_base["PRODUCT"])

        out = stock_base[["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3_VAL", "M0_VAL", "GP", "STOK"]].copy()
        out = out.rename(columns={"M3_VAL": "M3", "M0_VAL": "M0"})

        # skip M0 kosong / 0
        out = out[to_num(out["M0"]).fillna(0) > 0].copy()

        out = out[
            (to_num(out["GP"]).fillna(0) > 0) &
            (to_num(out["STOK"]).fillna(0) > 0)
        ].copy()

        out = out.sort_values(["GP", "STOK", "KODE BARANG"], ascending=[False, False, True]).reset_index(drop=True)
        return out[columns]


    def build_sku_top_gp_table(sales_pivot: pd.DataFrame, stock: pd.DataFrame, selected_products=None) -> pd.DataFrame:
        columns = ["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3", "M0", "QTY", "GP M0"]
        if sales_pivot.empty:
            return pd.DataFrame(columns=columns)

        sales_base = sales_pivot.copy()
        for col in ["PRODUCT", "SPESIFIKASI", "KODE BARANG", "M3", "M0", "QTY", "GP M0"]:
            if col not in sales_base.columns:
                sales_base[col] = np.nan

        sales_base["PRODUCT"] = normalize_text(sales_base["PRODUCT"])
        sales_base["KODE BARANG"] = normalize_text(sales_base["KODE BARANG"])
        sales_base["SPESIFIKASI"] = normalize_text(sales_base["SPESIFIKASI"])
        sales_base["M3_VAL"] = to_num(sales_base.get("M3", np.nan)).fillna(0)
        sales_base["M0_VAL"] = to_num(sales_base.get("M0", np.nan)).fillna(0)
        sales_base["GP_M0_VAL"] = to_num(sales_base.get("GP M0", np.nan)).fillna(0)
        sales_base["QTY"] = to_num(sales_base.get("QTY", 0)).fillna(0)

        if selected_products:
            sales_base = sales_base[sales_base["PRODUCT"].isin(selected_products)].copy()

        sales_base = sales_base[(sales_base["KODE BARANG"].notna()) & (sales_base["PRODUCT"].notna())].copy()
        sales_base = sales_base[sales_base["GP_M0_VAL"] > 0].copy()

        if sales_base.empty:
            return pd.DataFrame(columns=columns)

        # Samakan dengan Pivot Excel:
        # - filter PRODUCT dari file sales
        # - row label berdasarkan KODE BARANG
        # - value adalah SUM dari kolom GP M0 asli di file sales
        # Kolom lain diambil representatif per KODE BARANG agar tampilan card tetap informatif.
        meta = (
            sales_base.sort_values(["KODE BARANG", "QTY", "GP_M0_VAL"], ascending=[True, False, False])
            .drop_duplicates(subset=["KODE BARANG"], keep="first")
            [["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3_VAL", "M0_VAL"]]
            .copy()
        )

        pivot_like = (
            sales_base.groupby(["KODE BARANG"], dropna=False, as_index=False)
            .agg(QTY=("QTY", "sum"), **{"GP M0": ("GP_M0_VAL", "sum")})
        )

        out = pivot_like.merge(meta, how="left", on="KODE BARANG")
        out = out.rename(columns={"M3_VAL": "M3", "M0_VAL": "M0"})
        out = out[["KODE BARANG", "SPESIFIKASI", "PRODUCT", "M3", "M0", "QTY", "GP M0"]]
        out = out.sort_values(["GP M0", "QTY", "KODE BARANG"], ascending=[False, False, True]).reset_index(drop=True)
        return out[columns]


    def render_simple_card_table(df: pd.DataFrame, title: str):
        st.markdown(f"### {title}")
        if df.empty:
            st.info(f"{title} belum menemukan data.")
            return

        show_df = df.copy()
        for col in ["M3", "M0", "GP", "GP TOTAL", "GP M0"]:
            if col in show_df.columns:
                show_df[col] = show_df[col].apply(format_thousands_id)
        for col in ["QTY", "STOK"]:
            if col in show_df.columns:
                show_df[col] = show_df[col].apply(format_units_id)

        html = []
        html.append('<div class="main-fixed-wrap"><table class="main-fixed"><thead><tr>')
        for col in show_df.columns:
            width_style = ""
            if col in ["M3", "M0", "GP", "QTY", "STOK"]:
                width_style = ' style="min-width:70px;max-width:70px;"'
            html.append(f"<th{width_style}>{col}</th>")
        html.append("</tr></thead><tbody>")

        for _, row in show_df.iterrows():
            html.append("<tr>")
            for col in show_df.columns:
                width_style = ""
                if col in ["M3", "M0", "GP", "QTY", "STOK"]:
                    width_style = ' style="min-width:70px;max-width:70px;"'
                html.append(f"<td{width_style}>{'' if pd.isna(row[col]) else row[col]}</td>")
            html.append("</tr>")

        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)

    # =========================================================
    # BUILD TABLES
    # =========================================================
    def build_master(sales: pd.DataFrame, stock: pd.DataFrame) -> pd.DataFrame:
        df = sales.merge(stock, how="left", on="MERGE_KEY", suffixes=("_sales", "_stock"))

        for col in ["PRICE", "STOK_DIV03", "STOK_DIV04", "STOK_DIV05", "CATEGORY", "PRICE_SEGMENT"]:
            if col not in df.columns:
                df[col] = np.nan

        df["KODEBARANG"] = normalize_text(df["KODE BARANG"])
        df["SPESIFIKASI_FINAL"] = df["SPESIFIKASI_sales"].fillna(df.get("SPESIFIKASI_stock"))
        df["PRODUCT_FINAL"] = df["PRODUCT_sales"].fillna(df.get("PRODUCT_stock"))
        df["BRAND"] = normalize_text(df["BRAND"])
        df["QTY"] = to_num(df["QTY"]).fillna(0)

        stock_map = {"DIV03": "STOK_DIV03", "DIV04": "STOK_DIV04", "DIV05": "STOK_DIV05"}
        df["STOK_DIVISI"] = df.apply(lambda r: pd.to_numeric(r.get(stock_map[r["DIVISION"]]), errors="coerce"), axis=1).fillna(0)
        return df

    def build_segment_table(df, period, comparison_division_label="03 OLP"):
        tmp = df[df["PERIOD"] == period].copy()
        tmp["SEGMENT"] = tmp["PRICE"].apply(price_segment)
        seg = tmp.groupby(["SEGMENT", "DIVISION"])["QTY"].sum().unstack().fillna(0).reset_index()
        for div in DIVISIONS:
            if div not in seg.columns:
                seg[div] = 0
        seg = seg[["SEGMENT", "DIV03", "DIV04", "DIV05"]].copy()
        seg = seg.sort_values("SEGMENT", key=lambda s: s.map(segment_sort_key)).reset_index(drop=True)
        seg.columns = ["SEGMENT", "03 OLP", "04 MOD", "05 OLR"]

        compare_col = comparison_division_label if comparison_division_label in ["03 OLP", "04 MOD"] else "03 OLP"
        seg["DELTA"] = to_num(seg["05 OLR"]).fillna(0) - to_num(seg[compare_col]).fillna(0)
        return seg

    def build_brand_table(df, period, comparison_division_label="03 OLP"):
        brand = df[df["PERIOD"] == period].copy()
        brand = brand.groupby(["BRAND", "DIVISION"])["QTY"].sum().unstack().fillna(0).reset_index()
        for div in DIVISIONS:
            if div not in brand.columns:
                brand[div] = 0
        brand = brand[["BRAND", "DIV03", "DIV04", "DIV05"]].copy()
        brand["TOTAL"] = brand[["DIV03", "DIV04", "DIV05"]].sum(axis=1)
        brand = brand.sort_values(["TOTAL", "BRAND"], ascending=[False, True]).drop(columns=["TOTAL"])
        brand.columns = ["BRAND", "03 OLP", "04 MOD", "05 OLR"]

        compare_col = comparison_division_label if comparison_division_label in ["03 OLP", "04 MOD"] else "03 OLP"
        brand["DELTA"] = to_num(brand["05 OLR"]).fillna(0) - to_num(brand[compare_col]).fillna(0)
        return brand

    def render_left_table(df, title, selected_division="05 OLR", use_card=True):
        def is_number(v):
            return isinstance(v, (int, float, np.integer, np.floating)) and not pd.isna(v)

        def fmt_number(v):
            return f"{int(round(float(v))):,}".replace(",", ".")

        html = []
        if use_card:
            html.append("""
            <div class="table-card">
              <div style="font-weight:700;font-size:16px;margin-bottom:8px;">""" + title + """</div>
              <div style="overflow-x:auto;">
                <table style="border-collapse:collapse;width:100%;font-size:12px;">
            """)
        else:
            html.append("""
              <div style="font-weight:700;font-size:16px;margin-bottom:8px;">""" + title + """</div>
              <div style="overflow-x:auto;">
                <table style="border-collapse:collapse;width:100%;font-size:12px;">
            """)
        html.append("<thead><tr>")
        for col in df.columns:
            html.append(f'<th style="border:1px solid #2b2b2b;background:#f3f4f6;padding:6px;text-align:left;">{col}</th>')
        html.append("</tr></thead><tbody>")

        compare_cols = ["03 OLP", "04 MOD", "05 OLR"]
        has_compare_cols = all(c in df.columns for c in compare_cols)

        for _, row in df.iterrows():
            html.append("<tr>")

            losing_selected = False
            if has_compare_cols and selected_division in compare_cols:
                current_val = row[selected_division]
                other_vals = [row[c] for c in compare_cols if c != selected_division]
                if is_number(current_val):
                    numeric_others = [v for v in other_vals if is_number(v)]
                    if numeric_others:
                        losing_selected = any(float(current_val) < float(v) for v in numeric_others)

            for col in df.columns:
                val = row[col]
                try:
                    if pd.notna(val) and isinstance(val, (int, float, np.integer, np.floating)):
                        display = fmt_number(val)
                    else:
                        display = "" if pd.isna(val) else str(val)
                except Exception:
                    display = str(val)

                style = 'border:1px solid #2b2b2b;padding:6px;text-align:left;'
                if col == selected_division and losing_selected:
                    style += 'color:#c62828;font-weight:700;background:#ffebee;'
                if col == "DELTA" and is_number(val) and float(val) < 0:
                    style += 'color:#c62828;font-weight:700;background:#ffebee;'

                html.append(f'<td style="{style}">{display}</td>')
            html.append("</tr>")

        if use_card:
            html.append("</tbody></table></div></div>")
        else:
            html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)

    def build_main_table_filtered(
        df: pd.DataFrame,
        period: str,
        comparison_division_label: str,
        selected_segments=None,
        selected_brands=None,
        selected_products=None,
    ) -> pd.DataFrame:
        base = df[df["PERIOD"] == period].copy()

        if selected_segments:
            base = base[base["PRICE"].apply(price_segment).isin(selected_segments)]
        if selected_brands:
            base = base[base["BRAND"].isin(selected_brands)]
        if selected_products:
            base = base[base["PRODUCT_FINAL"].isin(selected_products)]

        qty = (
            base.groupby(["KODEBARANG", "SPESIFIKASI_FINAL", "PRICE", "DIVISION"], as_index=False)["QTY"]
            .sum()
            .pivot(index=["KODEBARANG", "SPESIFIKASI_FINAL", "PRICE"], columns="DIVISION", values="QTY")
            .fillna(0)
            .reset_index()
        )

        for div in DIVISIONS:
            if div not in qty.columns:
                qty[div] = 0

        product_brand = (
            base[["KODEBARANG", "PRODUCT_FINAL", "BRAND"]]
            .dropna(subset=["KODEBARANG"])
            .drop_duplicates(subset=["KODEBARANG"], keep="first")
            .copy()
        )

        stock_label_map = {"03 OLP": "STOK_DIV03", "04 MOD": "STOK_DIV04", "05 OLR": "STOK_DIV05"}
        stok_col = "STOK_DIV05"

        stock_df = (
            base[["KODEBARANG", "STOK_DIV03", "STOK_DIV04", "STOK_DIV05"]]
            .dropna(subset=["KODEBARANG"])
            .drop_duplicates(subset=["KODEBARANG"], keep="first")
            .copy()
        )
        stock_df["STOK"] = to_num(stock_df[stok_col]).fillna(0)

        out = qty.merge(product_brand, how="left", on="KODEBARANG")
        out = out.merge(stock_df, how="left", on="KODEBARANG")
        out["STOK"] = to_num(out["STOK"]).fillna(0)
        out["PRICE"] = to_num(out["PRICE"]).fillna(0)
        out["STOK_DIV03"] = to_num(out["STOK_DIV03"]).fillna(0)
        out["STOK_DIV04"] = to_num(out["STOK_DIV04"]).fillna(0)
        out["STOK_DIV05"] = to_num(out["STOK_DIV05"]).fillna(0)

        out = out.rename(columns={
            "PRODUCT_FINAL": "PRODUCT",
            "SPESIFIKASI_FINAL": "SPESIFIKASI",
            "PRICE": "M3",
            "DIV03": "03 OLP",
            "DIV04": "04 MOD",
            "DIV05": "05 OLR",
        })

        ordered_cols = [
            "KODEBARANG", "PRODUCT", "BRAND", "SPESIFIKASI", "M3",
            "03 OLP", "04 MOD", "05 OLR", "STOK",
            "STOK_DIV03", "STOK_DIV04", "STOK_DIV05"
        ]
        for col in ordered_cols:
            if col not in out.columns:
                out[col] = 0 if col not in ["KODEBARANG", "PRODUCT", "BRAND", "SPESIFIKASI"] else ""

        out["M3"] = (to_num(out["M3"]).fillna(0) / 1000).round(0)
        compare_map = {"03 OLP": "03 OLP", "04 MOD": "04 MOD"}
        compare_col = compare_map.get(comparison_division_label, "03 OLP")
        out["DELTA"] = to_num(out["05 OLR"]).fillna(0) - to_num(out[compare_col]).fillna(0)
        return out[ordered_cols + ["DELTA"]].sort_values(["DELTA", compare_col, "05 OLR"], ascending=[True, False, True]).reset_index(drop=True)

    def render_main_table_dynamic(df: pd.DataFrame, comparison_division_label: str):
        display_df = df.copy()

        compare_cols = ["03 OLP", "04 MOD", "05 OLR"]
        stock_hidden_map = {"03 OLP": "STOK_DIV03", "04 MOD": "STOK_DIV04", "05 OLR": "STOK_DIV05"}
        selected_stock_hidden = "STOK_DIV05"

        def losing_division(row):
            current_val = row.get("05 OLR", 0)
            compare_val = row.get(comparison_division_label, 0)
            try:
                return float(current_val) < float(compare_val)
            except Exception:
                return False

        def stok_problem(row):
            try:
                stok_selected = float(row.get("STOK", 0))
                qty_selected = float(row.get("05 OLR", 0))
                other_stock_cols = [stock_hidden_map[c] for c in compare_cols if c != "05 OLR"]
                other_stock_values = [float(row.get(c, 0)) for c in other_stock_cols]
                cond_a = stok_selected < qty_selected
                cond_b = stok_selected == 0 and qty_selected > 0 and any(v > 0 for v in other_stock_values)
                return cond_a or cond_b
            except Exception:
                return False

        display_df["_LOSS_DIVISION_FLAG"] = display_df.apply(losing_division, axis=1)
        display_df["_STOK_ALERT_FLAG"] = display_df.apply(stok_problem, axis=1)

        visible_df = display_df[["KODEBARANG", "PRODUCT", "BRAND", "SPESIFIKASI", "M3", "03 OLP", "04 MOD", "05 OLR", "DELTA", "STOK"]].copy()
        visible_df["M3"] = pd.to_numeric(visible_df["M3"], errors="coerce").fillna(0).round(0).astype(int)
        for col in ["03 OLP", "04 MOD", "05 OLR", "DELTA", "STOK"]:
            numeric_col = pd.to_numeric(visible_df[col], errors="coerce").fillna(0).round(0)
            visible_df[col] = numeric_col.astype(int)

        def fmt_value(val, col_name):
            if col_name in ["M3", "03 OLP", "04 MOD", "05 OLR", "DELTA", "STOK"]:
                try:
                    return f"{int(val)}"
                except Exception:
                    return "0"
            return "" if pd.isna(val) else str(val)

        html = []
        html.append('<div class="main-fixed-wrap"><table class="main-fixed"><thead><tr>')
        for col in visible_df.columns:
            html.append(f"<th>{col}</th>")
        html.append("</tr></thead><tbody>")

        for idx, row in visible_df.iterrows():
            html.append("<tr>")
            original = display_df.loc[idx]
            for col in visible_df.columns:
                cls = ""
                if col == "05 OLR" and bool(original["_LOSS_DIVISION_FLAG"]):
                    cls = ' class="bg-red"'
                if col == "DELTA":
                    try:
                        if float(original.get("DELTA", 0)) < 0:
                            cls = ' class="bg-red"'
                    except Exception:
                        pass
                if col == "STOK" and bool(original["_STOK_ALERT_FLAG"]):
                    cls = ' class="bg-red"'
                html.append(f"<td{cls}>{fmt_value(row[col], col)}</td>")
            html.append("</tr>")

        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)
        return visible_df



    def _render_analisa_produk_app_inner():
        st.title("Dashboard Analisa Produk")

        st.markdown('<div class="upload-card-wrap">', unsafe_allow_html=True)
        st.markdown("### Upload File")

        upload_col1, upload_col2, upload_col3 = st.columns(3)
        with upload_col1:
            st.markdown("**Upload MPLSSR**")
            mplssr_file = st.file_uploader("Upload MPLSSR", type=["xlsx", "xls"], key="upload_mplssr_main", label_visibility="collapsed")
            st.caption("200MB per file • XLSX, XLS")

        with upload_col2:
            st.markdown("**Upload Pricelist**")
            pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx", "xls"], key="upload_pricelist_main", label_visibility="collapsed")
            st.caption("200MB per file • XLSX, XLS")

        with upload_col3:
            st.markdown("**Upload Penjualan**")
            sales_pivot_file = st.file_uploader("Upload Penjualan", type=["xlsx", "xls"], key="upload_sales_pivot_main", label_visibility="collapsed")
            st.caption("200MB per file • XLSX, XLS")

        all_required_uploaded = all([mplssr_file is not None, pricelist_file is not None, sales_pivot_file is not None])

        if not all_required_uploaded:
            st.info("Silakan upload MPLSSR, Pricelist, dan Penjualan.")
        st.markdown('</div>', unsafe_allow_html=True)

        if all_required_uploaded:
            current_upload_signature = (
                getattr(mplssr_file, "name", None),
                getattr(pricelist_file, "name", None),
                getattr(sales_pivot_file, "name", None),
            )

            if st.session_state.get("processed_upload_signature") != current_upload_signature:
                try:
                    sales = load_mplssr(mplssr_file)
                    stock = load_pricelist(pricelist_file)
                    master = build_master(sales, stock)
                    pricelist_wh, warehouse_stock_cols = load_pricelist_with_warehouses(pricelist_file)
                    sales_pivot = load_sales_pivot(sales_pivot_file)
                    st.session_state["processed_data"] = {
                        "sales": sales,
                        "stock": stock,
                        "master": master,
                        "pricelist_wh": pricelist_wh,
                        "warehouse_stock_cols": warehouse_stock_cols,
                        "sales_pivot": sales_pivot,
                    }
                    st.session_state["processed_upload_signature"] = current_upload_signature
                except Exception as e:
                    st.error(f"Gagal membaca file: {e}")
                    st.stop()

        if "processed_data" not in st.session_state:
            st.stop()

        sales = st.session_state["processed_data"]["sales"]
        stock = st.session_state["processed_data"]["stock"]
        master = st.session_state["processed_data"]["master"]
        pricelist_wh = st.session_state["processed_data"]["pricelist_wh"]
        warehouse_stock_cols = st.session_state["processed_data"]["warehouse_stock_cols"]
        sales_pivot = st.session_state["processed_data"]["sales_pivot"]

        product_options = sorted(master["PRODUCT_FINAL"].dropna().unique().tolist())
        default_product = ["LAPTOP R"] if "LAPTOP R" in product_options else []

        st.markdown('<div class="upload-card-wrap">', unsafe_allow_html=True)
        st.markdown("### Filter Dashboard")
        with st.form("unified_filter_form"):
            fcol1, fcol2, fcol3, fcol4, fcol5, fcol6 = st.columns([1.2, 1.2, 1, 1.2, 1.1, 0.6])
            with fcol1:
                selected_products = st.multiselect("Product", product_options, default=default_product)
            with fcol2:
                selected_brands = st.multiselect("Brand", sorted(master["BRAND"].dropna().unique().tolist()))
            with fcol3:
                selected_period = st.selectbox("Period", PERIODS, index=0)
            with fcol4:
                selected_segments = st.multiselect("Range Harga", [s[2] for s in PRICE_SEGMENTS] + ["UNKNOWN"])
            with fcol5:
                comparison_division = st.selectbox("Perbandingan", ["03 OLP", "04 MOD", "05 OLR"], index=0)
            with fcol6:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                apply_filter = st.form_submit_button("PROSES")
        st.markdown('</div>', unsafe_allow_html=True)

        filtered = master.copy()
        if selected_products:
            filtered = filtered[filtered["PRODUCT_FINAL"].isin(selected_products)]
        if selected_brands:
            filtered = filtered[filtered["BRAND"].isin(selected_brands)]
        if selected_segments:
            filtered = filtered[filtered["PRICE"].apply(price_segment).isin(selected_segments)]

        if filtered.empty:
            st.warning("Data kosong setelah filter diterapkan.")
            st.stop()

        with st.container(border=True):
            st.markdown("### Analisa Segment vs Divisi Lain")
            left, right = st.columns(2)
            with left:
                render_left_table(build_segment_table(filtered, selected_period, comparison_division), f"Segmentasi Harga - {selected_period}", selected_division=comparison_division, use_card=False)
            with right:
                render_left_table(build_brand_table(filtered, selected_period, comparison_division), f"Segmentasi Brand - {selected_period}", selected_division=comparison_division, use_card=False)

        with st.container(border=True):
            st.markdown("### Analisa SKU vs Divisi Lain")
            main_table_export = build_main_table_filtered(
                filtered,
                selected_period,
                comparison_division,
                selected_segments=selected_segments,
                selected_brands=selected_brands,
                selected_products=selected_products,
            )
            main_table_export = render_main_table_dynamic(main_table_export, comparison_division)

        stok_kode_barang_options = sorted(sales_pivot["KODE BARANG"].dropna().unique().tolist()) if not sales_pivot.empty and "KODE BARANG" in sales_pivot.columns else []
        stok_team_options = sorted(normalize_text(sales_pivot["TEAM"]).dropna().unique().tolist()) if not sales_pivot.empty and "TEAM" in sales_pivot.columns else []
        stok_min_date = sales_pivot["TGL"].min() if not sales_pivot.empty and "TGL" in sales_pivot.columns else None
        stok_max_date = sales_pivot["TGL"].max() if not sales_pivot.empty and "TGL" in sales_pivot.columns else None
        stok_min_date_value = pd.to_datetime(stok_min_date, errors="coerce").date() if pd.notna(stok_min_date) else None
        stok_max_date_value = pd.to_datetime(stok_max_date, errors="coerce").date() if pd.notna(stok_max_date) else None

        if "stok_products" not in st.session_state:
            st.session_state["stok_products"] = selected_products
        if "stok_period" not in st.session_state:
            st.session_state["stok_period"] = selected_period if selected_period in PERIODS else PERIODS[0]
        if "stok_kode_barang" not in st.session_state:
            st.session_state["stok_kode_barang"] = []
        if "stok_teams" not in st.session_state:
            st.session_state["stok_teams"] = []

        period_default_start, period_default_end = get_period_date_range(sales_pivot, st.session_state["stok_period"])
        period_default_start_value = pd.to_datetime(period_default_start, errors="coerce").date() if pd.notna(period_default_start) else stok_min_date_value
        period_default_end_value = pd.to_datetime(period_default_end, errors="coerce").date() if pd.notna(period_default_end) else stok_max_date_value

        current_start = pd.to_datetime(st.session_state.get("stok_start_date"), errors="coerce")
        current_end = pd.to_datetime(st.session_state.get("stok_end_date"), errors="coerce")

        if pd.isna(current_start):
            st.session_state["stok_start_date"] = period_default_start_value
        else:
            current_start_value = current_start.date()
            if stok_min_date_value is not None and (current_start_value < stok_min_date_value or current_start_value > stok_max_date_value):
                st.session_state["stok_start_date"] = period_default_start_value

        if pd.isna(current_end):
            st.session_state["stok_end_date"] = period_default_end_value
        else:
            current_end_value = current_end.date()
            if stok_min_date_value is not None and (current_end_value < stok_min_date_value or current_end_value > stok_max_date_value):
                st.session_state["stok_end_date"] = period_default_end_value

        if (
            st.session_state.get("stok_start_date") is not None
            and st.session_state.get("stok_end_date") is not None
            and st.session_state["stok_start_date"] > st.session_state["stok_end_date"]
        ):
            st.session_state["stok_start_date"] = period_default_start_value
            st.session_state["stok_end_date"] = period_default_end_value

        with st.container(border=True):
            st.markdown("### Analisa Sales vs Stok 05 OLR")
            with st.form("stok_filter_form"):
                col1, col2 = st.columns(2)
                with col1:
                    stok_products = st.multiselect(
                        "Product",
                        sorted(pricelist_wh["PRODUCT"].dropna().unique()),
                        default=st.session_state.get("stok_products", selected_products),
                    )
                with col2:
                    current_stok_period = st.session_state.get("stok_period", selected_period if selected_period in PERIODS else PERIODS[0])
                    stok_period = st.selectbox(
                        "Period",
                        PERIODS,
                        index=PERIODS.index(current_stok_period) if current_stok_period in PERIODS else 0,
                    )

                col3, col4 = st.columns(2)
                with col3:
                    stok_kode_barang = st.multiselect(
                        "Kode Barang",
                        stok_kode_barang_options,
                        default=st.session_state.get("stok_kode_barang", []),
                    )
                with col4:
                    stok_teams = st.multiselect(
                        "Team",
                        stok_team_options,
                        default=st.session_state.get("stok_teams", []),
                    )

                col5 = st.columns(1)[0]
                with col5:
                    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                    process_stok = st.form_submit_button("PROSES")

            if process_stok:
                new_period_start, new_period_end = get_period_date_range(sales_pivot, stok_period)
                new_period_start_value = pd.to_datetime(new_period_start, errors="coerce").date() if pd.notna(new_period_start) else stok_min_date_value
                new_period_end_value = pd.to_datetime(new_period_end, errors="coerce").date() if pd.notna(new_period_end) else stok_max_date_value

                st.session_state["stok_products"] = stok_products
                st.session_state["stok_period"] = stok_period
                st.session_state["stok_kode_barang"] = stok_kode_barang
                st.session_state["stok_teams"] = stok_teams

                start_value = pd.to_datetime(st.session_state.get("stok_start_date"), errors="coerce")
                end_value = pd.to_datetime(st.session_state.get("stok_end_date"), errors="coerce")
                start_value = start_value.date() if pd.notna(start_value) else new_period_start_value
                end_value = end_value.date() if pd.notna(end_value) else new_period_end_value

                if stok_min_date_value is not None and (start_value < stok_min_date_value or start_value > stok_max_date_value):
                    start_value = new_period_start_value
                if stok_min_date_value is not None and (end_value < stok_min_date_value or end_value > stok_max_date_value):
                    end_value = new_period_end_value
                if start_value > end_value:
                    start_value, end_value = new_period_start_value, new_period_end_value

                st.session_state["stok_start_date"] = start_value
                st.session_state["stok_end_date"] = end_value

            sales_pivot_alerts = build_sales_pivot_alerts(
                sales_pivot,
                pricelist_wh,
                warehouse_stock_cols,
                period=st.session_state["stok_period"],
                selected_products=st.session_state["stok_products"],
                selected_kode_barang=st.session_state.get("stok_kode_barang", []),
                selected_teams=st.session_state.get("stok_teams", []),
        
            )

            render_sales_pivot_alert_table(sales_pivot_alerts)

        if "gp_products" not in st.session_state:
            st.session_state["gp_products"] = default_product.copy()
        top_gp_product_options = sorted(sales_pivot["PRODUCT"].dropna().unique().tolist()) if not sales_pivot.empty and "PRODUCT" in sales_pivot.columns else []
        default_top_gp_product = ["LAPTOP R"] if "LAPTOP R" in top_gp_product_options else []

        if "top_gp_products" not in st.session_state:
            st.session_state["top_gp_products"] = default_top_gp_product.copy()
        if "gp_cards_applied" not in st.session_state:
            st.session_state["gp_cards_applied"] = True

        card_col1, card_col2 = st.columns(2)

        with card_col1:
            with st.container(border=True):
                with st.form("gp_besar_form"):
                    gp_filter_col, gp_button_col = st.columns([4, 1])
                    with gp_filter_col:
                        gp_product_filter = st.multiselect(
                            "Filter Product - SKU Dengan GP Besar",
                            product_options,
                            default=st.session_state.get("gp_products", []),
                            key="gp_products_filter",
                        )
                    with gp_button_col:
                        st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
                        process_gp_besar = st.form_submit_button("PROSES")

                if process_gp_besar:
                    st.session_state["gp_products"] = gp_product_filter
                    st.session_state["gp_cards_applied"] = True

                if st.session_state.get("gp_cards_applied"):
                    render_simple_card_table(
                        build_sku_gp_besar_table(
                            sales_pivot=sales_pivot,
                            stock=stock,
                            selected_products=st.session_state.get("gp_products", []),
                        ),
                        "SKU Dengan GP Besar"
                    )
                else:
                    st.info("Pilih filter product lalu klik PROSES.")

        with card_col2:
            with st.container(border=True):
                with st.form("top_gp_form"):
                    top_filter_col, top_button_col = st.columns([4, 1])
                    with top_filter_col:
                        top_gp_product_filter = st.multiselect(
                            "Filter Product - SKU Top GP",
                            top_gp_product_options,
                            default=st.session_state.get("top_gp_products", default_top_gp_product),
                            key="top_gp_products_filter",
                        )
                    with top_button_col:
                        st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
                        process_top_gp = st.form_submit_button("PROSES")

                if process_top_gp:
                    st.session_state["top_gp_products"] = top_gp_product_filter
                    st.session_state["gp_cards_applied"] = True

                if st.session_state.get("gp_cards_applied"):
                    render_simple_card_table(
                        build_sku_top_gp_table(
                            sales_pivot=sales_pivot,
                            stock=stock,
                            selected_products=st.session_state.get("top_gp_products", []),
                        ),
                        "SKU Top GP"
                    )
                else:
                    st.info("Pilih filter product lalu klik PROSES.")

        st.markdown("<div style='height:120px;'></div>", unsafe_allow_html=True)

    _render_analisa_produk_app_inner()


# ============================================================
# PAGES
# ============================================================
def render_dashboard():
    st.title(APP_TITLE)
    st.markdown("Aplikasi all-in-one untuk **Update Stok**, **Harga Normal**, **Harga Coret**, **Submit Campaign**, **Analisa**, dan **Affiliate** marketplace.")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.subheader("Operasional Marketplace")
        st.write("- Update Stok\n- Update Harga Normal\n- Update Harga Coret\n- Submit Campaign")
    with c2:
        st.subheader("Analisa Penjualan")
        st.write("- Analisa Penjualan\n- Dashboard growth penjualan\n- Perbandingan periode & insight team")
    with c3:
        st.subheader("Analisa Produk")
        st.write("- Analisa Produk\n- Analisa produk vs divisi\n- Alert stok & SKU insight")
    st.info("Gunakan menu di sidebar untuk memilih fitur.")


def render_update_stok_shopee():
    page_header(
        "Update Stok Shopee (Mall & Star)",
        "Memproses file mass update Shopee (Mall & Star) berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update Shopee (.xlsx, Unprotect dulu)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Shopee (Mall & Star)", type=["xlsx"], accept_multiple_files=True, key="stock_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_shopee_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, process_disabled = render_stock_controls(
        area_key_prefix="stock_shopee",
        pricelist_file=pricelist_file,
        mode_key="stock_shopee_mode",
        loaded_areas_key="stock_shopee_areas_loaded",
        load_button_key="load_area_shopee",
    )

    if st.button("Proses", key="btn_stock_shopee", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_shopee_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback),
                "Memproses update stok Shopee...",
            )
            cache_downloads("stock_shopee", "hasil_update_stok_shopee.xlsx", result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_shopee")
    render_downloads("stock_shopee")


def render_update_stok_tiktokshop():
    page_header(
        "Update Stok TikTokShop",
        "Memproses file mass update TikTokShop berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update TikTokShop (.xlsx, Unprotect dulu)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update TikTokShop", type=["xlsx"], accept_multiple_files=True, key="stock_tiktokshop_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_tiktokshop_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, process_disabled = render_stock_controls(
        area_key_prefix="stock_tiktokshop",
        pricelist_file=pricelist_file,
        mode_key="stock_tiktokshop_mode",
        loaded_areas_key="stock_tiktokshop_areas_loaded",
        load_button_key="load_area_tiktokshop",
    )

    if st.button("Proses", key="btn_stock_tiktokshop", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_tiktokshop_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback),
                "Memproses update stok TikTokShop...",
            )
            cache_downloads("stock_tiktokshop", "hasil_update_stok_tiktokshop.xlsx", result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_tiktokshop")
    render_downloads("stock_tiktokshop")




def render_update_stok_mwh():
    page_header(
        "Update Stok Mwh",
        "Memproses file mass update Mwh berdasarkan stok dari pricelist. Baris yang tidak berubah tetap dipertahankan di output.",
        [
            "Mass Update Mwh (.xlsx, Unprotect dulu)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
            "Header stok Mwh harus berisi kolom Jumlah.",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Mwh", type=["xlsx"], accept_multiple_files=True, key="stock_mwh_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_mwh_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, process_disabled = render_stock_controls(
        area_key_prefix="stock_mwh",
        pricelist_file=pricelist_file,
        mode_key="stock_mwh_mode",
        loaded_areas_key="stock_mwh_areas_loaded",
        load_button_key="load_area_mwh",
    )

    if st.button("Proses", key="btn_stock_mwh", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_mwh_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback),
                "Memproses update stok Mwh...",
            )
            cache_downloads("stock_mwh", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_mwh")
    render_downloads("stock_mwh")

def render_update_stok_bigseller():
    page_header(
        "Update Stok Bigseller",
        "Mengubah stok Bigseller berdasarkan sheet pricelist LAPTOP, TELCO, dan PC HOM ELE. Output hanya baris yang berubah dan otomatis split 10.000 row per file.",
        [
            "Mass Update Bigseller (.xlsx, bisa banyak)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Bigseller", type=["xlsx"], accept_multiple_files=True, key="stock_bigseller_mass_v2")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_bigseller_pl_v2")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, process_disabled = render_stock_controls(
        area_key_prefix="stock_bigseller",
        pricelist_file=pricelist_file,
        mode_key="stock_bigseller_mode",
        loaded_areas_key="stock_bigseller_areas_loaded",
        load_button_key="load_area_bigseller",
    )

    if st.button("Proses", key="btn_stock_bigseller", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_bigseller_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback),
                "Memproses update stok Bigseller...",
            )
            cache_downloads("stock_bigseller", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_bigseller")
    render_downloads("stock_bigseller")




def render_update_stok_blibli():
    page_header(
        "Update Stok Blibli",
        "Memproses file mass update Blibli berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update Blibli (.xlsx)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Blibli", type=["xlsx"], accept_multiple_files=True, key="stock_blibli_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_blibli_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, process_disabled = render_stock_controls(
        area_key_prefix="stock_blibli",
        pricelist_file=pricelist_file,
        mode_key="stock_blibli_mode",
        loaded_areas_key="stock_blibli_areas_loaded",
        load_button_key="load_area_blibli",
    )

    if st.button("Proses", key="btn_stock_blibli", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_blibli_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback),
                "Memproses update stok Blibli...",
            )
            cache_downloads("stock_blibli", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_blibli")
    render_downloads("stock_blibli")


def render_update_stok_akulaku():
    page_header(
        "Update Stok Akulaku",
        "Memproses file mass update Akulaku berdasarkan stok dari sheet pricelist LAPTOP, TELCO, dan PC HOM ELE.",
        [
            "Mass Update Akulaku (.xlsx)",
            "Pricelist (.xlsx, tidak perlu ada yang di ubah)",
        ],
    )
    c1, c2 = st.columns(2)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Akulaku", type=["xlsx"], accept_multiple_files=True, key="stock_akulaku_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="stock_akulaku_pl")

    selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, process_disabled = render_stock_controls(
        area_key_prefix="stock_akulaku",
        pricelist_file=pricelist_file,
        mode_key="stock_akulaku_mode",
        loaded_areas_key="stock_akulaku_areas_loaded",
        load_button_key="load_area_akulaku",
    )

    if st.button("Proses", key="btn_stock_akulaku", disabled=process_disabled):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if pricelist_file is None:
            st.error("Upload Pricelist dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_akulaku_stock(mass_files, pricelist_file, selected_modes, chosen_areas, chosen_gudangs, zero_below, zero_if_missing, progress_callback),
                "Memproses update stok Akulaku...",
            )
            cache_downloads("stock_akulaku", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("stock_akulaku")
    render_downloads("stock_akulaku")


def render_harga_normal_shopee():
    page_header(
        "Harga Normal Shopee (Mall & Star)",
        "Mengubah harga normal Shopee (Mall & Star) berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update Shopee (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_shopee_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_shopee_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_shopee_disc")

    if st.button("Proses", key="btn_normal_shopee"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_shopee_price(
                    mass_files, pricelist_file, addon_file, discount_rp, progress_callback
                ),
                "Memproses harga normal Shopee...",
            )
            cache_downloads("normal_shopee", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_shopee")
    render_downloads("normal_shopee")




def render_harga_normal_blibli():
    page_header(
        "Harga Normal Blibli",
        "Mengubah harga normal Blibli berdasarkan sheet CHANGE di pricelist dan addon mapping. Kolom Harga (Rp) dan Harga Penjualan (Rp) akan diisi harga M3.",
        ["Template Mass Update Blibli (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_blibli_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_blibli_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_blibli_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_blibli_disc")

    if st.button("Proses", key="btn_normal_blibli"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_blibli_price(mass_files, pricelist_file, addon_file, discount_rp, progress_callback),
                "Memproses harga normal Blibli...",
            )
            cache_downloads("normal_blibli", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_blibli")
    render_downloads("normal_blibli")


def render_harga_normal_akulaku():
    page_header(
        "Harga Normal Akulaku",
        "Mengubah harga normal Akulaku berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update Akulaku (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_akulaku_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_akulaku_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_akulaku_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_akulaku_disc")

    if st.button("Proses", key="btn_normal_akulaku"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_akulaku_price(mass_files, pricelist_file, addon_file, discount_rp, progress_callback),
                "Memproses harga normal Akulaku...",
            )
            cache_downloads("normal_akulaku", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_akulaku")
    render_downloads("normal_akulaku")


def render_harga_coret_shopee():
    page_header(
        "Harga Coret Shopee (Mall & Star)",
        "Mengubah harga coret Shopee (Mall & Star) berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Discount Nominate Shopee (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Template Mass Update", type=["xlsx"], accept_multiple_files=True, key="coret_shopee_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_shopee_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_shopee_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_shopee_disc")

    if st.button("Proses", key="btn_coret_shopee"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_shopee_discount(
                    mass_files, pricelist_file, addon_file, discount_rp, progress_callback
                ),
                "Memproses harga coret Shopee...",
            )
            cache_downloads("coret_shopee", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_shopee")
    render_downloads("coret_shopee")


def render_harga_normal_tiktokshop():
    page_header(
        "Harga Normal TikTokShop",
        "Mengubah harga normal TikTokShop berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update TikTokShop (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_tiktokshop_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_tiktokshop_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_tiktokshop_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_tiktokshop_disc")

    if st.button("Proses", key="btn_normal_tiktokshop"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_tiktokshop_price(
                    mass_files, pricelist_file, addon_file, discount_rp, progress_callback
                ),
                "Memproses harga normal TikTokShop...",
            )
            cache_downloads("normal_tiktokshop", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_tiktokshop")
    render_downloads("normal_tiktokshop")




def render_harga_normal_mwh():
    page_header(
        "Harga Normal Mwh",
        "Mengubah harga normal Mwh berdasarkan pricelist dan addon mapping. Baris yang tidak berubah tetap dipertahankan di output.",
        ["Template Mass Update Mwh (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update Mwh", type=["xlsx"], accept_multiple_files=True, key="normal_mwh_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_mwh_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_mwh_add")
    price_key = st.radio(
        "Ambil Harga dari Pricelist",
        ["M3", "M4"],
        horizontal=True,
        key="normal_mwh_price_key",
    )
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_mwh_disc")

    if st.button("Proses", key="btn_normal_mwh"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_mwh_price(
                    mass_files, pricelist_file, addon_file, discount_rp, price_key, progress_callback
                ),
                "Memproses harga normal Mwh...",
            )
            cache_downloads("normal_mwh", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_mwh")
    render_downloads("normal_mwh")

def render_harga_coret_tiktokshop():
    page_header(
        "Harga Coret TikTokShop",
        "Membuat template output promo TikTokShop berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Input File TikTokShop (.xlsx)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        input_file = st.file_uploader("Upload File TikTokShop", type=["xlsx"], key="coret_tiktokshop_input")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_tiktokshop_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_tiktokshop_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_tiktokshop_disc")

    if st.button("Proses", key="btn_coret_tiktokshop"):
        if not input_file or not pricelist_file or not addon_file:
            st.error("Upload semua file yang dibutuhkan dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_tiktokshop_discount(
                    input_file, pricelist_file, addon_file, discount_rp, True, progress_callback
                ),
                "Memproses harga coret TikTokShop...",
            )
            cache_downloads("coret_tiktokshop", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_tiktokshop")
    render_downloads("coret_tiktokshop")


def render_harga_normal_powemerchant():
    page_header(
        "Harga Normal PowerMerchant",
        "Mengubah harga normal PowerMerchant berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update PowerMerchant (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_pm_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_pm_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_pm_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_pm_disc")

    if st.button("Proses", key="btn_normal_pm"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_powemerchant_price(
                    mass_files, pricelist_file, addon_file, discount_rp, progress_callback
                ),
                "Memproses harga normal PowerMerchant...",
            )
            cache_downloads("normal_pm", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_pm")
    render_downloads("normal_pm")


def render_harga_coret_powemerchant():
    page_header(
        "Harga Coret PowerMerchant",
        "Mengubah harga coret PowerMerchant berdasarkan sheet CHANGE di pricelist dan addon mapping.",
        ["Template Mass Update PowerMerchant (.xlsx, Unprotect dulu)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="coret_pm_mass")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="coret_pm_pl")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="coret_pm_add")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="coret_pm_disc")

    if st.button("Proses", key="btn_coret_pm"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_powemerchant_discount(
                    mass_files, pricelist_file, addon_file, discount_rp, progress_callback
                ),
                "Memproses harga coret PowerMerchant...",
            )
            cache_downloads("coret_pm", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("coret_pm")
    render_downloads("coret_pm")


def render_harga_normal_bigseller():
    page_header(
        "Harga Normal Bigseller",
        "Mengubah harga Bigseller, hanya output baris yang berubah, dan otomatis split 10.000 row per file.",
        ["Mass Update Bigseller (.xlsx, bisa banyak)", "Pricelist (.xlsx, tidak perlu ada yang di ubah)", "Addon Mapping (.xlsx)"],
    )
    c1, c2, c3 = st.columns(3)
    with c1:
        mass_files = st.file_uploader("Upload Mass Update", type=["xlsx"], accept_multiple_files=True, key="normal_bigseller_mass_v2")
    with c2:
        pricelist_file = st.file_uploader("Upload Pricelist", type=["xlsx"], key="normal_bigseller_pl_v2")
    with c3:
        addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"], key="normal_bigseller_add_v2")
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000, key="normal_bigseller_disc")
    price_key = st.radio(
        "Ambil harga dari Pricelist",
        ["M3", "M4"],
        horizontal=True,
        key="normal_bigseller_price_key",
    )

    if st.button("Proses", key="btn_normal_bigseller"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return
        if not pricelist_file or not addon_file:
            st.error("Upload Pricelist dan Addon Mapping dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_bigseller_price(
                    mass_files, pricelist_file, addon_file, discount_rp, price_key, progress_callback
                ),
                "Memproses harga normal Bigseller...",
            )
            cache_downloads("normal_bigseller", result_name, result_bytes, issues_bytes, summary=summary)
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("normal_bigseller")
    render_downloads("normal_bigseller")


def render_submit_campaign_shopee():
    page_header(
        "Submit Campaign Shopee",
        "Memfilter file campaign Shopee berdasarkan Kode Variasi dari file mass update yang SKU-nya ND-ALL-CAMPAIGN.",
        [
            "File Mass Update Shopee (.xlsx)",
            "File Campaign Shopee / Eligible Product List (.xlsx)",
        ],
    )

    col1, col2 = st.columns(2)
    with col1:
        mass_file = st.file_uploader(
            "Upload File Mass Update Shopee",
            type=["xlsx"],
            key="submit_campaign_shopee_mass",
        )
    with col2:
        campaign_file = st.file_uploader(
            "Upload File Campaign Shopee",
            type=["xlsx"],
            key="submit_campaign_shopee_campaign",
        )

    if st.button("Proses", key="btn_submit_campaign_shopee"):
        if not mass_file or not campaign_file:
            st.error("Upload File Mass Update dan File Campaign dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_shopee_campaign(mass_file=mass_file, campaign_file=campaign_file, progress_callback=progress_callback),
                "Memproses Submit Campaign Shopee...",
            )
            cache_downloads(
                "submit_campaign_shopee",
                result_name,
                result_bytes,
                issues_bytes,
                summary=summary,
            )
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("submit_campaign_shopee")
    render_downloads("submit_campaign_shopee")


def render_submit_campaign_tiktokshop():
    page_header(
        "Submit Campaign TikTokShop",
        "Cukup upload template campaign TikTokShop. Sistem hanya akan menyimpan baris yang memiliki karakter 'ND-ALL-CAMPAIGN' pada kolom SKU Name.",
        [
            "Template Campaign Tiktokshop (.xlsx)",
        ],
    )

    mass_files = st.file_uploader(
        "Upload Template Campaign Tiktokshop",
        type=["xlsx"],
        accept_multiple_files=True,
        key="submit_campaign_tiktokshop_mass",
    )

    if st.button("Proses", key="btn_submit_campaign_tiktokshop"):
        err = validate_mass_uploads(mass_files)
        if err:
            st.error(err)
            return

        try:
            result_bytes, result_name, issues_bytes, summary = run_with_loading_callback(
                lambda progress_callback: process_tiktokshop_campaign(mass_files=mass_files, progress_callback=progress_callback),
                "Memfilter template campaign TikTokShop...",
            )
            cache_downloads(
                "submit_campaign_tiktokshop",
                result_name,
                result_bytes,
                issues_bytes,
                summary=summary,
            )
        except Exception as e:
            st.error(f"Gagal memproses: {e}")

    render_cached_summary("submit_campaign_tiktokshop")
    render_downloads("submit_campaign_tiktokshop")


# ============================================================
# SIDEBAR ROUTER
# ============================================================


def render_analisa_penjualan():
    st.title("Analisa Penjualan")
    st.caption("Fitur analisa penjualan sudah tertanam langsung di file utama Codexid.")
    try:
        render_analisa_penjualan_app()
    except Exception as e:
        st.error(f"Gagal membuka fitur Analisa Penjualan: {e}")


def render_analisa_produk():
    st.title("Analisa Produk")
    st.caption("Fitur analisa produk sudah tertanam langsung di file utama Codexid.")
    try:
        render_analisa_produk_app()
    except Exception as e:
        st.error(f"Gagal membuka fitur Analisa Produk: {e}")


def render_progres_on():
    """Upload laporan progres bulanan dan tampilkan proyeksi serta grafiknya."""
    import calendar
    from datetime import date
    import plotly.graph_objects as go

    st.title("Progres ON")
    st.caption("Upload format progres seperti contoh. Kolom EST dihitung dari nilai bulan berjalan ÷ hari berjalan × jumlah hari dalam bulan.")
    uploaded = st.file_uploader("Upload file Progres ON (.xlsx)", type=["xlsx"], key="progres_on_upload")
    if not uploaded:
        st.info("Pilih file Excel untuk melihat tabel dan grafik proyeksi.")
        return
    try:
        workbook = pd.ExcelFile(uploaded)
    except Exception as e:
        st.error(f"File Excel tidak dapat dibaca: {e}")
        return

    sheet_name = st.selectbox("Sheet", workbook.sheet_names, key="progres_on_sheet")
    raw = pd.read_excel(workbook, sheet_name=sheet_name, header=None)
    if raw.empty or raw.shape[1] < 3:
        st.error("Format file belum cukup untuk dianalisa.")
        return

    detected_days = None
    for _, row in raw.iterrows():
        if s_clean(row.iloc[0]).upper() == "TGL":
            detected_days = to_int_or_none(row.iloc[1])
            break
    detected_days = detected_days or date.today().day
    today_days = st.number_input("Hari berjalan (today)", min_value=1, max_value=31, value=min(int(detected_days), 31), step=1, key="progres_on_days", help="Nilai ini dipakai untuk menghitung proyeksi EST bulan berjalan.")

    def cell_text(value) -> str:
        return s_clean(value).upper()

    def number_value(value):
        if isinstance(value, (int, float)) and not pd.isna(value):
            return float(value)
        parsed = parse_price_cell(value)
        return float(parsed) if parsed is not None else None

    sections = []
    for header_idx in range(1, len(raw)):
        header = [cell_text(value) for value in raw.iloc[header_idx].tolist()]
        if not any(value.startswith("EST") for value in header):
            continue
        month_positions = [idx for idx, value in enumerate(header) if value in {"JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"}]
        if len(month_positions) < 2 or header_idx + 2 >= len(raw):
            continue
        category = s_clean(raw.iloc[header_idx - 1, 0]) or f"Blok {header_idx + 1}"
        if category.upper() == "TGL":
            category = "QTY ALL"
        labels, series = [], []
        for data_idx in range(header_idx + 1, min(header_idx + 5, len(raw))):
            label = s_clean(raw.iloc[data_idx, 0])
            if not label or cell_text(label) in {"ACH", "MTD"}:
                continue
            values = [number_value(raw.iloc[data_idx, col]) for col in month_positions]
            if any(value is not None for value in values):
                labels.append(label)
                series.append(values)
        if len(series) >= 2:
            sections.append({"category": category, "months": [header[col].title() for col in month_positions], "labels": labels, "series": series})

    if not sections:
        st.error("Blok progres tidak ditemukan. Pastikan ada baris header bulan dan kolom seperti 'EST JUL'.")
        return
    st.success(f"{len(sections)} kategori progres terbaca. Proyeksi memakai {int(today_days)} hari berjalan.")
    for index, section in enumerate(sections):
        months = section["months"]
        current_month = months[-1]
        month_numbers = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6, "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}
        days_in_month = calendar.monthrange(date.today().year, month_numbers.get(current_month.upper(), date.today().month))[1]
        projected = [((values[-1] or 0) / int(today_days)) * days_in_month for values in section["series"]]
        table = pd.DataFrame(section["series"], index=section["labels"], columns=months)
        table[f"EST {current_month.upper()}"] = projected
        st.subheader(section["category"])
        st.dataframe(table.style.format("{:,.0f}"), use_container_width=True)
        fig = go.Figure()
        chart_months = months + [f"EST {current_month}"]
        for label, values, estimate in zip(section["labels"], section["series"], projected):
            fig.add_trace(go.Scatter(x=chart_months, y=values + [estimate], mode="lines+markers", name=label, line={"width": 3}))
        fig.update_layout(title=f"{section['category']} — progres vs proyeksi {current_month}", height=360, margin={"l": 20, "r": 20, "t": 55, "b": 20}, yaxis_title="Nilai", legend_title="Metrik", hovermode="x unified")
        st.plotly_chart(fig, use_container_width=True, key=f"progres_on_chart_{index}")


def render_progres_on_v2():
    """Dashboard progres sesuai KPI laporan ON."""
    import calendar
    from datetime import date, datetime
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    st.title("Progres ON")
    st.caption("Dashboard KPI dari laporan ON. EST = aktual bulan berjalan / hari berjalan x jumlah hari dalam bulan.")
    uploaded = st.file_uploader("Upload file Progres ON (.xlsx)", type=["xlsx"], key="progres_on_upload_v2")
    if not uploaded:
        st.info("Pilih file Excel untuk melihat dashboard progres.")
        return
    try:
        workbook = pd.ExcelFile(uploaded)
        sheet_name = st.selectbox("Sheet", workbook.sheet_names, key="progres_on_sheet_v2")
        raw = pd.read_excel(workbook, sheet_name=sheet_name, header=None)
    except Exception as e:
        st.error(f"File Excel tidak dapat dibaca: {e}")
        return

    detected_days = None
    for _, row in raw.iterrows():
        if s_clean(row.iloc[0]).upper() == "TGL":
            detected_days = to_int_or_none(row.iloc[1])
            break
    today_days = st.number_input("Hari berjalan (today)", min_value=1, max_value=31, value=min(int(detected_days or date.today().day), 31), step=1, key="progres_on_days_v2")
    month_numbers = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6, "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}

    def num(value):
        if isinstance(value, (int, float)) and not pd.isna(value):
            return float(value)
        value = parse_price_cell(value)
        return float(value) if value is not None else None

    def label(value):
        if value is None or pd.isna(value):
            return ""
        return value.strftime("%d %b") if isinstance(value, (datetime, date)) else s_clean(value)

    def add_year_to_decembers(labels):
        """Bedakan Desember tahun lalu dan Desember tahun berjalan pada grafik."""
        month_values = [month_numbers.get(item.upper()[:3]) for item in labels]
        rollover_count = sum(
            1 for previous, current in zip(month_values, month_values[1:])
            if previous is not None and current is not None and current < previous
        )
        year = date.today().year - rollover_count
        result = []
        previous = None
        for item, month in zip(labels, month_values):
            if previous is not None and month is not None and month < previous:
                year += 1
            result.append(f"{item} {year}" if month == 12 else item)
            if month is not None:
                previous = month
        return result

    def has_value(value):
        return value is not None and not pd.isna(value) and s_clean(value) != ""

    def read_section(name):
        start = next((i for i in range(len(raw)) if s_clean(raw.iloc[i, 0]).upper() == name), None)
        if start is None:
            return None
        header_row = start
        headers = raw.iloc[header_row].tolist()
        if not any(has_value(v) for v in headers[1:]):
            header_row += 1
            headers = raw.iloc[header_row].tolist()
        last_header = max((idx for idx, value in enumerate(headers) if has_value(value)), default=0)
        headers = headers[:last_header + 1]
        rows = {}
        for row_index in range(header_row + 1, len(raw)):
            row_label = s_clean(raw.iloc[row_index, 0])
            if not has_value(raw.iloc[row_index, 0]):
                break
            rows[row_label.upper()] = [num(value) for value in raw.iloc[row_index, 1:last_header + 1]]
        x_labels = add_year_to_decembers([label(v) for v in headers[1:]])
        return {"name": name.title(), "x": x_labels, "rows": rows}

    def has_future_actual_data(data):
        if not data:
            return False
        current_indexes = [idx for idx, x in enumerate(data["x"]) if month_numbers.get(x.upper()[:3]) == date.today().month]
        if not current_indexes:
            return False
        current_index = current_indexes[-1]
        future_indexes = [idx for idx in range(current_index + 1, len(data["x"])) if data["x"][idx].upper()[:3] in month_numbers]
        derived_rows = {"VS OLP", "MTD", "KONV%"}
        return any(
            idx < len(values) and values[idx] is not None
            for row_name, values in data["rows"].items()
            if row_name not in derived_rows
            for idx in future_indexes
        )

    def remove_estimate_columns(data):
        for est_index in reversed([idx for idx, x in enumerate(data["x"]) if x.upper().startswith("EST")]):
            data["x"].pop(est_index)
            for values in data["rows"].values():
                if est_index < len(values):
                    values.pop(est_index)
        return data

    def limit_to_current_month(data):
        """Jangan tampilkan data bulan masa depan walaupun template Excel sudah memuatnya."""
        if not data:
            return data
        if has_future_actual_data(data):
            return data
        current_indexes = [idx for idx, x in enumerate(data["x"]) if month_numbers.get(x.upper()[:3]) == date.today().month]
        if not current_indexes:
            return data
        current_index = current_indexes[-1]
        data["x"] = data["x"][:current_index + 1]
        for key, values in data["rows"].items():
            data["rows"][key] = values[:current_index + 1]
        return data

    def add_estimate(data):
        if not data:
            return data
        # Jika sudah ada data aktual setelah bulan kalender saat ini, laporan dianggap final/historis.
        if has_future_actual_data(data):
            return remove_estimate_columns(data)
        data = limit_to_current_month(data)
        current_month = date.today().month
        current_indexes = [idx for idx, x in enumerate(data["x"]) if month_numbers.get(x.upper()[:3]) == current_month]
        if not current_indexes:
            return data
        # Jika nama bulan muncul dua kali (mis. Dec tahun lalu dan Dec tahun ini), ambil yang terakhir.
        current_index = current_indexes[-1]
        days = calendar.monthrange(date.today().year, current_month)[1]
        existing_estimates = [idx for idx, x in enumerate(data["x"]) if x.upper().startswith("EST")]

        # Nilai bulan final tidak memerlukan proyeksi. Hapus EST template bila ada.
        if int(today_days) >= days:
            return remove_estimate_columns(data)

        # Hanya bulan kalender saat ini yang diproyeksikan; EST lain dari template diabaikan.
        for est_index in reversed(existing_estimates):
            data["x"].pop(est_index)
            for values in data["rows"].values():
                if est_index < len(values):
                    values.pop(est_index)
            if est_index < current_index:
                current_index -= 1

        est_index = current_index + 1
        data["x"].insert(est_index, f"EST {data['x'][current_index].upper()}")
        for values in data["rows"].values():
            values.insert(est_index, None)
            if current_index < len(values) and values[current_index] is not None:
                values[est_index] = values[current_index] / int(today_days) * days
        return data

    def complete_all_qty_from_categories(all_qty, laptop, phone_tab):
        """Isi bulan ALL QTY yang kosong dari Laptop + Pho+Tab tanpa mengganti angka aktual."""
        if not all([all_qty, laptop, phone_tab]):
            return all_qty
        reference_x = laptop["x"] if len(laptop["x"]) >= len(phone_tab["x"]) else phone_tab["x"]
        existing_x = all_qty["x"]
        all_qty["x"] = reference_x.copy()
        for row_name in ("03 OLP", "05 OLR"):
            existing = dict(zip(existing_x, all_qty["rows"].get(row_name, [])))
            laptop_values = dict(zip(laptop["x"], laptop["rows"].get(row_name, [])))
            phone_values = dict(zip(phone_tab["x"], phone_tab["rows"].get(row_name, [])))
            filled = []
            for month in reference_x:
                if existing.get(month) is not None:
                    filled.append(existing[month])
                    continue
                parts = [value for value in (laptop_values.get(month), phone_values.get(month)) if value is not None]
                filled.append(sum(parts) if parts else None)
            all_qty["rows"][row_name] = filled
        for row_name, values in all_qty["rows"].items():
            if row_name not in {"03 OLP", "05 OLR"}:
                all_qty["rows"][row_name] = values + [None] * max(0, len(reference_x) - len(values))
        return all_qty

    def line_chart(title, x, series, percent, key, compact=False, percent_decimals=0):
        fig = go.Figure()
        numeric_values = [value for _, values, _ in series for value in values if value is not None]
        low = min(0, min(numeric_values)) if numeric_values else 0
        high = max(numeric_values) if numeric_values else 1
        padding = max((high - low) * 0.16, abs(high) * 0.08, 0.002 if percent else 1)
        for name, values, color in series:
            def display_value(value):
                if value is None:
                    return ""
                if percent:
                    return f"{value:.{percent_decimals}%}"
                if compact and abs(value) >= 1_000_000:
                    return f"{value / 1_000_000:.0f} jt"
                return f"{value:,.0f}"
            texts = [display_value(value) for value in values]
            fig.add_trace(go.Scatter(x=x, y=values, mode="lines+markers", name=name, marker={"size": 7, "color": color}, line={"width": 3, "color": color, "shape": "spline", "smoothing": 1.1}))
            for point_x, point_y, point_text in zip(x, values, texts):
                if point_y is None:
                    continue
                fig.add_annotation(
                    x=point_x, y=point_y, text=point_text, showarrow=False, yshift=15,
                    bgcolor="rgba(255, 255, 255, 0.96)", bordercolor=color, borderwidth=1,
                    borderpad=3, font={"color": "#111827", "size": 10},
                )
        fig.update_layout(title=title, height=320, margin={"l": 16, "r": 16, "t": 66, "b": 16}, hovermode="x unified", legend={"orientation": "h", "y": 1.18})
        fig.update_yaxes(tickformat=f".{percent_decimals}%" if percent else ",.0f", range=[low - padding * 0.25, high + padding], automargin=True)
        fig.update_xaxes(tickfont={"color": "#374151"})
        fig.update_yaxes(tickfont={"color": "#374151"})
        st.plotly_chart(fig, width="stretch", key=key)

    def dual_line_chart(title, x, left_label, left, right_label, right, key):
        if not any(v is not None for v in left + right):
            st.info(f"{title}: data belum tersedia di Excel.")
            return
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        left_text = [f"{value:,.0f}" if value is not None else "" for value in left]
        right_text = [f"{value:,.0f}" if value is not None else "" for value in right]
        fig.add_trace(go.Scatter(x=x, y=left, text=left_text, textposition="top center", name=left_label, mode="lines+markers+text", line={"width": 3, "color": "#2563eb", "shape": "spline", "smoothing": 1.1}), secondary_y=False)
        fig.add_trace(go.Scatter(x=x, y=right, text=right_text, textposition="top center", name=right_label, mode="lines+markers+text", line={"width": 3, "color": "#f97316", "shape": "spline", "smoothing": 1.1}), secondary_y=True)
        fig.update_layout(title=title, height=360, margin={"l": 16, "r": 16, "t": 48, "b": 16}, hovermode="x unified", legend={"orientation": "h", "y": 1.12})
        fig.update_yaxes(title_text=left_label, tickformat=",.0f", secondary_y=False)
        fig.update_yaxes(title_text=right_label, tickformat=",.0f", secondary_y=True)
        st.plotly_chart(fig, width="stretch", key=key)

    raw_divisions = [read_section(name) for name in ["ALL QTY", "LAPTOP", "PHO + TAB"]]
    if all(raw_divisions):
        complete_all_qty_from_categories(*raw_divisions)
    divisions = [add_estimate(data) for data in raw_divisions]
    if not all(divisions):
        st.error("Blok ALL QTY, LAPTOP, dan PHO + TAB belum lengkap.")
        return
    website = add_estimate(read_section("WEBSITE"))
    # Meta Ads (klik dan spending) ditampilkan sebagai data aktual, tanpa EST Juli.
    meta = limit_to_current_month(read_section("META ADS"))
    google = limit_to_current_month(read_section("GOOGLE ADS") or read_section("GOOGLE AD"))
    live = read_section("HOST LIVE")
    lenovo = add_estimate(read_section("LAPTOP LENOVO"))
    st.success(f"Dashboard siap. Proyeksi memakai {int(today_days)} hari berjalan.")

    st.subheader("1. 05 OLR berapa persen dari 03 OLP")
    cols = st.columns(3)
    for i, data in enumerate(divisions):
        target, actual = data["rows"].get("03 OLP", []), data["rows"].get("05 OLR", [])
        ratio = [a / t if a is not None and t not in (None, 0) else None for t, a in zip(target, actual)]
        with cols[i]:
            line_chart(data["name"], data["x"], [("05 OLR / 03 OLP", ratio, "#2563eb")], True, f"ratio_{i}")

    st.subheader("2. MTD 05 OLR versus bulan sebelumnya")
    cols = st.columns(3)
    for i, data in enumerate(divisions):
        actual = data["rows"].get("05 OLR", [])
        est_index = next((idx for idx, label in enumerate(data["x"]) if label.upper().startswith("EST")), None)
        mtd = []
        for idx, current in enumerate(actual):
            # EST bulan berjalan harus dibandingkan dengan bulan sebelumnya (Juni),
            # bukan dengan nilai aktual Juli yang masih berjalan.
            # Setelah EST disisipkan, bulan-bulan berikutnya tetap dibandingkan dengan
            # nilai aktual bulan sebelumnya, bukan dengan nilai proyeksi.
            previous_index = idx - 2 if est_index is not None and idx >= est_index else idx - 1
            previous = actual[previous_index] if previous_index >= 0 else None
            mtd.append(current / previous - 1 if current is not None and previous not in (None, 0) else None)
        with cols[i]:
            line_chart(data["name"], data["x"], [("MTD 05 OLR", mtd, "#16a34a")], True, f"mtd_{i}")

    st.subheader("3. MTD 05 OLR versus bulan sebelumnya - Jumlah QTY")
    cols = st.columns(3)
    for i, data in enumerate(divisions):
        with cols[i]:
            line_chart(data["name"], data["x"], [("QTY 05 OLR", data["rows"].get("05 OLR", []), "#7c3aed")], False, f"mtd_qty_{i}")

    def conversion_rate(data, qty_row):
        if not data:
            return []
        qty = data["rows"].get(qty_row, [])
        clicks = data["rows"].get("KLIK", [])
        return [quantity / click if quantity is not None and click not in (None, 0) else None for quantity, click in zip(qty, clicks)]

    def indicator_cards(items, section_key):
        for start in range(0, len(items), 3):
            cols = st.columns(3)
            for col, item in zip(cols, items[start:start + 3]):
                title, data, row_name, color, percent, compact = item
                with col:
                    values = conversion_rate(data, "QTY" if data is website else "QTY ALL LAP") if row_name == "__CONVERSION__" else (data["rows"].get(row_name, []) if data else [])
                    if values and any(value is not None for value in values):
                        metric_name = "Konversi" if row_name == "__CONVERSION__" else row_name.title()
                        line_chart(title, data["x"], [(metric_name, values, color)], percent, f"{section_key}_{title}", compact=compact, percent_decimals=2 if row_name == "__CONVERSION__" else 0)
                    else:
                        st.info(f"{title}: data belum tersedia di Excel.")

    st.subheader("4. Website")
    indicator_cards([
        ("Website - QTY", website, "QTY", "#2563eb", False, False),
        ("Website - GMV", website, "GMV", "#f97316", False, True),
        ("Website - Klik", website, "KLIK", "#16a34a", False, False),
    ], "website")

    st.subheader("5. Meta Ads")
    indicator_cards([
        ("Meta Ads - Spend", meta, "SPEND", "#f97316", False, False),
        ("Meta Ads - Klik", meta, "KLIK", "#16a34a", False, False),
    ], "meta")

    st.subheader("6. Google Ads")
    indicator_cards([
        ("Google Ads - Spend", google, "SPEND", "#f97316", False, False),
        ("Google Ads - Klik", google, "KLIK", "#16a34a", False, False),
    ], "google")

    st.subheader("7. Host Live - QTY dan GMV per hari")
    indicator_cards([
        ("Host Live - QTY", live, "QTY", "#2563eb", False, False),
        ("Host Live - GMV", live, "GMV", "#f97316", False, True),
    ], "live")

    st.subheader("8. Laptop Lenovo")
    indicator_cards([
        ("Laptop Lenovo - QTY", lenovo, "05 OLR", "#2563eb", False, False),
    ], "lenovo")


def build_margin_df_for_affiliate(pl_bytes: bytes, harga_key: str = "M3") -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True, read_only=False)
    target_sheets = [name for name in ["LAPTOP", "TELCO", "PC HOM ELE"] if name in wb.sheetnames]
    rows: List[Dict[str, Any]] = []

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        skip_rows: Set[int] = set()
        if su(sheet_name) == "LAPTOP":
            r_start = find_row_contains(ws, "COMING", scan_rows=600)
            r_end = find_row_contains(ws, "END COMING", scan_rows=1200)
            if r_start and r_end and r_end >= r_start:
                skip_rows = set(range(r_start, r_end + 1))

        try:
            header_row, sku_col, price_cols = find_header_row_and_cols_pricelist_fixed(ws, ["M0", harga_key])
        except Exception:
            continue

        m0_col = price_cols["M0"]
        harga_col = price_cols[harga_key]
        spec_col = get_header_col_fuzzy(ws, header_row, ["SPESIFIKASI", "Specification", "Nama Barang", "Nama Produk"])

        for r in range(header_row + 1, ws.max_row + 1):
            if r in skip_rows:
                continue
            sku = norm_sku(ws.cell(row=r, column=sku_col).value)
            if not sku or sku in ("TOTAL", "KODEBARANG", "KODE BARANG"):
                continue
            m0 = parse_price_cell(ws.cell(row=r, column=m0_col).value)
            harga_jual = parse_price_cell(ws.cell(row=r, column=harga_col).value)
            if m0 is None or harga_jual is None or m0 <= 0 or harga_jual <= 0:
                continue
            m0 = apply_multiplier_if_needed(m0)
            harga_jual = apply_multiplier_if_needed(harga_jual)
            biaya = (harga_jual * 0.047) + 150
            margin_rp = harga_jual - m0 - biaya
            margin_pct = margin_rp / harga_jual
            rows.append({
                "KODEBARANG": sku,
                "SPESIFIKASI": s_clean(ws.cell(row=r, column=spec_col).value) if spec_col else "",
                "M0": int(m0),
                harga_key: int(harga_jual),
                "Margin %": float(margin_pct),
            })

    return pd.DataFrame(rows)


def find_product_id_col_any_row(ws: Worksheet, scan_rows: int = 10) -> Tuple[Optional[int], Optional[int]]:
    return find_col_contains_any_row(
        ws,
        ["ID PRODUK", "PRODUCT ID", "PRODUCT_ID", "ID PRODUCT"],
        scan_rows=scan_rows,
    )


def process_affiliate_tiktokshop(mass_file: Any, pl_mgr_file: Any, affiliate_file: Any, harga_key: str, margin_min: float, margin_max: float, progress_callback=None):
    issues: List[Dict[str, Any]] = []
    summary = init_summary(1)

    if progress_callback:
        progress_callback(5, "Membaca PL MGR dan menghitung margin...")
    margin_df = build_margin_df_for_affiliate(pl_mgr_file.getvalue(), harga_key)
    if margin_df.empty:
        raise ValueError("PL MGR tidak menghasilkan data margin. Pastikan ada KODEBARANG, M0, dan kolom harga yang dipilih.")

    selected_df = margin_df[(margin_df["Margin %"] >= margin_min) & (margin_df["Margin %"] <= margin_max)].copy()
    selected_skus = set(selected_df["KODEBARANG"].dropna().astype(str).map(norm_sku))
    if not selected_skus:
        issues.append({"file": pl_mgr_file.name, "reason": "Tidak ada KODEBARANG yang masuk range margin terpilih."})

    if progress_callback:
        progress_callback(25, "Membaca file mass update Affiliate...")
    mass_wb = load_workbook(io.BytesIO(mass_file.getvalue()), data_only=True, read_only=False)
    mass_ws = mass_wb["Template"] if "Template" in mass_wb.sheetnames else mass_wb.active
    pid_header_row, product_id_col = find_product_id_col_any_row(mass_ws, scan_rows=10)
    sku_header_row, seller_sku_col = find_sku_col_any_row(mass_ws, scan_rows=10)
    if product_id_col is None:
        raise ValueError("Kolom ID Produk / Product ID tidak ditemukan di File Mass Update.")
    if seller_sku_col is None:
        raise ValueError("Kolom SKU Penjual / Seller SKU tidak ditemukan di File Mass Update.")

    data_start = max([x for x in [pid_header_row, sku_header_row] if x is not None]) + 3
    product_ids: Set[str] = set()
    total_mass_rows = max(1, mass_ws.max_row - data_start + 1)
    for r in range(data_start, mass_ws.max_row + 1):
        if progress_callback and (r - data_start) % 2000 == 0:
            progress_tick(progress_callback, 35, 60, r - data_start, total_mass_rows, f"Scan mass update Affiliate: {r - data_start} baris...")
        seller_sku_full = s_clean(mass_ws.cell(row=r, column=seller_sku_col).value)
        if not seller_sku_full:
            continue
        summary["rows_scanned"] += 1
        base_sku, _ = split_sku_addons(seller_sku_full)
        if norm_sku(base_sku) not in selected_skus:
            continue
        product_id = parse_number_like_id(mass_ws.cell(row=r, column=product_id_col).value)
        if product_id:
            product_ids.add(product_id)

    if not product_ids:
        issues.append({"file": mass_file.name, "reason": "Tidak ada ID Produk di Mass Update yang cocok dengan KODEBARANG dari range margin."})

    if progress_callback:
        progress_callback(65, "Membaca file Affiliate dan update komisi...")
    aff_wb = load_workbook(io.BytesIO(affiliate_file.getvalue()), data_only=False, read_only=False)
    aff_ws = aff_wb.active
    aff_pid_header_row, aff_product_id_col = find_product_id_col_any_row(aff_ws, scan_rows=10)
    if aff_product_id_col is None:
        raise ValueError("Kolom ID Produk / Product ID tidak ditemukan di File Affiliate.")

    # Sesuai format Affiliate TikTokShop: F = YES, G sampai I = 100.
    col_yes = 6
    commission_cols = [7, 8, 9]
    data_start_aff = (aff_pid_header_row or 1) + 1

    for r in range(data_start_aff, aff_ws.max_row + 1):
        pid = parse_number_like_id(aff_ws.cell(row=r, column=aff_product_id_col).value)
        if not pid or pid.upper().startswith("PLEASE"):
            continue
        if pid not in product_ids:
            continue
        safe_set_cell_value(aff_ws, r, col_yes, "YES")
        for c in commission_cols:
            safe_set_cell_value(aff_ws, r, c, 100)
        summary["rows_written"] += 1

    if summary["rows_written"] == 0:
        issues.append({"file": affiliate_file.name, "reason": "Tidak ada ID Produk di File Affiliate yang cocok."})

    summary["rows_unmatched"] = max(0, len(product_ids) - summary["rows_written"])
    if progress_callback:
        progress_callback(92, "Menyimpan output Affiliate TikTokShop...")
    summary["issues_count"] = len(issues)
    return workbook_to_bytes(aff_wb), f"hasil_affiliate_tiktokshop_{affiliate_file.name}", issues_workbook_or_none(issues), summary, margin_df


def _uploaded_file_sig(f):
    if f is None:
        return None
    try:
        return (getattr(f, "name", ""), int(getattr(f, "size", 0) or len(f.getvalue())))
    except Exception:
        return (getattr(f, "name", ""), 0)


def _clear_affiliate_cache_if_input_changed(signature):
    sig_key = "affiliate_tiktokshop_input_signature"
    if st.session_state.get(sig_key) != signature:
        st.session_state.download_cache.pop("affiliate_tiktokshop", None)
        st.session_state.summary_cache.pop("affiliate_tiktokshop", None)
        st.session_state[sig_key] = signature


def render_affiliate_tiktokshop():
    page_header(
        "Affiliate TikTokShop",
        "Generate file Affiliate TikTokShop berdasarkan range margin dari PL MGR, lalu update File Affiliate: kolom F = YES dan kolom G-I = 100 untuk ID Produk yang cocok.",
        ["File Mass Update TikTokShop (.xlsx)", "PL MGR / Pricelist Margin (.xlsx)", "File Affiliate TikTokShop (.xlsx)"],
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        mass_file = st.file_uploader("Upload File Mass Update", type=["xlsx"], key="affiliate_mass")
    with c2:
        pl_mgr_file = st.file_uploader("Upload PL MGR", type=["xlsx"], key="affiliate_pl_mgr")
    with c3:
        affiliate_file = st.file_uploader("Upload File Affiliate", type=["xlsx"], key="affiliate_file")

    harga_key = st.radio("Harga acuan margin", ["M3", "M4"], index=0, horizontal=True, key="affiliate_harga_key")
    selected_pct = st.slider(
        "Pilih Range Margin (%)",
        min_value=-100.0,
        max_value=100.0,
        value=(-100.0, 100.0),
        step=0.1,
        key="affiliate_margin_slider",
    )
    margin_range = (selected_pct[0] / 100, selected_pct[1] / 100)
    st.caption("File hasil baru dibuat setelah klik Proses.")

    current_signature = (
        _uploaded_file_sig(mass_file),
        _uploaded_file_sig(pl_mgr_file),
        _uploaded_file_sig(affiliate_file),
        harga_key,
        tuple(selected_pct),
    )
    _clear_affiliate_cache_if_input_changed(current_signature)

    if st.button("Proses", key="btn_affiliate"):
        if not mass_file or not pl_mgr_file or not affiliate_file:
            st.error("Upload File Mass Update, PL MGR, dan File Affiliate dulu.")
            return
        try:
            result_bytes, result_name, issues_bytes, summary, margin_df = run_with_loading_callback(
                lambda progress_callback: process_affiliate_tiktokshop(mass_file, pl_mgr_file, affiliate_file, harga_key, margin_range[0], margin_range[1], progress_callback),
                "Memproses Affiliate TikTokShop...",
            )
            cache_downloads("affiliate_tiktokshop", result_name, result_bytes, issues_bytes, summary=summary)
            if margin_df is not None and not margin_df.empty:
                st.caption(f"SKU PL MGR yang masuk range margin: {int(summary.get('sku_selected', 0))}")
        except Exception as e:
            st.error(f"Gagal memproses Affiliate: {e}")

    render_cached_summary("affiliate_tiktokshop")
    render_downloads("affiliate_tiktokshop")


# Backward compatible route name.
def render_affiliate():
    render_affiliate_tiktokshop()


def render_analisa_margin():
    st.title("Analisa Margin")
    st.caption("Upload Pricelist untuk menghitung % margin per SKU dari sheet LAPTOP, TELCO, dan PC HOM ELE.")

    pricelist_file = st.file_uploader("Upload Pricelist (.xlsx)", type=["xlsx"], key="analisa_margin_pricelist")
    harga_margin_key = st.radio(
        "Pilih Harga untuk Analisa Margin",
        ["M3", "M4"],
        index=1,
        horizontal=True,
        key="analisa_margin_harga_key",
    )
    if pricelist_file is None:
        st.info("Upload Pricelist dulu.")
        return

    def _format_idr_local(x):
        try:
            return f"IDR {int(round(float(x))):,}".replace(",", ".")
        except Exception:
            return "-"

    def _format_pct_local(x):
        try:
            return f"{float(x) * 100:.2f}%"
        except Exception:
            return "-"

    def _coming_rows_to_skip(ws: Worksheet) -> Set[int]:
        skip: Set[int] = set()
        r_start = find_row_contains(ws, "COMING", scan_rows=600)
        r_end = find_row_contains(ws, "END COMING", scan_rows=1200)
        if r_start and r_end and r_end >= r_start:
            skip = set(range(r_start, r_end + 1))
        return skip

    def _build_margin_df(pl_bytes: bytes, harga_key: str) -> pd.DataFrame:
        wb = load_workbook(io.BytesIO(pl_bytes), data_only=True, read_only=False)
        target_sheets = ["LAPTOP", "TELCO", "PC HOM ELE"]
        rows: List[Dict[str, Any]] = []
        issues: List[str] = []

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                issues.append(f"Sheet {sheet_name} tidak ditemukan.")
                continue

            ws = wb[sheet_name]
            skip_rows = _coming_rows_to_skip(ws) if su(sheet_name) == "LAPTOP" else set()

            try:
                header_row, sku_col, price_cols = find_header_row_and_cols_pricelist_fixed(ws, ["M0", harga_key])
            except Exception as e:
                issues.append(f"Sheet {sheet_name}: {e}")
                continue

            spec_col = get_header_col_fuzzy(ws, header_row, ["SPESIFIKASI", "Specification", "Nama Barang", "Nama Produk"])
            product_col = get_header_col_fuzzy(ws, header_row, ["PRODUCT", "Produk", "Category"])
            m0_col = price_cols["M0"]
            harga_col = price_cols[harga_key]
            try:
                _, tot_col = find_tot_col(ws, header_row)
            except Exception:
                tot_col = None

            for r in range(header_row + 1, ws.max_row + 1):
                if r in skip_rows:
                    continue

                sku = norm_sku(ws.cell(row=r, column=sku_col).value)
                if not sku or sku in ("TOTAL", "KODEBARANG", "KODE BARANG"):
                    continue

                m0 = parse_price_cell(ws.cell(row=r, column=m0_col).value)
                harga_jual = parse_price_cell(ws.cell(row=r, column=harga_col).value)
                if m0 is None or m0 <= 0:
                    continue
                if harga_jual is None or harga_jual <= 0:
                    continue

                m0 = apply_multiplier_if_needed(m0)
                harga_jual = apply_multiplier_if_needed(harga_jual)
                if harga_jual <= 0:
                    continue

                biaya = (harga_jual * 0.047) + 150
                margin_rp = harga_jual - m0 - biaya
                margin_pct = margin_rp / harga_jual
                stok_tot = to_int_or_none(ws.cell(row=r, column=tot_col).value) if tot_col else None

                rows.append({
                    "KODEBARANG": sku,
                    "SPESIFIKASI": s_clean(ws.cell(row=r, column=spec_col).value) if spec_col else "",
                    "PRODUCT": s_clean(ws.cell(row=r, column=product_col).value) if product_col else "",
                    "STOK TOT": int(stok_tot) if stok_tot is not None else "",
                    "M0": int(m0),
                    harga_key: int(harga_jual),
                    "Biaya 4.7% + 150": float(biaya),
                    "Margin Rp": float(margin_rp),
                    "Margin %": float(margin_pct),
                })

        if issues:
            st.warning(" | ".join(issues))
        return pd.DataFrame(rows)

    try:
        df = _build_margin_df(pricelist_file.getvalue(), harga_margin_key)
    except Exception as e:
        st.error(f"Gagal membaca Pricelist: {e}")
        return

    if df.empty:
        st.warning(f"Tidak ada SKU valid. Pastikan sheet LAPTOP/TELCO/PC HOM ELE punya header KODEBARANG, M0, dan {harga_margin_key}, serta M0/{harga_margin_key} tidak kosong.")
        return

    product_options = sorted([
        x for x in df["PRODUCT"].dropna().astype(str).unique().tolist()
        if x.strip()
    ])

    if "analisa_margin_product_filter_applied" not in st.session_state:
        st.session_state["analisa_margin_product_filter_applied"] = []

    with st.form("analisa_margin_filter_form", clear_on_submit=False):
        selected_products_temp = st.multiselect(
            "Filter by PRODUCT",
            product_options,
            default=st.session_state["analisa_margin_product_filter_applied"],
            key="analisa_margin_product_filter_temp",
        )
        apply_filter = st.form_submit_button("Konfirmasi Filter", use_container_width=True)

    if apply_filter:
        st.session_state["analisa_margin_product_filter_applied"] = selected_products_temp

    selected_products = st.session_state["analisa_margin_product_filter_applied"]
    if selected_products:
        df = df[df["PRODUCT"].isin(selected_products)].copy()

    c1, c2, c3 = st.columns(3)
    c1.metric("SKU Terbaca", len(df))
    c2.metric("Margin Minus", int((df["Margin %"] < 0).sum()))
    c3.metric("Rata-rata Margin", _format_pct_local(df["Margin %"].mean()))

    sort_mode = st.radio(
        "Urutan Margin",
        ["Small to Large", "Large to Small"],
        horizontal=True,
        key="analisa_margin_sort_mode",
    )
    ascending = sort_mode == "Small to Large"
    view = df.sort_values("Margin %", ascending=ascending).copy()

    view_display = view.copy()
    for col in ["M0", harga_margin_key, "Biaya 4.7% + 150", "Margin Rp"]:
        view_display[col] = view_display[col].map(_format_idr_local)
    view_display["Margin %"] = view_display["Margin %"].map(_format_pct_local)

    st.dataframe(
        view_display,
        use_container_width=True,
        hide_index=True,
        height=650,
    )


# ============================================================
# POSTING - SHOPEE
# ============================================================
POSTING_SHOPEE_BASE_URL = "https://www.agres.id"
POSTING_SHOPEE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36 CodexPostingShopee/1.1",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
}


def ensure_posting_shopee_dependencies():
    missing = []
    if requests is None:
        missing.append("requests")
    if BeautifulSoup is None:
        missing.append("beautifulsoup4")
    if fuzz is None:
        missing.append("rapidfuzz")
    if missing:
        raise RuntimeError(
            "Fitur Posting Shopee butuh package tambahan: "
            + ", ".join(missing)
            + ". Tambahkan package tersebut ke requirements.txt lalu redeploy aplikasi."
        )


def posting_clean_text(text):
    if text is None:
        return ""
    try:
        if pd.isna(text):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(text or "")).strip()


def posting_sku_tokens(text: str) -> List[str]:
    raw = posting_clean_text(text)
    parts = [p for p in re.split(r"[-_\s/]+", raw) if p]
    tokens = []
    for p in parts:
        p = p.strip()
        if len(p) >= 4:
            tokens.append(p)
    return tokens


def posting_query_variants(kodebarang, spesifikasi) -> List[str]:
    kodebarang = posting_clean_text(kodebarang)
    spesifikasi = posting_clean_text(spesifikasi)
    tokens = posting_sku_tokens(kodebarang)

    variants = []
    if spesifikasi and kodebarang:
        variants.append(f"{spesifikasi} {kodebarang}")
    if kodebarang:
        variants.append(kodebarang)
    if tokens:
        variants.append(" ".join(tokens[-2:]))
        variants.append(tokens[-1])
    if spesifikasi:
        variants.append(spesifikasi)

    seen = set()
    clean_variants = []
    for q in variants:
        q = posting_clean_text(q)
        key = q.lower()
        if q and key not in seen:
            seen.add(key)
            clean_variants.append(q)
    return clean_variants


_POSTING_SHOPEE_HTTP_SESSION = None


def posting_get_http_session():
    """Pakai ulang koneksi HTTP selama proses batch supaya request AGRES.ID lebih ringan."""
    ensure_posting_shopee_dependencies()
    global _POSTING_SHOPEE_HTTP_SESSION
    if _POSTING_SHOPEE_HTTP_SESSION is None:
        _POSTING_SHOPEE_HTTP_SESSION = requests.Session()
        _POSTING_SHOPEE_HTTP_SESSION.headers.update(POSTING_SHOPEE_HEADERS)
    return _POSTING_SHOPEE_HTTP_SESSION


def posting_get_html(url, retries: int = 3):
    ensure_posting_shopee_dependencies()
    session = posting_get_http_session()
    last_error = None

    for attempt in range(max(1, int(retries or 1))):
        try:
            response = session.get(url, timeout=(8, 25))

            # AGRES.ID kadang membalas sementara penuh / terlalu banyak request.
            # Jangan langsung matikan proses, tunggu sebentar lalu coba ulang.
            if response.status_code in (429, 500, 502, 503, 504):
                last_error = RuntimeError(f"HTTP {response.status_code} dari AGRES.ID")
                time.sleep(2 + attempt * 2)
                continue

            response.raise_for_status()
            time.sleep(0.55)
            return response.text

        except Exception as error:
            last_error = error
            time.sleep(2 + attempt * 2)

    raise last_error


def posting_product_url_from_text(text: str) -> str:
    if not text:
        return ""
    match = re.search(r"https?://(?:www\.)?agres\.id/products/[^\s\"'<>]+", str(text), flags=re.I)
    if match:
        return match.group(0).rstrip("),.;")
    return ""


def posting_extract_product_links_from_html(html: str) -> List[Dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    candidates = []

    for a in soup.select("a[href]"):
        href = a.get("href", "")
        if "/products/" not in href:
            continue
        title = posting_clean_text(a.get_text(" "))
        full_url = urljoin(POSTING_SHOPEE_BASE_URL, href)
        candidates.append({"title": title, "url": full_url})

    # Next/React pages kadang menyimpan link di script, bukan anchor biasa.
    for href in re.findall(r"(?:https?://www\.agres\.id)?/products/[^\"'<>\s]+", html, flags=re.I):
        full_url = urljoin(POSTING_SHOPEE_BASE_URL, href.replace("\\/", "/"))
        candidates.append({"title": "", "url": full_url})

    seen = set()
    unique = []
    for item in candidates:
        url = item.get("url", "").split("?")[0]
        if not url or url in seen:
            continue
        seen.add(url)
        unique.append({"title": item.get("title", ""), "url": url})
    return unique


def posting_search_product(query):
    # Tetap langsung mengambil dari AGRES.ID, bukan Google/Bing/API luar.
    search_urls = [
        f"{POSTING_SHOPEE_BASE_URL}/catalogue?search={quote_plus(query)}",
        f"{POSTING_SHOPEE_BASE_URL}/search?q={quote_plus(query)}",
    ]

    candidates = []
    errors = []
    for url in search_urls:
        try:
            html = posting_get_html(url)
            candidates.extend(posting_extract_product_links_from_html(html))
        except Exception as error:
            errors.append(f"{url}: {error}")

    # Simpan error untuk dipakai kalau semua query gagal total.
    posting_search_product.last_errors = errors
    return candidates


posting_search_product.last_errors = []


def posting_score_candidate(query: str, kodebarang: str, spesifikasi: str, candidate: Dict[str, str]) -> int:
    compare_text = f'{candidate.get("title", "")} {candidate.get("url", "")}'.lower()
    query_score = fuzz.token_set_ratio(query.lower(), compare_text)

    kode_tokens = posting_sku_tokens(kodebarang)
    token_bonus = 0
    for token in kode_tokens:
        if token.lower() in compare_text:
            token_bonus += 18

    spec_tokens = [t for t in posting_sku_tokens(spesifikasi) if len(t) >= 5]
    for token in spec_tokens[:8]:
        if token.lower() in compare_text:
            token_bonus += 3

    return min(100, int(query_score + token_bonus))


def posting_find_best_match(kodebarang, spesifikasi):
    direct_url = posting_product_url_from_text(kodebarang) or posting_product_url_from_text(spesifikasi)
    if direct_url:
        return {"title": "", "url": direct_url, "score": 100, "source": "DIRECT_URL"}

    all_candidates = []
    all_errors = []
    queries = posting_query_variants(kodebarang, spesifikasi)

    for query in queries:
        candidates = posting_search_product(query)
        all_errors.extend(getattr(posting_search_product, "last_errors", []))
        for item in candidates:
            item = dict(item)
            item["query"] = query
            item["score"] = posting_score_candidate(query, kodebarang, spesifikasi, item)
            item["source"] = "AGRES_SEARCH"
            all_candidates.append(item)
        if any(int(c.get("score", 0)) >= 85 for c in all_candidates):
            break

    unique = {}
    for item in all_candidates:
        url = item.get("url", "")
        if not url:
            continue
        if url not in unique or int(item.get("score", 0)) > int(unique[url].get("score", 0)):
            unique[url] = item

    if not unique:
        return {"not_found_error": "; ".join(all_errors[-2:]) if all_errors else "Tidak ada link produk dari hasil pencarian AGRES.ID"}

    best = max(unique.values(), key=lambda x: int(x.get("score", 0)))
    return best


def posting_normalize_image_url(src):
    if not src:
        return ""

    src = str(src).strip()

    if src.startswith("data:"):
        return ""

    src = src.replace("\\/", "/")
    return urljoin(POSTING_SHOPEE_BASE_URL, src)


def posting_to_original_image_url(url):
    if not url:
        return ""

    url = posting_normalize_image_url(url)
    parsed = urlparse(url)

    if "/_next/image" in parsed.path:
        qs = parse_qs(parsed.query)

        if "url" in qs and qs["url"]:
            return unquote(qs["url"][0])

    return url


def posting_image_identity(url):
    original = posting_to_original_image_url(url)
    parsed = urlparse(original)
    path = parsed.path.lower()
    path = re.sub(r"/w_\d+[^/]*?/", "/", path)
    path = re.sub(r"/c_[^/]+/", "/", path)
    path = re.sub(r"/f_[^/]+/", "/", path)
    return path


def posting_looks_like_image_url(url):
    low = url.lower()

    # Jangan ambil logo/icon SVG. Shopee butuh direct bitmap image URL.
    if re.search(r"\.svg(\?|$)", low):
        return False

    if "/_next/image" in low:
        return True

    if re.search(r"\.(jpg|jpeg|png|webp|avif)(\?|$)", low):
        return True

    image_markers = [
        "/storage/",
        "/uploads/",
        "/upload/",
        "/products/",
        "/product/",
        "/produk/",
        "/catalog/",
        "/catalogue/",
        "res.cloudinary.com",
        "image/upload",
        "cdn",
        "cloudfront",
        "s3.",
    ]

    return any(marker in low for marker in image_markers)


def posting_is_bad_image(url, alt=""):
    low = f"{url} {alt}".lower()

    if re.search(r"\.svg(\?|$)", low):
        return True

    bad_words = [
        "brand-logo",
        "agres-red",
        "agrescare",
        "agreskomputer",
        "whatsapp",
        "email",
        "gmail",
        "facebook",
        "instagram",
        "youtube",
        "twitter",
        "tiktok",
        "favicon",
        "banner",
        "payment",
        "bank",
        "store-locator",
        "placeholder",
        "no-image",
        "background",
        "newsletter",
    ]

    return any(word in low for word in bad_words)


def posting_get_img_candidates(img):
    candidates = []

    srcset = img.get("srcset") or img.get("data-srcset")

    if srcset:
        for part in srcset.split(","):
            candidate = part.strip().split(" ")[0]
            candidates.append(candidate)

    for attr in [
        "src",
        "data-src",
        "data-lazy-src",
        "data-original",
        "data-zoom-image",
        "data-large_image",
        "data-image",
        "data-full",
    ]:
        candidates.append(img.get(attr))

    return candidates


def posting_extract_title(soup):
    h1 = soup.find("h1")

    if h1:
        return posting_clean_text(h1.get_text(" "))

    og_title = soup.find("meta", property="og:title")

    if og_title and og_title.get("content"):
        return posting_clean_text(og_title.get("content"))

    title_tag = soup.find("title")
    if title_tag:
        return posting_clean_text(title_tag.get_text(" "))

    return ""


def posting_get_page_before_related(html):
    markers = [
        "Related Items",
        "Produk Terkait",
        "Berlangganan Newsletter",
    ]

    cut_at = len(html)

    for marker_text in markers:
        pos = html.lower().find(marker_text.lower())

        if pos != -1:
            cut_at = min(cut_at, pos)

    return html[:cut_at]


def posting_add_image(images, identities, url, alt=""):
    if not url:
        return False

    final_url = posting_to_original_image_url(url)
    ident = posting_image_identity(final_url)

    if not final_url or not ident:
        return False

    if ident in identities:
        return False

    if posting_is_bad_image(final_url, alt):
        return False

    if not posting_looks_like_image_url(final_url):
        return False

    identities.add(ident)
    images.append(final_url)

    return True


def posting_extract_images(soup, html, limit=4):
    html_main = posting_get_page_before_related(html)
    main_soup = BeautifulSoup(html_main, "html.parser")

    images = []
    identities = set()

    # Ambil meta image lebih dulu. Ini biasanya gambar produk utama AGRES.ID.
    for meta_selector in [
        {"property": "og:image"},
        {"name": "twitter:image"},
    ]:
        meta = soup.find("meta", attrs=meta_selector)
        if meta and meta.get("content"):
            posting_add_image(images, identities, meta.get("content"), "meta")

    gallery_keywords = ["gallery", "thumbnail", "thumb", "swiper", "slick", "carousel"]
    image_rows = []

    for img in main_soup.find_all("img"):
        alt = posting_clean_text(img.get("alt") or img.get("title") or "")
        parent_html = str(img.parent).lower()[:2500] if img.parent else ""
        parent_score = 0

        if any(key in parent_html for key in gallery_keywords):
            parent_score += 80

        if alt:
            parent_score += 20

        for src in posting_get_img_candidates(img):
            url = posting_normalize_image_url(src)

            if not url:
                continue

            if posting_is_bad_image(url, alt):
                continue

            if not posting_looks_like_image_url(url):
                continue

            original = posting_to_original_image_url(url)
            score = parent_score

            if "/_next/image" in url.lower():
                score += 20

            if "res.cloudinary.com" in original.lower():
                score += 20

            if any(x in original.lower() for x in ["logo", "brand", "agres-red"]):
                score -= 150

            image_rows.append((score, original, alt))

    image_rows.sort(key=lambda x: x[0], reverse=True)

    for score, url, alt in image_rows:
        posting_add_image(images, identities, url, alt)

        if len(images) >= limit:
            break

    if len(images) < limit:
        for match in re.findall(r'/_next/image\?url=[^"\']+', html_main, flags=re.I):
            posting_add_image(images, identities, match)

            if len(images) >= limit:
                break

    images = images[:limit]

    while len(images) < limit:
        images.append("")

    return images


def posting_extract_specification(soup):
    page_text = soup.get_text("\n")
    page_text = re.sub(r"\n{2,}", "\n", page_text)

    patterns = [
        r"(SPECIFICATIONS.*?)(?:No additional information available|Bagikan:|Related Items|Berlangganan Newsletter|Produk Terkait)",
        r"(SPESIFIKASI.*?)(?:No additional information available|Bagikan:|Related Items|Berlangganan Newsletter|Produk Terkait)",
        r"(Informasi\s*&\s*Spesifikasi.*?)(?:No additional information available|Bagikan:|Related Items|Berlangganan Newsletter|Produk Terkait)",
        r"(Deskripsi\s+Deskripsi.*?)(?:Related Items|Berlangganan Newsletter|Produk Terkait)",
    ]

    spec_text = ""

    for pattern in patterns:
        match = re.search(pattern, page_text, re.I | re.S)

        if match:
            spec_text = match.group(1)
            break

    if not spec_text:
        possible_headers = soup.find_all(string=re.compile(r"SPECIFICATIONS|SPESIFIKASI|Informasi|Deskripsi", re.I))

        for header in possible_headers:
            parent = header.parent
            block = parent.get_text("\n") if parent else ""

            if len(block) < 100 and parent and parent.parent:
                block = parent.parent.get_text("\n")

            if len(block) > len(spec_text):
                spec_text = block

    lines = []

    for line in spec_text.splitlines():
        line = posting_clean_text(line)

        if not line:
            continue

        low = line.lower()

        skip_lines = [
            "deskripsi deskripsi",
            "informasi & spesifikasi info",
            "specifications",
            "spesifikasi",
            "informasi & spesifikasi",
            "info",
        ]

        if low in skip_lines:
            continue

        if "no additional information available" in low:
            continue

        if "bagikan" in low:
            continue

        lines.append(line)

    cleaned_lines = []
    seen = set()

    for line in lines:
        if line in seen:
            continue

        seen.add(line)
        cleaned_lines.append(line)

    raw_spec = "\n".join(cleaned_lines[:100])
    formatted_spec = posting_format_specification_text(raw_spec)
    return formatted_spec or raw_spec


POSTING_SPEC_LABELS = [
    "Brand",
    # Metadata AGRES yang sering ikut kebaca di awal halaman. Dipakai untuk split, lalu dilewati saat output.
    "Kategori",
    "Category",
    "SKU",
    "Subtotal",
    "Harga",
    "Price",
    "Stok",
    "Stock",
    "Tipe",
    "Type",
    "Model",
    "Color",
    "Operating System",
    "Processor",
    "Graphics",
    "VGA",
    "RAM",
    "Memory",
    "Storage",
    "SSD",
    "HDD",
    "Main Display Size",
    "Main Display Type",
    "Display Size",
    "Display Type",
    "Display Resolution",
    "Resolution",
    "Camera",
    "Audio",
    "Battery",
    "Charger",
    "Connectivity",
    "Ports",
    "Dimension",
    "Dimensions",
    "Weight",
    "Warranty",
    "Garansi",
]


def posting_parse_specification_pairs(spec_text: str) -> List[Tuple[str, str]]:
    """Ubah spesifikasi AGRES.ID dari paragraf panjang menjadi pasangan label-value.

    AGRES kadang merender spesifikasi sebagai teks panjang seperti:
    Brand: Asus Tipe: ... Processor: ...
    Fungsi ini mengembalikan rows supaya preview Codex rapi seperti tabel web.
    """
    text = posting_clean_text(spec_text)
    if not text:
        return []

    labels_sorted = sorted(POSTING_SPEC_LABELS, key=len, reverse=True)
    label_pattern = "|".join(re.escape(label) for label in labels_sorted)
    pattern = re.compile(rf"(?<![A-Za-z0-9])({label_pattern})\s*:", flags=re.I)
    matches = list(pattern.finditer(text))

    if not matches:
        pairs = []
        for line in str(spec_text or "").splitlines():
            line = posting_clean_text(line)
            if not line:
                continue
            if ":" in line:
                label, value = line.split(":", 1)
                pairs.append((posting_clean_text(label), posting_clean_text(value)))
        return [(k, v) for k, v in pairs if k and v]

    pairs: List[Tuple[str, str]] = []
    seen_keys: Set[str] = set()

    for idx, match in enumerate(matches):
        label = posting_clean_text(match.group(1))
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        value = posting_clean_text(text[start:end])

        if not label or not value:
            continue

        # Rapikan kapital label sesuai daftar label yang kita kenal.
        canonical = next((known for known in POSTING_SPEC_LABELS if known.lower() == label.lower()), label)
        dedupe_key = canonical.lower()
        if dedupe_key in seen_keys:
            continue

        seen_keys.add(dedupe_key)
        pairs.append((canonical, value))

    return pairs


POSTING_SPEC_GROUPS = [
    {"Brand", "Tipe", "Type", "Model", "Color"},
    {"Operating System"},
    {"Processor", "Graphics", "VGA"},
    {"RAM", "Memory", "Storage", "SSD", "HDD"},
    {"Main Display Size", "Main Display Type", "Display Size", "Display Type", "Display Resolution", "Resolution"},
    {"Camera", "Audio"},
    {"Battery", "Charger"},
    {"Connectivity", "Ports"},
    {"Dimension", "Dimensions", "Weight", "Warranty", "Garansi"},
]


def posting_spec_group_index(label: str) -> int:
    clean_label = posting_clean_text(label)
    for idx, group in enumerate(POSTING_SPEC_GROUPS):
        if clean_label in group:
            return idx
    return len(POSTING_SPEC_GROUPS)


def posting_format_ports_for_copy(value: str) -> str:
    """Bikin field Ports lebih enak dicopy seperti tampilan AGRES.ID."""
    value = posting_clean_text(value)
    if not value:
        return ""

    # Ubah "2x USB-A 2x Thunderbolt 1x HDMI" menjadi baris terpisah.
    parts = re.split(r"\s+(?=\d+\s*x\s+)", value, flags=re.I)
    parts = [posting_clean_text(part) for part in parts if posting_clean_text(part)]
    if len(parts) >= 2:
        return "Ports:\n" + "\n".join(parts)

    # Fallback kalau AGRES merender dengan koma.
    if "," in value:
        comma_parts = [posting_clean_text(part) for part in value.split(",") if posting_clean_text(part)]
        if len(comma_parts) >= 2:
            return "Ports:\n" + "\n".join(comma_parts)

    return f"Ports: {value}"


def posting_format_specification_text(spec_text: str) -> str:
    """Format spesifikasi menjadi teks siap copy, mirip layout AGRES.ID."""
    pairs = posting_parse_specification_pairs(spec_text)
    if not pairs:
        raw = str(spec_text or "").strip()
        raw = re.sub(r"\n{3,}", "\n\n", raw)
        return raw or posting_clean_text(spec_text)

    output_lines: List[str] = ["SPESIFIKASI", ""]
    previous_group = None

    skip_labels = {"kategori", "category", "sku", "subtotal", "harga", "price", "stok", "stock"}

    for label, value in pairs:
        label = posting_clean_text(label)
        value = posting_clean_text(value)
        if not label or not value:
            continue
        if label.lower() in skip_labels:
            continue

        current_group = posting_spec_group_index(label)
        if previous_group is not None and current_group != previous_group:
            if output_lines and output_lines[-1] != "":
                output_lines.append("")

        if label == "Ports":
            output_lines.append(posting_format_ports_for_copy(value))
        else:
            output_lines.append(f"{label}: {value}")

        previous_group = current_group

    formatted = "\n".join(output_lines).strip()
    formatted = re.sub(r"\n{3,}", "\n\n", formatted)
    return formatted


def posting_render_specification_table(spec_text: str):
    copy_text = posting_format_specification_text(spec_text)
    if not copy_text:
        st.warning("Deskripsi/spesifikasi belum terbaca dari halaman AGRES.ID untuk SKU ini.")
        return

    st.text_area(
        "Spesifikasi siap copy",
        value=copy_text,
        height=430,
        key=f"posting_spec_copy_single_{abs(hash(copy_text + str(time.time_ns()))) % 100000000}",
    )


def posting_extract_brand_category_sku(soup):
    text = soup.get_text("\n")
    result = {"brand": "", "kategori": "", "sku_web": ""}
    patterns = {
        "brand": r"Brand\s*:\s*([^\n]+)",
        "kategori": r"Kategori\s*:\s*([^\n]+)",
        "sku_web": r"SKU\s*:\s*([^\n]+)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, text, flags=re.I)
        if m:
            result[key] = posting_clean_text(m.group(1))
    return result


def posting_scrape_product(url):
    html = posting_get_html(url)
    soup = BeautifulSoup(html, "html.parser")
    images = posting_extract_images(soup, html, limit=5)
    extra = posting_extract_brand_category_sku(soup)

    return {
        "title": posting_extract_title(soup),
        "image_1": images[0],
        "image_2": images[1],
        "image_3": images[2],
        "image_4": images[3],
        "image_5": images[4],
        "spesifikasi_web": posting_extract_specification(soup),
        "brand": extra.get("brand", ""),
        "kategori": extra.get("kategori", ""),
        "sku_web": extra.get("sku_web", ""),
    }


POSTING_SHOPEE_OUTPUT_COLUMNS = [
    "KODEBARANG",
    "JUDUL",
    "DESKRIPSI",
    "direct image URL_1",
    "direct image URL_2",
    "direct image URL_3",
    "direct image URL_4",
    "direct image URL_5",
    "STOK",
]

POSTING_SHOPEE_TEMPLATE_HEADER_ROW = 3
POSTING_SHOPEE_TEMPLATE_DATA_START_ROW = 7
POSTING_SHOPEE_LAPTOP_ADDON_ANTIGORES = 200_000
POSTING_SHOPEE_LAPTOP_ADDON_ACCESSORIES = 200_000
POSTING_SHOPEE_LAPTOP_ADDON_CAMPAIGN = 5_000_000


def posting_shopee_found_only_df(df: pd.DataFrame) -> pd.DataFrame:
    """Hasil yang masuk ke Excel hanya SKU yang sukses FOUND."""
    if df is None or df.empty:
        return pd.DataFrame()
    if "STATUS" not in df.columns:
        return df.copy()
    return df[df["STATUS"].astype(str).str.upper() == "FOUND"].copy()


def posting_to_shopee_output_df(df):
    """Output ringkas hasil scan AGRES.ID. NOT_FOUND/ERROR tidak dimasukkan ke Excel."""
    df = posting_shopee_found_only_df(df)
    if df is None or df.empty:
        return pd.DataFrame(columns=POSTING_SHOPEE_OUTPUT_COLUMNS)

    out = pd.DataFrame()
    out["KODEBARANG"] = df.get("KODEBARANG", pd.Series(dtype=object)).fillna("").astype(str)
    out["JUDUL"] = df.get("PRODUCT", pd.Series(dtype=object)).fillna("").astype(str)

    # DESKRIPSI hanya spesifikasi siap copy. Tidak ikut Harga/Price/Subtotal.
    desc_src = df.get("SPESIFIKASI_WEB", pd.Series(dtype=object)).fillna("").astype(str)
    if "SPESIFIKASI_INPUT" in df.columns:
        desc_src = desc_src.where(desc_src.str.strip() != "", df["SPESIFIKASI_INPUT"].fillna("").astype(str))
    out["DESKRIPSI"] = desc_src.apply(posting_format_specification_text)

    image_map = {
        "direct image URL_1": "IMAGE_1",
        "direct image URL_2": "IMAGE_2",
        "direct image URL_3": "IMAGE_3",
        "direct image URL_4": "IMAGE_4",
        "direct image URL_5": "IMAGE_5",
    }
    for out_col, src_col in image_map.items():
        out[out_col] = df.get(src_col, pd.Series(dtype=object)).fillna("").astype(str)

    if "STOK" in df.columns:
        out["STOK"] = df["STOK"]
    else:
        out["STOK"] = ""

    return out[POSTING_SHOPEE_OUTPUT_COLUMNS]


def create_posting_shopee_excel_download(df):
    """Download hasil scan mentah/ringkas, bukan template mass upload Shopee."""
    output_df = posting_to_shopee_output_df(df)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="HASIL")

        ws = writer.book["HASIL"]
        ws.freeze_panes = "A2"

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        widths = {
            "A": 28,
            "B": 42,
            "C": 70,
            "D": 34,
            "E": 34,
            "F": 34,
            "G": 34,
            "H": 34,
            "I": 12,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.sheet_format.defaultRowHeight = 15
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 15

        try:
            from openpyxl.styles import Font, PatternFill, Border, Side
            header_fill = PatternFill("solid", fgColor="9DC3E6")
            thin = Side(style="thin", color="7F7F7F")
            header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(bold=False)
                cell.border = header_border
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 15
        except Exception:
            pass

    output.seek(0)
    return output


def posting_shopee_is_laptop_sku(kodebarang: str) -> bool:
    return "-LAP-" in su(kodebarang)


def posting_shopee_price_to_rupiah(value) -> Optional[int]:
    """Pricelist Posting menyimpan harga ribuan, jadi 10.000 menjadi 10.000.000."""
    price = parse_price_cell(value)
    if price is None:
        return None
    return int(price) * 1000


def posting_shopee_generated_integration_code(kodebarang: str) -> str:
    base = re.sub(r"[^A-Z0-9_\-]", "", su(kodebarang))
    if not base:
        base = "VARIASI"
    return f"COD-{base}"[:100]


def posting_shopee_product_title(row: pd.Series) -> str:
    title = posting_clean_text(row.get("PRODUCT", ""))
    if title:
        return title
    return posting_clean_text(row.get("KODEBARANG", ""))


def posting_shopee_product_description(row: pd.Series) -> str:
    spec_text = posting_clean_text(row.get("SPESIFIKASI_WEB", ""))
    if not spec_text:
        spec_text = posting_clean_text(row.get("SPESIFIKASI_INPUT", ""))
    return posting_format_specification_text(spec_text)


def posting_shopee_image_values(row: pd.Series, total_slots: int = 9) -> List[str]:
    images = []
    for col in ["IMAGE_1", "IMAGE_2", "IMAGE_3", "IMAGE_4", "IMAGE_5", "IMAGE_6", "IMAGE_7", "IMAGE_8", "IMAGE_9"]:
        val = posting_clean_text(row.get(col, ""))
        if val:
            images.append(val)
    while len(images) < total_slots:
        images.append("")
    return images[:total_slots]


def posting_shopee_detect_weight(row: pd.Series) -> str:
    kodebarang = posting_clean_text(row.get("KODEBARANG", ""))
    title = posting_clean_text(row.get("PRODUCT", ""))
    desc = posting_clean_text(row.get("SPESIFIKASI_WEB", "")) or posting_clean_text(row.get("SPESIFIKASI_INPUT", ""))
    kategori = posting_clean_text(row.get("KATEGORI_WEB", ""))
    text = f"{kodebarang} {title} {desc} {kategori}".lower()

    if posting_shopee_is_laptop_sku(kodebarang):
        # 15.6 / 15 inch dianggap laptop 15 inch.
        if re.search(r"(?<!\d)15(?:[\.,]6|[\.,]0)?\s*(?:inch|inches|inci|in\b|\"|”)", text):
            return "8000"
        if re.search(r"(?<!\d)14(?:[\.,]0)?\s*(?:inch|inches|inci|in\b|\"|”)", text):
            return "4000"
        # Default laptop kalau ukuran tidak kebaca dari hasil scan.
        return "4000"

    if any(word in text for word in ["phone", "smartphone", "handphone", "mobile phone", "ponsel", "telepon seluler"]):
        return "1000"

    return ""


def posting_shopee_variant_rows(row: pd.Series) -> List[Dict[str, Any]]:
    kodebarang = posting_clean_text(row.get("KODEBARANG", ""))
    stock = row.get("STOK", "")
    base_price = posting_shopee_price_to_rupiah(row.get("HARGA", ""))
    price_normal = "" if base_price is None else int(base_price)
    campaign_price = "" if base_price is None else int(base_price) + POSTING_SHOPEE_LAPTOP_ADDON_CAMPAIGN
    integration = posting_shopee_generated_integration_code(kodebarang)

    # Semua produk dibuat variasi supaya siap mass upload Shopee.
    # - SKU laptop (-LAP-) punya 4 variasi: Normal, +Antigores, +Accessories, Campaign.
    # - Produk lain punya 2 variasi: Normal dan Campaign.
    variants = [
        {
            "kode_variasi": kodebarang,
            "kode_integrasi": integration,
            "nama_variasi_1": "Pilihan",
            "varian_1": "Normal",
            "harga": price_normal,
            "stok": stock,
        }
    ]

    if posting_shopee_is_laptop_sku(kodebarang):
        variants.extend([
            {
                "kode_variasi": f"{kodebarang}+PC",
                "kode_integrasi": integration,
                "nama_variasi_1": "Pilihan",
                "varian_1": "+Antigores",
                "harga": "" if base_price is None else int(base_price) + POSTING_SHOPEE_LAPTOP_ADDON_ANTIGORES,
                "stok": stock,
            },
            {
                "kode_variasi": f"{kodebarang}+BA",
                "kode_integrasi": integration,
                "nama_variasi_1": "Pilihan",
                "varian_1": "+Accessories",
                "harga": "" if base_price is None else int(base_price) + POSTING_SHOPEE_LAPTOP_ADDON_ACCESSORIES,
                "stok": stock,
            },
        ])

    variants.append({
        "kode_variasi": "ND-ALL-CAMPAIGN",
        "kode_integrasi": integration,
        "nama_variasi_1": "Pilihan",
        "varian_1": "Campaign",
        "harga": campaign_price,
        "stok": stock,
    })

    return variants


def posting_find_template_header_columns(ws: Worksheet, header_row: int = POSTING_SHOPEE_TEMPLATE_HEADER_ROW) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        header = s_clean(ws.cell(row=header_row, column=c).value)
        if not header:
            continue
        header_map[header.lower()] = c

    wanted = [
        "Nama Produk",
        "Deskripsi Produk",
        "SKU Induk",
        "Kode Integrasi Variasi",
        "Nama Variasi 1",
        "Varian untuk Variasi 1",
        "Harga",
        "Stok",
        "Kode Variasi",
        "Foto Sampul",
        "Foto Produk 1",
        "Foto Produk 2",
        "Foto Produk 3",
        "Foto Produk 4",
        "Foto Produk 5",
        "Foto Produk 6",
        "Foto Produk 7",
        "Foto Produk 8",
        "Berat",
    ]

    found: Dict[str, int] = {}
    for name in wanted:
        col = header_map.get(name.lower())
        if col:
            found[name] = col

    required = [
        "Nama Produk",
        "Deskripsi Produk",
        "Kode Integrasi Variasi",
        "Nama Variasi 1",
        "Varian untuk Variasi 1",
        "Harga",
        "Stok",
        "Kode Variasi",
        "Foto Sampul",
        "Berat",
    ]
    missing = [name for name in required if name not in found]
    if missing:
        raise ValueError(
            "Header template Shopee tidak lengkap di row 3. Kolom tidak ditemukan: "
            + ", ".join(missing)
        )

    return found


def posting_clear_template_data_rows(ws: Worksheet, data_start: int = POSTING_SHOPEE_TEMPLATE_DATA_START_ROW):
    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)


def posting_count_shopee_template_rows(df: pd.DataFrame) -> int:
    df = posting_shopee_found_only_df(df)
    if df is None or df.empty:
        return 0
    total = 0
    for _, row in df.iterrows():
        total += len(posting_shopee_variant_rows(row))
    return total


def create_posting_shopee_template_download(df: pd.DataFrame, template_bytes: bytes):
    """Isi file template mass upload Shopee dari hasil FOUND saja."""
    df = posting_shopee_found_only_df(df)
    if df is None or df.empty:
        raise ValueError("Belum ada SKU FOUND untuk dibuat template Shopee.")
    if not template_bytes:
        raise ValueError("Template Shopee belum diupload.")

    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb["Template"] if "Template" in wb.sheetnames else wb.active
    cols = posting_find_template_header_columns(ws, POSTING_SHOPEE_TEMPLATE_HEADER_ROW)
    posting_clear_template_data_rows(ws, POSTING_SHOPEE_TEMPLATE_DATA_START_ROW)
    ws.sheet_format.defaultRowHeight = 15

    row_no = POSTING_SHOPEE_TEMPLATE_DATA_START_ROW
    for _, item in df.iterrows():
        kodebarang = posting_clean_text(item.get("KODEBARANG", ""))
        title = posting_shopee_product_title(item)
        description = posting_shopee_product_description(item)
        images = posting_shopee_image_values(item, total_slots=9)
        weight = posting_shopee_detect_weight(item)
        variants = posting_shopee_variant_rows(item)

        for variant_idx, variant in enumerate(variants):
            is_first_variant = variant_idx == 0

            # Kolom produk utama mengikuti panduan template: untuk variasi, isi di baris variasi pertama saja.
            if is_first_variant:
                safe_set_cell_value(ws, row_no, cols["Nama Produk"], title)
                safe_set_cell_value(ws, row_no, cols["Deskripsi Produk"], description)
                safe_set_cell_value(ws, row_no, cols["Foto Sampul"], images[0])
                for img_idx in range(1, 9):
                    col_name = f"Foto Produk {img_idx}"
                    if col_name in cols:
                        safe_set_cell_value(ws, row_no, cols[col_name], images[img_idx])
            else:
                safe_set_cell_value(ws, row_no, cols["Nama Produk"], "")
                safe_set_cell_value(ws, row_no, cols["Deskripsi Produk"], "")

            if "SKU Induk" in cols:
                safe_set_cell_value(ws, row_no, cols["SKU Induk"], "")

            safe_set_cell_value(ws, row_no, cols["Kode Integrasi Variasi"], variant["kode_integrasi"])
            safe_set_cell_value(ws, row_no, cols["Nama Variasi 1"], variant["nama_variasi_1"])
            safe_set_cell_value(ws, row_no, cols["Varian untuk Variasi 1"], variant["varian_1"])
            safe_set_cell_value(ws, row_no, cols["Harga"], variant["harga"])
            safe_set_cell_value(ws, row_no, cols["Stok"], variant["stok"])
            safe_set_cell_value(ws, row_no, cols["Kode Variasi"], variant["kode_variasi"])
            safe_set_cell_value(ws, row_no, cols["Berat"], weight)

            row_no += 1

    for r in range(1, max(POSTING_SHOPEE_TEMPLATE_DATA_START_ROW, row_no)):
        ws.row_dimensions[r].height = 15

    for r in range(POSTING_SHOPEE_TEMPLATE_DATA_START_ROW, max(POSTING_SHOPEE_TEMPLATE_DATA_START_ROW, row_no)):
        for c in cols.values():
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


POSTING_SHOPEE_BATCH_DEFAULT = 100
POSTING_SHOPEE_BATCH_MIN = 20
POSTING_SHOPEE_BATCH_MAX = 300
POSTING_SHOPEE_BATCH_DELAY_DEFAULT = 1.0


POSTING_SHOPEE_RESULT_COLUMNS = [
    "KODEBARANG",
    "SPESIFIKASI_INPUT",
    "HARGA",
    "PRODUCT",
    "IMAGE_1",
    "IMAGE_2",
    "IMAGE_3",
    "IMAGE_4",
    "IMAGE_5",
    "SPESIFIKASI_WEB",
    "STOK",
    "STATUS",
    "ERROR_MESSAGE",
    "AGRES_URL",
    "MATCH_SCORE",
    "MATCH_SOURCE",
    "BRAND_WEB",
    "KATEGORI_WEB",
    "SKU_WEB",
]


def posting_shopee_init_batch_state():
    defaults = {
        "posting_shopee_results": [],
        "posting_shopee_next_index": 0,
        "posting_shopee_running": False,
        "posting_shopee_done": False,
        "posting_shopee_upload_signature": "",
        "posting_shopee_last_message": "",
        "posting_shopee_batch_size": POSTING_SHOPEE_BATCH_DEFAULT,
        "posting_shopee_auto_continue": True,
        "posting_shopee_batch_delay": POSTING_SHOPEE_BATCH_DELAY_DEFAULT,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def posting_shopee_reset_batch_state(upload_signature: Optional[str] = None):
    st.session_state.posting_shopee_results = []
    st.session_state.posting_shopee_next_index = 0
    st.session_state.posting_shopee_running = False
    st.session_state.posting_shopee_done = False
    st.session_state.posting_shopee_last_message = ""
    if upload_signature is not None:
        st.session_state.posting_shopee_upload_signature = upload_signature


def posting_shopee_safe_rerun():
    if hasattr(st, "rerun"):
        st.rerun()
    else:
        st.experimental_rerun()


def posting_shopee_base_result(kodebarang: str, stok_input, harga_input="", spesifikasi_input: str = "") -> Dict[str, Any]:
    return {
        "KODEBARANG": kodebarang,
        "SPESIFIKASI_INPUT": spesifikasi_input,
        "HARGA": harga_input,
        "PRODUCT": "",
        "IMAGE_1": "",
        "IMAGE_2": "",
        "IMAGE_3": "",
        "IMAGE_4": "",
        "IMAGE_5": "",
        "SPESIFIKASI_WEB": "",
        "STOK": stok_input,
        "STATUS": "",
        "ERROR_MESSAGE": "",
        "AGRES_URL": "",
        "MATCH_SCORE": "",
        "MATCH_SOURCE": "",
        "BRAND_WEB": "",
        "KATEGORI_WEB": "",
        "SKU_WEB": "",
    }


def process_posting_shopee_batch(df, start_index: int = 0, batch_size: int = POSTING_SHOPEE_BATCH_DEFAULT) -> List[Dict[str, Any]]:
    """Proses sebagian SKU saja. Hasilnya disimpan oleh render_posting_shopee ke session_state."""
    results: List[Dict[str, Any]] = []
    total = len(df)

    if total <= 0:
        return results

    start_index = max(0, min(int(start_index or 0), total))
    batch_size = max(1, int(batch_size or POSTING_SHOPEE_BATCH_DEFAULT))
    end_index = min(start_index + batch_size, total)

    progress = st.progress(start_index / max(total, 1))
    status_box = st.empty()
    log_box = st.empty()
    logs: List[str] = []

    for pos in range(start_index, end_index):
        row = df.iloc[pos]
        no = pos + 1
        kodebarang = posting_clean_text(row.get("KODEBARANG", ""))
        spesifikasi_input = posting_clean_text(row.get("SPESIFIKASI", ""))
        harga_input = row.get("HARGA", "")
        stok_input = row.get("STOK", "")

        status_box.write(f"Memproses {no}/{total}: **{kodebarang}**")
        base_result = posting_shopee_base_result(kodebarang, stok_input, harga_input, spesifikasi_input)

        try:
            best = posting_find_best_match(kodebarang, spesifikasi_input)

            if not best or best.get("not_found_error"):
                item = dict(base_result)
                item.update({
                    "STATUS": "NOT_FOUND",
                    "ERROR_MESSAGE": best.get("not_found_error", "Produk tidak ditemukan di AGRES.ID") if isinstance(best, dict) else "Produk tidak ditemukan di AGRES.ID",
                })
                results.append(item)

                logs.append(f"❌ {kodebarang} - NOT_FOUND")
                progress.progress(no / max(total, 1))
                log_box.text("\n".join(logs[-12:]))
                continue

            product = posting_scrape_product(best["url"])

            item = dict(base_result)
            item.update({
                "PRODUCT": product.get("title", "") or best.get("title", ""),
                "IMAGE_1": product.get("image_1", ""),
                "IMAGE_2": product.get("image_2", ""),
                "IMAGE_3": product.get("image_3", ""),
                "IMAGE_4": product.get("image_4", ""),
                "IMAGE_5": product.get("image_5", ""),
                "SPESIFIKASI_WEB": product.get("spesifikasi_web", ""),
                "STATUS": "FOUND",
                "ERROR_MESSAGE": "",
                "AGRES_URL": best["url"],
                "MATCH_SCORE": best.get("score", ""),
                "MATCH_SOURCE": best.get("source", ""),
                "BRAND_WEB": product.get("brand", ""),
                "KATEGORI_WEB": product.get("kategori", ""),
                "SKU_WEB": product.get("sku_web", ""),
            })
            results.append(item)
            logs.append(f"✅ {kodebarang} - FOUND dari AGRES.ID")

        except Exception as error:
            item = dict(base_result)
            item.update({
                "STATUS": "ERROR",
                "ERROR_MESSAGE": str(error),
            })
            results.append(item)
            logs.append(f"⚠️ {kodebarang} - ERROR: {error}")

        progress.progress(no / max(total, 1))
        log_box.text("\n".join(logs[-12:]))

    status_box.write(f"Batch selesai: SKU {start_index + 1} sampai {end_index} dari {total}.")
    return results


def process_posting_shopee_bulk(df):
    """Kompatibel dengan pemanggilan lama, tapi render utama sekarang memakai batch otomatis."""
    all_results: List[Dict[str, Any]] = []
    total = len(df)
    next_index = 0

    while next_index < total:
        batch_results = process_posting_shopee_batch(
            df,
            start_index=next_index,
            batch_size=POSTING_SHOPEE_BATCH_DEFAULT,
        )
        if not batch_results:
            break
        all_results.extend(batch_results)
        next_index += len(batch_results)

    return pd.DataFrame(all_results, columns=POSTING_SHOPEE_RESULT_COLUMNS)

def render_posting_shopee_result_preview(result_df):
    """Preview tetap enak dilihat, tapi dibatasi supaya bulk tidak jebol."""
    if result_df is None or result_df.empty:
        return

    st.subheader("Preview Hasil AGRES.ID")

    total_rows = len(result_df)
    preview_limit = 10
    detail_limit = 3

    if total_rows > preview_limit:
        st.info(
            f"Preview tabel hanya {preview_limit} baris pertama dari {total_rows} SKU supaya Codex tidak berat. "
            "File XLSX tetap berisi semua hasil."
        )
    else:
        st.caption("Preview hasil dari AGRES.ID. File download tetap XLSX dengan format posting Shopee.")

    display_cols = [
        "KODEBARANG",
        "PRODUCT",
        "IMAGE_1",
        "IMAGE_2",
        "IMAGE_3",
        "IMAGE_4",
        "IMAGE_5",
        "STATUS",
        "AGRES_URL",
        "MATCH_SCORE",
        "ERROR_MESSAGE",
    ]
    display_cols = [col for col in display_cols if col in result_df.columns]

    column_config = {}
    for img_col in ["IMAGE_1", "IMAGE_2", "IMAGE_3", "IMAGE_4", "IMAGE_5"]:
        if img_col in display_cols:
            try:
                column_config[img_col] = st.column_config.ImageColumn(img_col, width="small")
            except Exception:
                pass

    if "AGRES_URL" in display_cols:
        try:
            column_config["AGRES_URL"] = st.column_config.LinkColumn("AGRES_URL", display_text="Buka AGRES")
        except Exception:
            pass

    if display_cols:
        st.dataframe(
            result_df[display_cols].head(preview_limit),
            use_container_width=True,
            hide_index=True,
            column_config=column_config,
            height=360 if total_rows > 4 else None,
        )

    found_df = result_df[result_df.get("STATUS", "") == "FOUND"] if "STATUS" in result_df.columns else result_df
    if found_df.empty:
        return

    if total_rows > detail_limit:
        st.caption(
            f"Detail gambar & deskripsi hanya dirender {detail_limit} SKU pertama supaya bulk tidak berat. "
            "Semua data lengkap tetap ada di file XLSX."
        )

    detail_rows = found_df.head(detail_limit)
    expanded = total_rows <= detail_limit
    with st.expander("Lihat detail gambar & deskripsi siap copy", expanded=expanded):
        for row_no, (idx, row) in enumerate(detail_rows.iterrows(), start=1):
            kode = posting_clean_text(row.get("KODEBARANG", ""))
            product_name = posting_clean_text(row.get("PRODUCT", ""))
            unique_base = re.sub(r"[^a-zA-Z0-9_]", "_", kode or str(idx))[:40]

            st.markdown(f"**{kode}**")
            if product_name:
                st.write(product_name)

            image_urls = [posting_clean_text(row.get(col, "")) for col in ["IMAGE_1", "IMAGE_2", "IMAGE_3", "IMAGE_4", "IMAGE_5"]]
            image_urls = [url for url in image_urls if url and not re.search(r"\.svg(\?|$)", url.lower())]

            if image_urls:
                cols = st.columns(min(5, len(image_urls)))
                for i, image_url in enumerate(image_urls[:5]):
                    with cols[i % len(cols)]:
                        try:
                            st.image(image_url, use_container_width=True, caption=f"Image {i + 1}")
                        except Exception:
                            st.write(image_url)

                st.caption("Direct image URL siap copy")
                st.code("\n".join(image_urls), language="text")
            else:
                st.warning("Gambar belum terbaca dari halaman AGRES.ID untuk SKU ini.")

            spec_text = posting_clean_text(row.get("SPESIFIKASI_WEB", ""))
            if spec_text:
                st.markdown("**Deskripsi / Spesifikasi AGRES.ID siap copy**")
                formatted_spec = posting_format_specification_text(spec_text)
                st.text_area(
                    "Deskripsi siap copy",
                    value=formatted_spec,
                    height=430,
                    key=f"posting_spec_copy_{row_no}_{unique_base}_{abs(hash(formatted_spec)) % 100000000}",
                )
            else:
                st.warning("Deskripsi/spesifikasi belum terbaca dari halaman AGRES.ID untuk SKU ini.")

            agres_url = posting_clean_text(row.get("AGRES_URL", ""))
            if agres_url:
                st.caption(agres_url)
            st.markdown("---")

def render_posting_shopee():
    st.title("Posting Shopee")

    try:
        ensure_posting_shopee_dependencies()
    except RuntimeError as error:
        st.error(str(error))
        st.code("requests\nbeautifulsoup4\nrapidfuzz")
        st.stop()

    posting_shopee_init_batch_state()

    st.caption(
        "Generate data posting bulk langsung dari AGRES.ID. "
        "Total SKU boleh besar, proses dibuat per batch supaya lebih tahan timeout."
    )

    st.write("Upload Excel dengan kolom:")
    st.code("KODEBARANG | SPESIFIKASI | HARGA | STOK")

    uploaded = st.file_uploader("Upload Pricelist", type=["xlsx", "xls"], key="posting_shopee_upload_pricelist")
    template_uploaded = st.file_uploader(
        "Upload Template Mass Upload Shopee",
        type=["xlsx", "xlsm"],
        key="posting_shopee_upload_template",
        help="Pakai template Shopee dengan header di row 3. Sheet 'Template' akan diisi mulai row 7.",
    )

    if uploaded:
        template_signature = f"{getattr(template_uploaded, 'name', '')}:{getattr(template_uploaded, 'size', '')}" if template_uploaded else "no-template"
        upload_signature = f"{getattr(uploaded, 'name', '')}:{getattr(uploaded, 'size', '')}:{template_signature}"
        if st.session_state.posting_shopee_upload_signature != upload_signature:
            posting_shopee_reset_batch_state(upload_signature)

        try:
            uploaded.seek(0)
        except Exception:
            pass

        df = pd.read_excel(uploaded)

        required = ["KODEBARANG", "SPESIFIKASI", "HARGA", "STOK"]
        missing = [col for col in required if col not in df.columns]

        if missing:
            st.error(f"Kolom tidak ditemukan: {missing}")
            st.stop()

        total_sku = len(df)
        if total_sku <= 0:
            st.warning("File tidak memiliki baris SKU untuk diproses.")
            st.stop()

        st.info(f"Total SKU terbaca: {total_sku}")

        with st.expander("Pengaturan Batch", expanded=False):
            st.number_input(
                "SKU per batch (bukan total SKU)",
                min_value=POSTING_SHOPEE_BATCH_MIN,
                max_value=POSTING_SHOPEE_BATCH_MAX,
                value=int(st.session_state.posting_shopee_batch_size),
                step=10,
                key="posting_shopee_batch_size",
                disabled=bool(st.session_state.posting_shopee_running),
                help="Total SKU boleh ribuan. Yang dibatasi adalah jumlah SKU yang diproses sekali jalan.",
            )
            st.checkbox(
                "Lanjut batch otomatis",
                value=bool(st.session_state.posting_shopee_auto_continue),
                key="posting_shopee_auto_continue",
            )
            st.number_input(
                "Jeda antar batch otomatis (detik)",
                min_value=0.0,
                max_value=10.0,
                value=float(st.session_state.posting_shopee_batch_delay),
                step=0.5,
                key="posting_shopee_batch_delay",
                help="Kasih napas sebentar ke AGRES.ID supaya tidak gampang kena limit.",
            )
            st.caption(
                "Rekomendasi aman: 100 SKU per batch. Kalau stabil boleh 150–200. "
                "Total 5000 SKU tetap bisa, prosesnya dicicil otomatis."
            )

        processed_count = min(int(st.session_state.posting_shopee_next_index or 0), total_sku)
        batch_size = int(st.session_state.posting_shopee_batch_size or POSTING_SHOPEE_BATCH_DEFAULT)
        next_start = processed_count + 1 if processed_count < total_sku else total_sku
        next_end = min(processed_count + batch_size, total_sku)

        st.progress(processed_count / max(total_sku, 1))
        st.write(
            f"Progress: {processed_count}/{total_sku} SKU "
            f"({round(processed_count / max(total_sku, 1) * 100, 1)}%)"
        )
        if processed_count < total_sku:
            st.caption(f"Batch berikutnya: SKU {next_start} sampai {next_end}.")

        col_start, col_pause, col_resume, col_reset = st.columns(4)

        with col_start:
            if st.button(
                "Mulai Generate Bulk",
                key="posting_shopee_generate_bulk",
                disabled=bool(st.session_state.posting_shopee_running),
            ):
                posting_shopee_reset_batch_state(upload_signature)
                st.session_state.posting_shopee_running = True
                st.session_state.posting_shopee_done = False
                posting_shopee_safe_rerun()

        with col_pause:
            if st.button(
                "Pause / Stop",
                key="posting_shopee_pause_bulk",
                disabled=not bool(st.session_state.posting_shopee_running),
            ):
                st.session_state.posting_shopee_running = False
                st.session_state.posting_shopee_last_message = "Proses dipause. Klik Lanjutkan untuk meneruskan dari SKU terakhir."
                posting_shopee_safe_rerun()

        with col_resume:
            can_resume = (
                not bool(st.session_state.posting_shopee_running)
                and not bool(st.session_state.posting_shopee_done)
                and int(st.session_state.posting_shopee_next_index or 0) < total_sku
                and bool(st.session_state.posting_shopee_results)
            )
            if st.button("Lanjutkan", key="posting_shopee_resume_bulk", disabled=not can_resume):
                st.session_state.posting_shopee_running = True
                st.session_state.posting_shopee_last_message = ""
                posting_shopee_safe_rerun()

        with col_reset:
            if st.button(
                "Reset Hasil",
                key="posting_shopee_reset_bulk",
                disabled=bool(st.session_state.posting_shopee_running),
            ):
                posting_shopee_reset_batch_state(upload_signature)
                posting_shopee_safe_rerun()

        if st.session_state.posting_shopee_last_message:
            st.info(st.session_state.posting_shopee_last_message)

        if st.session_state.posting_shopee_running:
            start_index = min(int(st.session_state.posting_shopee_next_index or 0), total_sku)

            if start_index >= total_sku:
                st.session_state.posting_shopee_running = False
                st.session_state.posting_shopee_done = True
            else:
                batch_size = int(st.session_state.posting_shopee_batch_size or POSTING_SHOPEE_BATCH_DEFAULT)
                with st.spinner(f"Memproses batch SKU {start_index + 1} sampai {min(start_index + batch_size, total_sku)}..."):
                    batch_results = process_posting_shopee_batch(
                        df,
                        start_index=start_index,
                        batch_size=batch_size,
                    )

                if not batch_results:
                    st.session_state.posting_shopee_running = False
                    st.session_state.posting_shopee_last_message = "Batch tidak menghasilkan data. Proses dihentikan supaya tidak loop kosong."
                else:
                    st.session_state.posting_shopee_results.extend(batch_results)
                    st.session_state.posting_shopee_next_index = start_index + len(batch_results)

                    if st.session_state.posting_shopee_next_index >= total_sku:
                        st.session_state.posting_shopee_running = False
                        st.session_state.posting_shopee_done = True
                        st.session_state.posting_shopee_last_message = "Selesai generate semua SKU."
                    else:
                        st.session_state.posting_shopee_last_message = (
                            f"Batch selesai. Sudah proses {st.session_state.posting_shopee_next_index}/{total_sku} SKU."
                        )
                        if st.session_state.posting_shopee_auto_continue:
                            time.sleep(float(st.session_state.posting_shopee_batch_delay or 0))
                            posting_shopee_safe_rerun()

        result_df = pd.DataFrame(st.session_state.posting_shopee_results, columns=POSTING_SHOPEE_RESULT_COLUMNS)

        if not result_df.empty:
            found_count = int((result_df["STATUS"] == "FOUND").sum())
            not_found_count = int((result_df["STATUS"] == "NOT_FOUND").sum())
            error_count = int((result_df["STATUS"] == "ERROR").sum())

            if st.session_state.posting_shopee_done:
                st.success("Selesai generate data.")
            elif not st.session_state.posting_shopee_running:
                st.warning("Proses belum selesai. Bisa download hasil sementara atau klik Lanjutkan.")
            else:
                st.info("Proses sedang berjalan per batch.")

            st.write(f"FOUND: {found_count} | NOT_FOUND: {not_found_count} | ERROR: {error_count}")

            if template_uploaded:
                try:
                    template_rows = posting_count_shopee_template_rows(result_df)
                    template_file = create_posting_shopee_template_download(result_df, template_uploaded.getvalue())
                    st.info(
                        f"Template Shopee siap dibuat: {template_rows} baris upload dari {found_count} SKU FOUND. "
                        "NOT_FOUND/ERROR tidak dimasukkan ke Excel. SKU -LAP- menjadi 4 variasi, selain -LAP- menjadi 2 variasi."
                    )
                    st.download_button(
                        "Download Template Shopee Siap Upload" if st.session_state.posting_shopee_done else "Download Template Shopee Sementara",
                        data=template_file,
                        file_name="hasil_mass_upload_shopee.xlsx" if st.session_state.posting_shopee_done else "hasil_mass_upload_shopee_sementara.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"posting_shopee_download_template_{len(result_df)}_{int(st.session_state.posting_shopee_done)}_{getattr(template_uploaded, 'size', 0)}",
                    )
                except Exception as error:
                    st.error(f"Gagal membuat template Shopee: {error}")
            else:
                st.warning("Upload Template Mass Upload Shopee untuk mendapatkan file siap upload ke Shopee.")

            excel_file = create_posting_shopee_excel_download(result_df)
            st.download_button(
                "Download Hasil Scan AGRES.ID" if st.session_state.posting_shopee_done else "Download Hasil Scan AGRES.ID Sementara",
                data=excel_file,
                file_name="hasil_scan_posting_shopee.xlsx" if st.session_state.posting_shopee_done else "hasil_scan_posting_shopee_sementara.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"posting_shopee_download_hasil_{len(result_df)}_{int(st.session_state.posting_shopee_done)}",
            )

            # Preview Hasil AGRES.ID sengaja tidak dirender supaya UI tetap ringan
            # dan tidak memunculkan error tinggi tabel di Streamlit Cloud.

def render_posting_tiktokshop():
    st.title("Posting TikTokShop")
    st.info("Coming Soon.")


# ============================================================
# ORACLE - VIDEO FINDER (EMBEDDED ENGINE DI DALAM CODEX)
# ============================================================
# Catatan teknis:
# - Ini mengganti bridge eksternal Oracle API. User cukup menjalankan Codex.
# - Chrome extension Oracle tetap boleh di-load manual. Extension lama tetap cocok
#   karena endpoint kompatibel tetap hidup di http://127.0.0.1:8000.
# - Kalau Codex dipasang di Streamlit Cloud, extension browser user TIDAK bisa
#   mengakses 127.0.0.1 milik server cloud. Mode extension lokal paling aman
#   untuk Codex yang jalan di komputer user / server lokal kantor.

ORACLE_EMBEDDED_API_HOST = os.environ.get("ORACLE_EMBEDDED_API_HOST", "127.0.0.1")
ORACLE_EMBEDDED_API_PORT = int(os.environ.get("ORACLE_EMBEDDED_API_PORT", "8000"))
ORACLE_API_BASE_DEFAULT = f"http://{ORACLE_EMBEDDED_API_HOST}:{ORACLE_EMBEDDED_API_PORT}"
ORACLE_EMBEDDED_ROOT = Path(os.environ.get("ORACLE_EMBEDDED_ROOT", "codex_oracle_runtime")).resolve()
ORACLE_OUTPUT_DIR = ORACLE_EMBEDDED_ROOT / "output"
ORACLE_VIDEO_DIR = ORACLE_OUTPUT_DIR / "videos"
ORACLE_IMAGE_DIR = ORACLE_OUTPUT_DIR / "images"
ORACLE_STATE_PATH = ORACLE_EMBEDDED_ROOT / "oracle_state.json"


def _oracle_now_text() -> str:
    import datetime as _dt
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _oracle_utc_iso() -> str:
    import datetime as _dt
    return _dt.datetime.utcnow().isoformat(timespec="seconds") + "Z"


def _oracle_new_state() -> Dict[str, Any]:
    return {
        "next_job_id": 1,
        "next_product_id": 1,
        "jobs": {},
        "products": {},
        "logs": [],
        "queue": {
            "running": False,
            "total": 0,
            "done": 0,
            "last_message": "Queue belum berjalan.",
            "updated_at": None,
            "interval_minutes": 20,
            "post_now": False,
        },
    }


def _oracle_prepare_dirs() -> None:
    ORACLE_VIDEO_DIR.mkdir(parents=True, exist_ok=True)
    ORACLE_IMAGE_DIR.mkdir(parents=True, exist_ok=True)


def _oracle_load_state() -> Dict[str, Any]:
    _oracle_prepare_dirs()
    if not ORACLE_STATE_PATH.exists():
        return _oracle_new_state()
    try:
        import json as _json
        with ORACLE_STATE_PATH.open("r", encoding="utf-8") as fh:
            data = _json.load(fh)
        base = _oracle_new_state()
        base.update(data if isinstance(data, dict) else {})
        base.setdefault("jobs", {})
        base.setdefault("products", {})
        base.setdefault("logs", [])
        base.setdefault("queue", _oracle_new_state()["queue"])
        return base
    except Exception:
        return _oracle_new_state()


def _oracle_save_state(state: Dict[str, Any]) -> None:
    _oracle_prepare_dirs()
    try:
        import json as _json
        tmp = ORACLE_STATE_PATH.with_suffix(".tmp")
        with tmp.open("w", encoding="utf-8") as fh:
            _json.dump(state, fh, ensure_ascii=False, indent=2)
        tmp.replace(ORACLE_STATE_PATH)
    except Exception:
        pass


def _oracle_with_state(mutator):
    import threading as _threading
    lock = getattr(_oracle_with_state, "_lock", None)
    if lock is None:
        lock = _threading.RLock()
        setattr(_oracle_with_state, "_lock", lock)
    with lock:
        state = _oracle_load_state()
        result = mutator(state)
        _oracle_save_state(state)
        return result


def oracle_log(message: str) -> None:
    def _mut(state):
        line = f"{_oracle_now_text()} · {str(message).strip()}"
        state.setdefault("logs", []).append(line)
        del state["logs"][:-80]
    _oracle_with_state(_mut)


def oracle_parse_number(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    txt = str(value).strip().lower().replace("rp", "").replace(" ", "")
    mult = 1
    if any(x in txt for x in ["jt", "juta"]):
        mult = 1_000_000
        txt = re.sub(r"jt|juta", "", txt)
    elif any(x in txt for x in ["rb", "ribu", "k"]):
        mult = 1_000
        txt = re.sub(r"rb|ribu|k", "", txt)
    if "." in txt and "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt and "." not in txt:
        txt = txt.replace(",", ".")
    elif txt.count(".") > 1:
        txt = txt.replace(".", "")
    txt = re.sub(r"[^0-9.]", "", txt)
    try:
        return float(txt) * mult if txt else None
    except Exception:
        return None


def oracle_parse_price_text(text: str) -> Optional[int]:
    matches = re.findall(r"Rp\s?[0-9][0-9.,]*(?:\s?(?:RB|JT|RIBU|JUTA|K))?", text or "", flags=re.I)
    values = [oracle_parse_number(x) for x in matches]
    values = [v for v in values if v is not None and v > 0]
    return int(min(values)) if values else None


def oracle_parse_sold_text(text: str) -> Optional[int]:
    m = re.search(r"([0-9.,]+)\s*(rb|jt|k|ribu|juta)?\+?\s*(terjual|terbeli)", text or "", flags=re.I)
    if not m:
        return None
    n = oracle_parse_number((m.group(1) or "") + (m.group(2) or ""))
    return int(n) if n is not None else None


def oracle_parse_rating_text(text: str) -> Optional[float]:
    raw = text or ""
    # Hindari angka harga seperti Rp87.420 kebaca sebagai rating 4.
    candidates = re.findall(r"(?<![\d.,])([45](?:[,.]\d)?)(?![\d.,])", raw)
    for item in candidates:
        try:
            value = float(item.replace(",", "."))
            if 4.0 <= value <= 5.0:
                return value
        except Exception:
            continue
    return None


def oracle_parse_discount_text(text: str) -> Optional[int]:
    vals = []
    for m in re.finditer(r"(\d{1,2})\s?%", text or ""):
        try:
            vals.append(int(m.group(1)))
        except Exception:
            pass
    return max(vals) if vals else None


def oracle_parse_product_ids(url: str) -> Tuple[str, str]:
    txt = str(url or "")
    m = re.search(r"-i\.(\d+)\.(\d+)", txt)
    if m:
        return m.group(2), m.group(1)
    m = re.search(r"(?:^|[/?&])i\.(\d+)\.(\d+)", txt)
    if m:
        return m.group(2), m.group(1)
    m = re.search(r"/(?:product|item)/(\d+)/(\d+)", txt)
    if m:
        return m.group(2), m.group(1)
    return str(abs(hash(txt)))[:16], "unknown"


def oracle_safe_filename(name: str, suffix: str = ".mp4") -> str:
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", str(name or "video")).strip("_")
    safe = safe[:140] or "video"
    if not safe.lower().endswith((".mp4", ".mov", ".webm", ".m4v", ".m3u8")):
        safe += suffix
    return safe


def oracle_public_video_path(path: Path) -> str:
    return "output/videos/" + path.name


def oracle_build_start_url(job: Dict[str, Any]) -> str:
    import json as _json
    from urllib.parse import quote, urlparse, parse_qsl, urlencode, urlunparse
    seller_type = (job.get("seller_type") or "any").lower()
    sort_by = (job.get("sort_by") or "relevance").lower()

    def add_params(url: str) -> str:
        try:
            parsed = urlparse(url)
            query = dict(parse_qsl(parsed.query, keep_blank_values=True))
            if seller_type == "mall":
                query["fe_filter_options"] = _json.dumps([{"group_name": "SHOP_TYPE", "values": ["OFFICIAL_MALL"]}], separators=(",", ":"))
                query["page"] = "0"
            if sort_by == "sales":
                query["sortBy"] = "sales"
            return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))
        except Exception:
            return url

    if job.get("search_url"):
        return add_params(job["search_url"])
    if job.get("category_url"):
        return add_params(job["category_url"])
    keyword = str(job.get("keyword") or "").strip()
    if keyword:
        pairs: List[Tuple[str, str]] = []
        if seller_type == "mall":
            pairs.append(("fe_filter_options", _json.dumps([{"group_name": "SHOP_TYPE", "values": ["OFFICIAL_MALL"]}], separators=(",", ":"))))
        pairs.append(("keyword", keyword))
        pairs.append(("page", "0"))
        if sort_by == "sales":
            pairs.append(("sortBy", "sales"))
        return "https://shopee.co.id/search?" + "&".join(f"{quote(k, safe='')}={quote(v, safe='')}" for k, v in pairs)
    return "https://shopee.co.id"


def oracle_parse_card(raw: Dict[str, Any], category: str = "") -> Optional[Dict[str, Any]]:
    text = re.sub(r"\s+", " ", str(raw.get("text") or "")).strip()
    if not text:
        return None
    rough_lines = [x.strip() for x in re.split(r"\n|\|", str(raw.get("text") or "")) if x.strip()]
    name = ""
    for line in rough_lines:
        low = line.lower()
        if "rp" in low or "terjual" in low or "rating" in low or low in {"iklan", "ad"}:
            continue
        if len(line) >= 8:
            name = line
            break
    if not name:
        name = rough_lines[0][:220] if rough_lines else text[:220]
    price = oracle_parse_price_text(text)
    sold = raw.get("sold_count")
    try:
        sold = int(float(sold)) if sold not in (None, "") else None
    except Exception:
        sold = None
    if sold is None:
        sold = oracle_parse_sold_text(text)
    rating = oracle_parse_rating_text(text)
    product_id, shop_id = oracle_parse_product_ids(raw.get("url") or "")
    seller_type = str(raw.get("seller_type") or "").strip()[:80]
    return {
        "platform": "Shopee",
        "product_id": product_id,
        "shop_id": shop_id,
        "name": name[:500],
        "url": raw.get("url") or "",
        "image_url": raw.get("image_url") or "",
        "price": price,
        "discount_percent": oracle_parse_discount_text(text),
        "rating": rating,
        "sold_count": sold,
        "category": category or "",
        "is_ad": bool(raw.get("is_ad")),
        "seller_type": seller_type,
        "has_komisixtra": bool(raw.get("has_komisixtra")),
        "has_video": bool(raw.get("has_video") or raw.get("video_url")),
        "video_url": raw.get("video_url") or "",
        "local_video_path": "",
        "status": "Scanned",
        "scraped_at": _oracle_utc_iso(),
    }


def oracle_seller_matches(filter_value: str, seller_value: str, job: Dict[str, Any]) -> bool:
    f = (filter_value or "any").lower()
    s_value = (seller_value or "").lower()
    if f in {"", "any", "semua"}:
        return True
    if f == "mall" and "official_mall" in oracle_build_start_url(job).lower():
        return True
    if f == "mall":
        return "mall" in s_value or "official" in s_value or "authorized" in s_value
    if f == "star_plus":
        return "star_plus" in s_value or "star+" in s_value or "plus" in s_value
    if f == "star":
        return "star" in s_value
    return f in s_value


def oracle_filter_product(job: Dict[str, Any], product: Dict[str, Any]) -> Optional[str]:
    if job.get("require_video") and not product.get("has_video"):
        return "bukan listing video"
    if not job.get("include_ads", True) and product.get("is_ad"):
        return "iklan disembunyikan"
    if job.get("require_komisixtra") and not product.get("has_komisixtra"):
        return "bukan KomisiXtra"
    if job.get("min_rating") is not None and (product.get("rating") is None or float(product.get("rating") or 0) < float(job["min_rating"])):
        return "rating kurang"
    if job.get("min_sold") is not None and (product.get("sold_count") is None or int(product.get("sold_count") or 0) < int(job["min_sold"])):
        return "sold kurang"
    if job.get("min_price") is not None and (product.get("price") is None or float(product.get("price") or 0) < float(job["min_price"])):
        return "harga di bawah minimum"
    if job.get("max_price") is not None and (product.get("price") is None or float(product.get("price") or 0) > float(job["max_price"])):
        return "harga di atas maksimum"
    if not oracle_seller_matches(job.get("seller_type", "any"), product.get("seller_type", ""), job):
        return "tipe penjual tidak cocok"
    return None


def oracle_format_price(value: Any) -> str:
    parsed = parse_price_cell(value)
    if parsed is None:
        return s_clean(value) or "-"
    return "Rp{:,.0f}".format(parsed).replace(",", ".")


def oracle_get_ready_products(limit: int = 100) -> List[Dict[str, Any]]:
    state = _oracle_load_state()
    products = []
    for product in state.get("products", {}).values():
        if not product.get("local_video_path"):
            continue
        if product.get("status") in {"Uploaded HP", "Upload HP Skip Resolusi"}:
            continue
        products.append(product)
    products.sort(key=lambda p: p.get("scraped_at") or "", reverse=True)
    return products[:limit]


def oracle_create_scan_job(form_values: Dict[str, Any]) -> Dict[str, Any]:
    def _fnum(key: str):
        value = form_values.get(key)
        if value in (None, ""):
            return None
        try:
            return float(str(value).replace(",", "."))
        except Exception:
            return None

    def _inum(key: str):
        value = form_values.get(key)
        if value in (None, ""):
            return None
        try:
            return int(float(str(value).replace(",", ".")))
        except Exception:
            return None

    def _mut(state):
        job_id = int(state.get("next_job_id") or 1)
        state["next_job_id"] = job_id + 1
        job = {
            "id": job_id,
            "keyword": s_clean(form_values.get("keyword")),
            "search_url": s_clean(form_values.get("search_url")),
            "category_url": s_clean(form_values.get("category_url")),
            "max_products": int(form_values.get("max_products") or 30),
            "max_pages": int(form_values.get("max_pages") or 1),
            "min_rating": _fnum("min_rating"),
            "min_sold": _inum("min_sold"),
            "min_price": _fnum("min_price"),
            "max_price": _fnum("max_price"),
            "include_ads": bool(form_values.get("include_ads")),
            "require_video": bool(form_values.get("require_video")),
            "download_videos": bool(form_values.get("download_videos")),
            "sort_by": form_values.get("sort_by") if form_values.get("sort_by") in {"relevance", "sales"} else "relevance",
            "seller_type": form_values.get("seller_type") if form_values.get("seller_type") in {"any", "mall", "star", "star_plus"} else "any",
            "require_komisixtra": bool(form_values.get("require_komisixtra")),
            "status": "Waiting Extension",
            "total_found": 0,
            "total_saved": 0,
            "started_at": _oracle_utc_iso(),
            "finished_at": None,
            "error_message": "Job dibuat dari Codex. Extension akan mengambil job otomatis.",
        }
        state.setdefault("jobs", {})[str(job_id)] = job
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · Job #{job_id} dibuat: {job.get('keyword') or job.get('search_url') or job.get('category_url')}")
        del state["logs"][:-80]
        return job
    return _oracle_with_state(_mut)


def oracle_get_active_job() -> Optional[Dict[str, Any]]:
    state = _oracle_load_state()
    jobs = list((state.get("jobs") or {}).values())
    active = [j for j in jobs if j.get("status") in {"Pending", "Waiting Extension", "Running", "Paused"}]
    if not active:
        return None
    active.sort(key=lambda j: int(j.get("id") or 0), reverse=True)
    return active[0]


def oracle_active_payload() -> Dict[str, Any]:
    job = oracle_get_active_job()
    if not job:
        return {"ok": False, "message": "Belum ada scan aktif. Buat job dari Codex > Oracle > Video Finder."}
    return {
        "ok": True,
        "job_id": job["id"],
        "keyword": job.get("keyword"),
        "search_url": job.get("search_url"),
        "category_url": job.get("category_url"),
        "start_url": oracle_build_start_url(job),
        "max_products": job.get("max_products", 30),
        "max_pages": job.get("max_pages", 1),
        "require_video": bool(job.get("require_video")),
        "download_videos": bool(job.get("download_videos")),
        "sort_by": job.get("sort_by", "relevance"),
        "seller_type": job.get("seller_type", "any"),
        "require_komisixtra": bool(job.get("require_komisixtra")),
        "total_saved": job.get("total_saved", 0),
        "status": job.get("status"),
    }


def oracle_ingest_cards(job_id: int, cards: List[Dict[str, Any]], source_url: str = "") -> Dict[str, Any]:
    skipped: Dict[str, int] = {}

    def _skip(reason: str):
        skipped[reason] = skipped.get(reason, 0) + 1

    def _mut(state):
        jobs = state.setdefault("jobs", {})
        products = state.setdefault("products", {})
        job = jobs.get(str(job_id))
        if not job:
            return {"ok": False, "message": "Job tidak ditemukan."}
        job["status"] = "Running"
        found = len(cards or [])
        saved = 0
        out_products: List[Dict[str, Any]] = []
        existing_keys = {(p.get("platform"), p.get("product_id"), p.get("shop_id")): pid for pid, p in products.items()}
        job_product_count = sum(1 for p in products.values() if int(p.get("scan_job_id") or 0) == int(job_id))
        for raw in cards or []:
            parsed = oracle_parse_card(raw, category=job.get("category_url") or source_url or "")
            if not parsed:
                _skip("kartu tidak terbaca")
                continue
            if job_product_count >= int(job.get("max_products") or 30):
                _skip("target produk tercapai")
                continue
            reason = oracle_filter_product(job, parsed)
            if reason:
                _skip(reason)
                continue
            key = (parsed.get("platform"), parsed.get("product_id"), parsed.get("shop_id"))
            product_id = existing_keys.get(key)
            if product_id:
                product = products[str(product_id)]
                product.update({k: v for k, v in parsed.items() if v not in (None, "")})
            else:
                product_id = int(state.get("next_product_id") or 1)
                state["next_product_id"] = product_id + 1
                parsed["id"] = product_id
                parsed["scan_job_id"] = job_id
                products[str(product_id)] = parsed
                existing_keys[key] = product_id
                product = parsed
                saved += 1
                job_product_count += 1
            out_products.append({
                "id": int(product_id),
                "name": product.get("name"),
                "url": product.get("url"),
                "local_video_path": product.get("local_video_path") or "",
            })
        job["total_found"] = int(job.get("total_found") or 0) + found
        job["total_saved"] = sum(1 for p in products.values() if int(p.get("scan_job_id") or 0) == int(job_id))
        if skipped:
            job["error_message"] = "Catatan filter: " + ", ".join(f"{k}: {v}" for k, v in sorted(skipped.items(), key=lambda x: -x[1])[:6])
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · Job #{job_id} ingest: found {found}, saved baru {saved}, total {job['total_saved']}")
        del state["logs"][:-80]
        return {
            "ok": True,
            "found": found,
            "saved": saved,
            "total_saved": job["total_saved"],
            "skipped": skipped,
            "status": job["status"],
            "products": out_products,
        }
    return _oracle_with_state(_mut)


def oracle_finish_job(job_id: int, ok: bool = True, message: str = "") -> Dict[str, Any]:
    def _mut(state):
        job = state.setdefault("jobs", {}).get(str(job_id))
        if not job:
            return {"ok": False, "message": "Job tidak ditemukan."}
        if job.get("status") != "Stopped":
            job["status"] = "Completed" if ok else "Failed"
            job["finished_at"] = _oracle_utc_iso()
            prev = (job.get("error_message") or "").strip()
            final = s_clean(message) or ("Scan otomatis selesai." if ok else "Scan otomatis gagal.")
            if prev and prev not in final and any(x in prev.lower() for x in ["filter:", "catatan"]):
                final += "\n\nDiagnosa: " + prev
            job["error_message"] = final
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · Job #{job_id} selesai: {job.get('status')} · {job.get('error_message') or ''}")
        del state["logs"][:-80]
        return {"ok": True, "status": job.get("status"), "message": job.get("error_message")}
    return _oracle_with_state(_mut)


def oracle_mark_product_video(product_id: int, video_path: str, source_url: str = "", video_url: str = "") -> Dict[str, Any]:
    def _mut(state):
        product = state.setdefault("products", {}).get(str(product_id))
        if not product:
            return {"ok": False, "message": "Produk tidak ditemukan."}
        product["local_video_path"] = video_path
        product["video_url"] = video_url or product.get("video_url") or ""
        product["status"] = "Video Downloaded"
        product["updated_at"] = _oracle_utc_iso()
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · Video produk #{product_id} tersimpan: {video_path}")
        del state["logs"][:-80]
        return {"ok": True, "product_id": product_id, "video_path": video_path, "message": "Video tersimpan."}
    return _oracle_with_state(_mut)


def oracle_download_video_url(product_id: int, video_url: str, source_url: str = "") -> Dict[str, Any]:
    if not requests:
        return {"ok": False, "message": "Package requests belum tersedia untuk server download."}
    if not video_url:
        return {"ok": False, "message": "Video URL kosong."}
    try:
        product_name = str(product_id)
        state = _oracle_load_state()
        product = state.get("products", {}).get(str(product_id)) or {}
        if product.get("name"):
            product_name = f"product_{product_id}_{product.get('name')}"
        filename = oracle_safe_filename(product_name, ".mp4")
        dest = ORACLE_VIDEO_DIR / filename
        headers = {"User-Agent": "Mozilla/5.0 (compatible; CodexOracle/1.0)"}
        with requests.get(video_url, headers=headers, stream=True, timeout=55) as resp:
            resp.raise_for_status()
            with dest.open("wb") as fh:
                total = 0
                for chunk in resp.iter_content(chunk_size=1024 * 256):
                    if not chunk:
                        continue
                    fh.write(chunk)
                    total += len(chunk)
                    if total > 80 * 1024 * 1024:
                        raise RuntimeError("Video terlalu besar, dihentikan di 80MB.")
        if dest.stat().st_size < 20_000:
            return {"ok": False, "message": "File video terlalu kecil."}
        return oracle_mark_product_video(product_id, oracle_public_video_path(dest), source_url, video_url)
    except Exception as exc:
        return {"ok": False, "message": str(exc)}


def oracle_save_uploaded_video(product_id: int, filename: str, content: bytes, source_url: str = "", video_url: str = "") -> Dict[str, Any]:
    try:
        if not content or len(content) < 20_000:
            return {"ok": False, "message": "File video kosong / terlalu kecil."}
        safe = oracle_safe_filename(f"source_product_{product_id}_{filename or 'video.mp4'}", ".mp4")
        dest = ORACLE_VIDEO_DIR / safe
        with dest.open("wb") as fh:
            fh.write(content)
        return oracle_mark_product_video(product_id, oracle_public_video_path(dest), source_url, video_url)
    except Exception as exc:
        return {"ok": False, "message": str(exc)}



# ============================================================
# ORACLE - PHONE / ADB AUTOMATION (EMBEDDED)
# ============================================================
# Bagian ini memindahkan phone_service Oracle lama ke dalam Codex.
# UI dan fitur Codex lain tidak disentuh. Chrome extension tetap load manual,
# sedangkan upload HP berjalan melalui ADB di komputer/server yang menjalankan Codex.

ORACLE_PHONE_DIR = "/sdcard/Movies/OracleUpload"
ORACLE_PHONE_ARCHIVE_DIR = "/sdcard/Movies/OracleUploaded"
ORACLE_TERMINAL_UPLOAD_STATUSES = {"Uploaded HP", "Upload HP Skip Resolusi"}


def oracle_adb_result(ok: bool, message: str, output: str = "") -> Dict[str, Any]:
    return {"ok": bool(ok), "message": str(message or ""), "output": str(output or "")}


def oracle_phone_log(message: str, ok: Optional[bool] = None) -> None:
    def _mut(state):
        now = _oracle_now_text()
        phone_status = state.setdefault("phone_status", {})
        phone_status["last_message"] = str(message or "")
        phone_status["last_ok"] = ok
        phone_status["updated_at"] = now
        line = f"{now} · {str(message or '').strip()}"
        state.setdefault("logs", []).append(line)
        del state["logs"][:-120]
    _oracle_with_state(_mut)


def oracle_clear_phone_logs() -> Dict[str, Any]:
    def _mut(state):
        now = _oracle_now_text()
        state["logs"] = [f"{now} · Log HP dibersihkan."]
        state["phone_status"] = {"last_message": "Log HP dibersihkan.", "last_ok": True, "updated_at": now}
        return {"ok": True, "message": "Log HP dibersihkan."}
    return _oracle_with_state(_mut)


def oracle_find_adb() -> Optional[str]:
    import shutil as _shutil
    candidates: List[str] = []
    env_path = os.environ.get("ORACLE_ADB_PATH") or os.environ.get("CODEX_ADB_PATH")
    if env_path:
        candidates.append(env_path)
    roots: List[Path] = [Path.cwd(), ORACLE_EMBEDDED_ROOT]
    try:
        roots.append(Path(__file__).resolve().parent)
    except Exception:
        pass
    for root in roots:
        candidates.extend([
            str(root / "platform-tools" / "adb.exe"),
            str(root / "platform-tools" / "adb"),
            str(root / "adb" / "adb.exe"),
            str(root / "adb" / "adb"),
        ])
    candidates.extend([
        str(Path.home() / "AppData" / "Local" / "Android" / "Sdk" / "platform-tools" / "adb.exe"),
        "adb",
        "adb.exe",
    ])
    seen: Set[str] = set()
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        if candidate in {"adb", "adb.exe"}:
            found = _shutil.which(candidate)
            if found:
                return found
        elif Path(candidate).exists():
            return candidate
    return None


def oracle_adb_run(args: Any, timeout: int = 30, device_id: Optional[str] = None) -> Dict[str, Any]:
    import subprocess as _subprocess
    adb = oracle_find_adb()
    if not adb:
        return oracle_adb_result(
            False,
            "ADB belum ketemu. Taruh folder platform-tools di folder repo Codex, atau set ORACLE_ADB_PATH ke adb.exe, atau install Android platform-tools ke PATH.",
        )
    cmd = [adb]
    if device_id:
        cmd += ["-s", str(device_id)]
    cmd += [str(x) for x in list(args)]
    try:
        proc = _subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, encoding="utf-8", errors="replace")
        out_text = (proc.stdout or "") + (proc.stderr or "")
        if proc.returncode != 0:
            return oracle_adb_result(False, out_text.strip() or f"ADB error code {proc.returncode}", out_text)
        return oracle_adb_result(True, out_text.strip() or "OK", out_text)
    except _subprocess.TimeoutExpired:
        return oracle_adb_result(False, "ADB timeout. Cek kabel USB, izin debugging di HP, dan layar HP jangan terkunci.")
    except Exception as exc:
        return oracle_adb_result(False, str(exc))


def oracle_adb_shell(command: str, timeout: int = 30, device_id: Optional[str] = None) -> Dict[str, Any]:
    # Kirim sebagai satu string supaya URL Shopee yang mengandung &, ?, (, ) tidak pecah di shell Android.
    return oracle_adb_run(["shell", str(command)], timeout=timeout, device_id=device_id)


def oracle_list_devices() -> List[Dict[str, str]]:
    res = oracle_adb_run(["devices"], timeout=12)
    if not res.get("ok"):
        return [{"id": "", "status": "error", "label": res.get("message") or "ADB error"}]
    rows: List[Dict[str, str]] = []
    for line in str(res.get("output") or "").splitlines()[1:]:
        line = line.strip()
        if not line:
            continue
        parts = line.split()
        if len(parts) >= 2:
            rows.append({"id": parts[0], "status": parts[1], "label": f"{parts[0]} · {parts[1]}"})
    if not rows:
        rows.append({"id": "", "status": "none", "label": "Tidak ada perangkat terbaca."})
    return rows


def oracle_product_video_path(product: Dict[str, Any]) -> Optional[Path]:
    raw = str(product.get("local_video_path") or "").replace("\\", "/")
    if not raw:
        return None
    candidates: List[Path] = []
    raw_path = Path(raw)
    if raw_path.is_absolute():
        candidates.append(raw_path)
    candidates.extend([
        ORACLE_EMBEDDED_ROOT / raw.lstrip("/"),
        ORACLE_VIDEO_DIR / Path(raw).name,
        Path.cwd() / raw.lstrip("/"),
    ])
    try:
        candidates.append(Path(__file__).resolve().parent / raw.lstrip("/"))
    except Exception:
        pass
    for path in candidates:
        try:
            if path.exists() and path.is_file():
                return path
        except Exception:
            continue
    return None


def oracle_latest_caption(product: Dict[str, Any]) -> str:
    caption = s_clean(product.get("caption") or product.get("caption_text"))
    if caption:
        return caption
    price = oracle_format_price(product.get("price")) if product.get("price") not in (None, "") else ""
    name = s_clean(product.get("name"))[:90]
    return f"{name} {price}\nCek keranjang kuning sebelum promo berubah. #shopeeaffiliate #racunshopee".strip()


def oracle_safe_remote_name(path: Path) -> str:
    name = re.sub(r"[^a-zA-Z0-9_.-]+", "_", path.name)
    if not name.lower().endswith((".mp4", ".mov", ".webm", ".m4v")):
        name += ".mp4"
    return name[:120]


def oracle_active_remote_name(product: Dict[str, Any], local_path: Path) -> str:
    base = oracle_safe_remote_name(local_path)
    stem = Path(base).stem[:80]
    suffix = Path(base).suffix or ".mp4"
    return f"ORACLE_ACTIVE_{product.get('id')}_{stem}{suffix}"


def oracle_screen_size(device_id: Optional[str] = None) -> Tuple[int, int]:
    res = oracle_adb_run(["shell", "wm", "size"], timeout=10, device_id=device_id)
    if res.get("ok"):
        m = re.search(r"(\d+)x(\d+)", str(res.get("output") or ""))
        if m:
            return int(m.group(1)), int(m.group(2))
    return 720, 1600


def oracle_tap_pct(x_pct: float, y_pct: float, device_id: Optional[str] = None, pause: float = 0.6) -> None:
    w, h = oracle_screen_size(device_id)
    oracle_adb_run(["shell", "input", "tap", str(int(w * x_pct)), str(int(h * y_pct))], timeout=8, device_id=device_id)
    time.sleep(pause)


def oracle_tap_pct_logged(label: str, x_pct: float, y_pct: float, device_id: Optional[str] = None, pause: float = 0.8) -> None:
    oracle_phone_log(f"Auto Upload HP: {label}", None)
    oracle_tap_pct(x_pct, y_pct, device_id=device_id, pause=pause)


def oracle_adb_text(text: str) -> str:
    text = str(text or "").replace("\n", " ")
    text = re.sub(r"[^0-9A-Za-z #@._,:%/+-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:180].replace(" ", "%s")


def oracle_parse_bounds(bounds: str) -> Optional[Tuple[int, int]]:
    m = re.match(r"\[(\d+),(\d+)\]\[(\d+),(\d+)\]", bounds or "")
    if not m:
        return None
    x1, y1, x2, y2 = map(int, m.groups())
    return (x1 + x2) // 2, (y1 + y2) // 2


def oracle_dump_ui_xml(device_id: Optional[str] = None) -> str:
    remote = "/sdcard/oracle_ui.xml"
    oracle_adb_run(["shell", "uiautomator", "dump", remote], timeout=10, device_id=device_id)
    res = oracle_adb_run(["shell", "cat", remote], timeout=10, device_id=device_id)
    if not res.get("ok"):
        return ""
    return str(res.get("output") or "")


def oracle_tap_text_logged(
    label: str,
    patterns: List[str],
    fallback: Optional[Tuple[float, float]],
    device_id: Optional[str] = None,
    pause: float = 1.0,
    prefer_bottom: bool = False,
) -> bool:
    import xml.etree.ElementTree as _ET
    oracle_phone_log(f"Auto Upload HP: {label}", None)
    xml = oracle_dump_ui_xml(device_id=device_id)
    matches: List[Tuple[int, int, str]] = []
    if xml:
        try:
            root = _ET.fromstring(xml)
            low_patterns = [str(p).lower() for p in patterns]
            for node in root.iter("node"):
                joined = " ".join([
                    node.attrib.get("text", ""),
                    node.attrib.get("content-desc", ""),
                    node.attrib.get("resource-id", ""),
                ]).strip()
                if not joined:
                    continue
                low = joined.lower()
                if any(p in low for p in low_patterns):
                    center = oracle_parse_bounds(node.attrib.get("bounds", ""))
                    if center:
                        matches.append((center[0], center[1], joined))
        except Exception:
            matches = []
    if matches:
        x, y, _ = sorted(matches, key=lambda item: item[1])[-1] if prefer_bottom else matches[0]
        oracle_adb_run(["shell", "input", "tap", str(x), str(y)], timeout=8, device_id=device_id)
        time.sleep(pause)
        return True
    if fallback:
        oracle_tap_pct(fallback[0], fallback[1], device_id=device_id, pause=pause)
        return False
    time.sleep(pause)
    return False


def oracle_ui_text_blob(device_id: Optional[str] = None) -> str:
    import xml.etree.ElementTree as _ET
    xml = oracle_dump_ui_xml(device_id=device_id)
    if not xml:
        return ""
    try:
        root = _ET.fromstring(xml)
    except Exception:
        return xml.lower()
    parts: List[str] = []
    for node in root.iter("node"):
        parts.append(node.attrib.get("text", ""))
        parts.append(node.attrib.get("content-desc", ""))
        parts.append(node.attrib.get("resource-id", ""))
    return " ".join(x for x in parts if x).lower()


def oracle_upload_rejection_reason(device_id: Optional[str] = None) -> Optional[str]:
    blob = oracle_ui_text_blob(device_id=device_id)
    if not blob:
        return None
    checks = [
        ("resolusi", "resolusi video ditolak"),
        ("resolution", "resolusi video ditolak"),
        ("terlalu rendah", "kualitas/resolusi video terlalu rendah"),
        ("format", "format video tidak didukung"),
        ("tidak didukung", "video tidak didukung"),
        ("tidak dapat", "video tidak dapat dipakai"),
        ("gagal memproses", "video gagal diproses"),
        ("pilih video lain", "Shopee meminta pilih video lain"),
        ("terlalu besar", "ukuran video terlalu besar"),
        ("durasi", "durasi video tidak sesuai"),
        ("minimum", "syarat minimum video tidak terpenuhi"),
        ("maksimum", "syarat maksimum video terlampaui"),
    ]
    for needle, reason in checks:
        if needle in blob:
            if needle == "durasi" and "pilih musik" in blob and "posting" in blob:
                continue
            return reason
    return None


def oracle_make_skip_result(reason: str, label: str, device_id: Optional[str] = None) -> Dict[str, Any]:
    msg = f"SKIP_NEXT: {reason}. Oracle skip produk ini dan lanjut produk berikutnya."
    oracle_phone_log(msg, False)
    return oracle_adb_result(False, msg, "SKIP_NEXT_RESOLUTION")


def oracle_abort_if_video_rejected(label: str, device_id: Optional[str] = None) -> Optional[Dict[str, Any]]:
    reason = oracle_upload_rejection_reason(device_id=device_id)
    if reason:
        return oracle_make_skip_result(f"Video ditolak Shopee ({reason})", label, device_id=device_id)
    return None


def oracle_skip_if_stuck_after_gallery_next(label: str, device_id: Optional[str] = None) -> Optional[Dict[str, Any]]:
    blob = oracle_ui_text_blob(device_id=device_id)
    if blob and "galeri" in blob and "lanjutkan" in blob and ("dipilih" in blob or "semua" in blob):
        return oracle_make_skip_result("Video masih tertahan di Galeri setelah klik Lanjutkan, kemungkinan resolusi/format ditolak", label, device_id=device_id)
    return None


def oracle_force_tap_caption_field(device_id: Optional[str] = None) -> None:
    oracle_tap_pct(0.58, 0.125, device_id=device_id, pause=0.9)


def oracle_tap_posting_button(device_id: Optional[str] = None, pause: float = 7.0) -> None:
    clicked_by_text = oracle_tap_text_logged(
        "klik Posting",
        ["Posting"],
        fallback=None,
        device_id=device_id,
        pause=0.6,
        prefer_bottom=True,
    )
    if not clicked_by_text:
        oracle_tap_pct_logged("klik Posting pakai koordinat bawah", 0.50, 0.925, device_id, pause=0.6)
    time.sleep(pause)


def oracle_clear_active_phone_video_folder(device_id: Optional[str] = None) -> Dict[str, Any]:
    import shlex as _shlex
    cmd = (
        f"mkdir -p {_shlex.quote(ORACLE_PHONE_DIR)}; "
        f"rm -f {_shlex.quote(ORACLE_PHONE_DIR)}/* 2>/dev/null || true"
    )
    res = oracle_adb_shell(cmd, timeout=15, device_id=device_id)
    if not res.get("ok"):
        return res
    oracle_adb_shell(
        "content delete --uri content://media/external/video/media "
        "--where \"_data LIKE '/storage/emulated/0/Movies/OracleUpload/%'\" >/dev/null 2>&1 || true",
        timeout=10,
        device_id=device_id,
    )
    return oracle_adb_result(True, "Folder OracleUpload di HP sudah dikosongkan.")


def oracle_push_video_to_phone(product: Dict[str, Any], device_id: Optional[str] = None) -> Dict[str, Any]:
    import shlex as _shlex
    local_path = oracle_product_video_path(product)
    if not local_path:
        return oracle_adb_result(False, "Video lokal produk belum ketemu. Pastikan video sudah berhasil didownload dan tombol Open Video muncul.")
    cleared = oracle_clear_active_phone_video_folder(device_id=device_id)
    if not cleared.get("ok"):
        oracle_phone_log(cleared.get("message") or "Gagal membersihkan folder OracleUpload.", False)
        return cleared
    remote_path = f"{ORACLE_PHONE_DIR}/{oracle_active_remote_name(product, local_path)}"
    res = oracle_adb_run(["shell", "mkdir", "-p", ORACLE_PHONE_DIR], timeout=10, device_id=device_id)
    if not res.get("ok"):
        oracle_phone_log(res.get("message") or "Gagal membuat folder OracleUpload.", False)
        return res
    res = oracle_adb_run(["push", str(local_path), remote_path], timeout=180, device_id=device_id)
    if not res.get("ok"):
        oracle_phone_log(res.get("message") or "Gagal push video ke HP.", False)
        return res
    oracle_adb_shell(f"touch {_shlex.quote(remote_path)}; chmod 0644 {_shlex.quote(remote_path)}", timeout=10, device_id=device_id)
    file_uri = f"file://{remote_path}"
    oracle_adb_shell(
        f"am broadcast -a android.intent.action.MEDIA_SCANNER_SCAN_FILE -d {_shlex.quote(file_uri)}",
        timeout=15,
        device_id=device_id,
    )
    oracle_adb_shell(
        f"cmd media scan --file {_shlex.quote(remote_path)} >/dev/null 2>&1 || true",
        timeout=10,
        device_id=device_id,
    )
    msg = f"Mode Aman: hanya 1 video aktif di HP untuk produk #{product.get('id')}: {remote_path}"
    oracle_phone_log(msg, True)
    return oracle_adb_result(True, f"Video pasangan produk sudah jadi satu-satunya video aktif Oracle di HP: {remote_path}", remote_path)


def oracle_open_product_on_phone(product: Dict[str, Any], device_id: Optional[str] = None) -> Dict[str, Any]:
    import shlex as _shlex
    url = s_clean(product.get("url"))
    if not url:
        return oracle_adb_result(False, "Produk belum punya URL Shopee.")
    res = oracle_adb_shell(
        f"am start -a android.intent.action.VIEW -d {_shlex.quote(url)}",
        timeout=20,
        device_id=device_id,
    )
    if not res.get("ok"):
        oracle_phone_log(res.get("message") or "Gagal membuka URL produk di HP.", False)
        return res
    oracle_phone_log(f"Produk dibuka di HP: {s_clean(product.get('name'))[:80]}", True)
    return oracle_adb_result(True, "Produk sudah dibuka di HP. Kalau Shopee minta pilihan app, pilih Shopee dan Always.")


def oracle_send_to_phone(product: Dict[str, Any], device_id: Optional[str] = None) -> Dict[str, Any]:
    pushed = oracle_push_video_to_phone(product, device_id=device_id)
    if not pushed.get("ok"):
        return pushed
    opened = oracle_open_product_on_phone(product, device_id=device_id)
    if not opened.get("ok"):
        return opened
    msg = "Video sudah masuk galeri HP dan listing Shopee sudah dibuka. Lanjutkan upload dari HP, atau pakai Auto Upload."
    oracle_phone_log(msg, True)
    return oracle_adb_result(True, msg, pushed.get("output") or "")


def oracle_auto_upload_from_phone(product: Dict[str, Any], caption: str, device_id: Optional[str] = None, post_now: bool = True) -> Dict[str, Any]:
    started = oracle_send_to_phone(product, device_id=device_id)
    if not started.get("ok"):
        return started

    oracle_phone_log("Auto Upload HP mulai. Jangan sentuh HP dulu.", None)
    time.sleep(4.0)

    oracle_tap_pct_logged("klik ikon share produk", 0.735, 0.075, device_id, pause=1.8)

    oracle_tap_text_logged(
        "klik Komisi Affiliate",
        ["Komisi Affiliate", "Komisi Afiliasi", "Affiliate"],
        fallback=(0.50, 0.64),
        device_id=device_id,
        pause=2.5,
    )

    oracle_tap_text_logged(
        "klik Bagikan & Dapatkan Komisi",
        ["Bagikan & Dapatkan Komisi", "Bagikan dan Dapatkan Komisi", "Dapatkan Komisi"],
        fallback=(0.75, 0.945),
        device_id=device_id,
        pause=1.8,
        prefer_bottom=True,
    )

    oracle_tap_text_logged(
        "pilih Shopee Video",
        ["Shopee Video"],
        fallback=(0.14, 0.745),
        device_id=device_id,
        pause=3.5,
    )

    oracle_tap_text_logged(
        "klik tab Video",
        ["Video"],
        fallback=(0.31, 0.902),
        device_id=device_id,
        pause=1.6,
        prefer_bottom=True,
    )

    oracle_tap_text_logged(
        "buka Galeri",
        ["Galeri", "Gallery"],
        fallback=(0.80, 0.775),
        device_id=device_id,
        pause=2.5,
        prefer_bottom=True,
    )

    oracle_tap_text_logged(
        "filter Galeri ke tab Video",
        ["Video"],
        fallback=(0.50, 0.132),
        device_id=device_id,
        pause=1.1,
    )

    oracle_tap_pct_logged("pilih video terbaru dari Galeri", 0.18, 0.205, device_id, pause=1.2)
    rejected = oracle_abort_if_video_rejected("08_video_dipilih", device_id=device_id)
    if rejected:
        return rejected

    oracle_tap_text_logged(
        "lanjut dari Galeri",
        ["Lanjutkan"],
        fallback=(0.83, 0.837),
        device_id=device_id,
        pause=5.5,
        prefer_bottom=True,
    )
    rejected = oracle_abort_if_video_rejected("09_preview_video", device_id=device_id) or oracle_skip_if_stuck_after_gallery_next("09_preview_video", device_id=device_id)
    if rejected:
        return rejected

    oracle_tap_text_logged(
        "lanjut dari Preview",
        ["Lanjutkan"],
        fallback=(0.86, 0.912),
        device_id=device_id,
        pause=4.0,
        prefer_bottom=True,
    )
    time.sleep(1.5)

    oracle_force_tap_caption_field(device_id=device_id)
    txt = oracle_adb_text(caption)
    if txt:
        oracle_phone_log("Auto Upload HP: input caption", None)
        oracle_adb_run(["shell", "input", "text", txt], timeout=12, device_id=device_id)
        time.sleep(0.7)
        oracle_adb_run(["shell", "input", "keyevent", "KEYCODE_BACK"], timeout=6, device_id=device_id)
        time.sleep(0.8)

    if post_now:
        oracle_tap_posting_button(device_id=device_id, pause=8.0)
        msg = "Auto Upload HP selesai dikirim. Cek Shopee Video di HP untuk memastikan statusnya berhasil."
    else:
        msg = "Mode Test selesai sampai halaman Posting. Cek video/caption di HP. Kalau sudah benar, pakai Auto Upload + Posting."
    oracle_phone_log(msg, True)
    return oracle_adb_result(True, msg)


def oracle_update_product_status(product_id: int, status: str) -> None:
    def _mut(state):
        product = state.setdefault("products", {}).get(str(product_id))
        if product:
            product["status"] = status
            product["updated_at"] = _oracle_utc_iso()
    _oracle_with_state(_mut)


def oracle_get_product(product_id: int) -> Optional[Dict[str, Any]]:
    product = (_oracle_load_state().get("products") or {}).get(str(product_id))
    return dict(product) if product else None


def oracle_queue_status(message: str, **extra) -> None:
    def _mut(state):
        queue = state.setdefault("queue", {})
        queue["last_message"] = message
        queue["updated_at"] = _oracle_now_text()
        for key, value in extra.items():
            queue[key] = value
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · {message}")
        del state["logs"][:-120]
    _oracle_with_state(_mut)


def oracle_queue_is_stop_requested() -> bool:
    queue = (_oracle_load_state().get("queue") or {})
    return bool(queue.get("stop_requested"))


def oracle_queue_worker(product_ids: List[int], device_id: Optional[str], post_now: bool, interval_min: int, interval_max: int) -> None:
    import random as _random
    interval_min = max(1, min(int(interval_min), 180))
    interval_max = max(1, min(int(interval_max), 180))
    if interval_max < interval_min:
        interval_min, interval_max = interval_max, interval_min
    oracle_queue_status(
        f"Queue Mode Aman mulai. Jeda aktif: {interval_min} menit. Jeda dihitung setelah 1 produk selesai.",
        running=True,
        stop_requested=False,
        total=len(product_ids),
        done=0,
        interval_min=interval_min,
        interval_max=interval_max,
        next_wait_minutes=None,
        next_run_at=None,
    )
    try:
        for index, product_id in enumerate(product_ids, start=1):
            if oracle_queue_is_stop_requested():
                oracle_queue_status("Queue dihentikan user.", running=False, current_product_id=None, next_wait_minutes=None, next_run_at=None)
                return
            product = oracle_get_product(product_id)
            if not product or not product.get("local_video_path"):
                oracle_queue_status(f"Lewati produk #{product_id}: video lokal tidak ada.", done=index - 1)
                continue
            if (product.get("status") or "") in ORACLE_TERMINAL_UPLOAD_STATUSES:
                oracle_queue_status(f"Lewati produk #{product_id}: status sudah {product.get('status')}.", done=index - 1)
                continue
            oracle_queue_status(
                f"Upload {index}/{len(product_ids)}: produk #{product_id}. Mode Aman mengirim 1 video aktif saja.",
                current_product_id=product_id,
                next_wait_minutes=None,
                next_run_at=None,
            )
            caption = oracle_latest_caption(product)
            res = oracle_auto_upload_from_phone(product, caption=caption, device_id=device_id, post_now=post_now)
            if not res.get("ok"):
                joined = (res.get("output") or "") + " " + (res.get("message") or "")
                if "SKIP_NEXT" in joined:
                    oracle_update_product_status(product_id, "Upload HP Skip Resolusi")
                    oracle_queue_status(
                        f"Skip {index}/{len(product_ids)}: produk #{product_id}. Video ditolak/mentok, lanjut produk berikutnya tanpa jeda.",
                        done=index,
                        current_product_id=None,
                        next_wait_minutes=None,
                        next_run_at=None,
                    )
                    continue
                oracle_update_product_status(product_id, "Upload HP Gagal")
                oracle_queue_status(
                    f"Queue berhenti karena gagal di produk #{product_id}: {res.get('message')}",
                    running=False,
                    done=index - 1,
                    current_product_id=None,
                    next_wait_minutes=None,
                    next_run_at=None,
                )
                return
            oracle_update_product_status(product_id, "Uploaded HP" if post_now else "Upload HP Tested")
            oracle_queue_status(f"Berhasil {index}/{len(product_ids)}: produk #{product_id}.", done=index)
            if index < len(product_ids):
                wait_minutes = _random.randint(interval_min, interval_max)
                next_ts = time.time() + wait_minutes * 60
                next_text = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(next_ts))
                oracle_queue_status(
                    f"Menunggu {wait_minutes} menit sebelum produk berikutnya. Stop bisa ditekan kapan saja.",
                    next_wait_minutes=wait_minutes,
                    next_run_at=next_text,
                    current_product_id=None,
                )
                for _ in range(wait_minutes * 60):
                    if oracle_queue_is_stop_requested():
                        oracle_queue_status("Queue dihentikan user saat jeda.", running=False, current_product_id=None, next_wait_minutes=None, next_run_at=None)
                        return
                    time.sleep(1)
        oracle_queue_status("Queue Mode Aman selesai.", running=False, current_product_id=None, next_wait_minutes=None, next_run_at=None)
    except Exception as exc:
        oracle_queue_status(f"Queue error: {exc}", running=False, current_product_id=None, next_wait_minutes=None, next_run_at=None)


def oracle_start_safe_upload_queue(max_items: int = 30, device_id: Optional[str] = None, post_now: bool = False, interval_min: int = 20, interval_max: int = 40) -> Dict[str, Any]:
    import threading as _threading
    state = _oracle_load_state()
    queue = state.get("queue", {}) or {}
    if queue.get("running"):
        active = f"{queue.get('interval_min') or queue.get('interval_minutes') or interval_min} menit"
        return oracle_adb_result(False, f"Queue masih berjalan dengan jeda aktif {active}. Tekan Stop dulu kalau ingin mulai ulang.")
    products = oracle_get_ready_products(1000)
    clean_ids = [int(p.get("id")) for p in products if p.get("id")][: max(1, min(int(max_items or 30), 30))]
    if not clean_ids:
        return oracle_adb_result(False, "Tidak ada produk yang bisa masuk queue.")
    interval_min = max(1, min(int(interval_min or 20), 180))
    interval_max = max(1, min(int(interval_max or interval_min), 180))
    def _mut(state):
        queue = state.setdefault("queue", {})
        queue.update({
            "running": True,
            "stop_requested": False,
            "total": len(clean_ids),
            "done": 0,
            "interval_min": interval_min,
            "interval_max": interval_max,
            "interval_minutes": interval_min,
            "post_now": post_now,
            "next_wait_minutes": None,
            "next_run_at": None,
            "last_message": f"Queue Mode Aman dimulai untuk {len(clean_ids)} produk. Jeda aktif: {interval_min} menit.",
            "updated_at": _oracle_now_text(),
        })
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · {queue['last_message']}")
        del state["logs"][:-120]
    _oracle_with_state(_mut)
    thread = _threading.Thread(target=oracle_queue_worker, args=(clean_ids, device_id, post_now, interval_min, interval_max), daemon=True)
    thread.start()
    return oracle_adb_result(True, f"Queue Mode Aman dimulai untuk {len(clean_ids)} produk. Jeda aktif: {interval_min} menit.")


def oracle_stop_safe_upload_queue() -> Dict[str, Any]:
    queue = (_oracle_load_state().get("queue") or {})
    if not queue.get("running"):
        def _mut(state):
            state.setdefault("queue", {})["stop_requested"] = False
        _oracle_with_state(_mut)
        return oracle_adb_result(True, "Queue sedang tidak berjalan.")
    def _mut(state):
        q = state.setdefault("queue", {})
        q["stop_requested"] = True
        q["last_message"] = "Stop diminta. Oracle akan berhenti setelah langkah/jeda saat ini selesai."
        q["updated_at"] = _oracle_now_text()
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · {q['last_message']}")
        del state["logs"][:-120]
    _oracle_with_state(_mut)
    return oracle_adb_result(True, "Stop queue diminta.")


def oracle_set_queue(running: bool, message: str, **extra) -> Dict[str, Any]:
    def _mut(state):
        queue = state.setdefault("queue", {})
        queue["running"] = running
        queue["last_message"] = message
        queue["updated_at"] = _oracle_now_text()
        for k, v in extra.items():
            queue[k] = v
        state.setdefault("logs", []).append(f"{_oracle_now_text()} · {message}")
        del state["logs"][:-120]
        return {"ok": True, "message": message}
    return _oracle_with_state(_mut)


def oracle_phone_action(product_id: int, action: str, post_now: bool = False, device_id: Optional[str] = None) -> Dict[str, Any]:
    if action == "mark-uploaded":
        oracle_update_product_status(product_id, "Uploaded HP")
        msg = f"Produk #{product_id} ditandai sudah uploaded."
        oracle_phone_log(msg, True)
        return oracle_adb_result(True, msg)
    if action == "reset-upload":
        product = oracle_get_product(product_id)
        oracle_update_product_status(product_id, "Video Downloaded" if product and product.get("local_video_path") else "Scanned")
        msg = f"Produk #{product_id} dikembalikan ke Produk Siap Upload."
        oracle_phone_log(msg, True)
        return oracle_adb_result(True, msg)
    product = oracle_get_product(product_id)
    if not product:
        return oracle_adb_result(False, "Produk tidak ditemukan.")
    if action == "send":
        res = oracle_send_to_phone(product, device_id=device_id)
        if res.get("ok"):
            oracle_update_product_status(product_id, "Sent HP")
        return res
    if action == "auto-upload":
        caption = oracle_latest_caption(product)
        res = oracle_auto_upload_from_phone(product, caption=caption, device_id=device_id, post_now=post_now)
        if res.get("ok"):
            oracle_update_product_status(product_id, "Uploaded HP" if post_now else "Upload HP Tested")
        elif "SKIP_NEXT" in ((res.get("output") or "") + " " + (res.get("message") or "")):
            oracle_update_product_status(product_id, "Upload HP Skip Resolusi")
        else:
            oracle_update_product_status(product_id, "Upload HP Gagal")
        return res
    return oracle_adb_result(False, f"Action tidak dikenal: {action}")


def oracle_phone_state_data() -> Dict[str, Any]:
    state = _oracle_load_state()
    logs = "\n".join(state.get("logs", [])[-50:]) or "Belum ada log."
    queue = state.get("queue", {}) or {}
    phone_status = state.get("phone_status", {}) or {}
    device_rows = oracle_list_devices()
    devices = [d.get("label") or "" for d in device_rows] or ["Tidak ada perangkat terbaca."]
    products = []
    for p in oracle_get_ready_products(100):
        video_path = p.get("local_video_path") or ""
        products.append({
            "id": str(p.get("id") or ""),
            "name": p.get("name") or "Produk",
            "image": p.get("image_url") or "",
            "price": oracle_format_price(p.get("price")),
            "video_href": "/" + video_path if video_path and not video_path.startswith("/") else video_path,
            "status": p.get("status") or "Video Downloaded",
            "detail_href": p.get("url") or "",
        })
    progress = queue.get("last_message") or "Queue belum berjalan."
    if queue.get("running"):
        progress += f" · {int(queue.get('done') or 0)}/{int(queue.get('total') or 0)}"
        if queue.get("next_run_at"):
            progress += f" · Next: {queue.get('next_run_at')}"
    return {
        "devices": devices,
        "last_status": phone_status.get("last_message") or "Mode ADB aktif. Sambungkan HP, aktifkan USB debugging, lalu pilih Allow di HP.",
        "logs": logs,
        "queue_badge": "Running" if queue.get("running") else "Idle",
        "queue_progress": progress,
        "products": products,
    }


def oracle_render_phone_html() -> str:
    data = oracle_phone_state_data()
    def esc(x):
        import html as _html
        return _html.escape(str(x or ""), quote=True)
    rows = []
    for p in data.get("products", []):
        img = f'<img src="{esc(p.get("image"))}" style="width:42px;height:42px;object-fit:cover">' if p.get("image") else ""
        video = f'<a href="{esc(p.get("video_href"))}">Open Video</a>' if p.get("video_href") else "-"
        detail = f'<a href="{esc(p.get("detail_href"))}">Detail</a>' if p.get("detail_href") else '<a href="#">Detail</a>'
        pid = esc(p.get("id"))
        rows.append(f'''
        <tr>
          <td>{img} {esc(p.get("name"))}</td>
          <td>{esc(p.get("price"))}</td>
          <td>{video}</td>
          <td>{esc(p.get("status"))}</td>
          <td>
            <form action="/phone/product/{pid}/send" method="post"><button>Kirim 1 Video + Buka Produk</button></form>
            <form action="/phone/product/{pid}/auto-upload" method="post"><button>Test Auto Upload</button></form>
            <form action="/phone/product/{pid}/auto-upload" method="post"><input type="hidden" name="post_now" value="on"><button>Auto Upload + Posting</button></form>
            {detail}
          </td>
        </tr>''')
    rows_html = "\n".join(rows)
    devices_html = "".join(f"<li>{esc(x)}</li>" for x in data.get("devices") or [])
    return f'''
    <html><body>
      <div class="card"><div class="card-header">Status HP</div><div class="card-body"><b>ADB:</b><ul>{devices_html}</ul><b>Status terakhir:</b> {esc(data.get("last_status"))}</div></div>
      <div class="card"><div class="card-header">Log HP</div><div id="phoneLogBox">{esc(data.get("logs"))}</div></div>
      <div class="card"><div class="card-header">Queue Upload <span class="badge">{esc(data.get("queue_badge"))}</span></div><div class="card-body">Progress: {esc(data.get("queue_progress"))}</div></div>
      <div class="card"><div class="card-header">Produk Siap Upload · {len(data.get("products") or [])}</div><table><tbody>{rows_html}</tbody></table></div>
    </body></html>
    '''


class _OracleEmbeddedHTTPHandler:
    """Factory supaya import http.server hanya saat engine dipakai."""

    @staticmethod
    def make():
        import json as _json
        import os as _os
        import mimetypes as _mimetypes
        from http.server import BaseHTTPRequestHandler
        from urllib.parse import parse_qs as _parse_qs, urlparse as _urlparse

        class Handler(BaseHTTPRequestHandler):
            server_version = "CodexOracleEmbedded/1.0"

            def log_message(self, format, *args):
                return

            def _headers(self, status=200, content_type="application/json"):
                self.send_response(status)
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
                self.send_header("Access-Control-Allow-Headers", "Content-Type")
                if content_type:
                    self.send_header("Content-Type", content_type)
                self.end_headers()

            def _json(self, payload, status=200):
                raw = _json.dumps(payload, ensure_ascii=False).encode("utf-8")
                self._headers(status, "application/json; charset=utf-8")
                self.wfile.write(raw)

            def _text(self, body, status=200, content_type="text/html; charset=utf-8"):
                raw = str(body or "").encode("utf-8")
                self._headers(status, content_type)
                self.wfile.write(raw)

            def _redirect_phone(self):
                self.send_response(303)
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Location", "/phone")
                self.end_headers()

            def _read_body(self):
                length = int(self.headers.get("Content-Length") or 0)
                return self.rfile.read(length) if length else b""

            def _parse_form(self):
                ctype = self.headers.get("Content-Type", "")
                raw = self._read_body()
                if "application/x-www-form-urlencoded" in ctype:
                    return {k: v[-1] if isinstance(v, list) else v for k, v in _parse_qs(raw.decode("utf-8", "replace")).items()}
                if "application/json" in ctype:
                    try:
                        return _json.loads(raw.decode("utf-8", "replace") or "{}")
                    except Exception:
                        return {}
                return {k: v[-1] if isinstance(v, list) else v for k, v in _parse_qs(raw.decode("utf-8", "replace")).items()}

            def do_OPTIONS(self):
                self._headers(204, None)

            def do_GET(self):
                parsed = _urlparse(self.path)
                path = parsed.path
                if path in {"/", "/__health"}:
                    self._json({"ok": True, "service": "Codex Embedded Oracle", "time": _oracle_now_text()})
                    return
                if path == "/scanner/active":
                    self._json(oracle_active_payload())
                    return
                if path == "/phone":
                    self._text(oracle_render_phone_html())
                    return
                if path.startswith("/output/videos/"):
                    name = _os.path.basename(path)
                    file_path = ORACLE_VIDEO_DIR / name
                    if not file_path.exists():
                        self._text("Not Found", status=404, content_type="text/plain; charset=utf-8")
                        return
                    ctype = _mimetypes.guess_type(str(file_path))[0] or "video/mp4"
                    self.send_response(200)
                    self.send_header("Access-Control-Allow-Origin", "*")
                    self.send_header("Content-Type", ctype)
                    self.send_header("Content-Length", str(file_path.stat().st_size))
                    self.end_headers()
                    with file_path.open("rb") as fh:
                        while True:
                            chunk = fh.read(1024 * 256)
                            if not chunk:
                                break
                            self.wfile.write(chunk)
                    return
                self._json({"ok": False, "message": "Endpoint tidak ditemukan."}, status=404)

            def do_POST(self):
                parsed = _urlparse(self.path)
                path = parsed.path
                try:
                    if path == "/scanner/start":
                        form = self._parse_form()
                        normalized = dict(form)
                        for key in ["include_ads", "require_video", "download_videos", "require_komisixtra"]:
                            normalized[key] = normalized.get(key) in {"on", "1", "true", True}
                        job = oracle_create_scan_job(normalized)
                        self.send_response(303)
                        self.send_header("Access-Control-Allow-Origin", "*")
                        self.send_header("Location", f"/scanner/{job['id']}")
                        self.end_headers()
                        return
                    if path == "/scanner/ingest":
                        payload = self._parse_form()
                        self._json(oracle_ingest_cards(int(payload.get("job_id") or 0), payload.get("cards") or [], payload.get("source_url") or ""))
                        return
                    if path == "/scanner/finish":
                        payload = self._parse_form()
                        ok_value = payload.get("ok", True)
                        ok_bool = ok_value if isinstance(ok_value, bool) else str(ok_value).lower() not in {"false", "0", "no"}
                        self._json(oracle_finish_job(int(payload.get("job_id") or 0), ok_bool, payload.get("message") or ""))
                        return
                    if path == "/scanner/save-video-url":
                        payload = self._parse_form()
                        self._json(oracle_download_video_url(int(payload.get("product_id") or 0), payload.get("video_url") or "", payload.get("source_url") or ""))
                        return
                    if path == "/scanner/upload-video":
                        ctype = self.headers.get("Content-Type", "")
                        if "multipart/form-data" not in ctype:
                            self._json({"ok": False, "message": "Expected multipart/form-data."}, status=400)
                            return
                        raw_body = self._read_body()
                        from email.parser import BytesParser as _BytesParser
                        from email.policy import default as _email_default_policy
                        pseudo = b"Content-Type: " + ctype.encode("utf-8", "replace") + b"\r\nMIME-Version: 1.0\r\n\r\n" + raw_body
                        msg = _BytesParser(policy=_email_default_policy).parsebytes(pseudo)
                        fields = {}
                        files = {}
                        for part in msg.iter_parts():
                            disp = part.get("Content-Disposition", "")
                            name = part.get_param("name", header="content-disposition")
                            if not name:
                                continue
                            filename = part.get_filename()
                            payload = part.get_payload(decode=True) or b""
                            if filename:
                                files[name] = {"filename": filename, "content": payload}
                            else:
                                fields[name] = payload.decode("utf-8", "replace")
                        product_id = int(fields.get("product_id") or 0)
                        source_url = fields.get("source_url") or ""
                        video_url = fields.get("video_url") or ""
                        fileitem = files.get("file")
                        if not fileitem:
                            self._json({"ok": False, "message": "File tidak ditemukan."}, status=400)
                            return
                        self._json(oracle_save_uploaded_video(product_id, fileitem.get("filename") or "video.mp4", fileitem.get("content") or b"", source_url, video_url))
                        return
                    if path == "/phone/logs/clear":
                        oracle_clear_phone_logs()
                        self._redirect_phone()
                        return
                    if path == "/phone/queue/start":
                        form = self._parse_form()
                        max_items = int(form.get("max_items") or 5)
                        interval = int(form.get("interval_minutes") or 20)
                        post_now = form.get("post_now") in {"on", "1", "true", True}
                        oracle_start_safe_upload_queue(max_items=max_items, device_id=form.get("device_id") or None, post_now=post_now, interval_min=interval, interval_max=interval)
                        self._redirect_phone()
                        return
                    if path == "/phone/queue/stop":
                        oracle_stop_safe_upload_queue()
                        self._redirect_phone()
                        return
                    m = re.match(r"^/phone/product/(\d+)/(send|auto-upload|mark-uploaded|reset-upload)$", path)
                    if m:
                        pid = int(m.group(1))
                        action = m.group(2)
                        form = self._parse_form()
                        oracle_phone_action(pid, action, post_now=form.get("post_now") in {"on", "1", "true", True}, device_id=form.get("device_id") or None)
                        self._redirect_phone()
                        return
                    self._json({"ok": False, "message": "Endpoint tidak ditemukan."}, status=404)
                except Exception as exc:
                    self._json({"ok": False, "message": str(exc)}, status=500)

        return Handler


def _oracle_port_is_open(host: str, port: int, timeout: float = 0.25) -> bool:
    import socket as _socket
    try:
        with _socket.create_connection((host, int(port)), timeout=timeout):
            return True
    except Exception:
        return False


def ensure_oracle_embedded_engine() -> Dict[str, Any]:
    _oracle_prepare_dirs()
    if _oracle_port_is_open(ORACLE_EMBEDDED_API_HOST, ORACLE_EMBEDDED_API_PORT):
        return {"ok": True, "message": f"Engine aktif di {ORACLE_API_BASE_DEFAULT}", "already_running": True}
    try:
        import threading as _threading
        from http.server import ThreadingHTTPServer
        Handler = _OracleEmbeddedHTTPHandler.make()
        server = ThreadingHTTPServer((ORACLE_EMBEDDED_API_HOST, ORACLE_EMBEDDED_API_PORT), Handler)
        thread = _threading.Thread(target=server.serve_forever, name="codex-oracle-embedded-api", daemon=True)
        thread.start()
        setattr(ensure_oracle_embedded_engine, "_server", server)
        setattr(ensure_oracle_embedded_engine, "_thread", thread)
        oracle_log(f"Codex Embedded Oracle engine aktif di {ORACLE_API_BASE_DEFAULT}")
        return {"ok": True, "message": f"Engine aktif di {ORACLE_API_BASE_DEFAULT}", "already_running": False}
    except Exception as exc:
        return {"ok": False, "message": str(exc)}


def oracle_get_api_base() -> str:
    st.session_state["oracle_api_base"] = ORACLE_API_BASE_DEFAULT
    return ORACLE_API_BASE_DEFAULT.rstrip("/")


def oracle_http_request(method: str, path: str, *, data: Optional[Dict[str, Any]] = None, timeout: int = 20, allow_redirects: bool = False) -> Dict[str, Any]:
    ensure_oracle_embedded_engine()
    if requests is None:
        return {"ok": False, "message": "Package requests belum terinstall.", "status_code": None, "text": ""}
    base = oracle_get_api_base()
    url = f"{base}/{path.lstrip('/')}"
    try:
        if method.upper() == "POST":
            response = requests.post(url, data=data or {}, timeout=timeout, allow_redirects=allow_redirects)
        else:
            response = requests.get(url, timeout=timeout, allow_redirects=allow_redirects)
        return {"ok": response.status_code < 400, "status_code": response.status_code, "headers": dict(response.headers), "text": response.text, "url": url}
    except Exception as exc:
        return {"ok": False, "message": str(exc), "status_code": None, "text": "", "url": url}


def oracle_status() -> Dict[str, Any]:
    engine = ensure_oracle_embedded_engine()
    if not engine.get("ok"):
        return {"online": False, "message": engine.get("message")}
    data = oracle_active_payload()
    return {"online": True, "data": data, "message": data.get("message") or "Codex Embedded Oracle engine online."}


def oracle_start_scan_job(form_values: Dict[str, Any]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "keyword": form_values.get("keyword", ""),
        "search_url": form_values.get("search_url", ""),
        "category_url": form_values.get("category_url", ""),
        "max_products": int(form_values.get("max_products") or 30),
        "max_pages": int(form_values.get("max_pages") or 1),
        "min_rating": form_values.get("min_rating", ""),
        "min_sold": form_values.get("min_sold", ""),
        "min_price": form_values.get("min_price", ""),
        "max_price": form_values.get("max_price", ""),
        "sort_by": form_values.get("sort_by", "relevance"),
        "seller_type": form_values.get("seller_type", "any"),
    }
    if form_values.get("include_ads"):
        payload["include_ads"] = "on"
    if form_values.get("require_video"):
        payload["require_video"] = "on"
    if form_values.get("download_videos"):
        payload["download_videos"] = "on"
    if form_values.get("require_komisixtra"):
        payload["require_komisixtra"] = "on"
    return oracle_http_request("POST", "/scanner/start", data=payload, timeout=60, allow_redirects=False)


def oracle_post_phone_action(path: str, payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    return oracle_http_request("POST", path, data=payload or {}, timeout=240, allow_redirects=False)


def oracle_fetch_phone_state() -> Dict[str, Any]:
    return {"ok": True, "message": "OK", "data": oracle_phone_state_data()}


def render_oracle_product_row(product: Dict[str, Any], idx: int):
    base = oracle_get_api_base()
    product_id = product.get("id") or f"row{idx}"
    cols = st.columns([4.6, 1.2, 1.0, 1.3, 3.8])
    with cols[0]:
        img_url = product.get("image") or ""
        if img_url and img_url.startswith("/"):
            img_url = base + img_url
        inner = st.columns([0.8, 4.2])
        with inner[0]:
            if img_url:
                st.image(img_url, width=42)
        with inner[1]:
            name = product.get("name") or "Produk"
            st.write(name[:140])
    with cols[1]:
        st.write(product.get("price") or "-")
    with cols[2]:
        video_href = product.get("video_href") or ""
        if video_href:
            if video_href.startswith("/"):
                video_href = base + video_href
            st.link_button("Open Video", video_href, use_container_width=True)
        else:
            st.caption("-")
    with cols[3]:
        st.caption(product.get("status") or "-")
    with cols[4]:
        action_cols = st.columns(4)
        with action_cols[0]:
            if st.button("Kirim", key=f"oracle_send_{product_id}_{idx}", use_container_width=True):
                res = oracle_post_phone_action(f"/phone/product/{product_id}/send", {"device_id": ""})
                st.toast("Perintah kirim diterima." if res.get("ok") else res.get("message", "Gagal kirim."))
        with action_cols[1]:
            if st.button("Test", key=f"oracle_test_{product_id}_{idx}", use_container_width=True):
                res = oracle_post_phone_action(f"/phone/product/{product_id}/auto-upload", {"device_id": ""})
                st.toast("Test auto upload diterima." if res.get("ok") else res.get("message", "Gagal test."))
        with action_cols[2]:
            if st.button("Posting", key=f"oracle_posting_{product_id}_{idx}", use_container_width=True):
                res = oracle_post_phone_action(f"/phone/product/{product_id}/auto-upload", {"device_id": "", "post_now": "on"})
                st.toast("Auto upload + posting diterima." if res.get("ok") else res.get("message", "Gagal posting."))
        with action_cols[3]:
            detail_href = product.get("detail_href") or ""
            if detail_href:
                st.link_button("Detail", detail_href, use_container_width=True)
            else:
                st.caption("-")


def render_oracle_video_finder():
    ensure_oracle_embedded_engine()
    st.title("Oracle Video Finder")
    st.caption("Scan keyword/URL Shopee, ambil listing video, download source video, lalu tampilkan hasil akhir di satu layar.")
    st.info(
        "Mesin Oracle sudah pindah ke Codex. User cukup menjalankan Codex, lalu load Chrome Extension Oracle secara manual. "
        "Extension tetap memakai endpoint lokal http://127.0.0.1:8000 yang sekarang dinyalakan oleh Codex."
    )
    with st.expander("Status Engine & Extension", expanded=False):
        status = oracle_status()
        if status.get("online"):
            st.success(f"Codex Embedded Oracle engine online: {ORACLE_API_BASE_DEFAULT}")
            active_data = status.get("data") or {}
            if active_data.get("ok"):
                st.caption(f"Scan aktif: #{active_data.get('job_id')} · {active_data.get('keyword') or active_data.get('start_url')}")
                st.code(active_data.get("start_url") or "", language="text")
            else:
                st.caption(active_data.get("message") or "Belum ada scan aktif.")
        else:
            st.error(f"Engine belum aktif: {status.get('message')}")
        st.caption("Extension: gunakan folder chrome_extension Oracle lama. Kalau belum bergerak setelah start scan, tunggu alarm extension ±30 detik atau reload extension.")

    left, right = st.columns([1.0, 2.05], gap="medium")
    with left:
        with st.container(border=True):
            st.subheader("Mulai Pencarian Video")
            keyword = st.text_input("Keyword", placeholder="contoh: rak dapur kayu", key="oracle_keyword")
            search_url = st.text_input("Search URL / halaman Shopee", placeholder="https://shopee.co.id/search?keyword=...", key="oracle_search_url")
            category_url = st.text_input("Category URL", placeholder="opsional", key="oracle_category_url")
            c1, c2 = st.columns(2)
            with c1:
                max_products = st.number_input("Target Video / Scan", min_value=1, max_value=500, value=30, step=1, key="oracle_max_products")
            with c2:
                max_pages = st.number_input("Target Page", min_value=1, max_value=50, value=1, step=1, key="oracle_max_pages")
            st.caption("Default 30. Bukan limit sistem, ini target per job.")
            c3, c4 = st.columns(2)
            with c3:
                sort_label = st.selectbox("Urutan", ["Relevansi Shopee", "Terlaris"], key="oracle_sort_label")
            with c4:
                seller_label = st.selectbox("Tipe Penjual", ["Semua", "Mall", "Star", "Star+"], key="oracle_seller_label")
            c5, c6 = st.columns(2)
            with c5:
                min_rating = st.text_input("Min Rating", value="4.5", key="oracle_min_rating")
                min_price = st.text_input("Min Price", value="", key="oracle_min_price")
            with c6:
                min_sold = st.text_input("Min Sold", value="100", key="oracle_min_sold")
                max_price = st.text_input("Max Price", value="", key="oracle_max_price")
            include_ads = st.checkbox("Include Ads", value=True, key="oracle_include_ads")
            require_video = st.checkbox("Hanya ambil listing yang ada video", value=True, key="oracle_require_video")
            require_komisixtra = st.checkbox("Hanya yang ada KomisiXtra jika badge terbaca di listing", value=False, key="oracle_require_komisixtra")
            download_videos = st.checkbox("Download video produk otomatis", value=True, key="oracle_download_videos")
            sort_by = "sales" if sort_label == "Terlaris" else "relevance"
            seller_map = {"Semua": "any", "Mall": "mall", "Star": "star", "Star+": "star_plus"}
            seller_type = seller_map.get(seller_label, "any")
            if st.button("Mulai Cari & Download Video", type="primary", use_container_width=True, key="oracle_start_scan"):
                if not any([s_clean(keyword), s_clean(search_url), s_clean(category_url)]):
                    st.error("Isi minimal Keyword, Search URL, atau Category URL dulu.")
                else:
                    payload = {
                        "keyword": keyword,
                        "search_url": search_url,
                        "category_url": category_url,
                        "max_products": max_products,
                        "max_pages": max_pages,
                        "min_rating": min_rating,
                        "min_sold": min_sold,
                        "min_price": min_price,
                        "max_price": max_price,
                        "include_ads": include_ads,
                        "require_video": require_video,
                        "require_komisixtra": require_komisixtra,
                        "download_videos": download_videos,
                        "sort_by": sort_by,
                        "seller_type": seller_type,
                    }
                    res = oracle_start_scan_job(payload)
                    if res.get("ok") or res.get("status_code") in {302, 303, 307}:
                        location = (res.get("headers") or {}).get("Location", "")
                        st.success(f"Job scan dibuat {location or ''}. Extension akan jalan otomatis.")
                        st.caption("Biarkan Chrome extension aktif. Biasanya tab Shopee terbuka dalam beberapa detik sampai ±30 detik.")
                    else:
                        st.error(res.get("message") or f"Gagal start scan. HTTP {res.get('status_code')}")
            st.caption("Pertama kali saja: Chrome > Extensions > Developer Mode > Load unpacked > pilih folder chrome_extension Oracle.")

    phone_state = oracle_fetch_phone_state()
    phone_data = phone_state.get("data") or {}
    with right:
        top1, top2 = st.columns([1.0, 1.55], gap="medium")
        with top1:
            with st.container(border=True):
                st.subheader("Upload HP")
                st.caption("Mode Aman v5.0")
                st.write("**ADB:**")
                devices = phone_data.get("devices") or ["Tidak ada perangkat terbaca."]
                for item in devices[:5]:
                    st.caption(f"• {item}")
                st.write("**Status terakhir:**")
                st.caption(phone_data.get("last_status") or "Belum ada proses HP.")
        with top2:
            with st.container(border=True):
                log_header = st.columns([3, 1])
                with log_header[0]:
                    st.subheader("Log HP")
                with log_header[1]:
                    if st.button("Bersihkan Log", key="oracle_clear_log"):
                        oracle_post_phone_action("/phone/logs/clear")
                        st.toast("Log HP dibersihkan.")
                st.text_area("", value=phone_data.get("logs") or "Belum ada log.", height=128, key="oracle_logs_view", label_visibility="collapsed")
        with st.container(border=True):
            qcols = st.columns([3, 1])
            with qcols[0]:
                st.subheader("Queue Upload")
            with qcols[1]:
                st.caption(phone_data.get("queue_badge") or "Idle")
            qc1, qc2, qc3, qc4 = st.columns([1.2, 1.2, 1.15, 1.2])
            with qc1:
                queue_max = st.number_input("Maksimal produk", min_value=1, max_value=30, value=5, step=1, key="oracle_queue_max")
            with qc2:
                queue_interval = st.number_input("Jeda / produk, menit", min_value=1, max_value=180, value=20, step=1, key="oracle_queue_interval")
            with qc3:
                st.write("")
                st.write("")
                if st.button("Start Queue Test", key="oracle_queue_test", use_container_width=True):
                    res = oracle_post_phone_action("/phone/queue/start", {"device_id": "", "max_items": int(queue_max), "interval_minutes": int(queue_interval)})
                    st.toast("Queue test dimulai." if res.get("ok") else res.get("message", "Gagal start queue."))
            with qc4:
                st.write("")
                st.write("")
                if st.button("Start Queue + Posting", key="oracle_queue_posting", use_container_width=True):
                    res = oracle_post_phone_action("/phone/queue/start", {"device_id": "", "max_items": int(queue_max), "interval_minutes": int(queue_interval), "post_now": "on"})
                    st.toast("Queue + posting dimulai." if res.get("ok") else res.get("message", "Gagal start queue."))
            stop_col, prog_col = st.columns([1, 4])
            with stop_col:
                if st.button("Stop Queue", key="oracle_queue_stop", use_container_width=True):
                    res = oracle_post_phone_action("/phone/queue/stop")
                    st.toast("Queue dihentikan." if res.get("ok") else res.get("message", "Gagal stop queue."))
            with prog_col:
                st.caption(phone_data.get("queue_progress") or "Queue belum berjalan.")
        products = phone_data.get("products") or []
        with st.container(border=True):
            st.subheader(f"Produk Siap Upload · {len(products)}")
            header = st.columns([4.6, 1.2, 1.0, 1.3, 3.8])
            header[0].markdown("**Produk**")
            header[1].markdown("**Harga**")
            header[2].markdown("**Video**")
            header[3].markdown("**Status**")
            header[4].markdown("**Upload HP**")
            st.divider()
            if products:
                for idx, product in enumerate(products, start=1):
                    render_oracle_product_row(product, idx)
                    st.divider()
            else:
                st.caption("Belum ada produk siap upload.")
        if st.checkbox("Auto refresh halaman Oracle setiap 5 detik", value=False, key="oracle_auto_refresh"):
            st.caption("Auto refresh aktif. Halaman akan memuat ulang supaya progress/log terbaru masuk.")
            st.markdown("<meta http-equiv='refresh' content='5'>", unsafe_allow_html=True)

def build_menu() -> str:
    st.sidebar.title(APP_TITLE)
    st.session_state["issue_output_mode"] = "Info saja"

    group = st.sidebar.radio(
        "Menu Utama",
        ["Dashboard", "Posting", "Oracle", "Update Stok", "Update Harga Normal", "Update Harga Coret", "Submit Campaign", "Analisa", "Affiliate"],
        key="sidebar_main_menu",
    )

    if group == "Dashboard":
        route = "dashboard"

    elif group == "Posting":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee", "TikTokShop"],
            key="sidebar_posting_menu",
        )
        if child == "Shopee":
            route = "posting_shopee"
        else:
            route = "posting_tiktokshop"

    elif group == "Oracle":
        child = st.sidebar.radio(
            "Menu Besar",
            ["Video Finder"],
            key="sidebar_oracle_menu",
        )
        if child == "Video Finder":
            route = "oracle_video_finder"
        else:
            route = "oracle_video_finder"

    elif group == "Update Stok":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "Mwh", "Bigseller", "Blibli", "Akulaku"],
            key="sidebar_update_stok_menu",
        )
        if child.startswith("Shopee"):
            route = "update_stok_shopee"
        elif child == "TikTokShop":
            route = "update_stok_tiktokshop"
        elif child == "Mwh":
            route = "update_stok_mwh"
        elif child == "Bigseller":
            route = "update_stok_bigseller"
        elif child == "Blibli":
            route = "update_stok_blibli"
        else:
            route = "update_stok_akulaku"

    elif group == "Update Harga Normal":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "Mwh", "PowerMerchant", "Bigseller", "Blibli", "Akulaku"],
            key="sidebar_harga_normal_menu",
        )
        if child.startswith("Shopee"):
            route = "harga_normal_shopee"
        elif child == "TikTokShop":
            route = "harga_normal_tiktokshop"
        elif child == "Mwh":
            route = "harga_normal_mwh"
        elif child == "PowerMerchant":
            route = "harga_normal_powermerchant"
        elif child == "Bigseller":
            route = "harga_normal_bigseller"
        elif child == "Blibli":
            route = "harga_normal_blibli"
        else:
            route = "harga_normal_akulaku"

    elif group == "Update Harga Coret":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee (Mall & Star)", "TikTokShop", "PowerMerchant"],
            key="sidebar_harga_coret_menu",
        )
        if child.startswith("Shopee"):
            route = "harga_coret_shopee"
        elif child == "TikTokShop":
            route = "harga_coret_tiktokshop"
        else:
            route = "harga_coret_powermerchant"

    elif group == "Submit Campaign":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["Shopee", "TikTokShop"],
            key="sidebar_submit_campaign_menu",
        )
        if child == "Shopee":
            route = "submit_campaign_shopee"
        else:
            route = "submit_campaign_tiktokshop"

    elif group == "Analisa":
        child = st.sidebar.radio(
            "Pilih Fitur Analisa",
            ["Analisa Penjualan", "Analisa Produk", "Progres ON"],
            key="sidebar_analisa_menu",
        )
        if child == "Analisa Penjualan":
            route = "analisa_penjualan"
        elif child == "Analisa Produk":
            route = "analisa_produk_stok"
        else:
            route = "progres_on"

    elif group == "Affiliate":
        child = st.sidebar.radio(
            "Pilih Platform",
            ["TikTokShop"],
            key="sidebar_affiliate_menu",
        )
        if child == "TikTokShop":
            route = "affiliate_tiktokshop"
        else:
            route = "affiliate_tiktokshop"

    else:
        route = "dashboard"

    st.sidebar.markdown("---")
    st.sidebar.markdown("<br><br><br>", unsafe_allow_html=True)
    st.sidebar.link_button(
        "Download File Addon",
        "https://drive.google.com/drive/u/0/folders/1r3qVqmm1ALfLGaLuvagAf5EQuMVT0iWI",
        use_container_width=True,
    )

    return route


def main():
    st.session_state["_rendered_cache_keys"] = set()
    route = build_menu()
    st.session_state.current_route = route
    if route == "dashboard":
        render_dashboard()
    elif route == "posting_shopee":
        render_posting_shopee()
    elif route == "posting_tiktokshop":
        render_posting_tiktokshop()
    elif route == "oracle_video_finder":
        render_oracle_video_finder()
    elif route == "update_stok_shopee":
        render_update_stok_shopee()
    elif route == "update_stok_tiktokshop":
        render_update_stok_tiktokshop()
    elif route == "update_stok_mwh":
        render_update_stok_mwh()
    elif route == "update_stok_bigseller":
        render_update_stok_bigseller()
    elif route == "update_stok_blibli":
        render_update_stok_blibli()
    elif route == "update_stok_akulaku":
        render_update_stok_akulaku()
    elif route == "harga_normal_shopee":
        render_harga_normal_shopee()
    elif route == "harga_normal_tiktokshop":
        render_harga_normal_tiktokshop()
    elif route == "harga_normal_mwh":
        render_harga_normal_mwh()
    elif route == "harga_normal_powermerchant":
        render_harga_normal_powemerchant()
    elif route == "harga_normal_bigseller":
        render_harga_normal_bigseller()
    elif route == "harga_normal_blibli":
        render_harga_normal_blibli()
    elif route == "harga_normal_akulaku":
        render_harga_normal_akulaku()
    elif route == "harga_coret_shopee":
        render_harga_coret_shopee()
    elif route == "harga_coret_tiktokshop":
        render_harga_coret_tiktokshop()
    elif route == "harga_coret_powermerchant":
        render_harga_coret_powemerchant()
    elif route == "submit_campaign_shopee":
        render_submit_campaign_shopee()
    elif route == "submit_campaign_tiktokshop":
        render_submit_campaign_tiktokshop()
    elif route == "analisa_penjualan":
        render_analisa_penjualan()
    elif route == "analisa_produk_stok":
        render_analisa_produk()
    elif route == "progres_on":
        render_progres_on_v2()
    elif route == "analisa_margin":
        render_analisa_margin()
    elif route in ("affiliate", "affiliate_tiktokshop"):
        render_affiliate_tiktokshop()
    else:
        st.error("Menu tidak dikenal.")

    render_last_result_panel()


if __name__ == "__main__":
    main()
